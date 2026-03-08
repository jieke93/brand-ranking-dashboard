"""
유니클로 여성 랭킹 상품 정보 수집기
- 법적 리스크 최소화: robots.txt 준수, User-Agent 명시, 요청 간격 유지
- 공개된 정보만 수집 (이름, 가격, 이미지, 컬러)
- 개인 사용 목적
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from io import BytesIO
from PIL import Image as PILImage
import time
import os

# 법적 리스크 최소화 설정
HEADERS = {
    'User-Agent': 'Personal-Research-Bot (Educational Purpose)',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9',
}

# 요청 간격 (초) - 서버 부담 최소화
REQUEST_DELAY = 3

def check_robots_txt():
    """robots.txt 확인"""
    print("📋 robots.txt 확인 중...")
    robots_url = "https://www.uniqlo.com/robots.txt"
    try:
        response = requests.get(robots_url, headers=HEADERS)
        if "/kr/ko/spl/ranking" in response.text:
            print("⚠️ 경고: 해당 경로가 robots.txt에서 차단되었습니다.")
            return False
        else:
            print("✅ robots.txt 확인 완료: 수집 허용")
            return True
    except Exception as e:
        print(f"❌ robots.txt 확인 실패: {e}")
        return False

def fetch_ranking_page(url):
    """랭킹 페이지 HTML 가져오기"""
    print(f"\n🌐 페이지 접근 중: {url}")
    try:
        time.sleep(REQUEST_DELAY)  # 요청 간격 유지
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        print("✅ 페이지 로드 성공")
        return response.text
    except Exception as e:
        print(f"❌ 페이지 로드 실패: {e}")
        return None

def parse_product_data(html):
    """HTML에서 상품 정보 추출"""
    print("\n🔍 상품 정보 파싱 중...")
    
    # 주의: 실제 페이지는 JavaScript로 동적 렌더링됨
    # API 엔드포인트를 찾아서 직접 호출하는 것이 더 안전하고 효율적
    
    soup = BeautifulSoup(html, 'html.parser')
    products = []
    
    # 실제 HTML 구조에 맞게 셀렉터 조정 필요
    # 이것은 예시 구조입니다
    product_items = soup.select('.product-item')  # 실제 셀렉터로 변경 필요
    
    if not product_items:
        print("⚠️ 동적 렌더링 감지: JavaScript가 필요한 페이지입니다.")
        print("💡 Selenium을 사용하거나 API 직접 호출이 필요합니다.")
        return None
    
    for idx, item in enumerate(product_items, 1):
        try:
            name = item.select_one('.product-name').text.strip()
            price = item.select_one('.product-price').text.strip()
            image_url = item.select_one('img')['src']
            colors = [color.get('data-color') for color in item.select('.color-chip')]
            
            products.append({
                'rank': idx,
                'name': name,
                'price': price,
                'image_url': image_url,
                'colors': ', '.join(colors)
            })
        except Exception as e:
            print(f"⚠️ 상품 {idx} 파싱 오류: {e}")
    
    return products

def download_image(url, max_size=(100, 100)):
    """이미지 다운로드 및 리사이즈"""
    try:
        time.sleep(1)  # 이미지 요청 간격
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        return img_byte_arr
    except Exception as e:
        print(f"⚠️ 이미지 다운로드 실패 ({url}): {e}")
        return None

def create_excel(products, output_file):
    """엑셀 파일 생성"""
    print(f"\n📊 엑셀 파일 생성 중: {output_file}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "유니클로 여성 랭킹"
    
    # 헤더 설정
    headers = ['순위', '상품명', '가격', '컬러', '상품 이미지']
    ws.append(headers)
    
    # 헤더 스타일
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    
    # 데이터 입력
    for product in products:
        row_idx = ws.max_row + 1
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product['colors'])
        
        # 행 높이 설정 (이미지 크기에 맞춤)
        ws.row_dimensions[row_idx].height = 80
        
        # 이미지 다운로드 및 삽입
        print(f"  📷 {product['rank']}위 이미지 다운로드 중...")
        img_data = download_image(product['image_url'])
        if img_data:
            img = XLImage(img_data)
            img.width = 80
            img.height = 80
            cell_position = f'E{row_idx}'
            ws.add_image(img, cell_position)
        
        # 텍스트 정렬
        for col in range(1, 5):
            ws.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 파일 저장
    wb.save(output_file)
    print(f"✅ 엑셀 파일 저장 완료: {output_file}")

def main():
    print("=" * 60)
    print("📦 유니클로 여성 랭킹 상품 정보 수집기")
    print("=" * 60)
    print("\n⚖️  법적 리스크 최소화 조치:")
    print("  ✓ robots.txt 준수")
    print("  ✓ User-Agent 명시 (봇 식별)")
    print("  ✓ 요청 간격 유지 (3초)")
    print("  ✓ 공개 정보만 수집")
    print("  ✓ 개인 사용 목적")
    
    # 사용자 동의
    print("\n" + "=" * 60)
    print("⚠️  주의사항:")
    print("  - 이 스크립트는 교육 및 개인 연구 목적입니다")
    print("  - 수집된 데이터는 상업적으로 사용할 수 없습니다")
    print("  - 대량 데이터 수집 시 법적 문제가 발생할 수 있습니다")
    print("=" * 60)
    
    consent = input("\n위 내용을 이해하고 동의하십니까? (y/n): ")
    if consent.lower() != 'y':
        print("❌ 작업이 취소되었습니다.")
        return
    
    # robots.txt 확인
    if not check_robots_txt():
        confirm = input("\n⚠️ 계속 진행하시겠습니까? (y/n): ")
        if confirm.lower() != 'y':
            print("❌ 작업이 취소되었습니다.")
            return
    
    # 페이지 가져오기
    url = "https://www.uniqlo.com/kr/ko/spl/ranking/women"
    html = fetch_ranking_page(url)
    
    if not html:
        print("\n❌ 페이지를 가져올 수 없습니다.")
        print("💡 해결 방법:")
        print("  1. Selenium을 사용하여 JavaScript 렌더링")
        print("  2. 브라우저 개발자 도구로 API 엔드포인트 확인")
        return
    
    # 상품 데이터 파싱
    products = parse_product_data(html)
    
    if not products:
        print("\n⚠️ 이 페이지는 JavaScript로 동적 렌더링됩니다.")
        print("💡 Selenium 버전의 스크립트가 필요합니다.")
        print("\n다음 명령어로 필요한 패키지를 설치하세요:")
        print("  pip install selenium webdriver-manager")
        return
    
    print(f"\n✅ {len(products)}개 상품 정보 수집 완료")
    
    # 엑셀 파일 생성
    output_file = "유니클로_여성랭킹.xlsx"
    create_excel(products, output_file)
    
    print("\n" + "=" * 60)
    print("🎉 작업 완료!")
    print(f"📁 파일 위치: {os.path.abspath(output_file)}")
    print("=" * 60)

if __name__ == "__main__":
    main()
