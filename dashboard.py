def get_baseline_product_set(baseline_date='20240308'):
    """3월 8일(혹은 baseline_date) 기준의 (브랜드, 상품명) 집합 반환"""
    history = _load_all_history_raw()
    product_set = set()
    for full_key, dates_data in history.items():
        if not isinstance(dates_data, dict) or baseline_date not in dates_data:
            continue
        items = dates_data[baseline_date]
        parts = full_key.split('_', 1)
        brand = parts[0]
        for name in items.keys():
            product_set.add((brand, name))
    return product_set

#!/usr/bin/env python3
"""
3사 브랜드 랭킹 대시보드 (Streamlit)
───────────────────────────────────
유니클로 · 아르켓 · 탑텐 랭킹 데이터를 누적 관리하며
웹 브라우저에서 인터랙티브하게 비교·분석할 수 있는 대시보드

실행: streamlit run dashboard.py
"""

import os
import io
import glob
import json
import re
import base64
import hashlib
from datetime import datetime
from collections import Counter, defaultdict

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ─── 경로 설정 ───
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
IMG_CACHE_DIR = os.path.join(WORK_DIR, 'product_images')
IMG_HD_DIR = os.path.join(WORK_DIR, 'product_images_hd')
IMG_ARCHIVE_DIR = os.path.join(WORK_DIR, 'image_archive')  # 상품명 기반 영구 보관

# ─── 브랜드 색상 ───
BRAND_COLORS = {
    '유니클로': '#C41E3A',  # 유니클로 레드
    '아르켓':   '#1A1A1A',  # 블랙
    '탑텐':     '#0066CC',  # 블루
    '미쏘':     '#E91E63',  # 핑크
    '스파오':   '#F39C12',  # 오렌지
}

# ─── 브랜드별 설정 ───
BRAND_CONFIG = {
    '유니클로': {
        'file_pattern': '유니클로_전체랭킹_이미지포함_V5_*.xlsx',
        'compare_sheets': ['WOMEN_모두보기', 'MEN_모두보기'],
        'category_map': {
            'WOMEN_모두보기': '여성', 'MEN_모두보기': '남성',
            'KIDS_모두보기': '키즈', 'BABY_모두보기': '베이비',
            'WOMEN_상의': '여성_상의', 'WOMEN_팬츠': '여성_팬츠',
            'WOMEN_드레스 & 스커트': '여성_드레스&스커트',
            'WOMEN_아우터': '여성_아우터', 'WOMEN_이너웨어': '여성_이너웨어',
            'WOMEN_홈웨어': '여성_홈웨어', 'WOMEN_악세서리': '여성_악세서리',
            'MEN_상의': '남성_상의', 'MEN_팬츠': '남성_팬츠',
            'MEN_아우터': '남성_아우터', 'MEN_이너웨어': '남성_이너웨어',
            'MEN_홈웨어': '남성_홈웨어', 'MEN_악세서리': '남성_악세서리',
            'KIDS_상의': '키즈_상의', 'KIDS_팬츠': '키즈_팬츠',
            'KIDS_아우터': '키즈_아우터', 'KIDS_이너웨어': '키즈_이너웨어',
            'BABY_모두보기': '베이비',
        },
    },
    '아르켓': {
        'file_pattern': '아르켓_인기상품_판매순_*.xlsx',
        'compare_sheets': ['WOMEN', 'MEN'],
        'category_map': {'WOMEN': '여성', 'MEN': '남성'},
    },
    '탑텐': {
        'file_pattern': '탑텐_주간베스트_이미지포함_V3_*.xlsx',
        'compare_sheets': ['여성', '남성'],
        'category_map': {
            '전체': '전체', '여성': '여성', '남성': '남성',
            '키즈': '키즈', '베이비': '베이비',
        },
    },
    '미쏘': {
        'file_pattern': '미쏘_이번주베스트_이미지포함_*.xlsx',
        'compare_sheets': ['여성'],
        'category_map': {'여성': '여성'},
    },
}

# 브랜드 목록 (자동 – BRAND_CONFIG 기반)
BRAND_LIST = list(BRAND_CONFIG.keys())
ALL_BRANDS = BRAND_LIST + ['스파오']  # 스파오 포함 전체 브랜드

# ─── 아이템타입 분류 규칙 ───
UNIQLO_ITEM_RULES = [
    (r'(?i)(ultra\s*light\s*down|울트라라이트다운|패딩|다운)', '패딩/다운'),
    (r'(?i)(coat|코트|트렌치)', '코트'),
    (r'(?i)(jacket|자켓|재킷|블루종|blouson|블레이저|blazer|점퍼|jumper|파카|parka)', '자켓/아우터'),
    (r'(?i)(cardigan|가디건)', '카디건'),
    (r'(?i)(vest|조끼|베스트)', '베스트'),
    (r'(?i)(fleece|후리스|플리스)', '플리스'),
    (r'(?i)(hoodie|후드|후디)', '후드'),
    (r'(?i)(sweat|맨투맨|스웨트)', '맨투맨/스웨트'),
    (r'(?i)(니트|knit|sweater|스웨터|터틀넥)', '니트/스웨터'),
    (r'(?i)(flannel|플란넬|셔츠|shirt|블라우스)', '셔츠'),
    (r'(?i)(polo|폴로)', '폴로'),
    (r'(?i)(airism|에어리즘)', '에어리즘'),
    (r'(?i)(히트텍|heattech)', '히트텍'),
    (r'(?i)(t-shirt|tee|티셔츠|반팔|크루넥T)', '티셔츠'),
    (r'(?i)(jeans|진|데님)', '진/데님'),
    (r'(?i)(chino|치노|팬츠|pants|슬랙스|이지팬츠|카고|바지|트라우저)', '팬츠'),
    (r'(?i)(jogger|조거|스웨트팬츠|트레이닝)', '조거'),
    (r'(?i)(shorts|숏|반바지)', '쇼츠'),
    (r'(?i)(skirt|스커트|치마)', '스커트'),
    (r'(?i)(legging|레깅스)', '레깅스'),
    (r'(?i)(dress|원피스|드레스)', '원피스'),
    (r'(?i)(inner|이너|속옷|브라탑|런닝|팬티)', '이너웨어'),
    (r'(?i)(sock|양말)', '양말'),
    (r'(?i)(pajama|파자마|라운지|loungewear)', '라운지웨어'),
    (r'(?i)(bag|가방|백|토트)', '가방'),
    (r'(?i)(cap|모자|hat|비니|버킷)', '모자'),
]

ARKET_ITEM_RULES = [
    (r'(?i)(coat|코트)', '코트'),
    (r'(?i)(puffer|padded|패딩)', '패딩/다운'),
    (r'(?i)(jacket|자켓|blazer|블레이저)', '자켓/아우터'),
    (r'(?i)(cardigan|카디건)', '카디건'),
    (r'(?i)(vest|gilet)', '베스트'),
    (r'(?i)(hoodie|후디|hooded)', '후드'),
    (r'(?i)(sweatshirt|스웨트)', '맨투맨/스웨트'),
    (r'(?i)(knit|니트|sweater|스웨터|jumper)', '니트/스웨터'),
    (r'(?i)(shirt|셔츠|blouse|블라우스)', '셔츠'),
    (r'(?i)(polo)', '폴로'),
    (r'(?i)(t-shirt|tee|티셔츠)', '티셔츠'),
    (r'(?i)(top|탑|tank)', '탑'),
    (r'(?i)(jeans|jean|denim)', '진/데님'),
    (r'(?i)(trouser|팬츠|pants|chino)', '팬츠'),
    (r'(?i)(shorts)', '쇼츠'),
    (r'(?i)(skirt|스커트)', '스커트'),
    (r'(?i)(legging)', '레깅스'),
    (r'(?i)(dress|드레스)', '원피스'),
    (r'(?i)(jumpsuit|romper)', '점프수트'),
    (r'(?i)(bag|백|tote|backpack|crossbody)', '가방'),
    (r'(?i)(scarf|머플러|muffler)', '스카프/머플러'),
    (r'(?i)(hat|cap|beanie)', '모자'),
    (r'(?i)(sneaker|스니커즈)', '스니커즈'),
    (r'(?i)(boot|부츠)', '부츠'),
    (r'(?i)(sandal|샌들)', '샌들'),
    (r'(?i)(loafer|로퍼)', '로퍼'),
    (r'(?i)(shoe|슈즈|flat|trainer)', '슈즈'),
    (r'(?i)(jewel|ring|necklace|earring|bracelet)', '주얼리'),
    (r'(?i)(wallet|card\s*holder)', '지갑'),
    (r'(?i)(belt|벨트)', '벨트'),
    (r'(?i)(sock|양말)', '양말'),
]

TOPTEN_ITEM_RULES = [
    (r'(?i)(패딩|다운|덕다운|구스|웰론)', '패딩/다운'),
    (r'(?i)(코트|트렌치)', '코트'),
    (r'(?i)(자켓|재킷|점퍼|블레이저|집업|바람막이|야상|아노락)', '자켓/아우터'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(조끼|베스트|vest)', '베스트'),
    (r'(?i)(후리스|플리스|fleece|뽀글이)', '플리스'),
    (r'(?i)(후드|후디|hoodie)', '후드'),
    (r'(?i)(맨투맨|스웨트|크루넥)', '맨투맨/스웨트'),
    (r'(?i)(니트|스웨터|터틀넥|목폴라|폴라)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스|남방)', '셔츠'),
    (r'(?i)(긴팔|롱슬리브)', '긴팔티'),
    (r'(?i)(반팔|반소매|티셔츠|t-shirt|tee)', '티셔츠'),
    (r'(?i)(청바지|데님)', '진/데님'),
    (r'(?i)(슬랙스|치노|팬츠|바지|트라우저|면바지|카고)', '팬츠'),
    (r'(?i)(조거|트레이닝|스웨트팬츠|이지팬츠)', '조거'),
    (r'(?i)(반바지|숏팬츠|쇼츠|shorts)', '쇼츠'),
    (r'(?i)(스커트|치마)', '스커트'),
    (r'(?i)(레깅스)', '레깅스'),
    (r'(?i)(원피스|드레스)', '원피스'),
    (r'(?i)(세트|set|셋업)', '세트'),
    (r'(?i)(내복|내의|발열|히트텍|웜)', '내의/발열'),
    (r'(?i)(속옷|팬티|브라|런닝|캐미)', '이너웨어'),
    (r'(?i)(양말|삭스|sock)', '양말'),
    (r'(?i)(모자|캡|비니|버킷)', '모자'),
    (r'(?i)(가방|백팩|토트|크로스|에코백|숄더)', '가방'),
    (r'(?i)(머플러|스카프|목도리)', '스카프/머플러'),
    (r'(?i)(장갑)', '장갑'),
    (r'(?i)(운동화|스니커즈|신발|슈즈)', '스니커즈'),
    (r'(?i)(슬리퍼|샌들)', '샌들'),
    (r'(?i)(파자마|라운지)', '라운지웨어'),
]

MIXXO_ITEM_RULES = [
    (r'(?i)(코트|하프코트|롱코트|트렌치)', '코트'),
    (r'(?i)(패딩|다운|무스탕|퍼)', '패딩/다운'),
    (r'(?i)(자켓|재킷|블루종|블레이저|워크자켓|트위드)', '자켓/아우터'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(베스트|조끼)', '베스트'),
    (r'(?i)(후드|후디)', '후드'),
    (r'(?i)(맨투맨|스웨트|스웻)', '맨투맨/스웨트'),
    (r'(?i)(니트|스웨터|터틀넥|폴라|풀오버)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스|남방)', '셔츠'),
    (r'(?i)(티셔츠|반팔|tee)', '티셔츠'),
    (r'(?i)(진|데님|청바지)', '진/데님'),
    (r'(?i)(팬츠|바지|슬랙스|트라우저)', '팬츠'),
    (r'(?i)(스커트|치마|플리츠)', '스커트'),
    (r'(?i)(원피스|드레스)', '원피스'),
    (r'(?i)(가방|백|토트)', '가방'),
]

SPAO_ITEM_RULES = [
    (r'(?i)(코트|발마칸|더플)', '코트'),
    (r'(?i)(패딩|푸퍼|다운)', '패딩/다운'),
    (r'(?i)(자켓|재킷|점퍼|블루종|ma-1|항공|무스탕|블레이저)', '자켓/아우터'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(베스트|조끼|패쪼)', '베스트'),
    (r'(?i)(플리스|퍼플리스|fleece)', '플리스'),
    (r'(?i)(후드|hoodie|집업)', '후드'),
    (r'(?i)(맨투맨|스웨트셔츠|스웨트|크루넥)', '맨투맨/스웨트'),
    (r'(?i)(니트|스웨터|터틀넥|폴라)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스)', '셔츠'),
    (r'(?i)(긴팔|롱슬리브)', '긴팔티'),
    (r'(?i)(반팔|티셔츠|tee|t-shirt|캐미솔)', '티셔츠'),
    (r'(?i)(진|데님|청바지)', '진/데님'),
    (r'(?i)(슬랙스|팬츠|바지|코듀로이|카고|코튼)', '팬츠'),
    (r'(?i)(스웨트팬츠|조거|트레이닝|이지팬츠)', '조거'),
    (r'(?i)(쇼츠|반바지|숏팬츠)', '쇼츠'),
    (r'(?i)(스커트|치마)', '스커트'),
    (r'(?i)(레깅스|타이즈)', '레깅스'),
    (r'(?i)(원피스|드레스)', '원피스'),
    (r'(?i)(내복|내의|발열|웜텍|WARMTECH)', '내의/발열'),
    (r'(?i)(양말|삭스)', '양말'),
]

BRAND_ITEM_RULES = {
    '유니클로': UNIQLO_ITEM_RULES,
    '아르켓':   ARKET_ITEM_RULES,
    '탑텐':     TOPTEN_ITEM_RULES,
    '미쏘':     MIXXO_ITEM_RULES,
    '스파오':   SPAO_ITEM_RULES,
}


# ══════════════════════════════════════════════════════
#  유틸리티
# ══════════════════════════════════════════════════════

def classify_item_type(name, brand):
    if not name:
        return '미분류'
    rules = BRAND_ITEM_RULES.get(brand, UNIQLO_ITEM_RULES)
    for pattern, item_type in rules:
        if re.search(pattern, name):
            return item_type
    return '미분류'


def parse_price(price_str):
    if not price_str:
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(price_str)))
    except:
        return 0


def format_date(d):
    """'20260223' → '2026.02.23'"""
    if len(d) == 8:
        return f"{d[:4]}.{d[4:6]}.{d[6:8]}"
    return d


def safe_filename(name):
    """파일명에 사용할 수 없는 문자 제거"""
    return re.sub(r'[\\/:*?"<>|]', '_', str(name))


# ══════════════════════════════════════════════════════
#  이미지 추출 / 캐싱 시스템
# ══════════════════════════════════════════════════════

@st.cache_data(ttl=300)
def extract_all_product_images():
    """이미지 로드 우선순위: product_images_hd/ → product_images/ → 엑셀 추출
    반환: {(브랜드, 시트명, 순위): base64_string}
    부수효과: image_archive/ 에 상품명 기반 영구 보관 (이탈 상품용)
    """
    try:
        for d in [IMG_CACHE_DIR, IMG_HD_DIR, IMG_ARCHIVE_DIR]:
            os.makedirs(d, exist_ok=True)
    except (OSError, PermissionError):
        pass  # Cloud 환경에서 디렉토리 생성 실패 시 무시


    # 3월 8일 기준 상품 집합
    baseline_set = get_baseline_product_set('20240308')

    image_map = {}  # (brand, sheet, rank) → base64
    name_map = {}   # (brand, name) → base64  (archive용)

    def _parse_cache_filename(fname):
        """파일명에서 (brand, sheet, rank, product_name) 추출"""
        parts = fname.split('_', 3)
        if len(parts) < 4:
            return None
        try:
            brand = parts[1]
            remainder = parts[2] + '_' + parts[3].rsplit('.', 1)[0]
            rem_parts = remainder.split('_')
            sheet_parts = []
            rank_val = None
            rest_start = 0
            for i, p in enumerate(rem_parts):
                if p.isdigit() and i > 0:
                    rank_val = int(p)
                    rest_start = i + 1
                    break
                sheet_parts.append(p)
            if rank_val is None:
                return None
            sheet = '_'.join(sheet_parts)
            prod_name = '_'.join(rem_parts[rest_start:]) if rest_start < len(rem_parts) else ''
            return (brand, sheet, rank_val, prod_name)
        except Exception:
            return None


    def _load_from_dir(directory, upscale=False):
        """디렉토리에서 이미지 파일 로드 (3/8 기준 상품은 archive 우선)"""
        files = glob.glob(os.path.join(directory, '*.jpg'))
        loaded = 0
        for fpath in files:
            fname = os.path.basename(fpath)
            parsed = _parse_cache_filename(fname)
            if not parsed:
                continue
            brand, sheet, rank_val, prod_name = parsed
            key = (brand, sheet, rank_val)
            # 3/8 기준 상품이면 archive 우선 적용
            if (brand, prod_name) in baseline_set:
                # archive에 있으면 archive 이미지 사용
                safe_name = safe_filename(f"{brand}_{prod_name[:30]}")
                archive_path = os.path.join(IMG_ARCHIVE_DIR, f"{safe_name}.jpg")
                if os.path.exists(archive_path):
                    with open(archive_path, 'rb') as f:
                        img_bytes = f.read()
                    b64 = base64.b64encode(img_bytes).decode('utf-8')
                    image_map[key] = b64
                    loaded += 1
                    continue  # archive 우선 사용, 아래 로직 skip

            if key in image_map:
                continue  # HD가 이미 있으면 건너뜀

            with open(fpath, 'rb') as f:
                img_bytes = f.read()

            if upscale and len(img_bytes) < 5000:
                # 80x107 저해상도 → PIL로 400x534 업스케일
                try:
                    from PIL import Image as PILImage
                    img = PILImage.open(io.BytesIO(img_bytes))
                    img = img.resize((400, 534), PILImage.Resampling.LANCZOS)
                    buf = io.BytesIO()
                    img.save(buf, format='JPEG', quality=92)
                    img_bytes = buf.getvalue()
                except Exception:
                    pass

            b64 = base64.b64encode(img_bytes).decode('utf-8')
            image_map[key] = b64
            loaded += 1

            # archive에도 저장 (상품명 기반 영구 보관)
            if prod_name:
                _archive_image(brand, prod_name, img_bytes)

        return loaded

    def _archive_image(brand, prod_name, img_bytes):
        """상품명 기반으로 이미지 영구 보관 (이탈 상품도 이미지 유지)"""
        safe_name = safe_filename(f"{brand}_{prod_name}")
        archive_path = os.path.join(IMG_ARCHIVE_DIR, f"{safe_name}.jpg")
        if not os.path.exists(archive_path):
            try:
                with open(archive_path, 'wb') as f:
                    f.write(img_bytes)
            except Exception:
                pass

    # 1단계: product_images_hd/ (고해상도 200x267, 크롤러가 생성)
    hd_loaded = _load_from_dir(IMG_HD_DIR)

    # 2단계: product_images/ (엑셀용 80x107 → 업스케일)
    sd_loaded = _load_from_dir(IMG_CACHE_DIR, upscale=True)

    if image_map:
        return image_map

    # 3단계: 캐시 없으면 엑셀에서 추출 (최초 1회) + 업스케일 저장
    try:
        from openpyxl import load_workbook
    except ImportError:
        return image_map

    for brand_name, config in BRAND_CONFIG.items():
        pattern = os.path.join(WORK_DIR, config['file_pattern'])
        files = sorted(glob.glob(pattern), reverse=True)
        if not files:
            continue

        filepath = files[0]
        file_hash = hashlib.md5(f"{os.path.basename(filepath)}_{os.path.getmtime(filepath)}".encode()).hexdigest()[:8]

        wb = load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Sheet':
                continue
            ws = wb[sheet_name]

            img_by_row = {}
            for img in ws._images:
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    row_0based = anchor._from.row
                    try:
                        img_data = img._data()
                        if img_data and len(img_data) > 100:
                            img_by_row[row_0based] = img_data
                    except Exception:
                        pass

            for r in range(2, ws.max_row + 1):
                rank = ws.cell(r, 1).value
                if rank is None:
                    continue
                if brand_name == '탑텐':
                    name = ws.cell(r, 4).value or ''
                else:
                    name = ws.cell(r, 3).value or ''

                row_0based = r - 1
                if row_0based in img_by_row:
                    raw_bytes = img_by_row[row_0based]

                    # PIL로 업스케일 (80x107 → 400x534)
                    try:
                        from PIL import Image as PILImage
                        img_obj = PILImage.open(io.BytesIO(raw_bytes))
                        if img_obj.mode in ('RGBA', 'P'):
                            img_obj = img_obj.convert('RGB')
                        img_obj = img_obj.resize((400, 534), PILImage.Resampling.LANCZOS)
                        buf = io.BytesIO()
                        img_obj.save(buf, format='JPEG', quality=92)
                        upscaled_bytes = buf.getvalue()
                    except Exception:
                        upscaled_bytes = raw_bytes

                    b64 = base64.b64encode(upscaled_bytes).decode('utf-8')
                    image_map[(brand_name, sheet_name, int(rank))] = b64

                    # SD 캐시에 저장
                    safe_name = safe_filename(f"{brand_name}_{sheet_name}_{rank}_{name[:20]}")
                    img_path = os.path.join(IMG_CACHE_DIR, f"{file_hash}_{safe_name}.jpg")
                    if not os.path.exists(img_path):
                        try:
                            with open(img_path, 'wb') as f:
                                f.write(raw_bytes)
                        except Exception:
                            pass

                    # archive에 영구 보관
                    _archive_image(brand_name, name[:30], upscaled_bytes)

        wb.close()

    return image_map


def get_archived_image_b64(brand, product_name):
    """이탈 상품 등 archive에서 상품명 기반으로 이미지 조회"""
    safe_name = safe_filename(f"{brand}_{product_name[:30]}")
    archive_path = os.path.join(IMG_ARCHIVE_DIR, f"{safe_name}.jpg")
    if os.path.exists(archive_path):
        with open(archive_path, 'rb') as f:
            return base64.b64encode(f.read()).decode('utf-8')
    # Cloud fallback: product_thumbnails.json
    thumbs = _load_thumbnail_json()
    if thumbs:
        thumb_key = _make_thumb_key(brand, product_name)
        if thumb_key in thumbs:
            return thumbs[thumb_key]
    return None


@st.cache_data(ttl=600)
def _load_thumbnail_json():
    """product_thumbnails.json 로드 (Cloud 환경용 이미지 fallback)"""
    fp = os.path.join(WORK_DIR, 'product_thumbnails.json')
    if not os.path.exists(fp):
        return {}
    try:
        with open(fp, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _make_thumb_key(brand, product_name):
    """상품명을 HD 파일명과 동일한 safe key로 변환"""
    name_short = str(product_name)[:20]
    safe = re.sub(r'[^\w\s-]', '', name_short).strip().replace(' ', '_')
    return f"{brand}_{safe}"


def augment_image_map_from_thumbnails(image_map, df):
    """product_thumbnails.json에서 이미지 보충 (Cloud 환경 fallback)
    
    로컬에 HD/archive 이미지가 없는 경우 썸네일 JSON에서 매칭
    """
    thumbs = _load_thumbnail_json()
    if not thumbs or df is None or df.empty:
        return image_map

    augmented = dict(image_map)
    filled = 0
    for _, row in df.iterrows():
        brand = row.get('brand', '')
        sheet = row.get('sheet', '')
        rank = row.get('rank', 0)
        name = row.get('name', '')
        if not brand or not name or not rank:
            continue

        key = (brand, sheet, int(rank))
        if key in augmented:
            continue

        thumb_key = _make_thumb_key(brand, name)
        if thumb_key in thumbs:
            augmented[key] = thumbs[thumb_key]
            filled += 1

    return augmented


@st.cache_data(ttl=600)
def load_spao_image_map():
    """SPAO 이미지 맵 반환: {(brand, sheet, rank): url_or_base64}
    
    1순위: image_archive/에서 base64로 로드
    2순위: product_images_hd/에서 base64로 로드
    3순위: image_url을 'url:https://...' 형태로 저장 (HTML에서 직접 사용)
    """
    spao_file = os.path.join(WORK_DIR, 'spao_history.json')
    if not os.path.exists(spao_file):
        return {}

    try:
        with open(spao_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
    except (json.JSONDecodeError, Exception):
        return {}

    # 3/8 기준 상품 집합
    baseline_set = get_baseline_product_set('20240308')

    spao_images = {}
    for cat_key, dates_data in history.items():
        if not dates_data:
            continue
        latest_date = max(dates_data.keys())
        items = dates_data[latest_date]
        parts = cat_key.split('_', 1)
        sheet = parts[1] if len(parts) > 1 else ''

        for name, info in items.items():
            if not isinstance(info, dict):
                continue
            rank = info.get('rank', 0)
            if not rank:
                continue

            key = ('스파오', sheet, int(rank))
            if key in spao_images:
                continue

            # 3/8 기준 상품이면 archive 우선 적용
            if ('스파오', name) in baseline_set:
                safe_name = safe_filename(f"스파오_{name[:30]}")
                archive_path = os.path.join(IMG_ARCHIVE_DIR, f"{safe_name}.jpg")
                if os.path.exists(archive_path):
                    try:
                        with open(archive_path, 'rb') as f:
                            spao_images[key] = base64.b64encode(f.read()).decode('utf-8')
                        continue
                    except Exception:
                        pass

            # 2순위: image_url을 URL 마커와 함께 저장
            image_url = info.get('image_url', '')
            if image_url:
                spao_images[key] = f'url:{image_url}'

    return spao_images


def augment_image_map_from_archive(image_map, df):
    """archive 이미지를 상품명 매칭으로 image_map에 보충 (HD 이미지 없는 브랜드/상품용)
    
    product_images_hd 에 이미지가 없는 유니클로·아르켓 등도
    image_archive 에 상품명 기반 이미지가 있으면 매칭하여 채워넣는다.
    """
    if df is None or df.empty:
        return image_map

    # 3/8 기준 상품 집합
    baseline_set = get_baseline_product_set('20240308')

    # archive 파일 → {(brand, safe_name): base64} 로드
    archive_files = glob.glob(os.path.join(IMG_ARCHIVE_DIR, '*.jpg'))
    if not archive_files:
        return image_map

    archive_cache = {}  # safe_filename_without_ext → base64
    for fpath in archive_files:
        fname = os.path.basename(fpath)
        name_key = fname.rsplit('.', 1)[0]  # 확장자 제거
        try:
            with open(fpath, 'rb') as f:
                archive_cache[name_key] = base64.b64encode(f.read()).decode('utf-8')
        except Exception:
            continue

    if not archive_cache:
        return image_map

    augmented = dict(image_map)
    filled = 0
    for _, row in df.iterrows():
        brand = row.get('brand', '')
        sheet = row.get('sheet', '')
        rank = row.get('rank', 0)
        name = row.get('name', '')
        if not brand or not name or not rank:
            continue

        key = (brand, sheet, int(rank))
        if key in augmented:
            continue  # 이미 HD 이미지 있음

        # 3/8 기준 상품이면 archive 우선 적용
        if (brand, name) in baseline_set:
            safe_key = safe_filename(f"{brand}_{name[:30]}")
            if safe_key in archive_cache:
                augmented[key] = archive_cache[safe_key]
                filled += 1
                continue

        # archive에서 상품명 매칭 시도 (기존 로직)
        safe_key = safe_filename(f"{brand}_{name[:30]}")
        if safe_key in archive_cache:
            augmented[key] = archive_cache[safe_key]
            filled += 1

    return augmented


def get_image_b64(image_map, brand, sheet, rank):
    """이미지 base64 조회 (rank 기반)"""
    try:
        return image_map.get((brand, sheet, int(rank)), None)
    except (ValueError, TypeError):
        return None


def render_image_table(df_display, image_map, brand_col=None, sheet_col=None, rank_col='순위',
                       name_col='상품명', height=500, key_prefix='tbl', brand_sheet_data=None):
    """이미지 미리보기 고정 패널이 있는 HTML 테이블 렌더링 (components.html iframe 사용)"""
    if df_display.empty:
        st.info("표시할 데이터가 없습니다.")
        return

    table_id = f"img_table_{key_prefix}"

    # 행별 이미지 base64 수집
    bsd_list = list(brand_sheet_data) if brand_sheet_data else []
    row_images = {}  # row_index → base64
    for i, (brand_val, sheet_val, rank_val) in enumerate(bsd_list):
        if brand_val and sheet_val and rank_val is not None:
            try:
                b64 = get_image_b64(image_map, brand_val, sheet_val, int(rank_val))
                if b64:
                    row_images[i] = b64
            except (ValueError, TypeError):
                pass

    # 이미지 데이터를 JS 객체로 별도 관리
    img_js_entries = ",".join(f'{k}:"{v}"' for k, v in row_images.items())
    img_js = f"var IMG={{{img_js_entries}}};"

    cols = list(df_display.columns)
    header = "".join(f"<th>{c}</th>" for c in cols)

    rows_html = []
    for row_num, (idx, row) in enumerate(df_display.iterrows()):
        cells = []
        has_img = row_num in row_images
        for c in cols:
            val = row[c] if pd.notna(row[c]) else ''
            if c == name_col and has_img:
                escaped_val = str(val).replace("'", "&#39;").replace('"', '&quot;')
                display_val = str(val).replace('<', '&lt;').replace('>', '&gt;')
                cells.append(
                    f'<td class="pn" data-row="{row_num}" '
                    f'onclick="selectRow({row_num},\'{escaped_val[:50]}\')">'
                    f'{display_val}</td>'
                )
            else:
                display_val = str(val).replace('<', '&lt;').replace('>', '&gt;')
                cells.append(f"<td>{display_val}</td>")
        row_class = ' class="has-img"' if has_img else ''
        rows_html.append(f'<tr data-row="{row_num}"{row_class}>{"".join(cells)}</tr>')

    full_html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size:13px; }}

.container {{ display:flex; height:{height - 10}px; }}

/* 왼쪽: 테이블 영역 */
.table-area {{ flex:1; overflow-y:auto; overflow-x:auto; min-width:0; }}
table {{ width:100%; border-collapse:collapse; }}
th {{ background:#1a1a2e; color:#fff; padding:8px 10px; text-align:center;
     position:sticky; top:0; z-index:10; font-weight:600; font-size:12px; }}
td {{ padding:5px 8px; border-bottom:1px solid #e9ecef; text-align:center; vertical-align:middle; }}
tr:hover {{ background:#f0f4ff; }}
tr.selected {{ background:#dbeafe !important; }}
tr.has-img {{ cursor:pointer; }}
.pn {{ text-align:left; color:#1a73e8; font-weight:500; cursor:pointer; }}
.pn:hover {{ text-decoration:underline; }}

/* 오른쪽: 이미지 미리보기 고정 패널 */
.preview-panel {{
    width:280px; min-width:280px; border-left:2px solid #e0e0e0;
    display:flex; flex-direction:column; align-items:center; justify-content:center;
    padding:15px; background:#fafbfc;
}}
.preview-empty {{
    text-align:center; color:#aaa;
}}
.preview-empty .icon {{ font-size:48px; margin-bottom:10px; }}
.preview-empty .msg {{ font-size:13px; line-height:1.6; }}
.preview-content {{
    display:none; text-align:center; width:100%;
}}
.preview-content img {{
    max-width:250px; width:100%; height:auto;
    border-radius:8px; box-shadow:0 4px 16px rgba(0,0,0,0.15);
    image-rendering:auto;
}}
.preview-content .pname {{
    margin-top:12px; font-weight:600; font-size:13px; color:#333;
    word-break:break-all; line-height:1.4;
}}
.preview-content .prank {{
    margin-top:4px; font-size:12px; color:#888;
}}
</style></head>
<body>
<div class="container">
  <div class="table-area">
    <table id="{table_id}">
    <thead><tr>{header}</tr></thead>
    <tbody>{''.join(rows_html)}</tbody>
    </table>
  </div>
  <div class="preview-panel">
    <div class="preview-empty" id="pe">
      <div class="icon">🖼️</div>
      <div class="msg">상품을 클릭하면<br>이미지가 표시됩니다</div>
    </div>
    <div class="preview-content" id="pc">
      <img id="pimg" src=""/>
      <div class="pname" id="pname"></div>
      <div class="prank" id="prank"></div>
    </div>
  </div>
</div>

<script>
{img_js}
function selectRow(r, name) {{
  if (!IMG[r]) return;
  var val = IMG[r];
  var src = val.startsWith('url:') ? val.substring(4) : 'data:image/jpeg;base64,' + val;
  document.getElementById('pimg').src = src;
  document.getElementById('pname').innerText = name;
  // 순위 정보 표시
  var row = document.querySelector('tr[data-row="'+r+'"]');
  if (row) {{
    var firstCell = row.querySelector('td');
    if (firstCell) document.getElementById('prank').innerText = '순위: ' + firstCell.innerText;
  }}
  document.getElementById('pe').style.display = 'none';
  document.getElementById('pc').style.display = 'block';
  // 선택 행 하이라이트
  document.querySelectorAll('tr.selected').forEach(function(el) {{ el.classList.remove('selected'); }});
  if (row) row.classList.add('selected');
}}
</script>
</body></html>"""

    components.html(full_html, height=height + 20, scrolling=True)


# ══════════════════════════════════════════════════════
#  데이터 로드 (캐싱)
# ══════════════════════════════════════════════════════

def _build_df_from_history():
    """xlsx 파일이 없을 때 JSON 히스토리에서 최신 날짜 데이터로 DataFrame 구성 (Cloud용 폴백)"""
    history = _load_all_history_raw()
    if not history:
        return pd.DataFrame()

    all_products = []
    for full_key, dates_data in history.items():
        if not dates_data:
            continue
        latest_date = max(dates_data.keys())
        items = dates_data[latest_date]

        # full_key 형태: "브랜드_카테고리" (예: 유니클로_WOMEN_모두보기)
        parts = full_key.split('_', 1)
        brand = parts[0]
        sheet = parts[1] if len(parts) > 1 else ''
        cat_map = BRAND_CONFIG.get(brand, {}).get('category_map', {})
        mapped_cat = cat_map.get(sheet, sheet)

        for name, info in items.items():
            if isinstance(info, dict):
                raw_price = info.get('price', 0)
                numeric_price = parse_price(raw_price) if isinstance(raw_price, str) else (raw_price if isinstance(raw_price, (int, float)) else 0)
                price_str_val = str(raw_price) if raw_price else ''
                all_products.append({
                    'brand': brand, 'category': mapped_cat,
                    'sheet': sheet, 'rank': info.get('rank', 0),
                    'name': name,
                    'item_type': info.get('item_type', classify_item_type(name, brand)),
                    'price': numeric_price,
                    'price_str': price_str_val,
                    'date': latest_date,
                })

    df = pd.DataFrame(all_products) if all_products else pd.DataFrame()
    return df


def _list_history_dates(history):
    """히스토리(dict)에서 사용 가능한 날짜 키 목록(내림차순)"""
    dates = set()
    for dates_data in (history or {}).values():
        if isinstance(dates_data, dict):
            dates.update(dates_data.keys())
    # YYYYMMDD 문자열을 기본으로 가정하고 문자열 정렬
    return sorted([d for d in dates if d], reverse=True)


def _build_df_from_history_for_date(date_key):
    """JSON 히스토리에서 특정 날짜(date_key)의 데이터로 DataFrame 구성"""
    history = _load_all_history_raw()
    if not history or not date_key:
        return pd.DataFrame()

    all_products = []
    for full_key, dates_data in history.items():
        if not isinstance(dates_data, dict) or date_key not in dates_data:
            continue
        items = dates_data.get(date_key) or {}
        if not isinstance(items, dict):
            continue

        parts = full_key.split('_', 1)
        brand = parts[0]
        sheet = parts[1] if len(parts) > 1 else ''
        cat_map = BRAND_CONFIG.get(brand, {}).get('category_map', {})
        mapped_cat = cat_map.get(sheet, sheet)

        for name, info in items.items():
            if not isinstance(info, dict):
                continue
            raw_price = info.get('price', 0)
            numeric_price = parse_price(raw_price) if isinstance(raw_price, str) else (raw_price if isinstance(raw_price, (int, float)) else 0)
            price_str_val = str(raw_price) if raw_price else ''
            all_products.append({
                'brand': brand,
                'category': mapped_cat,
                'sheet': sheet,
                'rank': info.get('rank', 0),
                'name': name,
                'item_type': info.get('item_type', classify_item_type(name, brand)),
                'price': numeric_price,
                'price_str': price_str_val,
                'date': date_key,
            })

    return pd.DataFrame(all_products) if all_products else pd.DataFrame()


def _load_all_history_raw():
    """모든 히스토리 JSON 통합 로드 (내부 구현, 캐시 없음)"""
    combined = {}

    # 1) 통합 히스토리
    fp = os.path.join(WORK_DIR, 'all_brands_history.json')
    if os.path.exists(fp):
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for k, v in data.items():
                    combined[k] = v
        except (json.JSONDecodeError, Exception):
            pass

    # 2) 개별 히스토리 (보완)
    individual = {
        'ranking_history.json': '유니클로',
        'arket_history.json': '아르켓',
        'topten_history.json': '탑텐',
        'spao_history.json': '스파오',
    }
    for filename, brand in individual.items():
        fp = os.path.join(WORK_DIR, filename)
        if not os.path.exists(fp):
            continue
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except (json.JSONDecodeError, Exception):
            continue
        for cat_key, dates_data in data.items():
            full_key = f"{brand}_{cat_key}"
            # 이미 더 세분화된 키(예: 유니클로_MEN_모두보기)가 있으면
            # 상위 키(예: 유니클로_MEN)는 중복이므로 스킵
            has_subkeys = any(k.startswith(full_key + '_') for k in combined)
            if has_subkeys:
                continue
            if full_key not in combined:
                combined[full_key] = dates_data
            else:
                for date_key, products in dates_data.items():
                    if date_key not in combined[full_key]:
                        combined[full_key][date_key] = products

    return combined


@st.cache_data(ttl=300)
def load_all_history():
    """모든 히스토리 JSON 통합 로드 (캐싱 래퍼)"""
    return _load_all_history_raw()


@st.cache_data(ttl=300)
def load_latest_excel_data():
    """최신 크롤링 엑셀 데이터 로드 → DataFrame (xlsx 없으면 JSON 히스토리에서 구성)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _build_df_from_history()

    # xlsx 파일이 하나도 없으면 JSON 히스토리에서 구성
    has_any_xlsx = False
    for config in BRAND_CONFIG.values():
        pattern = os.path.join(WORK_DIR, config['file_pattern'])
        if glob.glob(pattern):
            has_any_xlsx = True
            break
    if not has_any_xlsx:
        return _build_df_from_history()
    all_products = []

    for brand_name, config in BRAND_CONFIG.items():
        pattern = os.path.join(WORK_DIR, config['file_pattern'])
        files = sorted(glob.glob(pattern), reverse=True)
        if not files:
            continue

        filepath = files[0]
        date_match = re.search(r'(\d{8})', os.path.basename(filepath))
        file_date = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Sheet':
                continue
            ws = wb[sheet_name]
            mapped_cat = config['category_map'].get(sheet_name, sheet_name)

            if brand_name == '유니클로':
                for r in range(2, ws.max_row + 1):
                    rank = ws.cell(r, 1).value
                    if rank is None:
                        continue
                    name = ws.cell(r, 3).value or ''
                    item_type = ws.cell(r, 4).value or classify_item_type(name, brand_name)
                    price_str = ws.cell(r, 5).value or ''
                    all_products.append({
                        'brand': brand_name, 'category': mapped_cat,
                        'sheet': sheet_name, 'rank': rank,
                        'name': name, 'item_type': item_type,
                        'price': parse_price(price_str), 'price_str': price_str,
                        'date': file_date,
                    })

            elif brand_name == '아르켓':
                for r in range(2, ws.max_row + 1):
                    rank = ws.cell(r, 1).value
                    if rank is None:
                        continue
                    name = ws.cell(r, 3).value or ''
                    price_str = ws.cell(r, 5).value or ''
                    all_products.append({
                        'brand': brand_name, 'category': mapped_cat,
                        'sheet': sheet_name, 'rank': rank,
                        'name': name, 'item_type': classify_item_type(name, brand_name),
                        'price': parse_price(price_str), 'price_str': price_str,
                        'date': file_date,
                    })

            elif brand_name == '탑텐':
                for r in range(2, ws.max_row + 1):
                    rank = ws.cell(r, 1).value
                    if rank is None:
                        continue
                    name = ws.cell(r, 4).value or ''
                    price_str = ws.cell(r, 5).value or ''
                    sub_brand = ws.cell(r, 3).value or ''
                    all_products.append({
                        'brand': brand_name, 'category': mapped_cat,
                        'sheet': sheet_name, 'rank': rank,
                        'name': name, 'item_type': classify_item_type(name, brand_name),
                        'price': parse_price(price_str), 'price_str': price_str,
                        'sub_brand': sub_brand, 'date': file_date,
                    })

            elif brand_name == '미쏘':
                for r in range(2, ws.max_row + 1):
                    rank = ws.cell(r, 1).value
                    if rank is None:
                        continue
                    name = ws.cell(r, 3).value or ''
                    price_str = ws.cell(r, 4).value or ''
                    all_products.append({
                        'brand': brand_name, 'category': mapped_cat,
                        'sheet': sheet_name, 'rank': rank,
                        'name': name, 'item_type': classify_item_type(name, brand_name),
                        'price': parse_price(price_str), 'price_str': price_str,
                        'date': file_date,
                    })

        wb.close()

    df = pd.DataFrame(all_products) if all_products else pd.DataFrame()
    return df


def deduplicate_products(df):
    """동일 상품 중복 제거 (브랜드+상품명 기준, 세부카테고리 우선)"""
    if df.empty:
        return df

    # "모두보기"/"전체" 등 통합 시트 정의
    general_sheets = {'WOMEN_모두보기', 'MEN_모두보기', 'KIDS_모두보기', 'BABY_모두보기', '전체'}

    df = df.copy()
    # 세부 카테고리 우선 (0), 통합 카테고리 후순위 (1)
    df['_is_general'] = df['sheet'].isin(general_sheets).astype(int)

    # 세부 카테고리 우선 → 같은 수준이면 순위 낮은 번호(=높은 랭킹) 우선
    df = df.sort_values(['brand', 'name', '_is_general', 'rank'])

    # 브랜드+상품명 기준 중복 제거 (첫 번째 = 세부카테고리 + 최고순위)
    before_count = len(df)
    df = df.drop_duplicates(subset=['brand', 'name'], keep='first')
    after_count = len(df)

    df = df.drop(columns=['_is_general'])
    df = df.sort_values(['brand', 'sheet', 'rank']).reset_index(drop=True)

    if before_count != after_count:
        print(f"[중복제거] {before_count}개 → {after_count}개 ({before_count - after_count}개 중복 제거)")

    return df


def get_available_dates(history):
    """히스토리에서 모든 날짜 추출"""
    dates = set()
    for cat_data in history.values():
        dates.update(cat_data.keys())
    return sorted(dates)


def get_compare_data(df, brand, gender):
    """비교용 데이터 추출 (여성/남성) - 중복 제거 후 데이터 대응"""
    if brand == '유니클로':
        # 중복 제거 후 세부카테고리에 분산되어 있으므로 접두어로 매칭
        prefix_map = {'여성': 'WOMEN', '남성': 'MEN'}
        prefix = prefix_map.get(gender, '')
        return df[(df['brand'] == brand) & (df['sheet'].str.startswith(prefix))]
    elif brand == '아르켓':
        gender_map = {'여성': 'WOMEN', '남성': 'MEN'}
        return df[(df['brand'] == brand) & (df['sheet'] == gender_map.get(gender, ''))]
    elif brand == '탑텐':
        # 중복 제거 후 '전체'에만 남은 아이템도 포함
        gender_map = {'여성': '여성', '남성': '남성'}
        target = gender_map.get(gender, '')
        return df[(df['brand'] == brand) & (df['sheet'].isin([target, '전체']))]
    elif brand == '미쏘':
        # 미쏬는 여성 전용
        if gender == '여성':
            return df[df['brand'] == brand]
        return pd.DataFrame()
    elif brand == '스파오':
        gender_map = {'여성': '여성', '남성': '남성'}
        target = gender_map.get(gender, '')
        return df[(df['brand'] == brand) & (df['sheet'] == target)]
    return pd.DataFrame()


# ══════════════════════════════════════════════════════
#  페이지: 종합 대시보드
# ══════════════════════════════════════════════════════

def page_overview(df, history, image_map=None):
    st.header(f"📊 {len(ALL_BRANDS)}사 브랜드 종합 대시보드")

    if image_map is None:
        image_map = {}

    # 종합 대시보드에서는 중복 제거 적용
    if not df.empty:
        df = deduplicate_products(df)

    if df.empty:
        st.warning("크롤링 데이터가 없습니다. 먼저 크롤러를 실행해주세요.")
        return

    dates = get_available_dates(history)
    latest_date = dates[-1] if dates else '없음'
    st.caption(f"최신 데이터: {format_date(latest_date)} | 누적 수집: {len(dates)}회")

    # ── KPI 카드 ──
    cols = st.columns(len(ALL_BRANDS))
    for i, brand in enumerate(ALL_BRANDS):
        bdf = df[df['brand'] == brand]
        with cols[i]:
            color = BRAND_COLORS[brand]
            st.markdown(f"""
            <div style="background: {color}; color: white; padding: 20px; border-radius: 12px; text-align: center;">
                <h2 style="margin:0; color:white;">{brand}</h2>
                <h1 style="margin:5px 0; color:white;">{len(bdf)}개</h1>
                <p style="margin:0; color: rgba(255,255,255,0.8);">
                    카테고리 {bdf['sheet'].nunique()}개 |
                    평균 {bdf['price'].mean():,.0f}원
                </p>
            </div>
            """, unsafe_allow_html=True)

    st.divider()

    # ── 여성/남성 아이템타입 비교 ──
    for gender in ['여성', '남성']:
        st.subheader(f"👗 {gender} 카테고리 아이템타입 비중 비교" if gender == '여성' else f"👔 {gender} 카테고리 아이템타입 비중 비교")

        fig = make_subplots(
            rows=1, cols=len(ALL_BRANDS),
            subplot_titles=ALL_BRANDS,
            specs=[[{'type': 'pie'}] * len(ALL_BRANDS)]
        )

        for ci, brand in enumerate(ALL_BRANDS, 1):
            bdf = get_compare_data(df, brand, gender)
            if bdf.empty:
                continue
            type_counts = bdf['item_type'].value_counts()
            fig.add_trace(
                go.Pie(
                    labels=type_counts.index.tolist(),
                    values=type_counts.values.tolist(),
                    name=brand,
                    textinfo='label+percent',
                    textposition='inside',
                    hole=0.35,
                ),
                row=1, col=ci
            )

        fig.update_layout(height=420, showlegend=False,
                          margin=dict(t=50, b=20, l=20, r=20))
        st.plotly_chart(fig, use_container_width=True)

        # 표 비교
        compare_data = []
        for brand in ALL_BRANDS:
            bdf = get_compare_data(df, brand, gender)
            if bdf.empty:
                continue
            type_counts = bdf['item_type'].value_counts()
            total = len(bdf)
            for item_type, count in type_counts.items():
                compare_data.append({
                    '브랜드': brand,
                    '아이템타입': item_type,
                    '상품수': count,
                    '비중(%)': round(count / total * 100, 1),
                })

        if compare_data:
            cdf = pd.DataFrame(compare_data)
            fig_bar = px.bar(
                cdf, x='아이템타입', y='비중(%)', color='브랜드',
                barmode='group',
                color_discrete_map=BRAND_COLORS,
                title=f'{gender} 아이템타입별 비중 비교',
            )
            fig_bar.update_layout(height=400, xaxis_tickangle=-45)
            st.plotly_chart(fig_bar, use_container_width=True)

            # 피봇 테이블 (접이식)
            with st.expander(f"📋 {gender} 아이템타입 비중 상세 표 보기", expanded=False):
                pivot_count = cdf.pivot_table(
                    index='아이템타입', columns='브랜드',
                    values='상품수', fill_value=0, aggfunc='sum'
                )
                pivot_pct = cdf.pivot_table(
                    index='아이템타입', columns='브랜드',
                    values='비중(%)', fill_value=0.0, aggfunc='sum'
                )
                # 브랜드 순서 고정
                brand_order = [b for b in ALL_BRANDS if b in pivot_count.columns]
                pivot_count = pivot_count[brand_order]
                pivot_pct = pivot_pct[brand_order]

                # 정렬 기준 브랜드 선택
                sort_brand = st.selectbox(
                    "정렬 기준 브랜드",
                    brand_order,
                    key=f"sort_brand_{gender}",
                )

                # 선택된 브랜드 비중 기준 내림차순 정렬
                sort_series = pivot_pct[sort_brand].drop('합계', errors='ignore')
                sorted_index = sort_series.sort_values(ascending=False).index
                pivot_count = pivot_count.loc[sorted_index]
                pivot_pct = pivot_pct.loc[sorted_index]

                # 합친 표: 상품수(비중%)
                display_df = pivot_count.copy().astype(str)
                for b in brand_order:
                    display_df[b] = pivot_count[b].astype(int).astype(str) + '개 (' + pivot_pct[b].round(1).astype(str) + '%)'

                display_df.index.name = '아이템타입'

                # 합계 행 추가
                total_row = {}
                for b in brand_order:
                    tc = int(pivot_count[b].sum())
                    total_row[b] = f"{tc}개 (100.0%)"
                display_df.loc['합계'] = total_row

                st.dataframe(display_df, use_container_width=True, height=min(len(display_df) * 38 + 40, 600))

                # 아이템타입 선택 → 상품 리스트
                st.markdown("---")
                type_list = sorted_index.tolist()
                selected_type = st.selectbox(
                    "🔎 아이템타입 선택 → 상품 리스트 보기",
                    type_list,
                    key=f"type_detail_{gender}",
                )

                if selected_type:
                    for brand in brand_order:
                        bdf = get_compare_data(df, brand, gender)
                        items = bdf[bdf['item_type'] == selected_type].sort_values('rank')
                        if items.empty:
                            continue
                        st.markdown(f"**{brand}** — {selected_type} ({len(items)}개)")
                        item_table = items[['rank', 'name', 'price_str', 'sheet']].copy()
                        show_table = item_table[['rank', 'name', 'price_str']].copy()
                        show_table.columns = ['순위', '상품명', '가격']
                        bs_data = [(brand, s, r) for s, r in zip(item_table['sheet'], item_table['rank'])]
                        render_image_table(show_table, image_map, rank_col='순위', name_col='상품명',
                                           height=min(len(show_table)*38+60, 400), key_prefix=f'ov_{gender}_{brand}',
                                           brand_sheet_data=bs_data)


# ══════════════════════════════════════════════════════
#  페이지: 브랜드별 상세
# ══════════════════════════════════════════════════════

def page_brand_detail(df, history, image_map=None):
    st.header("🏷️ 브랜드별 상세 분석")

    if image_map is None:
        image_map = {}

    if df.empty:
        st.warning("데이터가 없습니다.")
        return

    brand = st.selectbox("브랜드 선택", ALL_BRANDS)
    bdf = df[df['brand'] == brand]

    if bdf.empty:
        st.info(f"{brand} 데이터가 없습니다.")
        return

    color = BRAND_COLORS[brand]

    # KPI
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("총 상품수", f"{len(bdf)}개")
    c2.metric("카테고리", f"{bdf['sheet'].nunique()}개")
    c3.metric("평균가격", f"{bdf['price'].mean():,.0f}원")
    c4.metric("아이템타입 종류", f"{bdf['item_type'].nunique()}개")

    # 카테고리 선택
    sheets = bdf['sheet'].unique().tolist()
    selected_sheet = st.selectbox("카테고리", ['전체'] + sheets)

    if selected_sheet != '전체':
        bdf = bdf[bdf['sheet'] == selected_sheet]

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        # 아이템타입 파이차트
        type_counts = bdf['item_type'].value_counts()
        fig_pie = px.pie(
            values=type_counts.values, names=type_counts.index,
            title='아이템타입 구성', hole=0.4,
        )
        fig_pie.update_traces(textposition='inside', textinfo='label+percent')
        fig_pie.update_layout(height=400, showlegend=False)
        st.plotly_chart(fig_pie, use_container_width=True)

    with col2:
        # 가격대 분포
        price_bins = [0, 10000, 20000, 30000, 50000, 70000, 100000, 150000, 200000, 300000, float('inf')]
        price_labels = ['~1만', '1~2만', '2~3만', '3~5만', '5~7만', '7~10만', '10~15만', '15~20만', '20~30만', '30만+']
        valid_prices = bdf[bdf['price'] > 0].copy()
        if not valid_prices.empty:
            valid_prices['가격대'] = pd.cut(valid_prices['price'], bins=price_bins, labels=price_labels, right=False)
            price_dist = valid_prices['가격대'].value_counts().sort_index()
            fig_bar = px.bar(
                x=price_dist.index.astype(str), y=price_dist.values,
                title='가격대 분포', labels={'x': '가격대', 'y': '상품수'},
                color_discrete_sequence=[color],
            )
            fig_bar.update_layout(height=400)
            st.plotly_chart(fig_bar, use_container_width=True)

    # 상품 테이블
    st.subheader("📋 상품 목록")
    display_df = bdf[['rank', 'name', 'item_type', 'price_str', 'price', 'category', 'sheet']].copy()
    display_df.columns = ['순위', '상품명', '아이템타입', '가격', '_price_num', '카테고리', '_sheet']

    # 필터링
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        type_options = sorted(display_df['아이템타입'].dropna().unique().tolist())
        type_filter = st.multiselect("아이템타입 필터", type_options, key='bd_type_filter')
    with fc2:
        cat_options = sorted(display_df['카테고리'].dropna().unique().tolist())
        cat_filter = st.multiselect("카테고리 필터", cat_options, key='bd_cat_filter')
    with fc3:
        name_search = st.text_input("상품명 검색", key='bd_name_search', placeholder="검색어 입력...")

    # 정렬 + 가격 범위
    sc1, sc2 = st.columns(2)
    with sc1:
        sort_options = ['순위 (오름차순)', '가격 (낮은순)', '가격 (높은순)', '아이템타입 → 순위', '카테고리 → 순위']
        sort_by = st.selectbox("정렬 기준", sort_options, key='bd_sort')
    with sc2:
        max_price = int(display_df['_price_num'].max()) if not display_df.empty and display_df['_price_num'].max() > 0 else 500000
        step = 10000 if max_price > 50000 else 5000
        price_range = st.slider("가격 범위 (원)", 0, max_price, (0, max_price), step=step, key='bd_price_range')

    if type_filter:
        display_df = display_df[display_df['아이템타입'].isin(type_filter)]
    if cat_filter:
        display_df = display_df[display_df['카테고리'].isin(cat_filter)]
    if name_search:
        display_df = display_df[display_df['상품명'].str.contains(name_search, case=False, na=False)]
    display_df = display_df[(display_df['_price_num'] >= price_range[0]) & (display_df['_price_num'] <= price_range[1])]

    # 정렬 적용
    if sort_by == '가격 (낮은순)':
        display_df = display_df.sort_values('_price_num')
    elif sort_by == '가격 (높은순)':
        display_df = display_df.sort_values('_price_num', ascending=False)
    elif sort_by == '아이템타입 → 순위':
        display_df = display_df.sort_values(['아이템타입', '순위'])
    elif sort_by == '카테고리 → 순위':
        display_df = display_df.sort_values(['카테고리', '순위'])
    else:  # 순위 (오름차순)
        display_df = display_df.sort_values('순위')

    st.caption(f"표시: {len(display_df)}개 상품")

    # 이미지 테이블 렌더링
    display_df['_brand'] = brand
    show_df = display_df[['순위', '상품명', '아이템타입', '가격', '카테고리']].copy()
    render_image_table(show_df, image_map, rank_col='순위', name_col='상품명',
                       height=400, key_prefix='bd',
                       brand_sheet_data=list(zip(display_df['_brand'], display_df['_sheet'], display_df['순위'])))


# ══════════════════════════════════════════════════════
#  페이지: 가격 비교
# ══════════════════════════════════════════════════════

def page_price_compare(df, image_map=None):
    st.header(f"💰 {len(ALL_BRANDS)}사 가격 비교 분석")

    if image_map is None:
        image_map = {}

    # 가격 비교에서는 중복 제거 적용
    if not df.empty:
        df = deduplicate_products(df)

    if df.empty:
        st.warning("데이터가 없습니다.")
        return

    gender = st.radio("카테고리", ['여성', '남성'], horizontal=True)

    rows = []
    for brand in ALL_BRANDS:
        bdf = get_compare_data(df, brand, gender)
        if bdf.empty:
            continue
        for _, p in bdf.iterrows():
            rows.append({
                '브랜드': brand, '상품명': p['name'],
                '아이템타입': p['item_type'], '가격': p['price'],
            })

    if not rows:
        st.info("비교할 데이터가 없습니다.")
        return

    cdf = pd.DataFrame(rows)

    # 아이템타입 필터
    all_types = sorted(cdf['아이템타입'].dropna().unique().tolist())
    price_type_filter = st.multiselect("아이템타입 필터", all_types, key='price_type_filter')
    if price_type_filter:
        cdf = cdf[cdf['아이템타입'].isin(price_type_filter)]
    st.caption(f"분석 대상: {len(cdf)}개 상품")

    # 평균가격 비교
    avg_df = cdf.groupby('브랜드')['가격'].agg(['mean', 'median', 'min', 'max']).reset_index()
    avg_df.columns = ['브랜드', '평균', '중앙값', '최저', '최고']

    cols = st.columns(len(ALL_BRANDS))
    for i, brand in enumerate(ALL_BRANDS):
        row = avg_df[avg_df['브랜드'] == brand]
        if row.empty:
            continue
        r = row.iloc[0]
        with cols[i]:
            st.markdown(f"""
            <div style="background: {BRAND_COLORS.get(brand, '#888')}; color: white; padding: 15px; border-radius: 10px; text-align: center;">
                <h3 style="margin:0; color:white;">{brand}</h3>
                <h2 style="margin:5px 0; color:white;">{r['평균']:,.0f}원</h2>
                <p style="margin:0; font-size:0.85em; color: rgba(255,255,255,0.8);">
                    중앙값 {r['중앙값']:,.0f}원<br>
                    {r['최저']:,.0f}원 ~ {r['최고']:,.0f}원
                </p>
            </div>
            """, unsafe_allow_html=True)

    st.divider()

    # 가격 분포 히스토그램
    fig_hist = px.histogram(
        cdf, x='가격', color='브랜드', nbins=30,
        barmode='overlay', opacity=0.6,
        color_discrete_map=BRAND_COLORS,
        title=f'{gender} 가격 분포 비교',
    )
    fig_hist.update_layout(height=400, xaxis_title='가격(원)', yaxis_title='상품 수')
    st.plotly_chart(fig_hist, use_container_width=True)

    # 아이템타입별 평균가격
    type_price = cdf.groupby(['브랜드', '아이템타입'])['가격'].mean().reset_index()
    type_price.columns = ['브랜드', '아이템타입', '평균가격']

    common_types = cdf['아이템타입'].value_counts()
    common_types = common_types[common_types >= 2].index.tolist()
    type_price_filtered = type_price[type_price['아이템타입'].isin(common_types)]

    if not type_price_filtered.empty:
        fig_type = px.bar(
            type_price_filtered, x='아이템타입', y='평균가격', color='브랜드',
            barmode='group', color_discrete_map=BRAND_COLORS,
            title='아이템타입별 평균가격 비교',
        )
        fig_type.update_layout(height=450, xaxis_tickangle=-45)
        st.plotly_chart(fig_type, use_container_width=True)

    # 가격대별 분포표
    st.subheader("가격대별 상품수")
    price_bins = [0, 10000, 20000, 30000, 50000, 70000, 100000, 150000, 200000, 300000, float('inf')]
    price_labels = ['~1만', '1~2만', '2~3만', '3~5만', '5~7만', '7~10만', '10~15만', '15~20만', '20~30만', '30만+']

    band_data = []
    for brand in ALL_BRANDS:
        bd = cdf[cdf['브랜드'] == brand]['가격']
        if bd.empty:
            continue
        cuts = pd.cut(bd, bins=price_bins, labels=price_labels, right=False)
        for label, count in cuts.value_counts().sort_index().items():
            if count > 0:
                band_data.append({'가격대': str(label), '브랜드': brand, '상품수': count})

    if band_data:
        band_df = pd.DataFrame(band_data)
        pivot = band_df.pivot_table(index='가격대', columns='브랜드', values='상품수', fill_value=0)
        st.dataframe(pivot, use_container_width=True)


# ══════════════════════════════════════════════════════
#  페이지: 핵심 아이템
# ══════════════════════════════════════════════════════

def page_top_items(df, image_map=None):
    st.header("🏆 핵심 아이템 TOP 10 비교")

    if image_map is None:
        image_map = {}

    # 날짜 선택: 과거 업데이트 시점별 핵심 아이템 조회
    history_raw = _load_all_history_raw()
    available_dates = _list_history_dates(history_raw)
    if available_dates:
        latest_date = available_dates[0]
        selected_date = st.selectbox("기준 날짜", available_dates, index=0, key='top_items_date')
        df = _build_df_from_history_for_date(selected_date)
    else:
        latest_date = None
        selected_date = None

    # 최신 날짜일 때만 직전 날짜와 비교해 변동 표시
    prev_date = None
    df_prev = pd.DataFrame()
    if available_dates and len(available_dates) >= 2 and selected_date == latest_date:
        prev_date = available_dates[1]
        df_prev = _build_df_from_history_for_date(prev_date)

    # 중복 제거는 _get_gender_overall_top20 이후 결과에서 name 기준으로 적용
    # (전체 df에 먼저 dedup하면 유니클로 "모두보기" 시트 항목이 세부카테고리로 대체되어 누락됨)
    df_original = df.copy() if not df.empty else df

    if df.empty:
        st.warning("데이터가 없습니다.")
        return

    gender = st.radio("카테고리", ['여성', '남성'], horizontal=True, key='top_gender')
    top_n = st.slider("표시 개수", 5, 30, 10)

    tabs = st.tabs(ALL_BRANDS + ['나란히 비교'])

    for ti, brand in enumerate(ALL_BRANDS):
        with tabs[ti]:
            # 전체 랭킹 시트에서 TOP N 추출 (원본 df 사용)
            bdf = _get_gender_overall_top20(df_original, brand, gender)
            if bdf.empty:
                # fallback: get_compare_data
                bdf = get_compare_data(df_original, brand, gender)
            if bdf.empty:
                st.info(f"{brand} 데이터 없음")
                continue

            top_raw = bdf.nsmallest(top_n, 'rank').drop_duplicates(subset='name').head(top_n).copy()
            top_raw = top_raw[['rank', 'name', 'item_type', 'price_str', 'sheet']].copy()
            # 핵심아이템 순위를 1~N으로 재부여
            top_raw = top_raw.reset_index(drop=True)
            top_raw['display_rank'] = range(1, len(top_raw) + 1)
            original_ranks = top_raw['rank'].values.copy()

            # 변동(상승/하강/신규) 계산: 최신 날짜에서만 직전 날짜와 비교
            change_labels = []
            if prev_date and not df_prev.empty:
                prev_bdf = _get_gender_overall_top20(df_prev, brand, gender)
                if prev_bdf.empty:
                    prev_bdf = get_compare_data(df_prev, brand, gender)
                prev_rank_by_name = {}
                if not prev_bdf.empty:
                    # 동명 상품이 여러 시트에 있으면 가장 좋은(rank 최소) 값 사용
                    for _, prow in prev_bdf[['name', 'rank']].dropna().iterrows():
                        pname = str(prow['name'])
                        prank = int(prow['rank']) if pd.notna(prow['rank']) else None
                        if prank is None:
                            continue
                        if pname not in prev_rank_by_name or prank < prev_rank_by_name[pname]:
                            prev_rank_by_name[pname] = prank

                for _, crow in top_raw.iterrows():
                    cname = str(crow['name'])
                    crank = int(crow['rank']) if pd.notna(crow['rank']) else 0
                    prank = prev_rank_by_name.get(cname)
                    if prank is None:
                        change_labels.append('신규')
                        continue
                    delta = prank - crank
                    if delta > 0:
                        change_labels.append(f'상승{delta}')
                    elif delta < 0:
                        change_labels.append(f'하강{abs(delta)}')
                    else:
                        change_labels.append('-')
            else:
                change_labels = ['-'] * len(top_raw)

            top_raw['change'] = change_labels

            top_df = top_raw[['display_rank', 'name', 'item_type', 'change', 'price_str']].copy()
            top_df.columns = ['순위', '상품명', '아이템타입', '변동', '가격']

            # 필터링
            ti_fc1, ti_fc2 = st.columns(2)
            with ti_fc1:
                ti_types = sorted(top_df['아이템타입'].dropna().unique().tolist())
                ti_type_f = st.multiselect("아이템타입 필터", ti_types, key=f'ti_type_{brand}')
            with ti_fc2:
                ti_name_f = st.text_input("상품명 검색", key=f'ti_name_{brand}', placeholder="검색어 입력...")
            if ti_type_f:
                mask = top_df['아이템타입'].isin(ti_type_f)
                top_df = top_df[mask]
                top_raw = top_raw[mask.values]
            if ti_name_f:
                mask = top_df['상품명'].str.contains(ti_name_f, case=False, na=False)
                top_df = top_df[mask]
                top_raw = top_raw[mask.values]

            # 이미지 테이블 (과거 날짜는 이미지 매칭 불일치 가능성이 있어 텍스트 테이블로 표시)
            show_images = (selected_date is None) or (latest_date is None) or (selected_date == latest_date)
            if show_images:
                bs_data = [(brand, s, r) for s, r in zip(top_raw['sheet'], top_raw['rank'])]
                render_image_table(
                    top_df,
                    image_map,
                    rank_col='순위',
                    name_col='상품명',
                    height=400,
                    key_prefix=f'ti_{brand}',
                    brand_sheet_data=bs_data,
                )
            else:
                st.dataframe(top_df, use_container_width=True)

            # 아이템타입 분포
            type_counts = bdf.nsmallest(top_n, 'rank')['item_type'].value_counts()
            fig = px.bar(
                x=type_counts.index, y=type_counts.values,
                color_discrete_sequence=[BRAND_COLORS[brand]],
                title=f'{brand} TOP{top_n} 아이템타입',
                labels={'x': '아이템타입', 'y': '상품수'},
            )
            fig.update_layout(height=300)
            st.plotly_chart(fig, use_container_width=True)

    with tabs[len(ALL_BRANDS)]:
        # 브랜드별 리스트를 나란히 표시
        brand_tops = {}
        cols = st.columns(len(ALL_BRANDS))
        for ci, brand in enumerate(ALL_BRANDS):
            with cols[ci]:
                color = BRAND_COLORS.get(brand, '#888')
                st.markdown(f"<div style='background:{color};color:white;padding:6px;border-radius:6px;text-align:center;font-weight:bold;'>{brand}</div>", unsafe_allow_html=True)
                bdf = _get_gender_overall_top20(df_original, brand, gender)
                if bdf.empty:
                    bdf = get_compare_data(df_original, brand, gender)
                if bdf.empty:
                    st.caption("데이터 없음")
                    continue
                top_raw = bdf.nsmallest(top_n, 'rank').drop_duplicates(subset='name').head(top_n).copy()
                top_raw = top_raw[['rank', 'name', 'item_type', 'price_str', 'sheet']].copy()
                top_raw = top_raw.reset_index(drop=True)
                top_raw['display_rank'] = range(1, len(top_raw) + 1)
                brand_tops[brand] = top_raw
                for _, row in top_raw.iterrows():
                    st.markdown(f"**{int(row['display_rank'])}**. {row['name'][:20]}  \n<span style='color:#888;font-size:0.85em;'>{row['item_type']} · {row['price_str']}</span>", unsafe_allow_html=True)

        # 아래에 이미지 갤러리
        st.divider()
        st.subheader("🖼️ 상품 이미지")
        img_cols = st.columns(len(ALL_BRANDS))
        for ci, brand in enumerate(ALL_BRANDS):
            with img_cols[ci]:
                st.markdown(f"**{brand}**")
                if brand not in brand_tops:
                    continue
                top_raw = brand_tops[brand]
                for _, row in top_raw.iterrows():
                    disp_rank = int(row['display_rank'])
                    orig_rank = int(row['rank'])
                    sheet_val = row['sheet']
                    key = (brand, sheet_val, orig_rank)
                    img_val = image_map.get(key, '')
                    if img_val:
                        if img_val.startswith('url:'):
                            src = img_val[4:]
                        else:
                            src = f'data:image/jpeg;base64,{img_val}'
                        st.markdown(
                            f"<div style='margin-bottom:8px;'>"
                            f"<img src='{src}' style='width:100%;max-width:120px;border-radius:4px;'>"
                            f"<br><span style='font-size:0.75em;'>{disp_rank}. {row['name'][:15]}</span>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                    else:
                        st.caption(f"{disp_rank}. {row['name'][:15]} (이미지 없음)")


# ══════════════════════════════════════════════════════
#  페이지: 랭킹 변동 추적
# ══════════════════════════════════════════════════════

def page_ranking_trend(history, image_map=None):
    st.header("📈 랭킹 변동 추적")

    if image_map is None:
        image_map = {}

    dates = get_available_dates(history)
    if len(dates) < 1:
        st.warning("히스토리 데이터가 없습니다. 크롤러를 실행하면 자동으로 누적됩니다.")
        return

    st.caption(f"수집 기간: {format_date(dates[0])} ~ {format_date(dates[-1])} ({len(dates)}회)")

    if len(dates) < 2:
        st.info("📌 2회 이상 수집되면 랭킹 변동을 추적할 수 있습니다. 크롤러를 매일 실행해보세요!")

        # 현재 데이터 테이블만 보여주기
        brand = st.selectbox("브랜드", ALL_BRANDS, key='trend_brand_1')
        cat_keys = [k for k in history.keys() if k.startswith(brand)]
        if cat_keys:
            cat = st.selectbox("카테고리", cat_keys, key='trend_cat_1')
            date = dates[-1]
            items = history[cat].get(date, {})
            rows = [{'상품명': n, '순위': v['rank'], '아이템타입': v['item_type'], '가격': v['price']}
                    for n, v in items.items()]
            if rows:
                rdf = pd.DataFrame(rows).sort_values('순위')
                # 필터링
                rf1, rf2 = st.columns(2)
                with rf1:
                    rt_types = sorted(rdf['아이템타입'].dropna().unique().tolist())
                    rt_type_f = st.multiselect("아이템타입 필터", rt_types, key='rt1_type')
                with rf2:
                    rt_name_f = st.text_input("상품명 검색", key='rt1_name', placeholder="검색어 입력...")
                if rt_type_f:
                    rdf = rdf[rdf['아이템타입'].isin(rt_type_f)]
                if rt_name_f:
                    rdf = rdf[rdf['상품명'].str.contains(rt_name_f, case=False, na=False)]
                st.caption(f"표시: {len(rdf)}개 상품")
                # 이미지 테이블
                sheet_key = cat.replace(f'{brand}_', '')
                bs_data = [(brand, sheet_key, r) for r in rdf['순위']]
                render_image_table(rdf, image_map, rank_col='순위', name_col='상품명',
                                   height=400, key_prefix='rt1', brand_sheet_data=bs_data)
        return

    # 랭킹 변동이 있을 때
    brand = st.selectbox("브랜드", ALL_BRANDS, key='trend_brand')

    cat_keys = sorted([k for k in history.keys() if k.startswith(brand)])
    if not cat_keys:
        st.info(f"{brand} 히스토리 없음")
        return

    cat = st.selectbox("카테고리", cat_keys, key='trend_cat',
                       format_func=lambda x: x.replace(f'{brand}_', ''))

    cat_data = history[cat]
    cat_dates = sorted(cat_data.keys())

    # 상품 선택
    all_products = set()
    for d in cat_dates:
        all_products.update(cat_data[d].keys())
    all_products = sorted(all_products)

    # 최신 날짜 기준 랭킹 TOP 10을 기본 선택값으로 설정
    latest_date = cat_dates[-1]
    latest_items = cat_data[latest_date]
    top10_products = sorted(latest_items.keys(), key=lambda x: latest_items[x].get('rank', 999))[:10]
    # 기본값이 전체 목록에 포함되는지 확인
    default_products = [p for p in top10_products if p in all_products]

    selected = st.multiselect("상품 선택 (복수 선택 가능)", all_products,
                               default=default_products)

    if not selected:
        st.info("추적할 상품을 선택하세요.")
        return

    # 라인 차트
    chart_data = []
    for prod in selected:
        for d in cat_dates:
            rank = cat_data[d].get(prod, {}).get('rank', None)
            chart_data.append({
                '날짜': format_date(d), '상품명': prod[:25], '순위': rank
            })

    chart_df = pd.DataFrame(chart_data)
    chart_df = chart_df.dropna(subset=['순위'])

    if not chart_df.empty:
        fig = px.line(
            chart_df, x='날짜', y='순위', color='상품명',
            markers=True, title='랭킹 추이',
        )
        fig.update_yaxes(autorange='reversed', title='순위 (1위가 위)')
        fig.update_layout(height=500)
        st.plotly_chart(fig, use_container_width=True)

    # ── 현재 순위 테이블 + 이미지 (항상 표시) ──
    st.subheader("📋 현재 순위 (이미지 포함)")
    sheet_key = cat.replace(f'{brand}_', '')
    latest_items_all = cat_data[cat_dates[-1]]
    current_rows = []
    for name, info in latest_items_all.items():
        current_rows.append({
            '순위': info['rank'], '상품명': name,
            '아이템타입': info.get('item_type', ''),
            '가격': info.get('price', ''),
        })
    if current_rows:
        current_df = pd.DataFrame(current_rows).sort_values('순위').reset_index(drop=True)
        # 필터링
        crf1, crf2 = st.columns(2)
        with crf1:
            cr_types = sorted(current_df['아이템타입'].dropna().unique().tolist())
            cr_type_f = st.multiselect("아이템타입 필터", cr_types, key='cr_type')
        with crf2:
            cr_name_f = st.text_input("상품명 검색", key='cr_name', placeholder="검색어 입력...")
        if cr_type_f:
            current_df = current_df[current_df['아이템타입'].isin(cr_type_f)]
        if cr_name_f:
            current_df = current_df[current_df['상품명'].str.contains(cr_name_f, case=False, na=False)]
        st.caption(f"표시: {len(current_df)}개 상품")
        bs_data_cr = [(brand, sheet_key, r) for r in current_df['순위']]
        render_image_table(current_df, image_map, rank_col='순위', name_col='상품명',
                           height=min(len(current_df) * 38 + 60, 600),
                           key_prefix='cr_tbl', brand_sheet_data=bs_data_cr)

    # 변동 테이블
    if len(cat_dates) >= 2:
        st.subheader("변동 상세")
        cur_date, prev_date = cat_dates[-1], cat_dates[-2]
        cur, prev = cat_data[cur_date], cat_data[prev_date]

        changes = []
        for name, info in cur.items():
            cr = info['rank']
            if name in prev:
                pr = prev[name]['rank']
                chg = pr - cr
                status = '🔺 상승' if chg > 0 else ('🔻 하락' if chg < 0 else '➖ 유지')
            else:
                chg = 0
                status = '🆕 신규'
            changes.append({
                '순위': cr, '상품명': name[:30], '아이템타입': info['item_type'],
                '가격': info['price'], '변동': chg, '상태': status,
            })

        dropped = [{'순위': '-', '상품명': n[:30], '아이템타입': prev[n]['item_type'],
                     '가격': prev[n]['price'], '변동': '-', '상태': '❌ 이탈'}
                    for n in prev if n not in cur]

        change_df = pd.DataFrame(changes + dropped)
        if not change_df.empty:
            change_df = change_df.sort_values('순위', key=lambda x: pd.to_numeric(x, errors='coerce'))

            # 필터링
            cf1, cf2, cf3 = st.columns(3)
            with cf1:
                status_options = sorted(change_df['상태'].unique().tolist())
                status_filter = st.multiselect("상태 필터", status_options, key='chg_status')
            with cf2:
                chg_types = sorted(change_df['아이템타입'].dropna().unique().tolist())
                chg_type_filter = st.multiselect("아이템타입 필터", chg_types, key='chg_type')
            with cf3:
                chg_name_search = st.text_input("상품명 검색", key='chg_name', placeholder="검색어 입력...")
            if status_filter:
                change_df = change_df[change_df['상태'].isin(status_filter)]
            if chg_type_filter:
                change_df = change_df[change_df['아이템타입'].isin(chg_type_filter)]
            if chg_name_search:
                change_df = change_df[change_df['상품명'].str.contains(chg_name_search, case=False, na=False)]
            st.caption(f"표시: {len(change_df)}개 항목")
            # 이미지 테이블 - 이탈 상품은 archive에서 이미지 조회
            sheet_key = cat.replace(f'{brand}_', '')
            # 이탈 상품 이미지를 image_map에 임시 추가
            augmented_map = dict(image_map)
            bs_data = []
            for _, row in change_df.iterrows():
                rank_val = row['순위']
                prod_name = row['상품명']
                if rank_val == '-' or pd.isna(pd.to_numeric(rank_val, errors='coerce')):
                    # 이탈 상품 → archive에서 조회
                    archived_b64 = get_archived_image_b64(brand, prod_name)
                    if archived_b64:
                        # 임시 키로 등록 (음수 사용)
                        temp_key = (brand, sheet_key, -hash(prod_name) % 100000)
                        augmented_map[temp_key] = archived_b64
                        bs_data.append((brand, sheet_key, temp_key[2]))
                    else:
                        bs_data.append((brand, sheet_key, rank_val))
                else:
                    bs_data.append((brand, sheet_key, rank_val))
            render_image_table(change_df, augmented_map, rank_col='순위', name_col='상품명',
                               height=500, key_prefix='chg', brand_sheet_data=bs_data)

            # 요약 (버튼으로 상품 리스트 연동)
            risen = [c for c in changes if '상승' in c['상태']]
            fallen = [c for c in changes if '하락' in c['상태']]
            new_items = [c for c in changes if '신규' in c['상태']]
            dropped_items = dropped

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                if st.button(f"🔺 상승 {len(risen)}개", key='btn_risen', use_container_width=True):
                    st.session_state['trend_detail'] = '상승'
            with c2:
                if st.button(f"🔻 하락 {len(fallen)}개", key='btn_fallen', use_container_width=True):
                    st.session_state['trend_detail'] = '하락'
            with c3:
                if st.button(f"🆕 신규 {len(new_items)}개", key='btn_new', use_container_width=True):
                    st.session_state['trend_detail'] = '신규'
            with c4:
                if st.button(f"❌ 이탈 {len(dropped_items)}개", key='btn_dropped', use_container_width=True):
                    st.session_state['trend_detail'] = '이탈'

            detail_key = st.session_state.get('trend_detail', None)
            if detail_key:
                label_map = {'상승': ('🔺 상승 상품', risen), '하락': ('🔻 하락 상품', fallen),
                             '신규': ('🆕 신규 상품', new_items), '이탈': ('❌ 이탈 상품', dropped_items)}
                label, items_list = label_map.get(detail_key, ('', []))
                if items_list:
                    st.markdown(f"#### {label} ({len(items_list)}개)")
                    detail_df = pd.DataFrame(items_list)
                    if detail_key != '이탈':
                        detail_df = detail_df.sort_values('변동', ascending=(detail_key == '하락'))

                    # 이미지 매핑 구성
                    detail_bs_data = []
                    detail_aug_map = dict(augmented_map)
                    for _, row in detail_df.iterrows():
                        rank_val = row['순위']
                        prod_name = row['상품명']
                        if rank_val == '-' or pd.isna(pd.to_numeric(rank_val, errors='coerce')):
                            archived_b64 = get_archived_image_b64(brand, prod_name)
                            if archived_b64:
                                temp_key = (brand, sheet_key, -hash(prod_name) % 100000)
                                detail_aug_map[temp_key] = archived_b64
                                detail_bs_data.append((brand, sheet_key, temp_key[2]))
                            else:
                                detail_bs_data.append((brand, sheet_key, rank_val))
                        else:
                            detail_bs_data.append((brand, sheet_key, rank_val))

                    render_image_table(detail_df, detail_aug_map, rank_col='순위', name_col='상품명',
                                       height=min(len(detail_df) * 38 + 40, 500),
                                       key_prefix=f'detail_{detail_key}', brand_sheet_data=detail_bs_data)
                else:
                    st.info(f"{label} 항목이 없습니다.")


# ══════════════════════════════════════════════════════
#  페이지: 상품 검색
# ══════════════════════════════════════════════════════

def page_search(df, image_map=None):
    st.header("🔍 상품 검색")

    if image_map is None:
        image_map = {}

    # SPAO 데이터는 이미 df에 통합됨

    if df.empty:
        st.warning("데이터가 없습니다.")
        return

    query = st.text_input("검색어 입력 (상품명, 아이템타입 등)", placeholder="예: 자켓, 패딩, shirt...")

    col1, col2, col3 = st.columns(3)
    with col1:
        brand_filter = st.multiselect("브랜드", ALL_BRANDS,
                                       default=ALL_BRANDS)
    with col2:
        types = df['item_type'].unique().tolist()
        type_filter = st.multiselect("아이템타입", sorted(types))
    with col3:
        price_range = st.slider("가격 범위 (원)", 0, 500000, (0, 500000), step=10000)

    # 카테고리 필터 + 정렬
    col4, col5 = st.columns(2)
    with col4:
        # 선택된 브랜드의 카테고리만 표시
        available_cats = sorted(df[df['brand'].isin(brand_filter)]['category'].unique().tolist()) if brand_filter else []
        cat_filter = st.multiselect("카테고리", available_cats, key='search_cat')
    with col5:
        sort_options = ['브랜드 → 순위', '순위 (오름차순)', '가격 (낮은순)', '가격 (높은순)', '카테고리 → 순위']
        sort_by = st.selectbox("정렬 기준", sort_options, key='search_sort')

    filtered = df.copy()
    if brand_filter:
        filtered = filtered[filtered['brand'].isin(brand_filter)]
    if type_filter:
        filtered = filtered[filtered['item_type'].isin(type_filter)]
    if cat_filter:
        filtered = filtered[filtered['category'].isin(cat_filter)]
    filtered = filtered[(filtered['price'] >= price_range[0]) & (filtered['price'] <= price_range[1])]

    if query:
        mask = (filtered['name'].str.contains(query, case=False, na=False) |
                filtered['item_type'].str.contains(query, case=False, na=False))
        filtered = filtered[mask]

    st.caption(f"검색 결과: {len(filtered)}개 상품")

    display = filtered[['brand', 'rank', 'name', 'item_type', 'price_str', 'category', 'sheet']].copy()
    display.columns = ['브랜드', '순위', '상품명', '아이템타입', '가격', '카테고리', '_sheet']

    # 정렬 적용
    if sort_by == '순위 (오름차순)':
        display = display.sort_values('순위')
    elif sort_by == '가격 (낮은순)':
        display = display.sort_values(by='순위')  # 임시
        display['_price_num'] = display['가격'].apply(lambda x: parse_price(x))
        display = display.sort_values('_price_num')
        display = display.drop(columns=['_price_num'])
    elif sort_by == '가격 (높은순)':
        display['_price_num'] = display['가격'].apply(lambda x: parse_price(x))
        display = display.sort_values('_price_num', ascending=False)
        display = display.drop(columns=['_price_num'])
    elif sort_by == '카테고리 → 순위':
        display = display.sort_values(['카테고리', '순위'])
    else:  # 브랜드 → 순위
        display = display.sort_values(['브랜드', '순위'])

    bs_data = list(zip(display['브랜드'], display['_sheet'], display['순위']))
    show_df = display[['브랜드', '순위', '상품명', '아이템타입', '가격', '카테고리']].copy()
    render_image_table(show_df, image_map, rank_col='순위', name_col='상품명',
                       height=600, key_prefix='search', brand_sheet_data=bs_data)

    # 검색 결과 통계
    if not filtered.empty:
        st.divider()
        c1, c2 = st.columns(2)
        with c1:
            brand_counts = filtered['brand'].value_counts()
            fig = px.pie(values=brand_counts.values, names=brand_counts.index,
                         title='브랜드 분포', color=brand_counts.index,
                         color_discrete_map=BRAND_COLORS, hole=0.4)
            fig.update_layout(height=350)
            st.plotly_chart(fig, use_container_width=True)

        with c2:
            type_counts = filtered['item_type'].value_counts().head(10)
            fig = px.bar(x=type_counts.index, y=type_counts.values,
                         title='아이템타입 분포 TOP10',
                         labels={'x': '아이템타입', 'y': '상품수'})
            fig.update_layout(height=350)
            st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════
#  페이지: AI 분석 인사이트
# ══════════════════════════════════════════════════════

def _load_analysis_history():
    """분석 히스토리 JSON 로드 (캐싱)"""
    fp = os.path.join(WORK_DIR, 'analysis_history.json')
    if os.path.exists(fp):
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, Exception):
            pass
    return {}


def _render_insight_card(ins, idx, expanded=False):
    """단일 인사이트 카드 렌더링"""
    emoji_map = {'상품': '📦', '유형': '👕', '브랜드': '🏢'}
    emoji = emoji_map.get(ins.get('category', ''), '💡')

    with st.expander(f"{emoji} **{ins['title']}**", expanded=expanded):
        st.markdown(f"**{ins['summary']}**")

        if ins.get('details'):
            st.markdown("---")
            for d in ins['details']:
                st.markdown(f"- {d}")

        if ins.get('sub_insights'):
            st.markdown("")
            for s in ins['sub_insights']:
                st.markdown(f"  ▸ {s}")


# ══════════════════════════════════════════════════════
#  SPAO 데이터 로드
# ══════════════════════════════════════════════════════

@st.cache_data(ttl=600)
def load_spao_data():
    """spao_history.json에서 SPAO 베스트 데이터 로드"""
    spao_file = os.path.join(WORK_DIR, 'spao_history.json')
    if not os.path.exists(spao_file):
        return pd.DataFrame()

    try:
        with open(spao_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
    except (json.JSONDecodeError, Exception):
        return pd.DataFrame()

    products = []
    for cat_key, dates_data in history.items():
        if not dates_data:
            continue
        latest_date = max(dates_data.keys())
        items = dates_data[latest_date]
        # cat_key 예: "스파오_여성"
        parts = cat_key.split('_', 1)
        gender = parts[1] if len(parts) > 1 else ''

        for name, info in items.items():
            if isinstance(info, dict):
                raw_price = info.get('price', 0)
                price = parse_price(raw_price) if isinstance(raw_price, str) else (raw_price if isinstance(raw_price, (int, float)) else 0)
                products.append({
                    'brand': '스파오',
                    'category': gender,
                    'sheet': gender,
                    'rank': info.get('rank', 0),
                    'name': name,
                    'item_type': classify_item_type(name, '스파오'),
                    'price': price,
                    'original_price': info.get('original_price', price),
                    'price_str': f"{price:,}원" if price else '',
                    'image_url': info.get('image_url', ''),
                    'review_count': info.get('review_count', 0),
                })

    df = pd.DataFrame(products) if products else pd.DataFrame()
    return df


# ══════════════════════════════════════════════════════
#  페이지: SPAO 비교 분석
# ══════════════════════════════════════════════════════

def _get_gender_overall_top20(df, brand, gender):
    """브랜드별 성별 전체 랭킹 시트에서 TOP 20만 추출"""
    # 브랜드별 성별-전체 랭킹 시트 매핑
    GENDER_SHEET_MAP = {
        '유니클로': {'여성': 'WOMEN_모두보기', '남성': 'MEN_모두보기'},
        '아르켓':   {'여성': 'WOMEN',          '남성': 'MEN'},
        '탑텐':     {'여성': '여성',            '남성': '남성'},
        '미쏘':     {'여성': '여성'},
        '스파오':   {'여성': '여성',            '남성': '남성'},
    }
    sheets = GENDER_SHEET_MAP.get(brand, {})
    target_sheet = sheets.get(gender)
    if not target_sheet:
        return pd.DataFrame()
    bdf = df[(df['brand'] == brand) & (df['sheet'] == target_sheet)].copy()
    bdf = bdf[bdf['rank'] <= 20].sort_values('rank')
    return bdf


def page_spao_compare(df, image_map=None):
    st.header("🆚 SPAO 비교 분석")
    st.caption("각 브랜드 성별 전체 랭킹 TOP 20 기준 · 유형 + 가격 + 이미지 종합 비교")

    if image_map is None:
        image_map = {}

    spao_in_df = df[df['brand'] == '스파오'] if not df.empty else pd.DataFrame()

    if spao_in_df.empty:
        st.warning("SPAO 데이터가 없습니다. `python spao_crawler.py`를 먼저 실행해주세요.")
        st.code("python spao_crawler.py", language="bash")
        return

    if df.empty:
        st.warning("기존 브랜드 데이터가 없습니다.")
        return

    # ── 성별 선택 ──
    gender = st.radio("성별", ["여성", "남성"], horizontal=True, key='spao_gender')

    # ── 각 브랜드 TOP 20 추출 ──
    brand_frames = {}

    for brand in ALL_BRANDS:
        bdf = _get_gender_overall_top20(df, brand, gender)
        if not bdf.empty:
            # item_type 빈 문자열 → '미분류' 변환 (Cloud 환경 방어)
            if 'item_type' in bdf.columns:
                bdf['item_type'] = bdf['item_type'].fillna('미분류').replace('', '미분류')
            else:
                bdf['item_type'] = '미분류'
            # price 숫자 보장
            if 'price' in bdf.columns:
                bdf['price'] = pd.to_numeric(bdf['price'], errors='coerce').fillna(0)
            if 'price_str' not in bdf.columns:
                bdf['price_str'] = bdf['price'].apply(lambda x: f"{int(x):,}원" if x else '')
            brand_frames[brand] = bdf

    if '스파오' not in brand_frames:
        st.info(f"SPAO {gender} 데이터가 없습니다.")
        return

    st.divider()

    # ══════════════  1. 요약 KPI  ══════════════
    try:
        cols = st.columns(len(brand_frames))
        for i, (brand, bdf) in enumerate(brand_frames.items()):
            avg_price = float(bdf['price'].mean()) if not bdf['price'].empty else 0
            min_price = float(bdf['price'].min()) if not bdf['price'].empty else 0
            max_price = float(bdf['price'].max()) if not bdf['price'].empty else 0
            type_count = bdf['item_type'].nunique() if 'item_type' in bdf.columns else 0
            color = BRAND_COLORS.get(brand, '#888888')
            spao_avg = float(brand_frames['스파오']['price'].mean()) if '스파오' in brand_frames else 0
            diff_pct = ((avg_price - spao_avg) / spao_avg * 100) if spao_avg > 0 and brand != '스파오' else 0
            diff_str = f"<span style='font-size:11px;color:{'#e74c3c' if diff_pct > 0 else '#27ae60'};'>SPAO 대비 {diff_pct:+.1f}%</span>" if brand != '스파오' else "<span style='font-size:11px;color:#ddd;'>기준</span>"
            cols[i].markdown(
                f"""<div style="background:{color};color:white;padding:12px;border-radius:8px;text-align:center;">
                <b style='font-size:15px;'>{brand}</b><br>
                <span style='font-size:20px;font-weight:bold;'>{len(bdf)}개</span><br>
                <span style='font-size:12px;'>평균 {avg_price:,.0f}원</span><br>
                <span style='font-size:11px;'>{min_price:,.0f}~{max_price:,.0f}원 · {type_count}유형</span><br>
                {diff_str}
                </div>""",
                unsafe_allow_html=True)
    except Exception as e:
        st.error(f"KPI 요약 오류: {e}")

    st.divider()

    # ═══════════════════════════════════════════
    #  탭 구조: 유형분석 | 가격분석 | 이미지비교 | 갭분석
    # ═══════════════════════════════════════════
    main_tabs = st.tabs(["📊 아이템유형 분석", "💰 가격 심층 분석", "🖼️ TOP 상품 이미지 비교", "🔍 SPAO 갭(Gap) 분석"])

    # ──────────────── 탭1: 아이템유형 분석 ────────────────
    with main_tabs[0]:
      try:
        st.subheader("📊 아이템유형 구성 비교")

        type_rows = []
        for brand, bdf in brand_frames.items():
            counts = bdf['item_type'].value_counts()
            total = len(bdf)
            for t, c in counts.items():
                type_rows.append({'브랜드': brand, '아이템타입': t if t else '미분류', '상품수': int(c),
                                  '비중(%)': round(c / total * 100, 1)})
        type_chart_df = pd.DataFrame(type_rows)

        if not type_chart_df.empty:
            # 상품 수 그룹 바 차트
            fig = px.bar(
                type_chart_df, x='아이템타입', y='상품수', color='브랜드', barmode='group',
                color_discrete_map=BRAND_COLORS,
                title=f'{gender} TOP20 — 아이템타입별 상품 수',
            )
            fig.update_layout(height=420, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)

            # 브랜드별 유형 비중 Sunburst
            st.markdown("#### 브랜드별 유형 비중")
            sun_df = type_chart_df[type_chart_df['상품수'] > 0].copy()
            if not sun_df.empty:
                fig_sun = px.sunburst(
                    sun_df, path=['브랜드', '아이템타입'], values='상품수',
                    color='브랜드', color_discrete_map=BRAND_COLORS,
                    title=f'{gender} TOP20 — 유형 구성 비율',
                )
                fig_sun.update_layout(height=480)
                st.plotly_chart(fig_sun, use_container_width=True)

            # 유형 비중 히트맵 테이블
            with st.expander("📋 유형 비중 상세 테이블", expanded=False):
                pivot = type_chart_df.pivot_table(
                    index='아이템타입', columns='브랜드', values='비중(%)', fill_value=0)
                brand_order = [b for b in ALL_BRANDS if b in pivot.columns]
                if brand_order:
                    pivot = pivot[brand_order]
                try:
                    st.dataframe(pivot.style.format("{:.1f}%").background_gradient(cmap='YlOrRd', axis=None),
                                 use_container_width=True)
                except ImportError:
                    st.dataframe(pivot.style.format("{:.1f}%"), use_container_width=True)
      except Exception as e:
        st.error(f"아이템유형 분석 오류: {e}")
        st.exception(e)

    # ──────────────── 탭2: 가격 심층 분석 ────────────────
    with main_tabs[1]:
      try:
        st.subheader("💰 가격 분포 비교")

        # 가격대 Box Plot
        price_rows = []
        for brand, bdf in brand_frames.items():
            for _, row in bdf.iterrows():
                price_rows.append({
                    '브랜드': brand, '가격': row['price'],
                    '상품명': str(row['name'])[:15], '아이템타입': row.get('item_type', '미분류')})
        price_df = pd.DataFrame(price_rows)

        if not price_df.empty:
            fig_box = px.box(
                price_df, x='브랜드', y='가격', color='브랜드',
                color_discrete_map=BRAND_COLORS,
                title=f'{gender} TOP20 가격 분포 (Box Plot)',
                points='all',
            )
            fig_box.update_layout(height=450, showlegend=False)
            st.plotly_chart(fig_box, use_container_width=True)

        # 아이템타입별 평균가격 비교
        st.markdown("#### 아이템타입별 평균가격 비교")
        type_price_rows = []
        for brand, bdf in brand_frames.items():
            for t, grp in bdf.groupby('item_type'):
                type_price_rows.append({
                    '브랜드': brand, '아이템타입': t,
                    '평균가격': int(grp['price'].mean()),
                    '상품수': len(grp)})
        type_price_df = pd.DataFrame(type_price_rows)

        if not type_price_df.empty:
            fig_tp = px.bar(
                type_price_df, x='아이템타입', y='평균가격', color='브랜드',
                barmode='group', color_discrete_map=BRAND_COLORS,
                title=f'{gender} TOP20 — 아이템타입별 평균가격',
                text='평균가격',
            )
            fig_tp.update_traces(texttemplate='%{text:,.0f}', textposition='outside')
            fig_tp.update_layout(height=480, xaxis_tickangle=-45, yaxis_title='평균가격(원)')
            st.plotly_chart(fig_tp, use_container_width=True)

            # SPAO 대비 가격 차이 테이블
            with st.expander("📋 SPAO 대비 아이템타입별 가격 차이", expanded=True):
                spao_type_prices = {}
                if '스파오' in brand_frames:
                    for t, grp in brand_frames['스파오'].groupby('item_type'):
                        spao_type_prices[t] = int(grp['price'].mean())

                diff_rows = []
                for brand in [b for b in ALL_BRANDS if b != '스파오' and b in brand_frames]:
                    for t, grp in brand_frames[brand].groupby('item_type'):
                        brand_avg = int(grp['price'].mean())
                        spao_avg_t = spao_type_prices.get(t, 0)
                        if spao_avg_t > 0:
                            diff = brand_avg - spao_avg_t
                            diff_pct_val = diff / spao_avg_t * 100
                            diff_rows.append({
                                '브랜드': brand, '아이템타입': t,
                                '브랜드 평균가': f"{brand_avg:,}원",
                                'SPAO 평균가': f"{spao_avg_t:,}원",
                                '차이': f"{diff:+,}원",
                                '차이율': f"{diff_pct_val:+.1f}%",
                            })
                        else:
                            diff_rows.append({
                                '브랜드': brand, '아이템타입': t,
                                '브랜드 평균가': f"{brand_avg:,}원",
                                'SPAO 평균가': '—',
                                '차이': '—', '차이율': '—',
                            })
                if diff_rows:
                    st.dataframe(pd.DataFrame(diff_rows), use_container_width=True, hide_index=True)

        # 가격대 분포 히스토그램
        st.markdown("#### 가격대 분포")
        if not price_df.empty:
            fig_hist = px.histogram(
                price_df, x='가격', color='브랜드', barmode='overlay',
                color_discrete_map=BRAND_COLORS,
                nbins=15, opacity=0.65,
                title=f'{gender} TOP20 가격대 분포',
            )
            fig_hist.update_layout(height=380, xaxis_title='가격(원)', yaxis_title='상품 수')
            st.plotly_chart(fig_hist, use_container_width=True)
      except Exception as e:
        st.error(f"가격 분석 오류: {e}")
        st.exception(e)

    # ──────────────── 탭3: TOP 상품 이미지 비교 ────────────────
    with main_tabs[2]:
      try:
        st.subheader("🖼️ TOP 상품 이미지 나란히 비교")
        st.caption("각 브랜드의 TOP 상품을 이미지와 함께 나란히 비교합니다.")

        img_top_n = st.slider("표시 개수", 3, 20, 10, key='spao_img_topn')
        brand_list = [b for b in ALL_BRANDS if b in brand_frames]

        if brand_list:
            img_cols = st.columns(len(brand_list))
            for ci, brand in enumerate(brand_list):
                with img_cols[ci]:
                    color = BRAND_COLORS.get(brand, '#888')
                    st.markdown(
                        f"<div style='background:{color};color:white;padding:6px;border-radius:6px;"
                        f"text-align:center;font-weight:bold;font-size:14px;'>{brand}</div>",
                        unsafe_allow_html=True)

                    bdf = brand_frames[brand]
                    top_items = bdf.nsmallest(img_top_n, 'rank')

                    for _, row in top_items.iterrows():
                        rank_val = int(row['rank'])
                        name = row['name']
                        price_s = row.get('price_str', '')
                        item_type = row.get('item_type', '')
                        sheet_val = row.get('sheet', '')

                        b64 = get_image_b64(image_map, brand, sheet_val, rank_val) if image_map else None

                        if b64:
                            st.markdown(
                                f"<div style='border:1px solid #eee;border-radius:8px;padding:8px;margin-bottom:8px;text-align:center;'>"
                                f"<img src='data:image/jpeg;base64,{b64}' style='width:100%;max-width:180px;border-radius:6px;'/>"
                                f"<div style='font-size:12px;font-weight:600;margin-top:4px;'>{rank_val}. {name[:18]}</div>"
                                f"<div style='font-size:11px;color:#666;'>{item_type} · {price_s}</div>"
                                f"</div>",
                                unsafe_allow_html=True)
                        else:
                            st.markdown(
                                f"<div style='border:1px solid #eee;border-radius:8px;padding:8px;margin-bottom:8px;text-align:center;"
                                f"background:#f9f9f9;min-height:80px;display:flex;align-items:center;justify-content:center;flex-direction:column;'>"
                                f"<div style='font-size:12px;font-weight:600;'>{rank_val}. {name[:18]}</div>"
                                f"<div style='font-size:11px;color:#666;'>{item_type} · {price_s}</div>"
                                f"</div>",
                                unsafe_allow_html=True)

        # 동일 아이템타입 이미지 매칭 비교
        st.divider()
        st.subheader("🔗 동일 아이템타입 상품 이미지 매칭")
        st.caption("같은 아이템타입(예: 팬츠, 아우터 등)을 가진 상품들을 브랜드 간 나란히 비교합니다.")

        # 공통 아이템타입 찾기
        all_types_sets = {b: set(bdf['item_type'].unique()) for b, bdf in brand_frames.items()}
        common_types = set()
        if all_types_sets:
            common_types = set.intersection(*all_types_sets.values()) if len(all_types_sets) > 1 else set()
            # SPAO와 다른 브랜드 간 공통
            if '스파오' in all_types_sets:
                for b, ts in all_types_sets.items():
                    if b != '스파오':
                        common_types = common_types | (ts & all_types_sets['스파오'])

        if common_types:
            selected_type = st.selectbox("아이템타입 선택", sorted(common_types), key='spao_type_match')

            match_cols = st.columns(len(brand_list))
            for ci, brand in enumerate(brand_list):
                with match_cols[ci]:
                    color = BRAND_COLORS.get(brand, '#888')
                    st.markdown(
                        f"<div style='background:{color};color:white;padding:4px;border-radius:4px;"
                        f"text-align:center;font-weight:bold;font-size:13px;margin-bottom:6px;'>{brand}</div>",
                        unsafe_allow_html=True)
                    bdf = brand_frames[brand]
                    type_items = bdf[bdf['item_type'] == selected_type].nsmallest(5, 'rank')

                    if type_items.empty:
                        st.caption("해당 유형 없음")
                        continue

                    for _, row in type_items.iterrows():
                        rank_val = int(row['rank'])
                        name = row['name']
                        price_s = row.get('price_str', '')
                        sheet_val = row.get('sheet', '')
                        b64 = get_image_b64(image_map, brand, sheet_val, rank_val) if image_map else None

                        if b64:
                            st.markdown(
                                f"<div style='border:1px solid #ddd;border-radius:6px;padding:6px;margin-bottom:6px;text-align:center;'>"
                                f"<img src='data:image/jpeg;base64,{b64}' style='width:100%;max-width:150px;border-radius:4px;'/>"
                                f"<div style='font-size:11px;font-weight:600;margin-top:3px;'>{rank_val}. {name[:16]}</div>"
                                f"<div style='font-size:10px;color:#888;'>{price_s}</div>"
                                f"</div>",
                                unsafe_allow_html=True)
                        else:
                            st.markdown(
                                f"<div style='border:1px solid #ddd;border-radius:6px;padding:6px;margin-bottom:6px;"
                                f"text-align:center;background:#f9f9f9;'>"
                                f"<div style='font-size:11px;font-weight:600;'>{rank_val}. {name[:16]}</div>"
                                f"<div style='font-size:10px;color:#888;'>{price_s}</div>"
                                f"</div>",
                                unsafe_allow_html=True)
        else:
            st.info("브랜드 간 공통 아이템타입이 없습니다.")
      except Exception as e:
        st.error(f"이미지 비교 오류: {e}")
        st.exception(e)

    # ──────────────── 탭4: SPAO 갭 분석 ────────────────
    with main_tabs[3]:
      try:
        st.subheader("🔍 SPAO 갭(Gap) 분석")
        st.caption("다른 브랜드 TOP20에는 있지만 SPAO TOP20에는 없는 아이템타입 · 상품을 분석합니다.")

        spao_types = set()
        if '스파오' in brand_frames:
            spao_types = set(brand_frames['스파오']['item_type'].unique())

        other_brands = [b for b in ALL_BRANDS if b != '스파오' and b in brand_frames]

        # 갭 요약 카드
        if other_brands:
            gap_summary = {}
            for brand in other_brands:
                brand_types = set(brand_frames[brand]['item_type'].unique())
                missing = brand_types - spao_types
                gap_summary[brand] = missing

            total_gap_types = set()
            for v in gap_summary.values():
                total_gap_types |= v

            if total_gap_types:
                st.markdown(f"**SPAO에 없는 아이템타입 총 {len(total_gap_types)}종**: {', '.join(sorted(total_gap_types))}")

                # 어느 브랜드에서 공통으로 나타나는지 분석
                gap_freq = {}
                for t in total_gap_types:
                    brands_with = [b for b in other_brands if t in gap_summary[b]]
                    gap_freq[t] = brands_with

                gap_importance = sorted(gap_freq.items(), key=lambda x: -len(x[1]))

                st.markdown("#### 🎯 갭 우선순위 (다수 브랜드에 있는 유형 = 시장 트렌드)")
                for item_type, brands_with in gap_importance:
                    cnt = len(brands_with)
                    bar_len = cnt * 20
                    st.markdown(
                        f"<div style='margin:4px 0;'>"
                        f"<span style='font-weight:bold;'>{item_type}</span> "
                        f"<span style='display:inline-block;background:#e74c3c;height:14px;width:{bar_len}px;border-radius:3px;vertical-align:middle;margin:0 6px;'></span>"
                        f"<span style='font-size:12px;color:#666;'>{cnt}개 브랜드 ({', '.join(brands_with)})</span>"
                        f"</div>",
                        unsafe_allow_html=True)
            else:
                st.success("SPAO TOP20에 다른 브랜드의 모든 아이템타입이 포함되어 있습니다!")

        st.divider()

        # 브랜드별 상세 (이미지 포함)
        st.markdown("#### 브랜드별 상세 — SPAO에 없는 상품 리스트")

        if not other_brands:
            st.info("비교할 다른 브랜드 데이터가 없습니다.")
        else:
            tabs = st.tabs(other_brands)
            for tab, brand in zip(tabs, other_brands):
                with tab:
                    bdf = brand_frames[brand]
                    brand_types = set(bdf['item_type'].unique())
                    missing_types = brand_types - spao_types

                    if not missing_types:
                        st.success(f"{brand} TOP20의 모든 아이템타입이 SPAO에도 있습니다!")
                        continue

                    st.markdown(f"**SPAO에 없는 아이템타입**: {', '.join(sorted(missing_types))}")

                    # 해당 아이템타입 상품들 리스트 + 이미지
                    missing_items = bdf[bdf['item_type'].isin(missing_types)].sort_values(['item_type', 'rank'])
                    show_cols = ['rank', 'name', 'item_type', 'price_str', 'sheet']
                    available_cols = [c for c in show_cols if c in missing_items.columns]
                    show_df = missing_items[available_cols].copy()
                    # 방어적 컬럼 확인
                    disp_cols = [c for c in ['rank', 'name', 'item_type', 'price_str'] if c in show_df.columns]
                    display_df = show_df[disp_cols].copy()
                    col_map = {'rank': '순위', 'name': '상품명', 'item_type': '아이템타입', 'price_str': '가격'}
                    display_df.columns = [col_map.get(c, c) for c in display_df.columns]
                    sheet_vals = show_df['sheet'].values if 'sheet' in show_df.columns else [''] * len(show_df)
                    rank_vals = show_df['rank'].values if 'rank' in show_df.columns else [0] * len(show_df)
                    bs_data = [(brand, s, r) for s, r in zip(sheet_vals, rank_vals)]
                    render_image_table(display_df, image_map, rank_col='순위', name_col='상품명',
                                       height=min(len(display_df) * 38 + 60, 500),
                                       key_prefix=f'spao_miss_{brand}', brand_sheet_data=bs_data)
      except Exception as e:
        st.error(f"갭 분석 오류: {e}")
        st.exception(e)


# ══════════════════════════════════════════════════════
#  페이지: AI 분석 인사이트
# ══════════════════════════════════════════════════════

def page_analysis():
    st.header("🤖 AI 분석 인사이트")

    analysis = _load_analysis_history()

    if not analysis:
        st.warning("분석 데이터가 없습니다. 크롤링 후 자동 생성됩니다.")
        st.info("수동 실행: `python brand_analysis.py`")
        return

    # ── 날짜 선택 ──
    dates = sorted(analysis.keys(), reverse=True)
    date_labels = {d: f"{d[:4]}.{d[4:6]}.{d[6:]}" for d in dates}

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        selected_date = st.selectbox(
            "분석 기준일", dates,
            format_func=lambda d: f"{date_labels[d]} ({analysis[d].get('generated_at', '')[:10]})",
            key='analysis_date'
        )
    with col2:
        total_insights = analysis[selected_date].get('total_insights', 0)
        st.metric("전체 인사이트", f"{total_insights}개")
    with col3:
        total_records = analysis[selected_date].get('total_records', 0)
        st.metric("분석 레코드", f"{total_records:,}건")

    entry = analysis[selected_date]

    product_insights = entry.get('product_insights', [])
    type_insights = entry.get('type_insights', [])
    brand_insights = entry.get('brand_insights', [])

    # ── 3축 탭 ──
    tab1, tab2, tab3 = st.tabs([
        f"📦 상품 단위 ({len(product_insights)})",
        f"👕 유형 단위 ({len(type_insights)})",
        f"🏢 브랜드 단위 ({len(brand_insights)})",
    ])

    with tab1:
        if product_insights:
            for i, ins in enumerate(product_insights):
                _render_insight_card(ins, i, expanded=(i == 0))
        else:
            st.info("상품 단위 인사이트가 없습니다.")

    with tab2:
        if type_insights:
            for i, ins in enumerate(type_insights):
                _render_insight_card(ins, i, expanded=(i == 0))
        else:
            st.info("유형 단위 인사이트가 없습니다.")

    with tab3:
        if brand_insights:
            for i, ins in enumerate(brand_insights):
                _render_insight_card(ins, i, expanded=(i == 0))
        else:
            st.info("브랜드 단위 인사이트가 없습니다.")

    # ── 인사이트 요약 테이블 ──
    all_ins = product_insights + type_insights + brand_insights
    if all_ins:
        st.divider()
        st.subheader("📋 전체 인사이트 요약")

        table_data = []
        for ins in all_ins:
            table_data.append({
                '분류': ins.get('category', ''),
                '제목': ins.get('title', ''),
                '요약': ins.get('summary', '')[:120],
            })

        keyword = st.text_input("키워드 필터", key='anal_keyword')
        table_df = pd.DataFrame(table_data)

        if keyword:
            mask = table_df['제목'].str.contains(keyword, case=False, na=False) | \
                   table_df['요약'].str.contains(keyword, case=False, na=False)
            table_df = table_df[mask]

        st.dataframe(table_df, use_container_width=True, hide_index=True, height=320)

    # ── 메타 정보 ──
    st.divider()
    st.caption(
        f"생성: {entry.get('generated_at', 'N/A')} | "
        f"데이터: {', '.join(entry.get('data_dates', []))} | "
        f"브랜드: {', '.join(entry.get('brands', []))} | "
        f"누적 분석: {len(dates)}회"
    )


# ══════════════════════════════════════════════════════
#  메인 앱
# ══════════════════════════════════════════════════════

def _check_login():
    """로그인 상태 확인 및 로그인 화면 표시. True면 인증됨."""
    if st.session_state.get('authenticated'):
        return True

    # 로그인 화면
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none; }
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<h1 style='text-align:center;'>📊 브랜드 랭킹 대시보드</h1>"
            f"<p style='text-align:center; color:#888;'>{'  ·  '.join(ALL_BRANDS)}</p>",
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        with st.form("login_form"):
            user_id = st.text_input("아이디", placeholder="아이디를 입력하세요")
            user_pw = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
            submitted = st.form_submit_button("로그인", use_container_width=True)

            if submitted:
                # 환경변수 또는 Streamlit secrets에서 인증정보 로드
                valid_id = ""
                valid_pw = ""
                try:
                    if hasattr(st, 'secrets') and st.secrets is not None:
                        valid_id = st.secrets.get("LOGIN_ID", "")
                        valid_pw = st.secrets.get("LOGIN_PW", "")
                except Exception:
                    valid_id = ""
                    valid_pw = ""
                if not valid_id:
                    valid_id = os.environ.get("LOGIN_ID", "")
                if not valid_pw:
                    valid_pw = os.environ.get("LOGIN_PW", "")
                if valid_id and valid_pw and user_id == valid_id and user_pw == valid_pw:
                    st.session_state['authenticated'] = True
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호가 올바르지 않습니다.")

    return False


# ═══════════════════════════════════════════════════
# 품평회 분석 페이지
# ═══════════════════════════════════════════════════

def page_survey_analysis():
    """품평회 설문 분석 페이지 – 엑셀 업로드 → 분석 엑셀 + PPT 자동 생성"""
    st.header("📋 품평회 설문 분석")
    st.caption("품평회 설문 RAW 엑셀 파일을 업로드하면 자동으로 분석 엑셀 + PPT를 생성합니다.")

    uploaded = st.file_uploader(
        "설문 RAW 파일 업로드 (.xlsx, .csv, .tsv, .xls, .ods)",
        type=["xlsx", "csv", "tsv", "xls", "ods"],
        help="Google Forms에서 다운로드한 모든 형식 지원 (스프레드시트, CSV, TSV 등)",
        key="survey_file_uploader",
    )

    if uploaded is None:
        st.info("💡 파일을 업로드해주세요. 아이템별 선호도/적정가격/컬러선호 등을 자동 분석합니다.")

        # ── RAW 파일 요건 가이드 ──
        with st.expander("📖 RAW 파일 요건 (클릭하여 펼치기)", expanded=False):
            st.markdown("""
### 파일 형식
- **`.xlsx`** (엑셀), **`.csv`**, **`.tsv`**, **`.xls`** (구형 엑셀), **`.ods`** (오픈도큐먼트) 모두 지원
- Google Forms → **"응답" 시트 → 스프레드시트에서 응답 보기 → 파일 → 다운로드** → 아무 형식으로 받으면 됩니다
- CSV/TSV: UTF-8 인코딩 기준 (Google 기본값)
- 엑셀 파일은 첫 번째 시트(활성 시트)만 읽습니다

---

### 필수 열 구조 (1행 = 헤더)

| 구분 | 열 위치 | 헤더 예시 | 비고 |
|------|---------|----------|------|
| **성별** | A~F열 사이 | `성별을 선택해주세요` | 헤더에 **"성별"** 텍스트가 반드시 포함 |
| **연령** | A~F열 사이 | `연령을 선택해주세요` | 헤더에 **"연령"** 또는 **"나이"** 텍스트가 반드시 포함 |

- 성별 값: `남성`, `여성` (또는 `남자`, `여자` → 자동 변환됨)
- 연령 값: `20~24세`, `25~29세`, `45~49세` 등 (연령대 텍스트 그대로)
- 성별·연령 열은 **A~F열(1~6번째 열)** 안에 있어야 합니다
- 타임스탬프, 성함, 연락처, 소속 등 기본 열은 자유 배치 가능

---

### 아이템 질문 구조 (G열 = 7번째 열부터)

**핵심 규칙: 각 아이템의 첫 질문 헤더는 반드시 `1.` 로 시작해야 합니다.**

```
1. [아이템명] 해당 상품의 선호도를 10점 만점으로 평가해주세요.  ← 아이템 시작
2. 해당 상품의 구매 의향 가격을 적어주세요. (숫자만 기입)
3. 해당 상품의 선호하는 "컬러"를 골라 주세요.
해당 스타일에 대한 만족/불만족 요소를 알려주세요.
1. [다음 아이템명] 해당 상품의 선호도를 10점 만점으로 ...     ← 새 아이템 시작
```

- `1. ` (숫자1 + 마침표 + 공백) 또는 `1.[` (숫자1 + 마침표 + 대괄호)로 시작하면 **새 아이템**으로 인식
- G열 이전(A~F)에 있는 질문/열은 무시됩니다
- Google Sheets의 빈 열(`열1`, `열2` 등)은 자동으로 무시됩니다

---

### 자동 인식되는 질문 유형

| 질문 유형 | 헤더에 포함해야 할 키워드 | 데이터 형식 | 분석 내용 |
|----------|------------------------|-----------|----------|
| **선호도** | `선호도` + (`10점` 또는 `만점`) | 숫자 (0~10) | 전체/성별/연령별 평균, 랭킹 |
| **적정가격** | `가격` 또는 `적정` | 숫자 (원 단위) | 전체/성별/연령별 평균 |
| **컬러선호** | `컬러` + (`골라` 또는 `구매의향`) | 텍스트 (컬러명) | TOP 10 컬러, 비율 |
| **컬러선호도** | `컬러 선호도` | 숫자 | 평균 분석 |
| **선호유형** | `선호하는` + (`유형` 또는 `디자인`) | 텍스트 | 선택 분포 |
| **구매희망** | `구매하고 싶은` 또는 `구매하고싶은` | 텍스트 | 선택 분포 |
| **주관식** | `만족/불만족`, `자유롭게`, `제안` | 텍스트 | 특수질문 요약 |
| **기타수치** | `2.`, `3.` 등 숫자로 시작 | 숫자 | 평균/분포 |

- 선호도 질문이 **없는** 아이템은 랭킹 요약에서 제외됩니다
- 적정가격의 `원`, `,`, `만` 등은 자동 변환됩니다 (예: `39,000원` → 39000)
- **0점 = 비구매 의사**로 간주 → 0포함/0제외 두 가지 랭킹 생성

---

### 예시 헤더 배치

| A | B | C | D | E | F | G | H | I | J | K | L | M | ... |
|---|---|---|---|---|---|---|---|---|---|---|---|---|-----|
| 타임스탬프 | 성함 | 연락처 | **성별** | **연령** | 소속 | **1. [아이템1] 선호도...** | 2. 적정가격... | 3. 컬러... | 주관식... | **1. [아이템2] 선호도...** | 2. 적정가격... | 3. 컬러... | ... |

---

### 주의사항
- 헤더(1행)가 **비어있는 열**은 건너뜁니다
- 데이터 행에서 1~6열이 모두 빈 행은 자동 스킵됩니다
- 아이템 번호는 `1.`로 시작하는 헤더 순서대로 자동 부여됩니다 (1, 2, 3...)
- 하나의 아이템 안에 질문이 여러 개 있어도 됩니다 (다음 `1.`이 나올 때까지 같은 아이템)
""")

        # 기존 결과 표시
        _show_existing_survey_results()
        return

    # 파일 저장
    import tempfile, shutil
    tmp_dir = tempfile.mkdtemp()
    input_path = os.path.join(tmp_dir, uploaded.name)
    with open(input_path, "wb") as f:
        f.write(uploaded.getvalue())

    base_name = os.path.splitext(uploaded.name)[0]
    output_excel = os.path.join(tmp_dir, f"{base_name}_분석결과_v7.xlsx")
    output_ppt = os.path.join(tmp_dir, f"{base_name}_분석결과_v7.pptx")

    try:
        import sys
        # survey_analyzer 모듈 캐시 완전 제거 후 재임포트
        if 'survey_analyzer' in sys.modules:
            del sys.modules['survey_analyzer']
        import survey_analyzer as sa

        with st.spinner("📊 데이터 로딩 중..."):
            headers, data = sa.load_raw_data(input_path)

        gender_col, age_col = sa.find_gender_age_columns(headers)
        items = sa.identify_items(headers)

        if not items:
            st.error("아이템을 식별할 수 없습니다. 헤더 형식을 확인해주세요.")
            return

        genders_set = set()
        ages_set = set()
        for row in data:
            g = sa.clean_gender(row[gender_col]) if gender_col is not None and row[gender_col] else "미응답"
            a = str(row[age_col]).strip() if age_col is not None and row[age_col] else "미응답"
            genders_set.add(g)
            ages_set.add(a)
        genders = sorted(genders_set)
        ages = sa.sort_age_groups(ages_set)

        # 요약 정보
        col1, col2, col3 = st.columns(3)
        col1.metric("응답자 수", f"{len(data)}명")
        col2.metric("아이템 수", f"{len(items)}개")
        col3.metric("성별/연령", f"{len(genders)}성별 × {len(ages)}연령")

        st.divider()

        # 아이템 목록
        with st.expander("🔍 아이템 식별 결과", expanded=True):
            for item in items:
                q_types = list(set(q["type"] for q in item["questions"]))
                st.write(f"**아이템 {item['item_no']}**: {len(item['questions'])}개 질문 ({', '.join(q_types)})")

        # 분석 실행
        with st.spinner("📗 분석 엑셀 생성 중..."):
            sa.create_summary_excel(headers, data, items, gender_col, age_col, output_excel)

        with st.spinner("📙 PPT 생성 중..."):
            sa.create_ppt(items, data, gender_col, age_col, genders, ages, output_ppt)

        st.success(f"✅ 분석 완료! (응답자 {len(data)}명, 아이템 {len(items)}개)")

        # 다운로드 버튼
        st.divider()
        dcol1, dcol2 = st.columns(2)
        with open(output_excel, "rb") as f:
            dcol1.download_button(
                label="📗 분석 엑셀 다운로드",
                data=f.read(),
                file_name=f"{base_name}_분석결과_v7.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with open(output_ppt, "rb") as f:
            dcol2.download_button(
                label="📙 PPT 다운로드",
                data=f.read(),
                file_name=f"{base_name}_분석결과_v7.pptx",
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                use_container_width=True,
            )

        # 워크스페이스에도 복사
        try:
            shutil.copy2(output_excel, os.path.join(WORK_DIR, f"{base_name}_분석결과_v7.xlsx"))
            shutil.copy2(output_ppt, os.path.join(WORK_DIR, f"{base_name}_분석결과_v7.pptx"))
        except Exception:
            pass

        # 선호도 랭킹 미리보기
        _show_ranking_preview(items, data, gender_col, age_col, ages)

    except Exception as e:
        st.error(f"분석 중 오류가 발생했습니다: {e}")
        import traceback
        st.code(traceback.format_exc())
    finally:
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass


def _show_existing_survey_results():
    """워크스페이스에 있는 기존 분석 결과 파일 표시 + 삭제 기능"""
    result_files = sorted(glob.glob(os.path.join(WORK_DIR, "*_분석결과_v7.xlsx")))
    ppt_files = sorted(glob.glob(os.path.join(WORK_DIR, "*_분석결과_v7.pptx")))
    if not result_files and not ppt_files:
        return

    # 삭제 처리
    if 'survey_delete_target' in st.session_state and st.session_state['survey_delete_target']:
        target = st.session_state.pop('survey_delete_target')
        deleted = []
        for ext in ['.xlsx', '.pptx']:
            fp = os.path.join(WORK_DIR, target + ext)
            if os.path.exists(fp):
                os.remove(fp)
                deleted.append(os.path.basename(fp))
        if deleted:
            st.toast(f"🗑️ 삭제됨: {', '.join(deleted)}")
            st.rerun()

    st.divider()
    st.subheader("📁 기존 분석 결과")

    # 파일을 base_name 기준으로 그룹핑
    groups = {}
    for fp in result_files + ppt_files:
        fname = os.path.basename(fp)
        # "XXX_분석결과_v7.xlsx" → base = "XXX_분석결과_v7"
        base = os.path.splitext(fname)[0]
        if base not in groups:
            groups[base] = []
        groups[base].append(fp)

    for base, files in sorted(groups.items()):
        cols = st.columns([5, 5, 2])
        for fp in files:
            fname = os.path.basename(fp)
            ext = os.path.splitext(fname)[1]
            with open(fp, "rb") as f:
                file_data = f.read()
            if ext == '.xlsx':
                cols[0].download_button(
                    label=f"📗 {fname}",
                    data=file_data,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{fname}",
                )
            elif ext == '.pptx':
                cols[1].download_button(
                    label=f"📙 {fname}",
                    data=file_data,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    key=f"dl_{fname}",
                )
        if cols[2].button("🗑️ 삭제", key=f"del_{base}"):
            st.session_state['survey_delete_target'] = base
            st.rerun()


def _show_ranking_preview(items, data, gender_col, age_col, ages):
    """선호도 랭킹 미리보기 테이블"""
    import sys
    if 'survey_analyzer' in sys.modules:
        del sys.modules['survey_analyzer']
    import survey_analyzer as sa
    ranking = []
    for item in items:
        pref_qs = [q for q in item["questions"] if q["type"] == "선호도"]
        if not pref_qs:
            continue
        avg = sa.calculate_averages(data, pref_qs[0]["col_idx"], gender_col, age_col)
        avg_ex = sa.calculate_averages(data, pref_qs[0]["col_idx"], gender_col, age_col, exclude_zero=True)
        ranking.append({
            "아이템": f"아이템 {item['item_no']}",
            "전체(0포함)": avg["전체"],
            "전체(0제외)": avg_ex["전체"],
        })
    if ranking:
        st.divider()
        st.subheader("🏆 선호도 랭킹 미리보기")
        df_rank = pd.DataFrame(ranking)
        df_rank = df_rank.sort_values("전체(0포함)", ascending=False).reset_index(drop=True)
        df_rank.index = df_rank.index + 1
        df_rank.index.name = "순위"
        st.dataframe(df_rank, width='stretch')


def main():
    try:
        st.set_page_config(
            page_title="3사 브랜드 랭킹 대시보드",
            page_icon="📊",
            layout="wide",
            initial_sidebar_state="expanded",
        )
    except Exception:
        pass  # 이미 set_page_config가 호출된 경우

    try:
        # 로그인 체크
        if not _check_login():
            return
    except Exception as e:
        st.error(f"로그인 처리 중 오류: {e}")
        return

    # CSS
    st.markdown("""
    <style>
    .stApp { font-family: 'Pretendard', 'Noto Sans KR', sans-serif; }
    [data-testid="stSidebar"] { background: #f8f9fa; }
    .stMetric { background: #f8f9fa; padding: 12px; border-radius: 8px; }
    h1, h2, h3 { color: #1a1a2e; }
    </style>
    """, unsafe_allow_html=True)

    # 사이드바
    with st.sidebar:
        st.image("https://img.icons8.com/fluency/48/analytics.png", width=48)
        st.title("랭킹 대시보드")
        st.caption(' · '.join(ALL_BRANDS))
        st.divider()

        page = st.radio(
            "페이지",
            ["📊 종합 대시보드", "🏷️ 브랜드별 상세", "💰 가격 비교",
             "🏆 핵심 아이템", "📈 랭킹 변동 추적", "🔍 상품 검색",
             "🆚 SPAO 비교 분석", "🤖 AI 분석", "📋 품평회 분석"],
            label_visibility="collapsed"
        )

        st.divider()
        if st.button("🔄 데이터 새로고침", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        st.caption("크롤러 실행 후 새로고침하면\n최신 데이터가 반영됩니다.")

        st.divider()
        st.markdown(
            "<div style='font-size:10px; color:#999; line-height:1.4;'>"
            "⚖️ <b>면책 고지</b><br>"
            "본 대시보드는 <b>내부 참고용</b>이며, "
            "각 브랜드의 공식 서비스가 아닙니다.<br>"
            "모든 상품 정보·이미지의 저작권은 "
            "해당 브랜드에 귀속됩니다.<br>"
            "상업적 사용 및 데이터 재배포를 금지합니다."
            "</div>",
            unsafe_allow_html=True,
        )

    # 품평회 분석은 별도 데이터 → 먼저 분기
    if "품평회" in page:
        try:
            page_survey_analysis()
        except Exception as e:
            st.error(f"페이지 렌더링 중 오류: {e}")
            st.exception(e)
        return

    # 데이터 로드
    try:
        with st.spinner("데이터 로드 중..."):
            history = load_all_history()
            df = load_latest_excel_data()
            # SPAO 데이터를 메인 df에 통합
            spao_df = load_spao_data()
            if not spao_df.empty:
                spao_merge = spao_df[['brand', 'category', 'sheet', 'rank', 'name', 'item_type', 'price', 'price_str']].copy()
                if 'date' not in spao_merge.columns:
                    spao_merge['date'] = ''
                df = pd.concat([df, spao_merge], ignore_index=True)
            image_map = extract_all_product_images()
            # archive 이미지로 보충 (HD 이미지 없는 유니클로·아르켓 등)
            image_map = augment_image_map_from_archive(image_map, df)
            # 썸네일 JSON으로 보충 (Cloud 환경 fallback)
            image_map = augment_image_map_from_thumbnails(image_map, df)
            # SPAO 이미지 통합 (archive → URL fallback)
            spao_images = load_spao_image_map()
            if spao_images:
                image_map.update(spao_images)
    except Exception as e:
        st.error(f"데이터 로드 중 오류가 발생했습니다: {e}")
        history = {}
        df = pd.DataFrame()
        image_map = {}

    # 라우팅
    try:
        if "종합" in page:
            page_overview(df, history, image_map)
        elif "브랜드" in page:
            page_brand_detail(df, history, image_map)
        elif "가격" in page:
            page_price_compare(df, image_map)
        elif "핵심" in page:
            page_top_items(df, image_map)
        elif "변동" in page:
            page_ranking_trend(history, image_map)
        elif "검색" in page:
            page_search(df, image_map)
        elif "SPAO" in page:
            page_spao_compare(df, image_map)
        elif "AI" in page:
            page_analysis()
    except Exception as e:
        st.error(f"페이지 렌더링 중 오류: {e}")
        st.exception(e)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        import traceback
        st.error(f"앱 실행 중 오류가 발생했습니다: {e}")
        st.code(traceback.format_exc())
