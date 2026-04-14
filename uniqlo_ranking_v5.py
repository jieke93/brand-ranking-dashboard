# -*- coding: utf-8 -*-
"""
유니클로 랭킹 크롤러 V5 (이미지 삽입 버전)
- 탭 클릭 로직 개선 (CSS 셀렉터 + JavaScript click)
- 이미지: 스크린샷 캡쳐 방식 (네트워크 요청 없이 빠름)
- openpyxl + Pillow 사용
"""
import sys
import io
import signal
import socket
import urllib.parse
import urllib.request
import urllib.robotparser
import requests
import threading
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from PIL import Image as PILImage

# KeyboardInterrupt 방지 (Python 3.14 호환성)
signal.signal(signal.SIGINT, signal.SIG_IGN)

# 전역 소켓 타임아웃 설정 (Selenium 통신에 영향주지 않도록 충분히 길게)
socket.setdefaulttimeout(120)

# 이미지 설정
IMG_WIDTH = 80  # 엑셀에 삽입할 이미지 너비 (픽셀)
IMG_HEIGHT = 107  # 3:4 비율로 계산된 높이
HD_IMG_WIDTH = 400   # 대시보드용 고해상도 이미지
HD_IMG_HEIGHT = 534  # 3:4 비율
ROW_HEIGHT = 85  # 행 높이 (포인트)
SKIP_IMAGES = False  # True로 설정하면 이미지 캡쳐 건너뛰기 (빠른 테스트용)
SAFE_MODE = False  # 법적 위험 최소화 모드 (robots.txt 준수)
REQUEST_DELAY = 1.5  # 요청 간격 (초)
IMG_DOWNLOAD_TIMEOUT = 3  # 이미지 다운로드 타임아웃 (초)
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

# 로그 파일 설정
LOG_FILE = "crawler_v5_log.txt"

def log(msg, end='\n'):
    """로그를 파일과 콘솔에 출력"""
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + end)
    try:
        print(msg, end=end, flush=True)
    except:
        pass

def check_robots_allowed(url):
    """robots.txt 기준으로 접근 가능 여부 확인"""
    if not SAFE_MODE:
        return True
    try:
        parsed = urllib.parse.urlparse(url)
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        rp = urllib.robotparser.RobotFileParser()
        with urllib.request.urlopen(robots_url, timeout=5) as resp:
            content = resp.read().decode('utf-8', 'ignore')
        rp.parse(content.splitlines())
        return rp.can_fetch(USER_AGENT, url)
    except Exception:
        # robots.txt를 확인할 수 없는 경우에는 허용으로 간주
        return True

def safe_get(driver, url):
    """robots.txt 준수 + 요청 간격 적용 후 페이지 로드"""
    if SAFE_MODE and not check_robots_allowed(url):
        log(f"  -> robots.txt 제한으로 접근 건너뜀: {url}")
        return False
    time.sleep(REQUEST_DELAY)
    driver.get(url)
    close_unexpected_windows(driver)
    return True


def _run_with_timeout(func, timeout_sec=10):
    """func를 별도 스레드에서 실행하고, timeout_sec 초 안에 끝나지 않으면 None 반환.
    Selenium 호출이 hang될 때 전체 크롤러가 멈추는 것을 방지."""
    result_box = [None]
    error_box = [None]
    def _worker():
        try:
            result_box[0] = func()
        except Exception as e:
            error_box[0] = e
    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=timeout_sec)
    if t.is_alive():
        log(f"  [TIMEOUT] Selenium 호출 {timeout_sec}초 타임아웃")
        return None  # 스레드가 살아있지만 daemon이므로 무시
    if error_box[0] is not None:
        raise error_box[0]
    return result_box[0]


def _safe_quit_driver(driver):
    """driver.quit()를 타임아웃 보호하여 호출. hang 시 프로세스 강제 종료."""
    import subprocess
    try:
        r = _run_with_timeout(lambda: driver.quit(), timeout_sec=10)
        if r is None:
            log("  [WARN] driver.quit() 타임아웃 → 프로세스 강제 종료")
            subprocess.run(['taskkill', '/F', '/IM', 'chromedriver.exe'], capture_output=True)
            subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'], capture_output=True)
    except Exception:
        try:
            subprocess.run(['taskkill', '/F', '/IM', 'chromedriver.exe'], capture_output=True)
            subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'], capture_output=True)
        except Exception:
            pass


def close_unexpected_windows(driver):
    """예상치 못한 새 탭/창(예: ftc.go.kr CAPTCHA)이 열리면 닫고 원래 탭으로 복귀"""
    def _inner():
        handles = driver.window_handles
        if len(handles) <= 1:
            return
        main_handle = handles[0]
        for handle in handles[1:]:
            try:
                driver.switch_to.window(handle)
                cur_url = driver.current_url or ''
                log(f"  -> 예상치 못한 창 감지, 닫는 중: {cur_url[:60]}")
                driver.close()
            except Exception:
                pass
        driver.switch_to.window(main_handle)
    try:
        _run_with_timeout(_inner, timeout_sec=8)
    except Exception:
        pass


def _norm_text(s):
    """탭/라벨 비교용 간단 정규화"""
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\u00a0', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    # 비교 안정화: 공백/기호 주변 차이 흡수
    s = s.replace(' & ', '&').replace('& ', '&').replace(' &', '&')
    return s


def _force_remove_onetrust_dom(driver):
    """OneTrust/쿠키 관련 배너/오버레이를 DOM에서 제거/숨김"""
    try:
        driver.execute_script(
            """
            const ids = [
              'onetrust-banner-sdk',
              'onetrust-consent-sdk',
              'onetrust-pc-sdk',
              'ot-sdk-btn-floating',
              'onetrust-policy',
              'onetrust-accept-btn-handler',
              'onetrust-reject-all-handler'
            ];
            for (const id of ids) {
              const el = document.getElementById(id);
              if (el) { try { el.remove(); } catch(e) { el.style.display='none'; el.style.visibility='hidden'; } }
            }

            const selectors = [
              '#onetrust-banner-sdk',
              '#onetrust-consent-sdk',
              '#onetrust-pc-sdk',
              '.onetrust-pc-dark-filter',
              '.ot-sdk-container',
              '.ot-overlay',
              '.ot-floating-button',
              '[class*="onetrust"]',
              '[id*="onetrust"]',
              '[class*="ot-sdk"]',
              '[id*="ot-sdk"]'
            ];
            document.querySelectorAll(selectors.join(',')).forEach(el => {
              try { el.remove(); } catch(e) { el.style.display='none'; el.style.visibility='hidden'; }
            });

            // 혹시 body 스크롤이 막혔으면 해제
            try { document.body.style.overflow = 'auto'; } catch(e) {}
            """
        )
    except Exception:
        pass


def _try_click_cookie_buttons_in_context(driver):
    """현재 컨텍스트(메인/iframe)에서 쿠키 버튼을 찾아 클릭 시도"""
    clicked = False

    # 1) OneTrust 표준 ID
    for btn_id in ['onetrust-accept-btn-handler', 'onetrust-reject-all-handler']:
        try:
            btn = driver.find_element(By.ID, btn_id)
            if btn and btn.is_displayed():
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.2)
                clicked = True
        except Exception:
            pass

    # 2) 텍스트 기반 (한국어/영문 혼용 대응)
    xpaths = [
        "//button[contains(., '동의') or contains(., '수락') or contains(., '확인') or contains(., 'Accept') or contains(., 'Agree') or contains(., 'OK') or contains(., '모두') or contains(., '거부') or contains(., 'Reject')]",
        "//a[contains(., '동의') or contains(., '수락') or contains(., '확인') or contains(., 'Accept') or contains(., 'Agree') or contains(., 'OK')]",
    ]
    for xp in xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
            for el in elems[:4]:
                try:
                    if el.is_displayed():
                        driver.execute_script("arguments[0].click();", el)
                        time.sleep(0.2)
                        clicked = True
                        break
                except Exception:
                    continue
        except Exception:
            pass

    return clicked


def _cookie_banner_present(driver):
    try:
        return bool(
            driver.execute_script(
                """
                const ids = ['onetrust-banner-sdk','onetrust-consent-sdk','onetrust-pc-sdk'];
                for (const id of ids) {
                  const el = document.getElementById(id);
                  if (el && el.offsetParent !== null) return true;
                }
                const any = document.querySelector('[class*="onetrust"], [id*="onetrust"], .ot-sdk-container, .onetrust-pc-dark-filter');
                return !!(any && any.offsetParent !== null);
                """
            )
        )
    except Exception:
        return False


def _find_product_tiles(driver):
    """페이지 구조 변경 대비: 여러 셀렉터로 상품 타일을 탐색"""
    selectors = [
        '.product-tile',
        "[data-testid='product-tile']",
        '.fr-ec-product-tile',
    ]
    best = []
    for sel in selectors:
        try:
            elems = driver.find_elements(By.CSS_SELECTOR, sel)
            if len(elems) > len(best):
                best = elems
        except Exception:
            continue
    return best


class BrowserCrashedError(Exception):
    """Chrome 창이 크래시/종료되어 더 이상 세션을 사용할 수 없을 때"""


def _is_driver_dead_error(exc) -> bool:
    msg = (str(exc) or '').lower()
    fatal_markers = [
        'no such window',
        'target window already closed',
        'web view not found',
        'session deleted',
        'invalid session id',
        'disconnected',
        'connectionreset',
        'connection reset',
        '10054',
    ]
    return any(m in msg for m in fatal_markers)

# 로그 파일 초기화
with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write("")

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import time
import os
import glob
import argparse
from datetime import datetime
import re

# 유니클로 컬러 코드 → 컬러명 매핑
COLOR_MAP = {
    '00': 'WHITE', '01': 'OFF WHITE', '02': 'LIGHT GRAY', '03': 'GRAY',
    '04': 'DARK GRAY', '05': 'LIGHT GRAY', '06': 'GRAY', '07': 'DARK GRAY',
    '08': 'DARK GRAY', '09': 'BLACK', '10': 'PINK', '11': 'PINK',
    '12': 'LIGHT PINK', '13': 'PINK', '14': 'PINK', '15': 'LIGHT PINK',
    '16': 'DARK PINK', '17': 'PURPLE', '18': 'PURPLE', '19': 'WINE',
    '20': 'ORANGE', '21': 'LIGHT ORANGE', '22': 'ORANGE', '23': 'BROWN',
    '24': 'DARK ORANGE', '25': 'ORANGE', '26': 'BROWN', '27': 'BROWN',
    '28': 'DARK BROWN', '29': 'BROWN', '30': 'NATURAL', '31': 'BEIGE',
    '32': 'BEIGE', '33': 'BEIGE', '34': 'LIGHT BEIGE', '35': 'BEIGE',
    '36': 'BROWN', '37': 'BROWN', '38': 'DARK BROWN', '39': 'KHAKI',
    '40': 'YELLOW', '41': 'LIGHT YELLOW', '42': 'YELLOW', '43': 'MUSTARD',
    '44': 'GOLD', '45': 'LIME', '46': 'OLIVE', '47': 'KHAKI', '48': 'OLIVE',
    '49': 'DARK GREEN', '50': 'LIGHT GREEN', '51': 'GREEN', '52': 'GREEN',
    '53': 'GREEN', '54': 'DARK GREEN', '55': 'GREEN', '56': 'GREEN',
    '57': 'GREEN', '58': 'MINT', '59': 'TURQUOISE', '60': 'LIGHT BLUE',
    '61': 'LIGHT BLUE', '62': 'SKY BLUE', '63': 'LIGHT BLUE', '64': 'BLUE',
    '65': 'BLUE', '66': 'BLUE', '67': 'BLUE', '68': 'BLUE', '69': 'NAVY',
    '70': 'DARK BLUE', '71': 'NAVY', '72': 'DARK BLUE', '73': 'NAVY',
    '74': 'DARK NAVY', '75': 'NAVY', '76': 'NAVY', '77': 'INDIGO',
    '78': 'INDIGO', '79': 'DENIM', '80': 'RED', '81': 'LIGHT RED',
    '82': 'RED', '83': 'RED', '84': 'DARK RED', '85': 'WINE', '86': 'WINE',
    '87': 'BURGUNDY', '88': 'DARK RED', '89': 'WINE', '90': 'SILVER',
    '91': 'GOLD', '92': 'MULTI', '93': 'MULTI', '94': 'PATTERN',
    '95': 'PATTERN', '96': 'STRIPE', '97': 'CHECK', '98': 'PRINT',
    '99': 'OTHER'
}

# ===== 아이템 타입 자동 분류 딕셔너리 =====
# 네이버 쇼핑 패션 카테고리 기준 + 유니클로 상품명 키워드 매핑
# 매칭 우선순위: 리스트 앞쪽 키워드가 먼저 매칭됨 (구체적 키워드 → 일반 키워드 순)
ITEM_TYPE_RULES = [
    # ── 아우터 ──
    ('경량패딩',     ['경량패딩', '경량 패딩', '라이트패딩', '라이트다운']),
    ('울트라라이트다운', ['울트라라이트다운', 'ultra light down', '울트라라이트 다운']),
    ('롱패딩',       ['롱패딩', '롱다운', '롱 패딩', '롱 다운']),
    ('숏패딩',       ['숏패딩', '숏 패딩', '쇼트패딩']),
    ('패딩',         ['패딩', '다운재킷', '다운 재킷', '퍼팩트다운', '퍼팩트 다운', '하이브리드다운']),
    ('플리스',       ['플리스', '후리스', 'fleece', '뽀글이']),
    ('트렌치코트',   ['트렌치코트', '트렌치 코트']),
    ('코트',         ['코트', '체스터', '울코트', '울 코트', '핸드메이드코트']),
    ('블레이저',     ['블레이저']),
    ('블루종',       ['블루종', 'MA-1', 'ma1', 'MA1']),
    ('재킷',         ['재킷', '자켓', '점퍼', 'jacket']),
    ('파카',         ['파카', 'parka']),
    ('카디건',       ['카디건', '가디건', 'cardigan']),
    ('바람막이',     ['바람막이', '윈드브레이커', '포켓터블', 'windbreaker']),
    ('베스트/조끼',  ['조끼', '베스트', '패딩베스트', '패딩조끼', 'vest']),

    # ── 상의 ──
    ('후드',         ['후드', '후디', '스웻후드', '풀짚후드', 'hoodie', 'hooded']),
    ('맨투맨',       ['맨투맨', '스웻셔츠', '스웨트셔츠', '스웻', 'sweatshirt', 'sweat']),
    ('니트/스웨터',  ['니트', '스웨터', '캐시미어', '메리노', '터틀넥', '크루넥니트', 'knit', 'sweater']),
    ('셔츠/블라우스',['셔츠', '블라우스', '옥스포드', '플란넬셔츠', '린넨셔츠', 'shirt', 'blouse']),
    ('폴로',         ['폴로', 'polo']),
    ('탱크탑',       ['탱크탑', '나시', '슬리브리스', 'tank top']),
    ('티셔츠',       ['티셔츠', 'T셔츠', '크루넥T', '반팔', '긴팔', 'UT ', 't-shirt', 'tee']),

    # ── 팬츠 ──
    ('진/데님',      ['진', '데님', '스키니', '와이드스트레이트', '와이드핏진', 'jeans', 'denim']),
    ('치노',         ['치노', 'chino']),
    ('카고팬츠',     ['카고', 'cargo']),
    ('조거팬츠',     ['조거', 'jogger']),
    ('스웻팬츠',     ['스웻팬츠', '이지팬츠', '이지 팬츠', '트레이닝팬츠']),
    ('슬랙스',       ['슬랙스', '앵클팬츠', '스마트앵클', '감탄팬츠']),
    ('레깅스',       ['레깅스', 'leggings']),
    ('숏팬츠',       ['숏팬츠', '쇼츠', '반바지', '버뮤다', '하프팬츠', 'shorts']),
    ('팬츠',         ['팬츠', '바지', 'pants', 'trousers']),

    # ── 원피스/스커트 ──
    ('원피스',       ['원피스', '드레스', 'dress', 'one piece']),
    ('스커트',       ['스커트', '치마', 'skirt']),

    # ── 이너웨어/기능성 ──
    ('히트텍',       ['히트텍', 'heattech', '히트택']),
    ('에어리즘',     ['에어리즘', 'airism', 'AIRism']),
    ('브라탑',       ['브라탑', '브라 탑', '브라캐미솔', 'bratop']),
    ('이너웨어',     ['팬티', '트렁크', '브리프', '속옷', '보정', '런닝']),

    # ── 홈웨어 ──
    ('파자마/라운지',['파자마', '잠옷', '라운지', '홈웨어', 'pajama', 'loungewear']),

    # ── 악세서리 ──
    ('양말',         ['양말', '삭스', 'socks']),
    ('모자',         ['모자', '캡', '버킷햇', '비니', 'hat', 'cap']),
    ('가방',         ['가방', '백팩', '토트', '숄더백', '에코백', 'bag']),
    ('머플러/스카프',['머플러', '스카프', '목도리', '숄', 'scarf', 'muffler']),
    ('벨트',         ['벨트', 'belt']),
    ('장갑',         ['장갑', '글러브', 'gloves']),
    ('우산',         ['우산', 'umbrella']),
    ('슬리퍼/샌들',  ['슬리퍼', '샌들', '룸슈즈', 'slipper', 'sandal']),
]

def classify_item_type(product_name, tab_name=''):
    """
    상품명 + 탭명을 분석하여 세부 아이템 타입을 자동 분류
    네이버 쇼핑 패션 카테고리 기준
    """
    if not product_name:
        return '미분류'
    
    text = (product_name + ' ' + tab_name).lower()
    
    for item_type, keywords in ITEM_TYPE_RULES:
        for kw in keywords:
            if kw.lower() in text:
                return item_type
    
    # 키워드 매칭 실패 시 탭명 기반 대분류
    tab_fallback = {
        '상의': '상의(기타)',
        '팬츠': '팬츠(기타)',
        '아우터': '아우터(기타)',
        '드레스 & 스커트': '원피스/스커트(기타)',
        '이너웨어': '이너웨어(기타)',
        '홈웨어': '홈웨어(기타)',
        '악세서리': '악세서리(기타)',
    }
    return tab_fallback.get(tab_name, '미분류')

# 수집할 URL + 탭 정보
CATEGORIES = {
    'WOMEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/women',
        'tabs': ['모두보기', '상의', '팬츠', '드레스 & 스커트', '아우터', '이너웨어', '홈웨어', '악세서리']
    },
    'MEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/men',
        'tabs': ['모두보기', '상의', '팬츠', '아우터', '이너웨어', '홈웨어', '악세서리']
    },
    'KIDS': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/kids',
        'tabs': ['모두보기', '상의', '팬츠', '아우터', '이너웨어']
    },
    'BABY': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/baby',
        'tabs': ['모두보기']  # BABY는 상품 수가 적어 모두보기만 사용
    }
}

def get_color_name(code):
    """컬러 코드를 컬러명으로 변환"""
    code = str(code).zfill(2)
    return COLOR_MAP.get(code, f'COLOR_{code}')

def _capture_image_inner(element, driver):
    """capture_image_from_element 내부 로직 (스레드 안에서 실행됨) - 간소화 버전"""
    # 요소를 뷰포트로 스크롤 + 오버레이 제거 (JS 한 번으로)
    if driver:
        driver.execute_script("""
            arguments[0].scrollIntoView({block: 'center'});
            var b = document.getElementById('onetrust-banner-sdk');
            if (b) b.style.display = 'none';
            var o = document.querySelector('.onetrust-pc-dark-filter');
            if (o) o.style.display = 'none';
        """, element)
        time.sleep(0.5)

    # PNG 스크린샷 캡처
    png_data = element.screenshot_as_png
    if not png_data:
        return None

    img = PILImage.open(io.BytesIO(png_data))
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')

    img_small = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
    buf_small = io.BytesIO()
    img_small.save(buf_small, format='JPEG', quality=85)
    buf_small.seek(0)

    img_hd = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
    buf_hd = io.BytesIO()
    img_hd.save(buf_hd, format='JPEG', quality=90)
    buf_hd.seek(0)

    return (buf_small, buf_hd)


def capture_image_from_element(element, driver=None):
    """요소를 스크린샷 캡처하여 (엑셀용, HD용) 튜플로 반환 (타임아웃 보호)"""
    if SKIP_IMAGES:
        return None

    try:
        # 전체 캡처를 15초 타임아웃으로 실행
        result = _run_with_timeout(
            lambda: _capture_image_inner(element, driver),
            timeout_sec=15
        )
        return result
    except Exception as e:
        if _is_driver_dead_error(e):
            raise BrowserCrashedError(str(e))
        return None

def _download_image_worker(url):
    """이미지 다운로드 실제 작업 (스레드에서 실행)"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    try:
        response = requests.get(url, headers=headers, timeout=(10, 15))
        if response.status_code == 200:
            img = PILImage.open(io.BytesIO(response.content))
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            hd_img = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
            hd_bytes = io.BytesIO()
            hd_img.save(hd_bytes, format='JPEG', quality=92)
            hd_bytes.seek(0)
            
            xl_img = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
            xl_bytes = io.BytesIO()
            xl_img.save(xl_bytes, format='JPEG', quality=80)
            xl_bytes.seek(0)
            
            return (xl_bytes, hd_bytes)
    except:
        pass
    return None

def download_image(url):
    """이미지 URL에서 이미지 다운로드 - 스레드 타임아웃 적용 (멈춤 방지)"""
    if SKIP_IMAGES:
        return None
    
    if not url or not url.startswith('http'):
        return None
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_download_image_worker, url)
            result = future.result(timeout=IMG_DOWNLOAD_TIMEOUT)
            return result
    except (FuturesTimeoutError, Exception):
        return None

def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)
    
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-dev-shm-usage')  # 메모리 문제 방지
    options.add_argument('--disable-extensions')  # 확장프로그램 비활성화
    options.add_argument('--disable-popup-blocking')  # 팝업 차단 비활성화 (직접 제어)
    options.add_argument('--disable-notifications')  # 알림 차단
    options.add_argument('--host-rules=MAP www.ftc.go.kr 127.0.0.1, MAP ftc.go.kr 127.0.0.1')  # ftc.go.kr 팝업 DNS 차단
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36')
    options.page_load_strategy = 'normal'  # 페이지 완전 로드 후 진행
    
    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작...")
    driver = webdriver.Chrome(service=service, options=options)
    
    # 페이지/스크립트 타임아웃 설정 (멈춤 방지)
    driver.set_page_load_timeout(90)  # 페이지 로드 타임아웃 증가
    driver.set_script_timeout(60)  # 스크립트 타임아웃 증가
    # implicitly_wait은 0으로 설정 (find_element가 블로킹되지 않도록)
    driver.implicitly_wait(0)
    
    # ftc.go.kr 팝업 방지: window.open 차단 (CDP)
    try:
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': 'window.open = function() { return null; };'
        })
    except Exception:
        pass
    
    # ftc.go.kr 네트워크 요청 자체를 차단 (CDP - enable 먼저!)
    try:
        driver.execute_cdp_cmd('Network.enable', {})
        driver.execute_cdp_cmd('Network.setBlockedURLs', {'urls': ['*ftc.go.kr*']})
    except Exception:
        pass
    
    log("  [OK] 완료!\n")
    return driver

def close_cookie_popup(driver):
    """쿠키 동의 팝업 닫기 (OneTrust 배너 완전 제거)"""
    try:
        # FAST PATH: 배너가 없으면 무거운 클릭/iframe 스캔을 하지 않음
        try:
            if not _cookie_banner_present(driver):
                _force_remove_onetrust_dom(driver)
                return False
        except Exception as e:
            if _is_driver_dead_error(e):
                raise BrowserCrashedError(str(e))

        clicked_any = False

        for _ in range(2):
            try:
                driver.switch_to.default_content()
            except Exception:
                pass

            # 메인 문서에서 클릭 시도
            if _try_click_cookie_buttons_in_context(driver):
                clicked_any = True

            # iframe 내부에서도 시도 (간헐적으로 iframe에 렌더링되는 케이스 대응)
            try:
                frames = driver.find_elements(By.CSS_SELECTOR, 'iframe')
            except Exception:
                frames = []

            for fr in frames[:6]:
                try:
                    driver.switch_to.default_content()
                    driver.switch_to.frame(fr)
                    if _try_click_cookie_buttons_in_context(driver):
                        clicked_any = True
                except Exception:
                    pass
                finally:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass

            # DOM 강제 제거/숨김
            _force_remove_onetrust_dom(driver)
            time.sleep(0.2)

            if not _cookie_banner_present(driver):
                return True

        # 배너가 남아있더라도, DOM 제거는 수행한 상태
        return clicked_any
        
    except Exception as e:
        if _is_driver_dead_error(e):
            raise BrowserCrashedError(str(e))
        pass
    return False

def get_available_tabs(driver):
    """페이지에서 실제 존재하는 탭 목록 가져오기"""
    available_tabs = ['모두보기']  # 모두보기는 항상 있음 (기본 상태)
    
    try:
        # a.fr-ec-tab 요소에서 탭 라벨 추출
        tab_links = driver.find_elements(By.CSS_SELECTOR, "a.fr-ec-tab")
        
        for tab in tab_links:
            try:
                label = tab.find_element(By.CSS_SELECTOR, "span.fr-ec-tab__label")
                label_text = label.text.strip()
                if label_text and label_text != '모두보기' and label_text not in available_tabs:
                    available_tabs.append(label_text)
            except:
                continue
        
        # swiper-slide 내부에서도 확인
        if len(available_tabs) <= 1:
            slides = driver.find_elements(By.CSS_SELECTOR, ".swiper-slide a.fr-ec-tab")
            for slide in slides:
                try:
                    label_text = slide.text.strip()
                    if label_text and label_text != '모두보기' and label_text not in available_tabs:
                        available_tabs.append(label_text)
                except:
                    continue
    except Exception as e:
        pass
    
    return available_tabs

def click_tab(driver, tab_name, timeout=10):
    """하위 탭 클릭 - 서브카테고리 탭 전환 + 상품 리로드 대기"""
    start_time = time.time()
    
    try:
        close_unexpected_windows(driver)
        log(f"      -> '{tab_name}' 탭 찾는 중...", end='')
        
        # 클릭 전 현재 상품 개수 기록 (탭 전환 감지용)
        old_tiles = _find_product_tiles(driver)
        old_count = len(old_tiles)
        # 첫 번째 상품의 텍스트를 기록 (같은 개수여도 내용 변경 감지)
        old_first_text = ''
        if old_tiles:
            try:
                old_first_text = old_tiles[0].text[:50]
            except:
                pass
        
        clicked = False
        
        # 방법 1: fr-ec-tab--small-height 클래스(서브 탭 전용)에서 텍스트 매칭
        sub_tabs = driver.find_elements(By.CSS_SELECTOR, 'a.fr-ec-tab.fr-ec-tab--small-height')
        for tab in sub_tabs:
            if time.time() - start_time > timeout:
                log(" 타임아웃")
                return False
            try:
                tab_text = _norm_text(tab.text)
                if tab_text == _norm_text(tab_name) or _norm_text(tab_name) in tab_text:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                    time.sleep(0.3)
                    close_cookie_popup(driver)
                    driver.execute_script("arguments[0].click();", tab)
                    log(" 클릭!", end='')
                    clicked = True
                    break
            except:
                continue
        
        # 방법 2: [role=tab]에서 서브카테고리(모두보기/상의/팬츠 등) 찾기
        if not clicked:
            log(" role=tab...", end='')
            all_role_tabs = driver.find_elements(By.CSS_SELECTOR, '[role="tab"]')
            for tab in all_role_tabs:
                if time.time() - start_time > timeout:
                    log(" 타임아웃")
                    return False
                try:
                    tab_text = _norm_text(tab.text)
                    tab_cls = tab.get_attribute('class') or ''
                    # 카테고리 탭(WOMEN/MEN 등)이 아닌 서브 탭만 대상
                    if tab_text == _norm_text(tab_name) and 'fr-ec-tab--boxed' not in tab_cls and tab_text not in ['WOMEN', 'MEN', 'KIDS', 'BABY']:
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                        time.sleep(0.3)
                        close_cookie_popup(driver)
                        driver.execute_script("arguments[0].click();", tab)
                        log(" 클릭!", end='')
                        clicked = True
                        break
                except:
                    continue
        
        # 방법 3: XPath로 정확한 텍스트 매칭
        if not clicked:
            log(" XPath...", end='')
            xpath_patterns = [
                f"//a[contains(@class, 'fr-ec-tab--small-height')]//span[text()='{tab_name}']/ancestor::a",
                f"//a[@role='tab'][not(contains(@class, 'boxed'))]//span[contains(text(), '{tab_name}')]/ancestor::a",
                f"//a[contains(@class, 'fr-ec-tab')][not(contains(@class, 'boxed'))][contains(., '{tab_name}')]",
            ]
            for xpath in xpath_patterns:
                if time.time() - start_time > timeout:
                    break
                try:
                    tabs = driver.find_elements(By.XPATH, xpath)
                    for tab in tabs:
                        if tab.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                            time.sleep(0.2)
                            close_cookie_popup(driver)
                            driver.execute_script("arguments[0].click();", tab)
                            log(" 클릭!", end='')
                            clicked = True
                            break
                except:
                    continue
                if clicked:
                    break
        
        if not clicked:
            log(" 실패")
            return False
        
        # === 탭 클릭 후 상품 리로드 대기 (핵심 수정) ===
        log(" 상품 로딩 대기...", end='')
        max_wait = 10  # 최대 10초 대기
        wait_start = time.time()
        loaded = False
        
        while time.time() - wait_start < max_wait:
            time.sleep(0.5)
            try:
                close_cookie_popup(driver)
                new_tiles = _find_product_tiles(driver)
                new_count = len(new_tiles)
                
                if new_count > 0:
                    # 상품이 있으면, 내용이 바뀌었는지 확인
                    new_first_text = ''
                    try:
                        new_first_text = new_tiles[0].text[:50]
                    except:
                        pass
                    
                    # 상품 개수가 달라지거나 첫 번째 상품 텍스트가 달라지면 로딩 완료
                    if new_count != old_count or new_first_text != old_first_text:
                        loaded = True
                        break
                    # 같은 개수+같은 내용이어도 1초 이상 지났으면 완료로 간주
                    if time.time() - wait_start > 2.0:
                        loaded = True
                        break
            except:
                continue
        
        if loaded:
            log(" OK!")
        else:
            log(" 대기완료(변경감지못함)")
        
        time.sleep(0.5)  # 렌더링 안정화
        return True
        
    except Exception as e:
        if _is_driver_dead_error(e):
            raise BrowserCrashedError(str(e))
        log(f" 오류: {e}")
        return False

# ── JS 일괄 추출용 스크립트 (skip_images 모드 전용) ───────────────
_JS_EXTRACT_ALL = """
var maxProducts = arguments[0];
var selectors = ['.product-tile', "[data-testid='product-tile']", '.fr-ec-product-tile'];
var tiles = [];
for (var s = 0; s < selectors.length; s++) {
    var found = document.querySelectorAll(selectors[s]);
    if (found.length > tiles.length) tiles = Array.from(found);
}
var results = [];
var limit = Math.min(maxProducts, tiles.length);
for (var i = 0; i < limit; i++) {
    var tile = tiles[i];
    var p = {rank: i+1, name:'', price:'', image_url:'', color_codes:[], rating:'없음', review_count:'없음'};

    /* 상품명 + 이미지 URL */
    var activeImg = tile.querySelector('.swiper-slide-active img.image__img');
    if (activeImg) {
        p.name = activeImg.getAttribute('alt') || '';
        p.image_url = activeImg.getAttribute('data-src') || activeImg.getAttribute('src') || '';
    }
    if (!p.image_url) {
        var imgs = tile.querySelectorAll('img.image__img');
        for (var j = 0; j < imgs.length; j++) {
            var alt = imgs[j].getAttribute('alt') || '';
            if (alt && !/^\\d+$/.test(alt)) {
                p.name = alt;
                var u = imgs[j].getAttribute('data-src') || imgs[j].getAttribute('src') || '';
                if (u) { p.image_url = u; break; }
            }
        }
    }
    if (!p.image_url) {
        var itoImgs = tile.querySelectorAll("[data-testid='ITOImage'] img");
        for (var j = 0; j < itoImgs.length; j++) {
            var u2 = itoImgs[j].getAttribute('data-src') || itoImgs[j].getAttribute('src');
            if (u2 && u2.indexOf('uniqlo') >= 0) {
                p.image_url = u2;
                if (!p.name) p.name = itoImgs[j].getAttribute('alt') || '';
                break;
            }
        }
    }
    if (!p.name) {
        var link = tile.querySelector('a.product-tile__link');
        if (link) p.name = link.textContent.trim().split('\\n')[0];
    }

    /* 가격 */
    var priceElems = tile.querySelectorAll("[data-testid='ITOTypography']");
    for (var j = 0; j < priceElems.length; j++) {
        var txt = priceElems[j].textContent.trim();
        if (txt.indexOf('원') >= 0 && txt.length < 20) { p.price = txt; break; }
    }
    if (!p.price) {
        var spans = tile.querySelectorAll('span');
        for (var j = 0; j < spans.length; j++) {
            var txt2 = spans[j].textContent.trim();
            if (txt2.indexOf('원') >= 0 && txt2.length < 20) { p.price = txt2; break; }
        }
    }

    /* 컬러 코드 */
    var chips = tile.querySelectorAll('.product-tile__image-chip-group-item img');
    for (var j = 0; j < chips.length; j++) {
        var ca = chips[j].getAttribute('alt');
        if (ca && /^\\d+$/.test(ca) && p.color_codes.indexOf(ca) < 0) p.color_codes.push(ca);
    }

    /* 평점/리뷰 */
    var re = tile.querySelector('.fr-ec-rating-static, [role="figure"]');
    if (re) {
        var rv = re.getAttribute('reviews');
        if (rv) p.review_count = rv;
        var fs = tile.querySelectorAll('.fr-ec-star--full').length;
        var hs = tile.querySelectorAll('.fr-ec-star--half').length;
        if (fs > 0) p.rating = String(fs + hs * 0.5);
    }
    if (p.rating === '없음') {
        var rt = tile.querySelector('.fr-ec-rating-average-product-tile');
        if (rt) { var m = rt.textContent.match(/(\\d+\\.?\\d*)/); if (m) p.rating = m[1]; }
    }

    if (p.name) results.push(p);
}
return results;
"""


def _extract_products_fast_js(driver, max_products=30):
    """JavaScript 한 번 호출로 상품 데이터를 일괄 추출 (Phase 1 skip_images 전용).
    개별 Selenium find_element/get_attribute 호출을 제거하여 hang 방지."""
    log("      [DEBUG] JS 일괄 추출 모드")

    # 빠른 스크롤 (lazy loading 트리거)
    r = _run_with_timeout(
        lambda: driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight); return true;"
        ), timeout_sec=10)
    if r is None:
        raise BrowserCrashedError("스크롤 타임아웃 - 브라우저 응답 없음")

    threading.Event().wait(timeout=1.5)

    _run_with_timeout(
        lambda: driver.execute_script("window.scrollTo(0, 0); return true;"),
        timeout_sec=10)

    # JS 일괄 추출
    raw = _run_with_timeout(
        lambda: driver.execute_script(_JS_EXTRACT_ALL, max_products),
        timeout_sec=20)

    if raw is None:
        raise BrowserCrashedError("JS 추출 타임아웃 - 브라우저 응답 없음")

    products = []
    for item in raw:
        color_codes = item.get('color_codes', [])
        color_names = []
        seen = set()
        for c in color_codes:
            name = get_color_name(c)
            if name not in seen:
                color_names.append(name)
                seen.add(name)
        products.append({
            'rank': item['rank'],
            'name': item['name'],
            'price': item.get('price', ''),
            'item_type': '미분류',
            'color_count': len(color_names),
            'colors': ', '.join(color_names) if color_names else '정보없음',
            'rating': item.get('rating', '없음'),
            'review_count': item.get('review_count', '없음'),
            'image_url': item.get('image_url', ''),
            'image_data': None,
        })

    log(f"      -> {len(products)}개 수집 (JS 일괄)")
    return products


def extract_products(driver, max_products=30, skip_images=False):
    """상품 데이터 추출 - skip_images=True이면 JS 일괄 추출 (빠르고 안전)"""
    if skip_images:
        return _extract_products_fast_js(driver, max_products)

    products = []
    log("      [DEBUG] 상품 추출 시작...")
    
    # 상품 타일이 로드될 때까지 대기 (탭 전환 후 DOM 갱신 대기)
    log("      [DEBUG] 상품 로딩 대기 중...")
    product_tiles = []
    for wait_i in range(10):  # 최대 5초 대기
        close_unexpected_windows(driver)
        close_cookie_popup(driver)
        product_tiles = _find_product_tiles(driver)
        if len(product_tiles) > 0:
            break
        time.sleep(0.5)
    
    # 스크롤하여 상품 로드 (lazy-load 이미지 트리거)
    try:
        close_unexpected_windows(driver)
        if skip_images:
            log("      [DEBUG] 빠른 스크롤 (이미지 건너뜀)")
            # 이미지 불필요 → 빠르게 스크롤만 하고 대기 최소화
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            threading.Event().wait(timeout=1.5)
            driver.execute_script("window.scrollTo(0, 0);")
        else:
            log("      [DEBUG] 스크롤 시작 (이미지 로딩 대기)")
            driver.execute_script("""
                (function() {
                    var steps = 3;
                    var i = 0;
                    function doScroll() {
                        if (i < steps) {
                            window.scrollTo(0, document.body.scrollHeight);
                            i++;
                            setTimeout(doScroll, 1000);
                        } else {
                            window.scrollTo(0, 0);
                        }
                    }
                    doScroll();
                })();
            """)
            WebDriverWait(driver, 10).until(lambda d: True)
            threading.Event().wait(timeout=5.0)
        close_unexpected_windows(driver)
        log("      [DEBUG] 스크롤 완료")
    except BaseException as e:
        log(f"      [WARN] 스크롤 오류 (브라우저 재시작 필요): {type(e).__name__}: {e}")
        raise BrowserCrashedError(str(e))
    
    # 상품 타일 다시 찾기 (스크롤 후 추가 로드될 수 있음)
    log("      [DEBUG] 상품 타일 검색 시작")
    try:
        close_cookie_popup(driver)
        product_tiles = _find_product_tiles(driver)
        log(f"      [DEBUG] 상품 타일 {len(product_tiles)}개 발견")
    except Exception as e:
        if _is_driver_dead_error(e):
            raise BrowserCrashedError(str(e))
        log(f"      [ERROR] 상품 타일 찾기 실패: {e}")
        return []
    
    if not product_tiles:
        log("      [DEBUG] 상품 타일 없음, 반환")
        return []
    
    log(f"      [DEBUG] {min(max_products, len(product_tiles))}개 상품 처리 시작")
    for idx, tile in enumerate(product_tiles[:max_products], 1):
        # 이미지 모드일 때만 매 5개마다 예상치 못한 창 닫기 (속도 최적화)
        if not skip_images and idx % 5 == 1:
            close_unexpected_windows(driver)
        try:
            product = {
                'rank': idx,
                'name': '',
                'price': '',
                'item_type': '미분류',  # 세부 아이템 타입 (자동 분류)
                'color_count': 0,
                'colors': '',
                'rating': '없음',
                'review_count': '없음',
                'image_url': '',
                'image_data': None  # 실제 이미지 데이터
            }
            
            # 상품명 + 이미지 URL (개선: data-src 우선 확인)
            try:
                # 활성 슬라이드 내 이미지 우선 시도
                try:
                    active_img = tile.find_element(By.CSS_SELECTOR, ".swiper-slide-active img.image__img")
                    product['name'] = active_img.get_attribute("alt") or ""
                    # data-src 먼저, 없으면 src
                    img_url = active_img.get_attribute("data-src") or active_img.get_attribute("src") or ""
                    product['image_url'] = img_url
                except:
                    pass
                
                # 활성 슬라이드 못 찾으면 일반 이미지에서
                if not product['image_url']:
                    imgs = tile.find_elements(By.CSS_SELECTOR, "img.image__img")
                    for img in imgs:
                        alt = img.get_attribute("alt") or ""
                        # 컬러코드 이미지(숫자만 있는 alt)가 아닌 것 찾기
                        if alt and not alt.isdigit():
                            product['name'] = alt
                            img_url = img.get_attribute("data-src") or img.get_attribute("src") or ""
                            if img_url:
                                product['image_url'] = img_url
                                break
                
                # 그래도 없으면 첫 번째 이미지 사용
                if not product['image_url']:
                    all_imgs = tile.find_elements(By.CSS_SELECTOR, "[data-testid='ITOImage'] img")
                    for img in all_imgs:
                        img_url = img.get_attribute("data-src") or img.get_attribute("src")
                        if img_url and "uniqlo.com" in img_url:
                            product['image_url'] = img_url
                            product['image_element'] = img  # 이미지 요소 저장
                            if not product['name']:
                                product['name'] = img.get_attribute("alt") or ""
                            break
                
                # 이미지 스크린샷 캡쳐 (네트워크 요청 없이 빠름!)
                if not SKIP_IMAGES and not skip_images:
                    img_elem = product.get('image_element')
                    if not img_elem:
                        # 이미지 요소 찾기
                        try:
                            img_elem = tile.find_element(By.CSS_SELECTOR, ".swiper-slide-active img.image__img")
                        except:
                            try:
                                img_elem = tile.find_element(By.CSS_SELECTOR, "img.image__img")
                            except:
                                pass
                    
                    if img_elem:
                        img_data = capture_image_from_element(img_elem, driver)
                        if img_data:
                            product['image_data'] = img_data[0]      # 엑셀용
                            product['hd_image_data'] = img_data[1]   # 대시보드용 고해상도
                            
            except Exception as e:
                if _is_driver_dead_error(e):
                    raise BrowserCrashedError(str(e))
                pass
            
            # 상품명 백업 - 링크 텍스트에서
            if not product['name']:
                try:
                    link = tile.find_element(By.CSS_SELECTOR, "a.product-tile__link")
                    product['name'] = link.text.strip().split('\n')[0]
                except:
                    pass
            
            # 가격 - ITOTypography div에서 "원" 포함 텍스트
            try:
                price_elements = tile.find_elements(By.CSS_SELECTOR, "[data-testid='ITOTypography']")
                for elem in price_elements:
                    txt = elem.text.strip()
                    if '원' in txt and len(txt) < 20:
                        product['price'] = txt
                        break
            except:
                pass
            
            # 가격 백업 - span 태그에서
            if not product['price']:
                try:
                    spans = tile.find_elements(By.TAG_NAME, "span")
                    for span in spans:
                        txt = span.text.strip()
                        if '원' in txt and len(txt) < 20:
                            product['price'] = txt
                            break
                except:
                    pass
            
            # 컬러 정보 - 칩 이미지에서 컬러 코드 추출 후 컬러명 변환
            try:
                color_chips = tile.find_elements(By.CSS_SELECTOR, ".product-tile__image-chip-group-item img")
                color_codes = []
                for chip in color_chips:
                    alt = chip.get_attribute("alt")
                    if alt and alt.isdigit():
                        color_name = get_color_name(alt)
                        if color_name not in color_codes:
                            color_codes.append(color_name)
                
                product['color_count'] = len(color_codes)
                product['colors'] = ', '.join(color_codes) if color_codes else '정보없음'
            except:
                product['color_count'] = 0
                product['colors'] = '정보없음'
            
            # 평점 - reviews 속성에서
            try:
                rating_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-static, [role='figure']")
                reviews_attr = rating_elem.get_attribute("reviews")
                if reviews_attr:
                    product['review_count'] = reviews_attr
                
                # 별점 계산 (full star 개수)
                full_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--full")
                half_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--half")
                product['rating'] = str(len(full_stars) + len(half_stars) * 0.5) if full_stars else '없음'
            except:
                pass
            
            # 평점 텍스트에서 추출 시도
            if product['rating'] == '없음':
                try:
                    rating_text_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-average-product-tile")
                    rating_text = rating_text_elem.text.strip()
                    rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                    if rating_match:
                        product['rating'] = rating_match.group(1)
                except:
                    pass
            
            if product['name']:
                products.append(product)
                
        except Exception as e:
            if _is_driver_dead_error(e):
                raise BrowserCrashedError(str(e))
            continue
    
    return products

def classify_products(products, tab_name=''):
    """수집된 상품 리스트에 아이템 타입 자동 분류 적용"""
    for p in products:
        p['item_type'] = classify_item_type(p['name'], tab_name)
    return products

def scrape_category_with_tabs(driver, category, url, tabs, skip_images=False):
    """카테고리와 모든 탭 크롤링 (skip_images=True이면 이미지 없이 빠르게)"""
    all_data = {}
    
    log(f"\n{'='*60}")
    log(f"[수집] {category} 카테고리")
    log(f"{'='*60}")
    log(f"  URL: {url}")
    log(f"  탭: {', '.join(tabs)}")
    
    # 메인 페이지 접속
    try:
        if not safe_get(driver, url):
            return {}
        log(f"  -> 페이지 로딩 대기...", end='')
        # 대기 시간 단축 (eager 모드라 DOM 로드되면 바로 진행)
        for i in range(5):
            time.sleep(0.8)
            print(".", end='', flush=True)
        log(" OK!")
        
        # 쿠키 팝업 닫기
        if close_cookie_popup(driver):
            log("  -> 쿠키 팝업 닫음")
            time.sleep(0.5)
        # 예상치 못한 새 창(ftc.go.kr 등) 닫기
        close_unexpected_windows(driver)
    except BrowserCrashedError:
        raise
    except Exception as e:
        if _is_driver_dead_error(e):
            raise BrowserCrashedError(str(e))
        log(f" 오류: {e}")
        return {}
    
    # 실제 존재하는 탭만 필터링 (라벨 미세 변경 대응: 정규화 비교)
    close_unexpected_windows(driver)
    available_tabs = get_available_tabs(driver)

    def _tab_exists(desired, avail_list):
        nd = _norm_text(desired)
        for a in avail_list:
            if _norm_text(a) == nd:
                return True
        return False

    actual_tabs = [t for t in tabs if t == '모두보기' or _tab_exists(t, available_tabs)]
    
    if len(actual_tabs) < len(tabs):
        missing = set(tabs) - set(actual_tabs)
        log(f"  -> 일부 탭 없음: {', '.join(missing)} (건너뜀)")
        log(f"  -> 실제 탭: {', '.join(actual_tabs)}")
    
    for tab_idx, tab_name in enumerate(actual_tabs):
        log(f"\n  [{tab_idx+1}/{len(actual_tabs)}] 탭: {tab_name}")
        
        # 첫 번째 탭(모두보기)이 아니면 탭 클릭
        if tab_idx > 0:
            # 탭 클릭 전 쿠키 팝업 재확인 (다른 카테고리 전환 후 다시 나타날 수 있음)
            close_cookie_popup(driver)
            close_unexpected_windows(driver)
            if not click_tab(driver, tab_name):
                log(f"      -> 탭 클릭 실패, 건너뜀")
                continue
            time.sleep(1)  # click_tab 내부에서 이미 상품 로딩 대기함
            close_unexpected_windows(driver)
        
        # 상품 추출 (try-except로 개별 탭 오류 처리)
        try:
            products = extract_products(driver, max_products=30, skip_images=skip_images)
            
            # 아이템 타입 자동 분류 적용
            if products:
                products = classify_products(products, tab_name)
                sheet_name = f"{category}_{tab_name}"
                all_data[sheet_name] = products
                
                # 이미지 다운로드 성공 개수 계산
                img_count = sum(1 for p in products if p.get('image_data'))
                if skip_images:
                    log(f"      -> {len(products)}개 수집 (이미지: Phase 2/3에서 처리)")
                else:
                    log(f"      -> {len(products)}개 수집 (이미지: {img_count}개)")
                
                # 샘플 출력 (첫 2개만)
                if len(products) >= 2:
                    for i, p in enumerate(products[:2], 1):
                        name_short = p['name'][:15] if len(p['name']) > 15 else p['name']
                        has_img = "O" if p['image_data'] else "X"
                        log(f"        {i}. {name_short:15s} | {p['price']:10s} | img:{has_img}")
            else:
                log(f"      -> 상품 없음")
        except BrowserCrashedError:
            raise
        except Exception as e:
            log(f"      -> 추출 오류: {str(e)[:30]}")
        
        time.sleep(0.8)
    
    return all_data


def _find_latest_uniqlo_v5_excel(work_dir):
    pattern = os.path.join(work_dir, '유니클로_전체랭킹_이미지포함_V5_*.xlsx')
    files = sorted(glob.glob(pattern), reverse=True)
    return files[0] if files else None


def _load_sheets_from_excel(filepath, skip_prefixes=None):
    """기존 엑셀에서 시트 데이터를 로드 (이미지는 보존하지 않고 값만)"""
    skip_prefixes = skip_prefixes or []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        log(f"  [WARN] 기존 엑셀 로드 실패: {e}")
        return {}

    loaded = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        if any(sheet_name.startswith(p) for p in skip_prefixes):
            continue
        ws = wb[sheet_name]
        products = []
        for r in range(2, ws.max_row + 1):
            rank_val = ws.cell(r, 1).value
            if rank_val is None:
                continue
            try:
                rank_int = int(rank_val)
            except Exception:
                rank_int = r - 1

            name = ws.cell(r, 3).value or ''
            if not str(name).strip():
                continue

            product = {
                'rank': rank_int,
                'name': str(name).strip(),
                'item_type': (ws.cell(r, 4).value or '미분류'),
                'price': str(ws.cell(r, 5).value or ''),
                'color_count': ws.cell(r, 6).value or 0,
                'colors': str(ws.cell(r, 7).value or ''),
                'rating': str(ws.cell(r, 8).value or '없음'),
                'review_count': str(ws.cell(r, 9).value or '없음'),
                'image_url': '',
                'image_data': None,
                'hd_image_data': None,
            }
            products.append(product)

        if products:
            loaded[sheet_name] = products
    return loaded

def create_excel(all_data, filename):
    """이미지가 포함된 엑셀 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성 (이미지 포함)")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")
    
    wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, products in all_data.items():
        # 시트명 길이 제한 (Excel은 31자까지)
        safe_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(safe_name)
        log(f"  -> 시트 [{safe_name}]: {len(products)}개 상품 (이미지 삽입 중...)")
        
        # 헤더 (이미지 + 아이템타입 컬럼 추가)
        headers = ['순위', '이미지', '상품명', '아이템타입', '가격', '컬러수', '컬러목록', '평점', '리뷰수']
        ws.append(headers)
        
        for col in range(1, 10):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 6   # 순위
        ws.column_dimensions['B'].width = 12  # 이미지
        ws.column_dimensions['C'].width = 35  # 상품명
        ws.column_dimensions['D'].width = 14  # 아이템타입
        ws.column_dimensions['E'].width = 12  # 가격
        ws.column_dimensions['F'].width = 8   # 컬러수
        ws.column_dimensions['G'].width = 30  # 컬러목록
        ws.column_dimensions['H'].width = 8   # 평점
        ws.column_dimensions['I'].width = 8   # 리뷰수
        
        # 데이터 행 추가
        for row_idx, p in enumerate(products, 2):
            ws.append([
                p['rank'], '',  # 이미지 셀은 비워둠
                p['name'], p['item_type'],
                p['price'], p['color_count'],
                p['colors'], p['rating'],
                p['review_count']
            ])
            
            # 행 높이 설정
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
            
            # 이미지 삽입
            if p['image_data']:
                try:
                    img = XLImage(p['image_data'])
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    # B열에 이미지 삽입
                    cell_ref = f'B{row_idx}'
                    ws.add_image(img, cell_ref)
                except Exception as e:
                    log(f"    [WARN] 이미지 삽입 실패 row {row_idx}: {e}")
            
            # 셀 스타일
            for col in range(1, 10):
                cell = ws.cell(row_idx, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
    
    wb.save(filename)
    log(f"  [OK] 저장 완료!")
    
    # 대시보드용 고해상도 이미지 저장
    _save_hd_images(all_data, '유니클로', filename)
    
    return filename


def _normalize_product_name(name):
    """이미지 매칭용 상품명 정규화 (특수문자·공백·괄호 내용 제거)"""
    s = re.sub(r'\(.*?\)', '', name)           # 괄호 안 내용 제거
    s = re.sub(r'[^\w]', '', s)                # 특수문자·공백 제거
    return s.lower().strip()


def _build_existing_image_index(hd_dir):
    """product_images_hd/ 폴더에서 유니클로 이미지를 상품명 → 파일경로 딕셔너리로 구축"""
    index = {}  # { 정규화된_상품명: 파일경로 }
    if not os.path.isdir(hd_dir):
        return index
    for fname in os.listdir(hd_dir):
        if not fname.endswith('.jpg') or '유니클로' not in fname:
            continue
        # 파일명: {hash}_유니클로_{category}_{tab}_{rank}_{상품명}.jpg
        parts = fname.rsplit('.', 1)[0].split('_', 5)
        if len(parts) >= 6:
            product_name_part = parts[5]
            norm = _normalize_product_name(product_name_part)
            if norm and norm not in index:
                index[norm] = os.path.join(hd_dir, fname)
    return index


def _load_image_for_excel(file_path):
    """HD 이미지 파일 → (엑셀용 BytesIO, HD용 BytesIO) 튜플"""
    try:
        img = PILImage.open(file_path)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        img_small = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        buf_small = io.BytesIO()
        img_small.save(buf_small, format='JPEG', quality=85)
        buf_small.seek(0)

        img_hd = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        buf_hd = io.BytesIO()
        img_hd.save(buf_hd, format='JPEG', quality=90)
        buf_hd.seek(0)

        return (buf_small, buf_hd)
    except Exception:
        return None


def _reuse_existing_images(all_data, hd_dir):
    """Phase 2: 기존 HD 이미지에서 동일 상품명 이미지를 재활용"""
    log("\n  [Phase 2] 기존 이미지 매칭 중...")
    image_index = _build_existing_image_index(hd_dir)
    log(f"  -> 기존 이미지 인덱스: {len(image_index)}개 상품")
    
    matched = 0
    missing_products = []  # (sheet_name, product_index, product) 튜플 리스트
    
    for sheet_name, products in all_data.items():
        for i, p in enumerate(products):
            if p.get('image_data'):
                continue  # 이미 이미지가 있음
            norm_name = _normalize_product_name(p.get('name', ''))
            if norm_name in image_index:
                result = _load_image_for_excel(image_index[norm_name])
                if result:
                    p['image_data'] = result[0]
                    p['hd_image_data'] = result[1]
                    matched += 1
                else:
                    missing_products.append((sheet_name, i, p))
            else:
                missing_products.append((sheet_name, i, p))
    
    total = sum(len(prods) for prods in all_data.values())
    log(f"  -> 기존 이미지 매칭: {matched}/{total}개 재활용")
    log(f"  -> 이미지 누락: {len(missing_products)}개 (다운로드 필요)")
    
    return missing_products


def _save_hd_image_immediately(prod, sheet_name, hd_dir):
    """Phase 3 도중 HD 이미지를 즉시 디스크에 저장 (다음 실행 시 Phase 2에서 재활용)"""
    hd_data = prod.get('hd_image_data')
    if not hd_data:
        return
    try:
        name = prod.get('name', '')[:20]
        safe = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
        fname = f"p3_{sheet_name}_{prod['rank']}_{safe}.jpg"
        fpath = os.path.join(hd_dir, fname)
        if not os.path.exists(fpath):
            hd_data.seek(0)
            with open(fpath, 'wb') as f:
                f.write(hd_data.read())
    except Exception:
        pass


def _upgrade_image_url(url):
    """유니클로 CDN URL에서 고해상도 이미지 URL로 변환"""
    if not url:
        return url
    # /w/100 → /w/600 등 해상도 파라미터 업그레이드
    upgraded = re.sub(r'/w/\d+', '/w/600', url)
    upgraded = re.sub(r'/h/\d+', '/h/600', upgraded)
    # 쿼리 파라미터의 width/height도 업그레이드
    upgraded = re.sub(r'width=\d+', 'width=600', upgraded)
    upgraded = re.sub(r'height=\d+', 'height=600', upgraded)
    return upgraded


def _download_image_direct(url):
    """이미지 URL에서 직접 다운로드하여 (엑셀용, HD용) 튜플 반환"""
    if not url or not url.startswith('http'):
        return None
    try:
        # 고해상도 URL로 업그레이드
        hd_url = _upgrade_image_url(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36',
            'Referer': 'https://www.uniqlo.com/kr/ko/',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
        }
        response = requests.get(hd_url, headers=headers, timeout=(10, 15))
        if response.status_code != 200:
            # 고해상도 실패 시 원본 URL로 재시도
            response = requests.get(url, headers=headers, timeout=(10, 15))
        if response.status_code != 200:
            return None
        
        img = PILImage.open(io.BytesIO(response.content))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # 엑셀용 
        img_small = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        buf_small = io.BytesIO()
        img_small.save(buf_small, format='JPEG', quality=85)
        buf_small.seek(0)
        
        # HD용
        img_hd = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        buf_hd = io.BytesIO()
        img_hd.save(buf_hd, format='JPEG', quality=92)
        buf_hd.seek(0)
        
        return (buf_small, buf_hd)
    except Exception:
        return None


def _screenshot_missing_images(driver, all_data, missing_products):
    """Phase 3: 누락 이미지를 URL 직접 다운로드로 수집 (스크린샷 fallback)
    - 1단계: image_url로 HTTP 직접 다운로드 (정확도 100%, 팝업 영향 없음)
    - 2단계: 다운로드 실패 건만 Selenium 스크린샷 fallback
    - HD 이미지를 즉시 디스크 저장
    - driver를 반환 (재시작될 수 있으므로)
    """
    if not missing_products:
        log("\n  [Phase 3] 누락 이미지 없음 - 건너뜀")
        return driver
    
    log(f"\n  [Phase 3] 누락 이미지 {len(missing_products)}개 처리 시작")
    
    hd_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'product_images_hd')
    os.makedirs(hd_dir, exist_ok=True)
    
    # ── 1단계: image_url로 HTTP 직접 다운로드 ──
    log("\n    [3-1] 이미지 URL 직접 다운로드 시도...")
    downloaded = 0
    still_missing = []  # 다운로드 실패 목록
    
    for sheet_name, prod_idx, prod in missing_products:
        img_url = prod.get('image_url', '')
        if img_url:
            result = _download_image_direct(img_url)
            if result:
                prod['image_data'] = result[0]
                prod['hd_image_data'] = result[1]
                _save_hd_image_immediately(prod, sheet_name, hd_dir)
                downloaded += 1
            else:
                still_missing.append((sheet_name, prod_idx, prod))
        else:
            still_missing.append((sheet_name, prod_idx, prod))
    
    log(f"    -> URL 다운로드 성공: {downloaded}/{len(missing_products)}개")
    
    if not still_missing:
        log(f"\n  [Phase 3 완료] 전체 {downloaded}개 다운로드 성공 (스크린샷 불필요!)")
        return driver
    
    # ── 2단계: 실패 건만 Selenium 스크린샷 fallback ──
    log(f"\n    [3-2] 스크린샷 fallback: {len(still_missing)}개 남음")
    
    from collections import defaultdict
    by_sheet = defaultdict(list)
    for sheet_name, prod_idx, prod in still_missing:
        by_sheet[sheet_name].append((prod_idx, prod))
    
    captured_total = 0
    
    for sheet_name, items in by_sheet.items():
        parts = sheet_name.split('_', 1)
        if len(parts) != 2:
            continue
        category, tab_name = parts
        if category not in CATEGORIES:
            continue
        
        info = CATEGORIES[category]
        url = info['url']
        
        log(f"\n    [{sheet_name}] {len(items)}개 누락 → 페이지 방문")
        
        try:
            if not safe_get(driver, url):
                log(f"      -> safe_get 실패, Chrome 재시작")
                _safe_quit_driver(driver)
                driver = setup_driver()
                if not safe_get(driver, url):
                    log(f"      -> 재시작 후에도 실패, 건너뜀")
                    continue
            
            for _ in range(5):
                time.sleep(0.8)
            
            close_cookie_popup(driver)
            close_unexpected_windows(driver)
            
            alive = _run_with_timeout(lambda: driver.execute_script("return true;"), timeout_sec=10)
            if alive is None:
                log(f"      -> 드라이버 응답 없음, Chrome 재시작")
                _safe_quit_driver(driver)
                driver = setup_driver()
                if not safe_get(driver, url):
                    continue
                time.sleep(4)
                alive = _run_with_timeout(lambda: driver.execute_script("return true;"), timeout_sec=10)
                if alive is None:
                    continue
            
            if tab_name != '모두보기':
                if not click_tab(driver, tab_name):
                    log(f"      -> 탭 클릭 실패, 건너뜀")
                    continue
                time.sleep(1)
                close_unexpected_windows(driver)
            
            # 스크롤하여 이미지 로드
            scroll_r = _run_with_timeout(lambda: driver.execute_script("""
                (function() {
                    var steps = 3, i = 0;
                    function doScroll() {
                        if (i < steps) {
                            window.scrollTo(0, document.body.scrollHeight);
                            i++;
                            setTimeout(doScroll, 1000);
                        } else { window.scrollTo(0, 0); }
                    }
                    doScroll();
                })(); return true;
            """), timeout_sec=15)
            if scroll_r is None:
                _safe_quit_driver(driver)
                driver = setup_driver()
                continue
            
            threading.Event().wait(timeout=5.0)
            close_unexpected_windows(driver)
            
            product_tiles = _run_with_timeout(lambda: _find_product_tiles(driver), timeout_sec=10)
            if product_tiles is None:
                _safe_quit_driver(driver)
                driver = setup_driver()
                continue
            if not product_tiles:
                continue
            
            captured = 0
            consecutive_fail = 0
            for prod_idx, prod in items:
                if consecutive_fail >= 3:
                    log(f"      -> 연속 3회 스크린샷 실패, Chrome 재시작")
                    _safe_quit_driver(driver)
                    driver = setup_driver()
                    break
                
                rank = prod.get('rank', 0)
                if rank < 1 or rank > len(product_tiles):
                    continue
                
                tile = product_tiles[rank - 1]
                
                def _find_img(t=tile):
                    try:
                        return t.find_element(By.CSS_SELECTOR, ".swiper-slide-active img.image__img")
                    except Exception:
                        try:
                            return t.find_element(By.CSS_SELECTOR, "img.image__img")
                        except Exception:
                            return None
                
                img_elem = _run_with_timeout(_find_img, timeout_sec=10)
                
                if img_elem:
                    # fallback: 먼저 요소에서 src URL을 추출해 다운로드 시도
                    try:
                        src_url = img_elem.get_attribute('data-src') or img_elem.get_attribute('src') or ''
                        if src_url:
                            dl_result = _download_image_direct(src_url)
                            if dl_result:
                                prod['image_data'] = dl_result[0]
                                prod['hd_image_data'] = dl_result[1]
                                _save_hd_image_immediately(prod, sheet_name, hd_dir)
                                captured += 1
                                consecutive_fail = 0
                                continue
                    except Exception:
                        pass
                    
                    # 최후 수단: 요소 스크린샷
                    img_data = capture_image_from_element(img_elem, driver)
                    if img_data:
                        prod['image_data'] = img_data[0]
                        prod['hd_image_data'] = img_data[1]
                        _save_hd_image_immediately(prod, sheet_name, hd_dir)
                        captured += 1
                        consecutive_fail = 0
                    else:
                        consecutive_fail += 1
                else:
                    consecutive_fail += 1
            
            captured_total += captured
            log(f"      -> {captured}/{len(items)}개 캡처 완료")
            
        except BrowserCrashedError:
            log(f"      -> BrowserCrashedError, Chrome 재시작")
            try:
                _safe_quit_driver(driver)
            except Exception:
                pass
            driver = setup_driver()
            continue
        except Exception as e:
            if _is_driver_dead_error(e):
                try:
                    _safe_quit_driver(driver)
                except Exception:
                    pass
                driver = setup_driver()
                continue
            log(f"      -> 오류: {str(e)[:50]}")
    
    log(f"\n  [Phase 3 완료] URL 다운로드 {downloaded}개 + 스크린샷 {captured_total}개 = 총 {downloaded + captured_total}/{len(missing_products)}개")
    return driver


def _save_hd_images(all_data, brand_name, excel_filename):
    """대시보드용 고해상도 이미지를 product_images_hd/ 폴더에 저장"""
    import hashlib
    hd_dir = os.path.join(os.path.dirname(os.path.abspath(excel_filename)), 'product_images_hd')
    os.makedirs(hd_dir, exist_ok=True)
    
    file_hash = hashlib.md5(f"{os.path.basename(excel_filename)}_{os.path.getmtime(excel_filename)}".encode()).hexdigest()[:8]
    saved = 0
    for sheet_name, products in all_data.items():
        for p in products:
            hd_data = p.get('hd_image_data')
            if hd_data:
                name = p.get('name', '')[:20]
                safe = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
                fname = f"{file_hash}_{brand_name}_{sheet_name}_{p['rank']}_{safe}.jpg"
                fpath = os.path.join(hd_dir, fname)
                if not os.path.exists(fpath):
                    try:
                        hd_data.seek(0)
                        with open(fpath, 'wb') as f:
                            f.write(hd_data.read())
                        saved += 1
                    except:
                        pass
    if saved > 0:
        log(f"  [HD] 고해상도 이미지 {saved}개 저장 → {hd_dir}")

def main():
    log("\n" + "=" * 60)
    log("  유니클로 랭킹 크롤러 V5 (3단계 고속 버전)")
    log("=" * 60)
    log("  * Phase 1: 이미지 없이 전체 랭킹 빠르게 수집")
    log("  * Phase 2: 기존 이미지에서 동일 상품명 매칭")
    log("  * Phase 3: 누락 이미지만 스크린샷 캡쳐")
    log("=" * 60)

    parser = argparse.ArgumentParser(description='유니클로 랭킹 크롤러 V5')
    parser.add_argument('--only', type=str, default='', help='수집할 카테고리만 지정 (예: WOMEN,MEN)')
    parser.add_argument('--preserve-missing-from-latest', action='store_true', help='지정하지 않은 카테고리 시트는 최신 V5 엑셀에서 값만 보존')
    parser.add_argument('--skip-images', action='store_true', help='이미지 캡쳐/삽입을 건너뜀 (디버깅용)')
    args = parser.parse_args()

    global SKIP_IMAGES
    if args.skip_images:
        SKIP_IMAGES = True

    selected_categories = list(CATEGORIES.keys())
    if args.only.strip():
        selected_categories = [c.strip().upper() for c in args.only.split(',') if c.strip()]
        selected_categories = [c for c in selected_categories if c in CATEGORIES]
        if not selected_categories:
            log("  [WARN] --only 값이 유효하지 않아 전체 카테고리 수집으로 진행합니다")
            selected_categories = list(CATEGORIES.keys())
    
    driver = setup_driver()
    all_data = {}
    
    try:
        # ============================================================
        # Phase 1: 이미지 없이 전체 랭킹 빠르게 수집
        # ============================================================
        log("\n" + "=" * 60)
        log("[Phase 1] 이미지 없이 전체 랭킹 빠르게 수집")
        log("=" * 60)
        
        import time as _time
        phase1_start = _time.time()
        
        for category in selected_categories:
            info = CATEGORIES[category]
            for attempt in range(2):
                try:
                    data = scrape_category_with_tabs(driver, category, info['url'], info['tabs'], skip_images=True)
                    all_data.update(data)
                    break
                except BrowserCrashedError as e:
                    log(f"  [WARN] 브라우저 세션 오류 감지: {str(e)[:80]}")
                    if attempt >= 1:
                        raise
                    log("  [WARN] Chrome 재시작 후 카테고리 재시도...")
                    _safe_quit_driver(driver)
                    driver = setup_driver()
        
        phase1_elapsed = _time.time() - phase1_start
        total_p1 = sum(len(prods) for prods in all_data.values())
        log(f"\n  [Phase 1 완료] {len(all_data)}개 시트, {total_p1}개 상품 수집 ({phase1_elapsed:.0f}초)")

        # 선택 수집 시, 나머지 시트는 최신 엑셀에서 값만 보존
        if args.preserve_missing_from_latest and len(selected_categories) < len(CATEGORIES):
            work_dir = os.path.dirname(os.path.abspath(__file__))
            latest = _find_latest_uniqlo_v5_excel(work_dir)
            if latest:
                skip_prefixes = [f"{c}_" for c in selected_categories]
                log(f"\n  -> 기존 데이터 보존: {os.path.basename(latest)}")
                preserved = _load_sheets_from_excel(latest, skip_prefixes=skip_prefixes)
                for k, v in preserved.items():
                    if k not in all_data:
                        all_data[k] = v
                log(f"  -> 보존 시트 {len(preserved)}개 병합 완료")
            else:
                log("  [WARN] 보존할 최신 V5 엑셀을 찾지 못했습니다")
        
        # ============================================================
        # Phase 2: 기존 이미지에서 동일 상품명 매칭
        # ============================================================
        if not SKIP_IMAGES:
            log("\n" + "=" * 60)
            log("[Phase 2] 기존 이미지 매칭")
            log("=" * 60)
            
            hd_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'product_images_hd')
            missing_products = _reuse_existing_images(all_data, hd_dir)
            
            # ============================================================
            # Phase 3: 누락 이미지만 스크린샷
            # ============================================================
            if missing_products:
                log("\n" + "=" * 60)
                log(f"[Phase 3] 누락 이미지 {len(missing_products)}개 다운로드+캡처")
                log("=" * 60)
                
                phase3_start = _time.time()
                try:
                    driver = _screenshot_missing_images(driver, all_data, missing_products)
                except BrowserCrashedError as e:
                    log(f"  [WARN] Phase 3 브라우저 오류: {str(e)[:80]}")
                    log("  -> Chrome 재시작 후 재시도...")
                    _safe_quit_driver(driver)
                    driver = setup_driver()
                    try:
                        # 재시도 시 아직 이미지 없는 것만 필터
                        still_missing = [(s, i, p) for s, i, p in missing_products if not p.get('image_data')]
                        if still_missing:
                            driver = _screenshot_missing_images(driver, all_data, still_missing)
                    except Exception as e2:
                        log(f"  [WARN] Phase 3 재시도 실패: {str(e2)[:50]}")
                
                phase3_elapsed = _time.time() - phase3_start
                log(f"  -> Phase 3 소요시간: {phase3_elapsed:.0f}초")
            else:
                log("\n  [Phase 3] 모든 이미지 매칭 완료 - 스크린샷 불필요!")
        
        # ============================================================
        # 엑셀 저장
        # ============================================================
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_전체랭킹_이미지포함_V5_{timestamp}.xlsx"
        create_excel(all_data, filename)
        
        # 최종 통계
        log(f"\n{'='*60}")
        log(f"[완료] 수집 통계")
        log(f"{'='*60}")
        
        total_products = sum(len(p) for p in all_data.values())
        total_with_price = sum(1 for prods in all_data.values() for p in prods if p['price'])
        total_with_img = sum(1 for prods in all_data.values() for p in prods if p.get('image_data'))
        total_with_rating = sum(1 for prods in all_data.values() for p in prods if p['rating'] != '없음')
        total_with_review = sum(1 for prods in all_data.values() for p in prods if p['review_count'] != '없음')
        
        log(f"  총 시트: {len(all_data)}개")
        log(f"  총 상품: {total_products}개")
        log(f"  가격: {total_with_price}/{total_products}개 수집됨")
        log(f"  이미지 삽입: {total_with_img}/{total_products}개")
        log(f"  평점: {total_with_rating}/{total_products}개 수집됨")
        log(f"  리뷰: {total_with_review}/{total_products}개 수집됨")
        log(f"\n  파일: {filename}")
        log("=" * 60)
        
    finally:
        _safe_quit_driver(driver)
        log("\n브라우저 종료")

if __name__ == "__main__":
    main()
