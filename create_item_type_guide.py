# -*- coding: utf-8 -*-
"""
유니클로 아이템 타입 분류 가이드 생성기
- 네이버 쇼핑 패션 카테고리 기준
- 분류 딕셔너리를 보기 좋은 엑셀 자료로 출력
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===== 분류 체계 데이터 =====
CLASSIFICATION = {
    '아우터': {
        'color': 'D6EAF8',  # 연한 파랑
        'items': {
            '경량패딩':       {'keywords': '경량패딩, 경량 패딩, 라이트패딩, 라이트다운', 'desc': '가벼운 패딩류'},
            '울트라라이트다운': {'keywords': '울트라라이트다운, ultra light down', 'desc': '유니클로 시그니처 초경량 다운'},
            '롱패딩':        {'keywords': '롱패딩, 롱다운, 롱 패딩', 'desc': '무릎 아래 기장 패딩'},
            '숏패딩':        {'keywords': '숏패딩, 숏 패딩, 쇼트패딩', 'desc': '허리~엉덩이 기장 패딩'},
            '패딩':         {'keywords': '패딩, 다운재킷, 퍼펙트다운, 하이브리드다운', 'desc': '일반 패딩/다운 제품'},
            '플리스':        {'keywords': '플리스, 후리스, fleece, 뽀글이', 'desc': '플리스 소재 아우터'},
            '트렌치코트':     {'keywords': '트렌치코트, 트렌치 코트', 'desc': '트렌치코트'},
            '코트':         {'keywords': '코트, 체스터, 울코트, 핸드메이드코트', 'desc': '울/혼방 코트류'},
            '블레이저':      {'keywords': '블레이저', 'desc': '정장 스타일 블레이저'},
            '블루종':        {'keywords': '블루종, MA-1', 'desc': '블루종/봄버 재킷'},
            '재킷':         {'keywords': '재킷, 자켓, 점퍼, jacket', 'desc': '일반 재킷/점퍼'},
            '파카':         {'keywords': '파카, parka', 'desc': '파카/후드 방한 아우터'},
            '카디건':        {'keywords': '카디건, 가디건, cardigan', 'desc': '카디건 (아우터 겸용)'},
            '바람막이':      {'keywords': '바람막이, 윈드브레이커, 포켓터블', 'desc': '바람막이/윈드브레이커'},
            '베스트/조끼':    {'keywords': '조끼, 베스트, 패딩베스트, vest', 'desc': '소매 없는 아우터'},
        }
    },
    '상의': {
        'color': 'D5F5E3',  # 연한 초록
        'items': {
            '후드':         {'keywords': '후드, 후디, 스웻후드, 풀짚후드, hoodie', 'desc': '후드 달린 맨투맨/짚업'},
            '맨투맨':        {'keywords': '맨투맨, 스웻셔츠, 스웨트셔츠, sweatshirt', 'desc': '라운드넥 기모/스웻'},
            '니트/스웨터':    {'keywords': '니트, 스웨터, 캐시미어, 메리노, 터틀넥, 크루넥', 'desc': '니트/스웨터류'},
            '셔츠/블라우스':  {'keywords': '셔츠, 블라우스, 옥스포드, 플란넬셔츠, 린넨셔츠', 'desc': '셔츠/블라우스 전체'},
            '폴로':         {'keywords': '폴로, polo', 'desc': '폴로 셔츠'},
            '탱크탑':        {'keywords': '탱크탑, 나시, 슬리브리스, tank top', 'desc': '민소매 상의'},
            '티셔츠':        {'keywords': '티셔츠, T셔츠, 크루넥T, 반팔, 긴팔, UT, tee', 'desc': '일반 반팔/긴팔 티셔츠'},
        }
    },
    '팬츠': {
        'color': 'FCF3CF',  # 연한 노랑
        'items': {
            '진/데님':       {'keywords': '진, 데님, 스키니, 와이드스트레이트, 와이드핏진', 'desc': '데님/청바지류'},
            '치노':         {'keywords': '치노, chino', 'desc': '치노 팬츠'},
            '카고팬츠':      {'keywords': '카고, cargo', 'desc': '포켓 카고 팬츠'},
            '조거팬츠':      {'keywords': '조거, jogger', 'desc': '조거/트레이닝 팬츠'},
            '스웻팬츠':      {'keywords': '스웻팬츠, 이지팬츠, 트레이닝팬츠', 'desc': '편안한 스웻/이지 팬츠'},
            '슬랙스':        {'keywords': '슬랙스, 앵클팬츠, 스마트앵클, 감탄팬츠', 'desc': '정장/세미정장 팬츠'},
            '레깅스':        {'keywords': '레깅스, leggings', 'desc': '레깅스'},
            '숏팬츠':        {'keywords': '숏팬츠, 쇼츠, 반바지, 버뮤다, 하프팬츠, shorts', 'desc': '반바지/숏 팬츠'},
            '팬츠':         {'keywords': '팬츠, 바지, pants, trousers', 'desc': '기타 팬츠류'},
        }
    },
    '원피스/스커트': {
        'color': 'F5CBA7',  # 연한 주황
        'items': {
            '원피스':        {'keywords': '원피스, 드레스, dress, one piece', 'desc': '원피스/드레스'},
            '스커트':        {'keywords': '스커트, 치마, skirt', 'desc': '스커트/치마'},
        }
    },
    '이너웨어/기능성': {
        'color': 'FADBD8',  # 연한 빨강
        'items': {
            '히트텍':        {'keywords': '히트텍, heattech', 'desc': '유니클로 발열 이너웨어'},
            '에어리즘':      {'keywords': '에어리즘, airism, AIRism', 'desc': '유니클로 쿨링 이너웨어'},
            '브라탑':        {'keywords': '브라탑, 브라 탑, 브라캐미솔, bratop', 'desc': '브라 내장 탑'},
            '이너웨어':      {'keywords': '팬티, 트렁크, 브리프, 속옷, 보정, 런닝', 'desc': '속옷/언더웨어'},
        }
    },
    '홈웨어': {
        'color': 'E8DAEF',  # 연한 보라
        'items': {
            '파자마/라운지':  {'keywords': '파자마, 잠옷, 라운지, 홈웨어, pajama', 'desc': '실내복/파자마'},
        }
    },
    '악세서리': {
        'color': 'D5DBDB',  # 연한 회색
        'items': {
            '양말':         {'keywords': '양말, 삭스, socks', 'desc': '양말류'},
            '모자':         {'keywords': '모자, 캡, 버킷햇, 비니, hat, cap', 'desc': '모자/캡/비니'},
            '가방':         {'keywords': '가방, 백팩, 토트, 숄더백, 에코백, bag', 'desc': '가방/백류'},
            '머플러/스카프':  {'keywords': '머플러, 스카프, 목도리, 숄, scarf', 'desc': '머플러/스카프'},
            '벨트':         {'keywords': '벨트, belt', 'desc': '벨트'},
            '장갑':         {'keywords': '장갑, 글러브, gloves', 'desc': '장갑'},
            '우산':         {'keywords': '우산, umbrella', 'desc': '우산'},
            '슬리퍼/샌들':   {'keywords': '슬리퍼, 샌들, 룸슈즈, slipper, sandal', 'desc': '실내화/샌들'},
        }
    },
}

def create_guide():
    wb = openpyxl.Workbook()
    
    # ===== 시트1: 분류 체계 총괄표 =====
    ws1 = wb.active
    ws1.title = '아이템분류_총괄표'
    
    # 공통 스타일
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=12, name='맑은 고딕', color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    cat_font = Font(bold=True, size=11, name='맑은 고딕')
    normal_font = Font(size=10, name='맑은 고딕')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 제목
    ws1.merge_cells('A1:E1')
    title_cell = ws1['A1']
    title_cell.value = '유니클로 아이템 타입 분류 가이드 (네이버쇼핑 패션카테고리 기준)'
    title_cell.font = Font(bold=True, size=14, name='맑은 고딕', color='FFFFFF')
    title_cell.fill = PatternFill(start_color='E60012', end_color='E60012', fill_type='solid')
    title_cell.alignment = center_align
    ws1.row_dimensions[1].height = 35
    
    # 부제
    ws1.merge_cells('A2:E2')
    ws1['A2'].value = f'작성일: {datetime.now().strftime("%Y-%m-%d")}  |  총 {sum(len(v["items"]) for v in CLASSIFICATION.values())}개 아이템 타입  |  {len(CLASSIFICATION)}개 대분류'
    ws1['A2'].font = Font(size=9, name='맑은 고딕', color='666666')
    ws1['A2'].alignment = center_align
    ws1.row_dimensions[2].height = 22
    
    # 헤더
    headers = ['대분류', '아이템 타입', '매칭 키워드', '설명', '키워드 수']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(3, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws1.row_dimensions[3].height = 25
    
    # 열 너비
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 55
    ws1.column_dimensions['D'].width = 28
    ws1.column_dimensions['E'].width = 10
    
    # 데이터 입력
    row = 4
    for cat_name, cat_data in CLASSIFICATION.items():
        cat_fill = PatternFill(start_color=cat_data['color'], end_color=cat_data['color'], fill_type='solid')
        cat_start = row
        
        for item_name, item_data in cat_data['items'].items():
            kw_list = [k.strip() for k in item_data['keywords'].split(',')]
            
            ws1.cell(row, 1).value = ''  # 대분류는 나중에 병합
            ws1.cell(row, 2, item_name).font = Font(bold=True, size=10, name='맑은 고딕')
            ws1.cell(row, 3, item_data['keywords']).font = normal_font
            ws1.cell(row, 4, item_data['desc']).font = normal_font
            ws1.cell(row, 5, len(kw_list)).font = normal_font
            
            for col in range(1, 6):
                cell = ws1.cell(row, col)
                cell.border = thin_border
                cell.fill = cat_fill
                if col in (1, 2, 5):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            ws1.row_dimensions[row].height = 22
            row += 1
        
        # 대분류 셀 병합
        cat_end = row - 1
        if cat_start < cat_end:
            ws1.merge_cells(f'A{cat_start}:A{cat_end}')
        cat_cell = ws1.cell(cat_start, 1, cat_name)
        cat_cell.font = cat_font
        cat_cell.alignment = center_align
    
    # ===== 시트2: 통계 요약 =====
    ws2 = wb.create_sheet('분류_통계')
    
    ws2.merge_cells('A1:D1')
    ws2['A1'].value = '대분류별 아이템 타입 통계'
    ws2['A1'].font = Font(bold=True, size=13, name='맑은 고딕', color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 30
    
    stat_headers = ['대분류', '아이템 타입 수', '총 키워드 수', '대표 아이템']
    for col, h in enumerate(stat_headers, 1):
        cell = ws2.cell(2, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 45
    
    stat_row = 3
    total_types = 0
    total_kw = 0
    for cat_name, cat_data in CLASSIFICATION.items():
        cat_fill = PatternFill(start_color=cat_data['color'], end_color=cat_data['color'], fill_type='solid')
        item_count = len(cat_data['items'])
        kw_count = sum(len(v['keywords'].split(',')) for v in cat_data['items'].values())
        top_items = ', '.join(list(cat_data['items'].keys())[:4])
        
        ws2.cell(stat_row, 1, cat_name).font = cat_font
        ws2.cell(stat_row, 2, item_count).font = normal_font
        ws2.cell(stat_row, 3, kw_count).font = normal_font
        ws2.cell(stat_row, 4, top_items).font = normal_font
        
        for col in range(1, 5):
            cell = ws2.cell(stat_row, col)
            cell.border = thin_border
            cell.fill = cat_fill
            cell.alignment = center_align if col <= 3 else left_align
        
        total_types += item_count
        total_kw += kw_count
        stat_row += 1
    
    # 합계
    total_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    total_font = Font(bold=True, size=11, name='맑은 고딕', color='FFFFFF')
    ws2.cell(stat_row, 1, '합계').font = total_font
    ws2.cell(stat_row, 2, total_types).font = total_font
    ws2.cell(stat_row, 3, total_kw).font = total_font
    ws2.cell(stat_row, 4, '').font = total_font
    for col in range(1, 5):
        cell = ws2.cell(stat_row, col)
        cell.border = thin_border
        cell.fill = total_fill
        cell.alignment = center_align

    # ===== 시트3: 유니클로 탭 ↔ 대분류 매핑 =====
    ws3 = wb.create_sheet('탭_매핑')
    
    ws3.merge_cells('A1:D1')
    ws3['A1'].value = '유니클로 랭킹 탭 → 아이템 대분류 매핑'
    ws3['A1'].font = Font(bold=True, size=13, name='맑은 고딕', color='FFFFFF')
    ws3['A1'].fill = PatternFill(start_color='E60012', end_color='E60012', fill_type='solid')
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 30
    
    map_headers = ['유니클로 탭', '→ 대분류 매핑', '포함 아이템 타입', '비고']
    for col, h in enumerate(map_headers, 1):
        cell = ws3.cell(2, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 25
    
    tab_mapping = [
        ('모두보기', '전체', '모든 아이템 타입', '기본 탭 (필터 없음)'),
        ('상의', '상의', '후드, 맨투맨, 니트/스웨터, 셔츠/블라우스, 폴로, 탱크탑, 티셔츠', ''),
        ('팬츠', '팬츠', '진/데님, 치노, 카고, 조거, 스웻팬츠, 슬랙스, 레깅스, 숏팬츠', ''),
        ('드레스 & 스커트', '원피스/스커트', '원피스, 스커트', 'WOMEN 전용'),
        ('아우터', '아우터', '경량패딩~베스트/조끼 (15개 타입)', '시즌별 구성 변동'),
        ('이너웨어', '이너웨어/기능성', '히트텍, 에어리즘, 브라탑, 이너웨어', '기능성 제품 포함'),
        ('홈웨어', '홈웨어', '파자마/라운지', ''),
        ('악세서리', '악세서리', '양말, 모자, 가방, 머플러, 벨트, 장갑, 우산, 슬리퍼', ''),
    ]
    
    for i, (tab, mapping, items, note) in enumerate(tab_mapping, 3):
        ws3.cell(i, 1, tab).font = Font(bold=True, size=10, name='맑은 고딕')
        ws3.cell(i, 2, mapping).font = normal_font
        ws3.cell(i, 3, items).font = normal_font
        ws3.cell(i, 4, note).font = Font(size=9, name='맑은 고딕', color='888888')
        for col in range(1, 5):
            cell = ws3.cell(i, col)
            cell.border = thin_border
            cell.alignment = center_align if col <= 2 else left_align
            if i % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # 저장
    filename = f'유니클로_아이템분류_가이드_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    print(f'\n✅ 분류 가이드 생성 완료!')
    print(f'   파일: {filename}')
    print(f'   시트1: 아이템분류_총괄표 ({total_types}개 아이템 타입)')
    print(f'   시트2: 분류_통계 ({len(CLASSIFICATION)}개 대분류)')
    print(f'   시트3: 탭_매핑 (유니클로 탭 ↔ 대분류)')
    return filename

if __name__ == '__main__':
    create_guide()
