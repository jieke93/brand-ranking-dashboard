# -*- coding: utf-8 -*-
"""
유니클로 카테고리별 상품 정보 수집기
- WOMEN 카테고리의 하위 카테고리별 수집
- 컬러 개수 및 컬러명 추출
- 판매 데이터 수집 (평점, 리뷰 수 등)
"""

import sys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
from PIL import Image as PILImage
import time
import os
from datetime import datetime
import re

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9',
}

# 하위 카테고리 정의 (이미지에서 확인한 카테고리)
SUBCATEGORIES = {
    'all': '모두 보기',
    'tops': '상의',
    'bottoms': '팬츠',
    'dresses-skirts': '드레스 & 스커트',
    'outerwear': '아우터',
    'innerwear': '이너웨어',
    'loungewear': '홈웨어',
    'accessories': '악세서리'
}

def setup_driver():
    """Chrome 드라이버 설정"""
    print("=" * 70, flush=True)
    print("Chrome 드라이버 초기화 중...", flush=True)
    
    chrome_options = Options()
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    try:
        print("  ㄴ ChromeDriver 다운로드 중...", flush=True)
        service = Service(ChromeDriverManager().install())
        print("  ㄴ Chrome 브라우저 시작 중...", flush=True)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        print("[OK] 드라이버 초기화 완료\n", flush=True)
        sys.stdout.flush()
        return driver
    except Exception as e:
        print(f"[ERROR] 드라이버 초기화 실패: {e}")
        return None

def get_subcategories(driver):
    """WOMEN 카테고리의 하위 카테고리 URL 수집"""
    print("\n[단계 1/4] 하위 카테고리 URL 수집 중...", flush=True)
    print("  ㄴ WOMEN 메인 페이지 접속", flush=True)
    sys.stdout.flush()
    
    base_url = "https://www.uniqlo.com/kr/ko/women"
    
    try:
        driver.get(base_url)
        print("  ㄴ 페이지 로딩 중 (5초 대기)...", flush=True)
        sys.stdout.flush()
        for i in range(5):
            time.sleep(1)
            print(".", end='', flush=True)
        print(" 완료!", flush=True)
        print("  ㄴ 카테고리 링크 찾는 중...", flush=True)
        sys.stdout.flush()
        
        # 하위 카테고리 링크 찾기
        category_links = {}
        
        print("     → 페이지 소스 분석 중...", flush=True)
        
        # 여러 선택자 시도
        selectors = [
            "a[href*='/women/']",
            "nav a",
            ".category a",
            ".sub-nav a",
            "a[class*='category']",
            "header a[href*='/women/']"
        ]
        
        all_links = []
        for selector in selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                print(f"     → '{selector}': {len(elements)}개 링크 발견", flush=True)
                all_links.extend(elements)
            except Exception as e:
                continue
        
        print(f"     → 총 {len(all_links)}개 링크 분석 중...", flush=True)
        
        for elem in all_links:
            try:
                text = elem.text.strip()
                href = elem.get_attribute("href")
                
                if href and '/women/' in href and text and len(text) < 30:
                    # 중복 제거하고 메인 페이지 제외
                    if text not in category_links.values() and href.count('/') > 5:
                        category_links[href] = text
                        print(f"     ✓ 발견: {text}", flush=True)
            except:
                continue
        
        # 카테고리가 없으면 기본 카테고리 사용
        if not category_links:
            print("     → 자동 탐지 실패, 기본 카테고리 사용", flush=True)
            category_links = {
                "https://www.uniqlo.com/kr/ko/women/tops": "상의",
                "https://www.uniqlo.com/kr/ko/women/bottoms": "팬츠",
                "https://www.uniqlo.com/kr/ko/women/outerwear": "아우터"
            }
        
        # 확인된 카테고리 출력
        print(f"\n  ㄴ [성공] {len(category_links)}개 하위 카테고리 발견:", flush=True)
        for idx, (url, name) in enumerate(category_links.items(), 1):
            print(f"     {idx}. {name}", flush=True)
        sys.stdout.flush()
        
        return category_links
    
    except Exception as e:
        print(f"\n[ERROR] 카테고리 수집 실패: {e}", flush=True)
        import traceback
        traceback.print_exc()
        sys.stdout.flush()
        return {}

def scrape_category_products(driver, category_url, category_name, max_products=20):
    """특정 카테고리의 상품 정보 수집"""
    print(f"\n{'='*70}", flush=True)
    print(f"[수집 중] {category_name}", flush=True)
    print(f"{'='*70}", flush=True)
    sys.stdout.flush()
    
    try:
        print(f"  1/5) 페이지 접속 중...", flush=True)
        sys.stdout.flush()
        driver.get(category_url)
        print(f"  2/5) 페이지 로딩 중 (8초 대기)...", end='', flush=True)
        sys.stdout.flush()
        for i in range(8):
            time.sleep(1)
            print(".", end='', flush=True)
        print(" OK!", flush=True)
        
        # 스크롤하여 상품 로드
        print(f"  3/5) 상품 로딩을 위한 스크롤 중...", flush=True)
        sys.stdout.flush()
        for i in range(3):
            print(f"     ㄴ 스크롤 {i+1}/3", flush=True)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
        
        print("     ㄴ 맨 위로 이동", flush=True)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(2)
        
        print(f"  4/5) 상품 정보 추출 중...", flush=True)
        sys.stdout.flush()
        
        # 상품 요소 찾기
        product_selectors = [
            "li.fr-grid-item",
            "div.fr-grid-item",
            "article.product",
            "li.productTile",
            "div[class*='product-card']"
        ]
        
        products = []
        product_elements = []
        
        for selector in product_selectors:
            product_elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if len(product_elements) > 0:
                print(f"[OK] '{selector}'로 {len(product_elements)}개 상품 발견")
                break
        
        if not product_elements:
            print("     [경고] 상품을 찾을 수 없습니다.", flush=True)
            sys.stdout.flush()
            return []
        
        print(f"     ㄴ 총 {min(len(product_elements), max_products)}개 상품 처리 예정\n", flush=True)
        sys.stdout.flush()
        
        # 각 상품 정보 추출
        for idx, elem in enumerate(product_elements[:max_products], 1):
            try:
                print(f"     [{idx}/{min(len(product_elements), max_products)}] 처리 중...", end=' ', flush=True)
                sys.stdout.flush()
                product_data = {
                    'category': category_name,
                    'rank': idx
                }
                
                # 1. 상품명 추출
                name_selectors = [
                    ".fr-product-name",
                    ".productName",
                    "h3",
                    "p[class*='name']",
                    ".title"
                ]
                
                for sel in name_selectors:
                    try:
                        name_elem = elem.find_element(By.CSS_SELECTOR, sel)
                        product_data['name'] = name_elem.text.strip()
                        if product_data['name']:
                            break
                    except:
                        continue
                
                # 2. 가격 추출
                price_selectors = [
                    ".fr-product-price",
                    ".productPrice",
                    ".price",
                    "span[class*='price']"
                ]
                
                for sel in price_selectors:
                    try:
                        price_elem = elem.find_element(By.CSS_SELECTOR, sel)
                        product_data['price'] = price_elem.text.strip()
                        if product_data['price']:
                            break
                    except:
                        continue
                
                # 3. 이미지 URL 추출
                try:
                    img_elem = elem.find_element(By.TAG_NAME, "img")
                    product_data['image_url'] = img_elem.get_attribute("src") or img_elem.get_attribute("data-src") or ""
                    if product_data['image_url'] and not product_data['image_url'].startswith("http"):
                        product_data['image_url'] = "https://image.uniqlo.com" + product_data['image_url']
                except:
                    product_data['image_url'] = ""
                
                # 4. 컬러 정보 추출 (중요!)
                color_elements = elem.find_elements(By.CSS_SELECTOR, "div[class*='color'] img, img[class*='color'], .color-chip img, ul[class*='color'] img")
                
                colors = []
                for color_elem in color_elements:
                    try:
                        # alt 속성에서 컬러명 추출
                        color_name = color_elem.get_attribute("alt") or color_elem.get_attribute("title") or ""
                        if color_name and color_name not in colors:
                            colors.append(color_name)
                    except:
                        continue
                
                product_data['color_count'] = len(colors)
                product_data['colors'] = ', '.join(colors) if colors else '정보 없음'
                
                # 5. 판매 데이터 추출 (평점, 리뷰 수)
                # 평점
                try:
                    rating_elem = elem.find_element(By.CSS_SELECTOR, "[class*='rating'], [class*='star'], .review-rating")
                    rating_text = rating_elem.text.strip()
                    # 숫자 추출
                    rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                    if rating_match:
                        product_data['rating'] = rating_match.group(1)
                    else:
                        product_data['rating'] = rating_text
                except:
                    product_data['rating'] = '없음'
                
                # 리뷰 수
                try:
                    review_elem = elem.find_element(By.CSS_SELECTOR, "[class*='review-count'], [class*='review']")
                    review_text = review_elem.text.strip()
                    # 숫자 추출
                    review_match = re.search(r'\((\d+)\)', review_text)
                    if review_match:
                        product_data['review_count'] = review_match.group(1)
                    else:
                        product_data['review_count'] = review_text
                except:
                    product_data['review_count'] = '없음'
                
                # 6. 상품 코드 추출 (URL에서)
                try:
                    link_elem = elem.find_element(By.TAG_NAME, "a")
                    link_href = link_elem.get_attribute("href")
                    if link_href:
                        code_match = re.search(r'/products/(E\d{6}-\d{3})', link_href)
                        if code_match:
                            product_data['product_code'] = code_match.group(1)
                except:
                    product_data['product_code'] = ''
                
                # 필수 정보가 있는 경우만 추가
                if product_data.get('name'):
                    products.append(product_data)
                    print(f"OK (컬러: {product_data['color_count']}개)", flush=True)
                else:
                    print("실패 (상품명 없음)", flush=True)
                sys.stdout.flush()
            
            except Exception as e:
                print(f"실패 ({str(e)[:30]})", flush=True)
                sys.stdout.flush()
                continue
        
        print(f"\n  5/5) [완료] {category_name}: {len(products)}개 상품 수집 성공", flush=True)
        sys.stdout.flush()
        return products
    
    except Exception as e:
        print(f"\n[ERROR] 카테고리 '{category_name}' 수집 실패: {e}", flush=True)
        import traceback
        traceback.print_exc()
        return []

def download_image(url):
    """이미지 다운로드"""
    if not url:
        return None
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        
        if img.mode == 'RGBA':
            bg = PILImage.new('RGB', img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        
        img.thumbnail((80, 80), PILImage.Resampling.LANCZOS)
        
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        return img_bytes
    except:
        return None

def create_excel_report(all_products, filename):
    """엑셀 리포트 생성"""
    print(f"\n{'='*70}", flush=True)
    print("[단계 4/4] 엑셀 리포트 생성 중...", flush=True)
    print(f"  ㄴ 파일명: {filename}", flush=True)
    sys.stdout.flush()
    
    wb = openpyxl.Workbook()
    
    # 카테고리별로 시트 생성
    categories = {}
    for product in all_products:
        cat = product.get('category', '기타')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(product)
    
    # 기본 시트 삭제
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 카테고리별 시트 생성
    for cat_name, products in categories.items():
        # 시트명은 최대 31자
        sheet_name = cat_name[:31]
        ws = wb.create_sheet(sheet_name)
        
        print(f"  ㄴ 시트 [{sheet_name}] 생성 중 ({len(products)}개 상품)...", end=' ', flush=True)
        sys.stdout.flush()
        
        # 헤더
        headers = ['순번', '상품명', '가격', '컬러 수', '컬러 목록', '평점', '리뷰 수', '상품코드', '이미지']
        ws.append(headers)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF", name='맑은 고딕')
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col in range(1, 10):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 열 너비
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 15
        
        ws.row_dimensions[1].height = 30
        
        # 데이터 입력
        for prod_idx, product in enumerate(products, 1):
            if prod_idx % 5 == 0:
                print(".", end='', flush=True)
            row_idx = ws.max_row + 1
            
            ws.cell(row_idx, 1, product.get('rank', ''))
            ws.cell(row_idx, 2, product.get('name', ''))
            ws.cell(row_idx, 3, product.get('price', ''))
            ws.cell(row_idx, 4, product.get('color_count', 0))
            ws.cell(row_idx, 5, product.get('colors', ''))
            ws.cell(row_idx, 6, product.get('rating', ''))
            ws.cell(row_idx, 7, product.get('review_count', ''))
            ws.cell(row_idx, 8, product.get('product_code', ''))
            
            ws.row_dimensions[row_idx].height = 70
            
            # 스타일
            for col in range(1, 9):
                cell = ws.cell(row_idx, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                cell.font = Font(name='맑은 고딕', size=10)
            
            # 컬러 수 강조
            color_count_cell = ws.cell(row_idx, 4)
            if product.get('color_count', 0) > 5:
                color_count_cell.font = Font(bold=True, color="E60012", size=11, name='맑은 고딕')
            
            # 이미지
            if product.get('image_url'):
                img_data = download_image(product['image_url'])
                if img_data:
                    try:
                        img = XLImage(img_data)
                        img.width = 60
                        img.height = 60
                        ws.add_image(img, f'I{row_idx}')
                    except:
                        pass
            
            ws.cell(row_idx, 9).border = border
        
        # 통계 추가
        stats_row = ws.max_row + 2
        ws.cell(stats_row, 1, "통계")
        ws.cell(stats_row, 1).font = Font(bold=True, size=11, name='맑은 고딕')
        ws.merge_cells(f'A{stats_row}:B{stats_row}')
        
        ws.cell(stats_row + 1, 1, f"총 상품 수: {len(products)}개")
        
        avg_colors = sum(p.get('color_count', 0) for p in products) / len(products) if products else 0
        ws.cell(stats_row + 2, 1, f"평균 컬러 수: {avg_colors:.1f}개")
        
        print("완료!", flush=True)
        sys.stdout.flush()
        
        # 판매 데이터 피드백
        has_rating = sum(1 for p in products if p.get('rating') != '없음')
        has_review = sum(1 for p in products if p.get('review_count') != '없음')
        
        ws.cell(stats_row + 3, 1, f"평점 정보: {has_rating}개 상품")
        ws.cell(stats_row + 4, 1, f"리뷰 정보: {has_review}개 상품")
    
    # 종합 시트
    summary_ws = wb.create_sheet("종합 요약", 0)
    summary_ws.append(['카테고리', '상품 수', '평균 컬러 수', '평점 정보', '리뷰 정보'])
    
    for col in range(1, 6):
        cell = summary_ws.cell(1, col)
        cell.font = Font(bold=True, size=12, color="FFFFFF", name='맑은 고딕')
        cell.fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 12
    summary_ws.column_dimensions['C'].width = 15
    summary_ws.column_dimensions['D'].width = 15
    summary_ws.column_dimensions['E'].width = 15
    
    for cat_name, products in categories.items():
        avg_colors = sum(p.get('color_count', 0) for p in products) / len(products) if products else 0
        has_rating = sum(1 for p in products if p.get('rating') != '없음')
        has_review = sum(1 for p in products if p.get('review_count') != '없음')
        
        summary_ws.append([
            cat_name,
            len(products),
            f"{avg_colors:.1f}개",
            f"{has_rating}/{len(products)}",
            f"{has_review}/{len(products)}"
        ])
    
    # 저장
    try:
        wb.save(filename)
        print(f"\n[완료] 엑셀 파일 저장 완료!")
        return True
    except Exception as e:
        print(f"\n[오류] 파일 저장 실패: {e}")
        return False

def main():
    print("=" * 70, flush=True)
    print("유니클로 WOMEN 카테고리별 상품 정보 수집기", flush=True)
    print("=" * 70, flush=True)
    print("\n기능:", flush=True)
    print("  1. WOMEN 카테고리 하위 카테고리별 수집", flush=True)
    print("  2. 각 상품의 컬러 개수 및 컬러명 추출", flush=True)
    print("  3. 판매 데이터 수집 (평점, 리뷰 수)", flush=True)
    print(flush=True)
    print("법적 리스크 최소화: 공개 정보만 수집, 개인 사용 목적", flush=True)
    print("=" * 70, flush=True)
    sys.stdout.flush()
    
    driver = setup_driver()
    if not driver:
        print("[ERROR] 드라이버 초기화 실패")
        return
    
    try:
        # 1. 하위 카테고리 수집
        categories = get_subcategories(driver)
        
        if not categories:
            print("\n[WARNING] 하위 카테고리를 찾을 수 없어 기본 URL 사용")
            # 기본 카테고리
            categories = {
                "https://www.uniqlo.com/kr/ko/women": "전체",
            }
        
        # 2. 각 카테고리별 상품 수집
        all_products = []
        
        category_list = list(categories.items())[:5]
        total_categories = len(category_list)
        
        print(f"\n[단계 2/4] 카테고리별 상품 수집 시작 (총 {total_categories}개 카테고리)", flush=True)
        print(f"{'='*70}\n", flush=True)
        sys.stdout.flush()
        
        for cat_idx, (cat_url, cat_name) in enumerate(category_list, 1):
            print(f"\n>>> 진행: {cat_idx}/{total_categories} 카테고리 <<<", flush=True)
            sys.stdout.flush()
            products = scrape_category_products(driver, cat_url, cat_name, max_products=10)
            all_products.extend(products)
            print(f"\n  ㄴ 누적 수집: {len(all_products)}개 상품", flush=True)
            if cat_idx < total_categories:
                print(f"  ㄴ 다음 카테고리까지 2초 대기...", flush=True)
                sys.stdout.flush()
                time.sleep(2)
        
        if not all_products:
            print("\n[ERROR] 수집된 상품이 없습니다.")
            return
        
        print(f"\n{'='*70}", flush=True)
        print(f"[단계 3/4] 데이터 분석", flush=True)
        print(f"{'='*70}", flush=True)
        print(f"  ㄴ 총 수집 상품: {len(all_products)}개\n", flush=True)
        sys.stdout.flush()
        
        # 판매 데이터 피드백
        print("[판매 데이터 수집 결과]", flush=True)
        has_rating = sum(1 for p in all_products if p.get('rating') != '없음')
        has_review = sum(1 for p in all_products if p.get('review_count') != '없음')
        
        print(f"평점 정보: {has_rating}/{len(all_products)}개 상품에서 수집")
        print(f"리뷰 정보: {has_review}/{len(all_products)}개 상품에서 수집")
        
        if has_rating > 0 or has_review > 0:
            print("[OK] 판매 데이터가 공개되어 있어 수집되었습니다!")
        else:
            print("[INFO] 판매 데이터가 공개되지 않았거나 찾을 수 없습니다.")
        
        # 3. 엑셀 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_WOMEN_카테고리별_{timestamp}.xlsx"
        
        if create_excel_report(all_products, filename):
            print(f"\n{'='*70}")
            print("[SUCCESS] 작업 완료!")
            print(f"{'='*70}")
            print(f"파일: {os.path.abspath(filename)}")
            print(f"총 상품: {len(all_products)}개")
            print(f"평균 컬러 수: {sum(p.get('color_count', 0) for p in all_products) / len(all_products):.1f}개")
            print(f"{'='*70}")
    
    except KeyboardInterrupt:
        print("\n\n[중단] 사용자에 의해 중단되었습니다.")
    
    except Exception as e:
        print(f"\n[ERROR] 오류 발생: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if driver:
            print("\n브라우저 종료 중...")
            driver.quit()
            print("[OK] 종료 완료")

if __name__ == "__main__":
    # 실시간 출력을 위해 버퍼링 비활성화
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
    main()
