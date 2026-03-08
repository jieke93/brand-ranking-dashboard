# -*- coding: utf-8 -*-
"""
유니클로 여성 랭킹 상품 정보 수집기
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
from PIL import Image as PILImage
import time
import os
from datetime import datetime

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9',
}

def setup_driver():
    """Selenium 드라이버 설정"""
    print("Chrome 드라이버 설정 중...")
    
    chrome_options = Options()
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        print("[OK] 드라이버 설정 완료")
        return driver
    except Exception as e:
        print(f"[ERROR] 드라이버 설정 실패: {e}")
        return None

def scrape_ranking_data(driver):
    """랭킹 데이터 수집"""
    url = "https://www.uniqlo.com/kr/ko/spl/ranking/women"
    print(f"\n페이지 접근: {url}")
    
    try:
        driver.get(url)
        print("페이지 초기 로딩...")
        time.sleep(10)  # 초기 로딩 대기 시간 증가
        
        print("상품 목록 로딩 중...")
        
        # 페이지 끝까지 스크롤 (더 많이, 더 천천히)
        for i in range(5):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)
            print(f"  스크롤 {i+1}/5")
        
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(2)
        
        print("\n상품 정보 추출 중...")
        
        # 방법 1: 모든 링크에서 상품 링크 찾기
        all_links = driver.find_elements(By.TAG_NAME, "a")
        product_links = []
        
        for link in all_links:
            href = link.get_attribute("href")
            if href and "/products/E" in href:
                if href not in product_links:
                    product_links.append(href)
        
        print(f"[OK] {len(product_links)}개 상품 링크 발견")
        
        # 방법 2: 기존 셀렉터로도 시도
        product_elements = []
        selectors = [
            "li.productTile",
            "div.fr-grid-item",
            "article[class*='product']",
            "li[class*='product']",
            "div[class*='product-card']"
        ]
        
        for selector in selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if elements and len(elements) > 5:
                    product_elements = elements
                    print(f"[OK] 셀렉터 '{selector}'로 {len(elements)}개 상품 발견")
                    break
            except:
                continue
        
        products = []
        
        # 링크로 상품 정보 추출
        if product_links and not product_elements:
            print("링크 기반 데이터 추출...")
            for rank, link_href in enumerate(product_links[:50], 1):
                try:
                    # 링크에서 상품 코드 추출
                    import re
                    match = re.search(r'/products/(E\d{6}-\d{3})/(\d{2})\?colorDisplayCode=(\d+)', link_href)
                    if match:
                        product_code = match.group(1)
                        color_code = match.group(3)
                        
                        # 상품명은 링크 엘리먼트에서 찾기
                        name = f"{product_code}"
                        try:
                            link_elem = driver.find_element(By.CSS_SELECTOR, f'a[href*="{product_code}"]')
                            name = link_elem.get_attribute("aria-label") or link_elem.get_attribute("title") or product_code
                        except:
                            pass
                        
                        # 이미지 URL 생성
                        base_code = product_code.split('-')[0][1:]
                        image_url = f'https://image.uniqlo.com/UQ/ST3/AsianCommon/imagesgoods/{base_code}/item/goods_{color_code}_{base_code}.jpg'
                        
                        products.append({
                            'rank': rank,
                            'name': name[:100],
                            'price': '가격 정보 확인 필요',
                            'image_url': image_url,
                            'colors': '컬러 정보 확인 필요'
                        })
                        print(f"  [{rank}위] {name[:50]}")
                except Exception as e:
                    continue
        
        # 셀렉터로 상품 정보 추출 (기존 방식)
        elif product_elements:
            print("셀렉터 기반 데이터 추출...")
            for rank, elem in enumerate(product_elements[:50], 1):
                try:
                    # 상품명
                    name = None
                    for sel in [".productName", "h3", ".title", "p[class*='name']", "a[class*='name']"]:
                        try:
                            name = elem.find_element(By.CSS_SELECTOR, sel).text.strip()
                            if name:
                                break
                        except:
                            pass
                    
                    # 가격
                    price = None
                    for sel in [".productPrice", ".price", "span[class*='price']", "p[class*='price']"]:
                        try:
                            price = elem.find_element(By.CSS_SELECTOR, sel).text.strip()
                            if price:
                                break
                        except:
                            pass
                    
                    # 이미지
                    image_url = None
                    try:
                        img = elem.find_element(By.TAG_NAME, "img")
                        image_url = img.get_attribute("src") or img.get_attribute("data-src")
                        if image_url and not image_url.startswith("http"):
                            image_url = "https://image.uniqlo.com" + image_url
                    except:
                        pass
                    
                    # 컬러
                    colors = []
                    try:
                        color_elems = elem.find_elements(By.CSS_SELECTOR, "img[alt], .color-chip")
                        for ce in color_elems[:8]:
                            color = ce.get_attribute("alt") or ce.get_attribute("title")
                            if color:
                                colors.append(color)
                    except:
                        pass
                    
                    if name and price:
                        products.append({
                            'rank': rank,
                            'name': name,
                            'price': price,
                            'image_url': image_url or '',
                            'colors': ', '.join(colors) if colors else '정보 없음'
                        })
                        print(f"  [{rank}위] {name[:40]}")
                
                except Exception as e:
                    continue
        
        else:
            with open("debug_source.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("[WARNING] 상품을 찾을 수 없습니다. debug_source.html 확인")
            return []
        
        return products
    
    except Exception as e:
        print(f"[ERROR] 수집 실패: {e}")
        return []

def download_image(url):
    """이미지 다운로드"""
    if not url:
        return None
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        img = PILImage.open(BytesIO(response.content))
        
        if img.mode == 'RGBA':
            bg = PILImage.new('RGB', img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        
        img.thumbnail((100, 100), PILImage.Resampling.LANCZOS)
        
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        return img_bytes
    except:
        return None

def create_excel(products, filename):
    """엑셀 파일 생성"""
    print(f"\n엑셀 생성: {filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "여성 랭킹"
    
    # 헤더
    headers = ['순위', '상품명', '가격', '컬러 옵션', '상품 이미지']
    ws.append(headers)
    
    # 스타일
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 열 너비
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    
    # 데이터
    for product in products:
        row_idx = ws.max_row + 1
        
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product['colors'])
        
        ws.row_dimensions[row_idx].height = 85
        
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 이미지
        if product['image_url']:
            print(f"  [{product['rank']}위] 이미지 다운로드...")
            img_data = download_image(product['image_url'])
            if img_data:
                try:
                    img = XLImage(img_data)
                    img.width = 80
                    img.height = 80
                    ws.add_image(img, f'E{row_idx}')
                except:
                    pass
        
        ws.cell(row_idx, 5).border = border
    
    # 메타정보
    meta_row = ws.max_row + 2
    ws.cell(meta_row, 1, f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(meta_row + 1, 1, "출처: 유니클로 공식 온라인스토어")
    ws.cell(meta_row + 2, 1, "개인 사용 목적, 상업적 사용 금지")
    
    wb.save(filename)
    print(f"[OK] 저장 완료!")

def main():
    print("=" * 70)
    print("유니클로 여성 랭킹 상품 정보 수집기")
    print("=" * 70)
    print("\n개인 사용 목적 / 법적 리스크 최소화")
    print("상업적 사용 금지\n")
    
    driver = setup_driver()
    if not driver:
        print("[ERROR] Chrome 드라이버 실패.")
        return
    
    try:
        products = scrape_ranking_data(driver)
        
        if not products:
            print("\n[ERROR] 데이터 수집 실패")
            return
        
        print(f"\n[OK] {len(products)}개 상품 수집 완료")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_여성랭킹_{timestamp}.xlsx"
        
        create_excel(products, filename)
        
        print("\n" + "=" * 70)
        print("작업 완료!")
        print(f"파일: {os.path.abspath(filename)}")
        print(f"상품: {len(products)}개")
        print("=" * 70)
    
    except Exception as e:
        print(f"\n[ERROR] 오류: {e}")
    
    finally:
        if driver:
            driver.quit()

if __name__ == "__main__":
    main()
