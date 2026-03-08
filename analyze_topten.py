#!/usr/bin/env python3
"""
탑텐(TOPTEN) 랭킹 분석 도구
- 크롤링된 엑셀 데이터를 읽어 아이템타입별 비중 분석
- 브랜드별 분석, 할인율 분석, 가격대별 분석
- 날짜별 랭킹 변동 추적 (히스토리 누적)
- 분석 결과를 엑셀로 출력
"""

import os
import glob
import json
import re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ─── 설정 ───
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_FILE = os.path.join(WORK_DIR, 'topten_history.json')
HISTORY_BACKUP = os.path.join(WORK_DIR, 'topten_history_backup.json')
BRAND_NAME = '탑텐'
FILE_PATTERN = '탑텐_주간베스트_이미지포함_V3_*.xlsx'
CATEGORIES = ['전체', '여성', '남성', '키즈', '베이비']

# 스타일 (탑텐 블루)
HEADER_FILL = PatternFill('solid', fgColor='0066CC')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUB_HEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUB_HEADER_FONT = Font(bold=True, size=10)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
UP_FONT = Font(color='FF0000', bold=True)
DOWN_FONT = Font(color='0070C0', bold=True)
NEW_FONT = Font(color='00B050', bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ─── 아이템타입 분류 규칙 (탑텐 특화 - 한글 상품명 중심) ───
ITEM_TYPE_RULES = [
    # 아우터
    (r'(?i)(패딩|다운|덕다운|구스|웰론)', '패딩'),
    (r'(?i)(코트|트렌치)', '코트'),
    (r'(?i)(자켓|재킷|점퍼|블레이저|집업|바람막이|야상|아노락)', '자켓/점퍼'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(조끼|베스트|vest)', '조끼/베스트'),
    (r'(?i)(후리스|플리스|fleece|뽀글이)', '플리스'),
    # 상의
    (r'(?i)(후드|후디|hoodie)', '후드'),
    (r'(?i)(맨투맨|스웨트|크루넥)', '맨투맨'),
    (r'(?i)(니트|스웨터|터틀넥|목폴라|폴라)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스|남방)', '셔츠'),
    (r'(?i)(긴팔|롱슬리브|long\s*sleeve)', '긴팔티'),
    (r'(?i)(반팔|반소매|티셔츠|t-shirt|tee)', '티셔츠'),
    # 하의
    (r'(?i)(청바지|데님\s*팬|진\s)', '청바지'),
    (r'(?i)(슬랙스|치노|팬츠|바지|트라우저|면바지|카고)', '팬츠'),
    (r'(?i)(조거|트레이닝|스웨트팬츠|이지팬츠)', '조거/트레이닝'),
    (r'(?i)(반바지|숏팬츠|쇼츠|shorts)', '반바지'),
    (r'(?i)(스커트|치마)', '스커트'),
    (r'(?i)(레깅스)', '레깅스'),
    # 원피스
    (r'(?i)(원피스|드레스)', '원피스'),
    # 세트
    (r'(?i)(세트|set)', '세트'),
    # 내의/언더웨어
    (r'(?i)(내복|내의|발열|히트텍|웜)', '내의/발열'),
    (r'(?i)(속옷|팬티|브라|런닝)', '속옷'),
    (r'(?i)(양말|삭스|sock)', '양말'),
    # 잡화
    (r'(?i)(모자|캡|비니|버킷)', '모자'),
    (r'(?i)(가방|백팩|토트|크로스|에코백|숄더)', '가방'),
    (r'(?i)(벨트)', '벨트'),
    (r'(?i)(머플러|스카프|목도리)', '머플러'),
    (r'(?i)(장갑)', '장갑'),
    (r'(?i)(마스크)', '마스크'),
    # 신발
    (r'(?i)(운동화|스니커즈|신발|슈즈)', '신발'),
    (r'(?i)(슬리퍼|샌들)', '슬리퍼/샌들'),
]


def log(msg):
    print(msg)


def classify_item_type(name):
    """상품명으로 아이템타입 자동 분류"""
    if not name:
        return '미분류'
    for pattern, item_type in ITEM_TYPE_RULES:
        if re.search(pattern, name):
            return item_type
    return '미분류'


def find_latest_ranking_file():
    """가장 최근 크롤링 엑셀 파일 찾기"""
    pattern = os.path.join(WORK_DIR, FILE_PATTERN)
    files = glob.glob(pattern)
    if not files:
        log(f"[ERROR] 크롤링 결과 파일을 찾을 수 없습니다. ({FILE_PATTERN})")
        return None
    files.sort(reverse=True)
    return files[0]


def load_ranking_data(filepath):
    """엑셀 파일에서 랭킹 데이터 로드 (탑텐 컬럼 구조)"""
    log(f"  파일 로드: {os.path.basename(filepath)}")
    wb = load_workbook(filepath, data_only=True)

    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]

        products = []
        for row_idx in range(2, ws.max_row + 1):
            rank = ws.cell(row_idx, 1).value
            if rank is None:
                continue

            # 탑텐 컬럼: 순위(1), 이미지(2), 브랜드(3), 상품명(4),
            #             할인가(5), 정가(6), 할인율(7), 평점(8), 리뷰수(9)
            brand = ws.cell(row_idx, 3).value or ''
            name = ws.cell(row_idx, 4).value or ''
            sale_price = ws.cell(row_idx, 5).value or ''
            original_price = ws.cell(row_idx, 6).value or ''
            discount_rate = ws.cell(row_idx, 7).value or ''
            rating = ws.cell(row_idx, 8).value or '없음'
            review_count = ws.cell(row_idx, 9).value or '없음'
            item_type = classify_item_type(name)

            products.append({
                'rank': rank,
                'brand': brand,
                'name': name,
                'item_type': item_type,
                'sale_price': sale_price,
                'original_price': original_price,
                'discount_rate': discount_rate,
                'rating': rating,
                'review_count': review_count,
                'sheet': sheet_name
            })

        all_data[sheet_name] = products

    wb.close()
    return all_data


def parse_price(price_str):
    """가격 문자열을 숫자로 변환"""
    if not price_str:
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(price_str)))
    except:
        return 0


def parse_discount_rate(rate_str):
    """할인율 문자열을 숫자로 변환"""
    if not rate_str:
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(rate_str)))
    except:
        return 0


def parse_review(review_str):
    """리뷰 수 파싱"""
    if not review_str or review_str == '없음':
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(review_str)))
    except:
        return 0


def parse_rating(rating_str):
    """평점 파싱"""
    if not rating_str or rating_str == '없음':
        return 0.0
    try:
        return float(str(rating_str).replace('★', '').strip())
    except:
        return 0.0


def extract_date_from_filename(filepath):
    """파일명에서 날짜 추출"""
    basename = os.path.basename(filepath)
    match = re.search(r'(\d{8})_(\d{6})', basename)
    if match:
        return match.group(1)
    return datetime.now().strftime('%Y%m%d')


# ─── 분석 함수들 ───

def analyze_item_type_distribution(all_data):
    """아이템타입별 비중 분석 (카테고리별)"""
    log("\n  [분석1] 아이템타입별 비중 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        type_counter = Counter()
        type_products = defaultdict(list)

        for p in products:
            item_type = p['item_type']
            type_counter[item_type] += 1
            type_products[item_type].append(p)

        total = len(products)
        type_analysis = []
        for item_type, count in type_counter.most_common():
            pct = round(count / total * 100, 1) if total > 0 else 0

            prices = [parse_price(p['sale_price']) for p in type_products[item_type]]
            ratings = [parse_rating(p['rating']) for p in type_products[item_type] if parse_rating(p['rating']) > 0]
            reviews = [parse_review(p['review_count']) for p in type_products[item_type] if parse_review(p['review_count']) > 0]

            avg_price = round(sum(prices) / len(prices)) if prices else 0
            avg_rating = round(sum(ratings) / len(ratings), 1) if ratings else 0
            avg_review = round(sum(reviews) / len(reviews)) if reviews else 0

            top_products = sorted(type_products[item_type], key=lambda x: x['rank'])[:3]
            top_names = ', '.join([f"{p['rank']}위:{p['name'][:15]}" for p in top_products])

            type_analysis.append({
                'item_type': item_type,
                'count': count,
                'pct': pct,
                'avg_price': avg_price,
                'avg_rating': avg_rating,
                'avg_review': avg_review,
                'top_products': top_names,
                'best_rank': min([p['rank'] for p in type_products[item_type]]),
            })

        results[sheet_name] = {
            'total': total,
            'analysis': sorted(type_analysis, key=lambda x: x['count'], reverse=True)
        }

    return results


def analyze_top_items(all_data):
    """핵심 아이템 분석"""
    log("  [분석2] 핵심 아이템 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        unique_items = []
        for p in products:
            unique_items.append({
                'name': p['name'],
                'brand': p['brand'],
                'item_type': p['item_type'],
                'sale_price': p['sale_price'],
                'original_price': p['original_price'],
                'discount_rate': p['discount_rate'],
                'rating': p['rating'],
                'review_count': p['review_count'],
                'rank': p['rank'],
                'impact_score': max(31 - p['rank'], 1),
            })
        unique_items.sort(key=lambda x: x['impact_score'], reverse=True)
        results[sheet_name] = unique_items[:30]

    return results


def analyze_brand_distribution(all_data):
    """브랜드별 분석 (탑텐 특화 - 멀티 브랜드)"""
    log("  [분석3] 브랜드별 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        brand_counter = Counter()
        brand_products = defaultdict(list)

        for p in products:
            brand = p['brand'] or '미표기'
            brand_counter[brand] += 1
            brand_products[brand].append(p)

        total = len(products)
        brand_analysis = []
        for brand, count in brand_counter.most_common():
            pct = round(count / total * 100, 1) if total > 0 else 0
            prices = [parse_price(p['sale_price']) for p in brand_products[brand]]
            avg_price = round(sum(prices) / len(prices)) if prices else 0

            best_rank = min([p['rank'] for p in brand_products[brand]])
            top_names = ', '.join([p['name'][:15] for p in
                                   sorted(brand_products[brand], key=lambda x: x['rank'])[:3]])

            brand_analysis.append({
                'brand': brand,
                'count': count,
                'pct': pct,
                'avg_price': avg_price,
                'best_rank': best_rank,
                'top_products': top_names,
            })

        results[sheet_name] = {
            'total': total,
            'brands': sorted(brand_analysis, key=lambda x: x['count'], reverse=True)
        }

    return results


def analyze_discount(all_data):
    """할인율 분석 (탑텐 특화)"""
    log("  [분석4] 할인율 분석...")

    discount_bands = [
        (0, 1, '정가 판매'),
        (1, 10, '10% 미만'),
        (10, 20, '10~20%'),
        (20, 30, '20~30%'),
        (30, 40, '30~40%'),
        (40, 50, '40~50%'),
        (50, 101, '50% 이상'),
    ]

    results = {}
    for sheet_name, products in all_data.items():
        band_counter = Counter()
        band_items = defaultdict(list)

        for p in products:
            rate = parse_discount_rate(p['discount_rate'])
            for low, high, label in discount_bands:
                if low <= rate < high:
                    band_counter[label] += 1
                    band_items[label].append(f"{p['name'][:15]}({p['discount_rate']})")
                    break

        total = len(products)
        results[sheet_name] = {
            'total': total,
            'bands': [(label, band_counter.get(label, 0),
                       round(band_counter.get(label, 0) / total * 100, 1) if total > 0 else 0,
                       band_items.get(label, [])[:3])
                      for _, _, label in discount_bands if band_counter.get(label, 0) > 0]
        }

    return results


def analyze_price_band(all_data):
    """가격대별 분석"""
    log("  [분석5] 가격대별 분석...")

    price_bands = [
        (0, 10000, '1만원 이하'),
        (10000, 20000, '1~2만원'),
        (20000, 30000, '2~3만원'),
        (30000, 50000, '3~5만원'),
        (50000, 70000, '5~7만원'),
        (70000, 100000, '7~10만원'),
        (100000, 200000, '10~20만원'),
        (200000, float('inf'), '20만원 이상'),
    ]

    results = {}
    for sheet_name, products in all_data.items():
        band_counter = Counter()
        band_items = defaultdict(list)

        for p in products:
            price = parse_price(p['sale_price'])
            for low, high, label in price_bands:
                if low <= price < high:
                    band_counter[label] += 1
                    band_items[label].append(p['name'][:20])
                    break

        total = len(products)
        results[sheet_name] = {
            'total': total,
            'bands': [(label, band_counter.get(label, 0),
                       round(band_counter.get(label, 0) / total * 100, 1) if total > 0 else 0,
                       band_items.get(label, [])[:3])
                      for _, _, label in price_bands if band_counter.get(label, 0) > 0]
        }

    return results


# ─── 히스토리 관리 ───

def load_history():
    """랭킹 히스토리 로드"""
    for fpath in [HISTORY_FILE, HISTORY_BACKUP]:
        if os.path.exists(fpath):
            try:
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if data:
                    return data
            except:
                continue
    return {}


def save_history(history):
    """랭킹 히스토리 저장 (메인 + 백업)"""
    if os.path.exists(HISTORY_FILE):
        try:
            import shutil
            shutil.copy2(HISTORY_FILE, HISTORY_BACKUP)
        except:
            pass

    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def recover_history_from_excel_files():
    """이전 크롤링 엑셀 파일들에서 히스토리 역추출"""
    log("  [히스토리 복구] 이전 크롤링 파일에서 데이터 복구 시도...")

    pattern = os.path.join(WORK_DIR, FILE_PATTERN)
    files = sorted(glob.glob(pattern))

    if not files:
        log("  [히스토리 복구] 크롤링 파일 없음 - 신규 시작")
        return {}

    recovered = {}
    for filepath in files:
        date_key = extract_date_from_filename(filepath)
        try:
            wb = load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Sheet':
                    continue

                category = sheet_name
                if category not in recovered:
                    recovered[category] = {}

                ws = wb[sheet_name]
                day_data = {}
                for row_idx in range(2, ws.max_row + 1):
                    rank = ws.cell(row_idx, 1).value
                    if rank is None:
                        continue
                    brand = str(ws.cell(row_idx, 3).value or '')
                    name = str(ws.cell(row_idx, 4).value or '')[:40]
                    sale_price = str(ws.cell(row_idx, 5).value or '')
                    discount_rate = str(ws.cell(row_idx, 7).value or '')
                    rating = str(ws.cell(row_idx, 8).value or '없음')
                    review_count = str(ws.cell(row_idx, 9).value or '없음')
                    item_type = classify_item_type(name)

                    if name:
                        day_data[name] = {
                            'rank': rank,
                            'brand': brand,
                            'item_type': item_type,
                            'sale_price': sale_price,
                            'discount_rate': discount_rate,
                            'rating': rating,
                            'review_count': review_count,
                        }

                if day_data:
                    recovered[category][date_key] = day_data

            wb.close()
            log(f"    -> {os.path.basename(filepath)} ({date_key}) 복구 완료")
        except Exception as e:
            log(f"    -> {os.path.basename(filepath)} 복구 실패: {e}")

    return recovered


def merge_history(existing, new_data):
    """두 히스토리 데이터를 병합"""
    merged = {}
    all_categories = set(list(existing.keys()) + list(new_data.keys()))

    for category in all_categories:
        merged[category] = {}
        if category in existing:
            merged[category].update(existing[category])
        if category in new_data:
            for date_key, day_data in new_data[category].items():
                if date_key not in merged[category]:
                    merged[category][date_key] = day_data

    return merged


def update_history(all_data, date_key):
    """현재 데이터를 히스토리에 추가"""
    log("  [히스토리] 랭킹 변동 데이터 업데이트...")

    history = load_history()

    existing_dates = set()
    for cat in history.values():
        existing_dates.update(cat.keys())

    if not existing_dates:
        recovered = recover_history_from_excel_files()
        if recovered:
            history = merge_history(history, recovered)
            rec_dates = set()
            for cat in recovered.values():
                rec_dates.update(cat.keys())
            log(f"  [히스토리 복구] {len(rec_dates)}일치 데이터 복구됨")

    for sheet_name, products in all_data.items():
        category = sheet_name
        if category not in history:
            history[category] = {}

        day_data = {}
        for p in products:
            day_data[p['name'][:40]] = {
                'rank': p['rank'],
                'brand': p['brand'],
                'item_type': p['item_type'],
                'sale_price': str(p['sale_price']),
                'discount_rate': str(p['discount_rate']),
                'rating': str(p['rating']),
                'review_count': str(p['review_count']),
            }

        history[category][date_key] = day_data

    save_history(history)

    all_dates = set()
    for cat in history.values():
        all_dates.update(cat.keys())
    sorted_dates = sorted(all_dates)

    log(f"  [히스토리] 총 {len(sorted_dates)}일치 데이터 보유")
    if len(sorted_dates) > 1:
        log(f"  [히스토리] 기간: {sorted_dates[0]} ~ {sorted_dates[-1]}")

    return history


def analyze_ranking_changes(history):
    """랭킹 변동 분석 (최근 2회 비교)"""
    log("  [분석6] 랭킹 변동 분석...")

    results = {}

    for category, date_dict in history.items():
        dates = sorted(date_dict.keys())
        if len(dates) < 2:
            results[category] = {
                'status': 'first_data',
                'message': f'첫 번째 수집 데이터 (비교 불가, {dates[0] if dates else "N/A"})',
                'changes': []
            }
            if dates:
                current = date_dict[dates[-1]]
                for name, info in current.items():
                    results[category]['changes'].append({
                        'name': name[:30],
                        'current_rank': info['rank'],
                        'rank_change': 0,
                        'status': '초회',
                        'brand': info.get('brand', ''),
                        'item_type': info['item_type'],
                        'sale_price': info.get('sale_price', ''),
                    })
                results[category]['changes'].sort(key=lambda x: x['current_rank'])
            continue

        current_date = dates[-1]
        prev_date = dates[-2]
        current = date_dict[current_date]
        prev = date_dict[prev_date]

        changes = []
        for name, info in current.items():
            current_rank = info['rank']
            if name in prev:
                prev_rank = prev[name]['rank']
                rank_change = prev_rank - current_rank
                status = '상승' if rank_change > 0 else ('하락' if rank_change < 0 else '유지')
            else:
                rank_change = 0
                status = '신규진입'

            changes.append({
                'name': name[:30],
                'current_rank': current_rank,
                'rank_change': rank_change,
                'status': status,
                'brand': info.get('brand', ''),
                'item_type': info['item_type'],
                'sale_price': info.get('sale_price', ''),
            })

        dropped = []
        for name in prev:
            if name not in current:
                dropped.append({
                    'name': name[:30],
                    'prev_rank': prev[name]['rank'],
                    'brand': prev[name].get('brand', ''),
                    'item_type': prev[name]['item_type'],
                })

        changes.sort(key=lambda x: x['current_rank'])

        results[category] = {
            'status': 'compared',
            'current_date': current_date,
            'prev_date': prev_date,
            'changes': changes,
            'dropped': dropped,
            'total_dates': len(dates),
        }

    return results


# ─── 엑셀 출력 ───

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def create_analysis_excel(type_dist, top_items, brand_dist, discount_analysis,
                           price_band, ranking_changes, date_key):
    """분석 결과를 엑셀로 출력"""
    log("\n============================================================")
    log("[출력] 분석 결과 엑셀 생성")
    log("============================================================")

    wb = Workbook()
    wb.remove(wb.active)

    # ═══ 시트1: 종합 대시보드 ═══
    ws = wb.create_sheet('종합_대시보드')
    log("  -> 시트 [종합_대시보드] 생성")

    row = 1
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row, 1, f'TOPTEN 랭킹 분석 보고서 ({date_key[:4]}.{date_key[4:6]}.{date_key[6:8]})')
    ws.cell(row, 1).font = Font(bold=True, size=14, color='0066CC')
    ws.cell(row, 1).alignment = Alignment(horizontal='center')

    for category in CATEGORIES:
        if category not in type_dist:
            continue

        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row, 1, f'■ {category} 카테고리 ({type_dist[category]["total"]}개 상품)')
        ws.cell(row, 1).font = Font(bold=True, size=12, color='C00000')

        row += 1
        headers = ['순번', '아이템타입', '상품수', '비중(%)', '평균가격', '평균평점', '평균리뷰수', '대표상품(순위:상품명)']
        for col, h in enumerate(headers, 1):
            ws.cell(row, col, h)
        apply_header_style(ws, row, len(headers))

        data = type_dist[category]['analysis']
        for idx, item in enumerate(data, 1):
            row += 1
            ws.cell(row, 1, idx)
            ws.cell(row, 2, item['item_type'])
            ws.cell(row, 3, item['count'])
            ws.cell(row, 4, item['pct'])
            ws.cell(row, 5, f"{item['avg_price']:,}원" if item['avg_price'] > 0 else '-')
            ws.cell(row, 6, item['avg_rating'] if item['avg_rating'] > 0 else '-')
            ws.cell(row, 7, item['avg_review'] if item['avg_review'] > 0 else '-')
            ws.cell(row, 8, item['top_products'])
            apply_data_style(ws, row, len(headers))

            if item['pct'] >= 10:
                for col in range(1, len(headers) + 1):
                    ws.cell(row, col).fill = HIGHLIGHT_FILL

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 55

    # ═══ 시트2: 핵심 아이템 ═══
    ws2 = wb.create_sheet('핵심_아이템')
    log("  -> 시트 [핵심_아이템] 생성")

    row = 1
    for category in CATEGORIES:
        if category not in top_items:
            continue

        if row > 1:
            row += 1
        ws2.merge_cells(f'A{row}:I{row}')
        ws2.cell(row, 1, f'■ {category} 핵심 아이템 TOP 20')
        ws2.cell(row, 1).font = Font(bold=True, size=12, color='C00000')

        row += 1
        headers = ['순위', '브랜드', '상품명', '아이템타입', '할인가', '정가', '할인율', '평점', '영향력점수']
        for col, h in enumerate(headers, 1):
            ws2.cell(row, col, h)
        apply_header_style(ws2, row, len(headers))

        items = top_items[category][:20]
        for item in items:
            row += 1
            ws2.cell(row, 1, item['rank'])
            ws2.cell(row, 2, item['brand'])
            ws2.cell(row, 3, item['name'][:35])
            ws2.cell(row, 4, item['item_type'])
            ws2.cell(row, 5, item['sale_price'])
            ws2.cell(row, 6, item['original_price'])
            ws2.cell(row, 7, item['discount_rate'])
            ws2.cell(row, 8, item['rating'])
            ws2.cell(row, 9, item['impact_score'])
            apply_data_style(ws2, row, len(headers))

            if item['rank'] <= 5:
                for col in range(1, len(headers) + 1):
                    ws2.cell(row, col).fill = HIGHLIGHT_FILL

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 38
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 8
    ws2.column_dimensions['I'].width = 10

    # ═══ 시트3: 브랜드별 분석 ═══
    ws3 = wb.create_sheet('브랜드별_분석')
    log("  -> 시트 [브랜드별_분석] 생성")

    row = 1
    ws3.merge_cells(f'A{row}:F{row}')
    ws3.cell(row, 1, '브랜드별 랭킹 분석')
    ws3.cell(row, 1).font = Font(bold=True, size=12, color='0066CC')

    for category in CATEGORIES:
        if category not in brand_dist:
            continue

        row += 2
        ws3.merge_cells(f'A{row}:F{row}')
        ws3.cell(row, 1, f'■ {category}')
        ws3.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        for col, h in enumerate(['브랜드', '상품수', '비중(%)', '평균가격', '최고순위', '대표상품'], 1):
            ws3.cell(row, col, h)
        apply_header_style(ws3, row, 6)

        for b in brand_dist[category]['brands']:
            row += 1
            ws3.cell(row, 1, b['brand'])
            ws3.cell(row, 2, b['count'])
            ws3.cell(row, 3, b['pct'])
            ws3.cell(row, 4, f"{b['avg_price']:,}원" if b['avg_price'] > 0 else '-')
            ws3.cell(row, 5, b['best_rank'])
            ws3.cell(row, 6, b['top_products'])
            apply_data_style(ws3, row, 6)

    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 8
    ws3.column_dimensions['C'].width = 8
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 8
    ws3.column_dimensions['F'].width = 50

    # ═══ 시트4: 할인율 분석 ═══
    ws4 = wb.create_sheet('할인율_분석')
    log("  -> 시트 [할인율_분석] 생성")

    row = 1
    ws4.merge_cells(f'A{row}:E{row}')
    ws4.cell(row, 1, '할인율 분포 분석')
    ws4.cell(row, 1).font = Font(bold=True, size=12, color='0066CC')

    for category in CATEGORIES:
        if category not in discount_analysis:
            continue

        row += 2
        ws4.merge_cells(f'A{row}:E{row}')
        ws4.cell(row, 1, f'■ {category}')
        ws4.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        for col, h in enumerate(['할인구간', '상품수', '비중(%)', '구성바', '대표상품'], 1):
            ws4.cell(row, col, h)
        apply_header_style(ws4, row, 5)

        for label, count, pct, items in discount_analysis[category]['bands']:
            row += 1
            ws4.cell(row, 1, label)
            ws4.cell(row, 2, count)
            ws4.cell(row, 3, pct)
            bar = '█' * int(pct / 5) + '░' * max(0, 20 - int(pct / 5))
            ws4.cell(row, 4, bar)
            ws4.cell(row, 5, ', '.join(items) if items else '')
            apply_data_style(ws4, row, 5)

    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 8
    ws4.column_dimensions['C'].width = 8
    ws4.column_dimensions['D'].width = 25
    ws4.column_dimensions['E'].width = 55

    # ═══ 시트5: 가격대별 분석 ═══
    ws5 = wb.create_sheet('가격대별_분석')
    log("  -> 시트 [가격대별_분석] 생성")

    row = 1
    ws5.merge_cells(f'A{row}:E{row}')
    ws5.cell(row, 1, '가격대별 랭킹 분포')
    ws5.cell(row, 1).font = Font(bold=True, size=12, color='0066CC')

    for category in CATEGORIES:
        if category not in price_band:
            continue

        row += 2
        ws5.merge_cells(f'A{row}:E{row}')
        ws5.cell(row, 1, f'■ {category}')
        ws5.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        for col, h in enumerate(['가격대', '상품수', '비중(%)', '구성바', '대표상품'], 1):
            ws5.cell(row, col, h)
        apply_header_style(ws5, row, 5)

        for label, count, pct, items in price_band[category]['bands']:
            row += 1
            ws5.cell(row, 1, label)
            ws5.cell(row, 2, count)
            ws5.cell(row, 3, pct)
            bar = '█' * int(pct / 5) + '░' * max(0, 20 - int(pct / 5))
            ws5.cell(row, 4, bar)
            ws5.cell(row, 5, ', '.join(items) if items else '')
            apply_data_style(ws5, row, 5)

    ws5.column_dimensions['A'].width = 14
    ws5.column_dimensions['B'].width = 8
    ws5.column_dimensions['C'].width = 8
    ws5.column_dimensions['D'].width = 25
    ws5.column_dimensions['E'].width = 50

    # ═══ 시트6: 랭킹 변동 현황 ═══
    ws6 = wb.create_sheet('랭킹_변동현황')
    log("  -> 시트 [랭킹_변동현황] 생성")

    row = 1
    ws6.merge_cells(f'A{row}:G{row}')
    ws6.cell(row, 1, '랭킹 변동 현황')
    ws6.cell(row, 1).font = Font(bold=True, size=12, color='0066CC')

    for category in CATEGORIES:
        if category not in ranking_changes:
            continue

        rc = ranking_changes[category]
        row += 2
        ws6.merge_cells(f'A{row}:G{row}')

        if rc['status'] == 'first_data':
            ws6.cell(row, 1, f'■ {category} - {rc["message"]}')
            ws6.cell(row, 1).font = Font(bold=True, size=11, color='808080')

            row += 1
            ws6.cell(row, 1, '※ 2회 이상 수집 시 변동 비교가 가능합니다.')
            ws6.cell(row, 1).font = Font(italic=True, color='666666')

            row += 1
            headers = ['현재순위', '브랜드', '상품명', '아이템타입', '가격', '변동', '상태']
            for col, h in enumerate(headers, 1):
                ws6.cell(row, col, h)
            apply_header_style(ws6, row, len(headers))

            for item in rc.get('changes', []):
                row += 1
                ws6.cell(row, 1, item['current_rank'])
                ws6.cell(row, 2, item.get('brand', ''))
                ws6.cell(row, 3, item['name'])
                ws6.cell(row, 4, item['item_type'])
                ws6.cell(row, 5, item.get('sale_price', ''))
                ws6.cell(row, 6, '-')
                ws6.cell(row, 7, '초회')
                apply_data_style(ws6, row, len(headers))
            continue

        ws6.cell(row, 1, f'■ {category} ({rc["prev_date"]} → {rc["current_date"]}, 총 {rc["total_dates"]}회 수집)')
        ws6.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        ups = sum(1 for c in rc['changes'] if c['status'] == '상승')
        downs = sum(1 for c in rc['changes'] if c['status'] == '하락')
        news = sum(1 for c in rc['changes'] if c['status'] == '신규진입')
        sames = sum(1 for c in rc['changes'] if c['status'] == '유지')
        drops = len(rc.get('dropped', []))

        summary = f'상승 {ups}개 | 하락 {downs}개 | 유지 {sames}개 | 신규진입 {news}개 | 이탈 {drops}개'
        ws6.cell(row, 1, summary)
        ws6.cell(row, 1).font = Font(bold=True, size=10)

        row += 1
        headers = ['현재순위', '브랜드', '상품명', '아이템타입', '가격', '변동', '상태']
        for col, h in enumerate(headers, 1):
            ws6.cell(row, col, h)
        apply_header_style(ws6, row, len(headers))

        for item in rc['changes']:
            row += 1
            ws6.cell(row, 1, item['current_rank'])
            ws6.cell(row, 2, item.get('brand', ''))
            ws6.cell(row, 3, item['name'])
            ws6.cell(row, 4, item['item_type'])
            ws6.cell(row, 5, item.get('sale_price', ''))

            change = item['rank_change']
            if item['status'] == '신규진입':
                ws6.cell(row, 6, 'NEW')
                ws6.cell(row, 6).font = NEW_FONT
                ws6.cell(row, 7, '신규')
                ws6.cell(row, 7).font = NEW_FONT
            elif change > 0:
                ws6.cell(row, 6, f'▲{change}')
                ws6.cell(row, 6).font = UP_FONT
                ws6.cell(row, 7, '상승')
                ws6.cell(row, 7).font = UP_FONT
            elif change < 0:
                ws6.cell(row, 6, f'▼{abs(change)}')
                ws6.cell(row, 6).font = DOWN_FONT
                ws6.cell(row, 7, '하락')
                ws6.cell(row, 7).font = DOWN_FONT
            else:
                ws6.cell(row, 6, '-')
                ws6.cell(row, 7, '유지')

            apply_data_style(ws6, row, len(headers))

        if rc.get('dropped'):
            row += 1
            ws6.merge_cells(f'A{row}:G{row}')
            ws6.cell(row, 1, '  ▼ 이탈 상품')
            ws6.cell(row, 1).font = Font(bold=True, size=10, color='808080')

            for item in rc['dropped']:
                row += 1
                ws6.cell(row, 1, '-')
                ws6.cell(row, 2, item.get('brand', ''))
                ws6.cell(row, 3, item['name'])
                ws6.cell(row, 4, item['item_type'])
                ws6.cell(row, 5, '')
                ws6.cell(row, 6, f'이전: {item["prev_rank"]}위')
                ws6.cell(row, 7, '이탈')
                ws6.cell(row, 7).font = Font(color='808080')
                apply_data_style(ws6, row, 7)

    ws6.column_dimensions['A'].width = 10
    ws6.column_dimensions['B'].width = 15
    ws6.column_dimensions['C'].width = 35
    ws6.column_dimensions['D'].width = 14
    ws6.column_dimensions['E'].width = 12
    ws6.column_dimensions['F'].width = 10
    ws6.column_dimensions['G'].width = 10

    # ═══ 시트7: 아이템비중 차트 ═══
    ws7 = wb.create_sheet('아이템비중_차트')
    log("  -> 시트 [아이템비중_차트] 생성")

    col_offset = 0
    for category in CATEGORIES:
        if category not in type_dist:
            continue

        start_col = col_offset * 4 + 1
        ws7.cell(1, start_col, f'{category} 아이템타입 비중')
        ws7.cell(1, start_col).font = Font(bold=True, size=11)

        ws7.cell(2, start_col, '아이템타입')
        ws7.cell(2, start_col + 1, '상품수')
        ws7.cell(2, start_col + 2, '비중(%)')

        for r, item in enumerate(type_dist[category]['analysis'], 3):
            ws7.cell(r, start_col, item['item_type'])
            ws7.cell(r, start_col + 1, item['count'])
            ws7.cell(r, start_col + 2, item['pct'])

        data_count = len(type_dist[category]['analysis'])
        if data_count > 0:
            pie = PieChart()
            pie.title = f'{category} 아이템타입 비중'
            pie.style = 26
            pie.width = 18
            pie.height = 13

            data_ref = Reference(ws7, min_col=start_col + 1, min_row=2, max_row=2 + data_count)
            cats_ref = Reference(ws7, min_col=start_col, min_row=3, max_row=2 + data_count)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)

            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showVal = False
            pie.dataLabels.showCatName = True

            chart_row = data_count + 5
            ws7.add_chart(pie, f'{get_column_letter(start_col)}{chart_row}')

        col_offset += 1

    # ═══ 시트8: 브랜드비중 차트 ═══
    ws8 = wb.create_sheet('브랜드비중_차트')
    log("  -> 시트 [브랜드비중_차트] 생성")

    col_offset = 0
    for category in CATEGORIES:
        if category not in brand_dist:
            continue

        start_col = col_offset * 4 + 1
        ws8.cell(1, start_col, f'{category} 브랜드 비중')
        ws8.cell(1, start_col).font = Font(bold=True, size=11)

        ws8.cell(2, start_col, '브랜드')
        ws8.cell(2, start_col + 1, '상품수')
        ws8.cell(2, start_col + 2, '비중(%)')

        for r, brand in enumerate(brand_dist[category]['brands'][:10], 3):
            ws8.cell(r, start_col, brand['brand'])
            ws8.cell(r, start_col + 1, brand['count'])
            ws8.cell(r, start_col + 2, brand['pct'])

        data_count = min(len(brand_dist[category]['brands']), 10)
        if data_count > 0:
            pie = PieChart()
            pie.title = f'{category} 브랜드 비중'
            pie.style = 26
            pie.width = 18
            pie.height = 13

            data_ref = Reference(ws8, min_col=start_col + 1, min_row=2, max_row=2 + data_count)
            cats_ref = Reference(ws8, min_col=start_col, min_row=3, max_row=2 + data_count)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)

            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showVal = False
            pie.dataLabels.showCatName = True

            chart_row = data_count + 5
            ws8.add_chart(pie, f'{get_column_letter(start_col)}{chart_row}')

        col_offset += 1

    # ═══ 시트9: 히스토리 원본 데이터 ═══
    ws9 = wb.create_sheet('수집_히스토리')
    log("  -> 시트 [수집_히스토리] 생성")

    history = load_history()
    row = 1
    ws9.cell(row, 1, '날짜별 수집 히스토리 (랭킹 추적용)')
    ws9.cell(row, 1).font = Font(bold=True, size=12, color='0066CC')

    row += 1
    ws9.cell(row, 1, f'총 수집 횟수: {len(set(d for cat in history.values() for d in cat))}회')
    ws9.cell(row, 1).font = Font(italic=True, color='666666')

    for category in CATEGORIES:
        if category not in history:
            continue

        dates = sorted(history[category].keys())
        row += 2
        ws9.merge_cells(f'A{row}:{get_column_letter(3 + len(dates))}{row}')
        ws9.cell(row, 1, f'■ {category} 랭킹 추이')
        ws9.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        ws9.cell(row, 1, '상품명')
        ws9.cell(row, 2, '브랜드')
        ws9.cell(row, 3, '아이템타입')
        for di, d in enumerate(dates):
            ws9.cell(row, 4 + di, f'{d[4:6]}/{d[6:8]}')
        apply_header_style(ws9, row, 3 + len(dates))

        all_products_set = {}
        for d in dates:
            for name, info in history[category][d].items():
                if name not in all_products_set:
                    all_products_set[name] = {
                        'brand': info.get('brand', ''),
                        'item_type': info['item_type']
                    }

        for name in sorted(all_products_set.keys()):
            row += 1
            ws9.cell(row, 1, name[:35])
            ws9.cell(row, 2, all_products_set[name]['brand'])
            ws9.cell(row, 3, all_products_set[name]['item_type'])
            for di, d in enumerate(dates):
                if name in history[category][d]:
                    ws9.cell(row, 4 + di, history[category][d][name]['rank'])
                else:
                    ws9.cell(row, 4 + di, '-')
            apply_data_style(ws9, row, 3 + len(dates))

    ws9.column_dimensions['A'].width = 38
    ws9.column_dimensions['B'].width = 15
    ws9.column_dimensions['C'].width = 14

    # ═══ 저장 ═══
    filename = f'탑텐_랭킹분석_{date_key}.xlsx'
    filepath = os.path.join(WORK_DIR, filename)

    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime('%H%M%S')
        filename = f'탑텐_랭킹분석_{date_key}_{ts}.xlsx'
        filepath = os.path.join(WORK_DIR, filename)
        wb.save(filepath)

    log(f"\n  [OK] 저장 완료: {filename}")
    return filepath


# ─── 콘솔 요약 출력 ───

def print_summary(type_dist, top_items, brand_dist, price_band):
    """콘솔에 핵심 요약 출력"""
    log("\n" + "=" * 60)
    log("  TOPTEN 랭킹 분석 요약")
    log("=" * 60)

    for category in CATEGORIES:
        if category not in type_dist:
            continue

        log(f"\n  ■ {category} ({type_dist[category]['total']}개)")
        log(f"  {'─' * 50}")

        log("  [아이템타입 비중 TOP5]")
        for item in type_dist[category]['analysis'][:5]:
            bar = '█' * int(item['pct'] / 3)
            log(f"    {item['item_type']:12s} {item['count']:2d}개 ({item['pct']:4.1f}%) {bar}")

        if category in brand_dist:
            log(f"\n  [브랜드 비중 TOP5]")
            for b in brand_dist[category]['brands'][:5]:
                log(f"    {b['brand']:15s} {b['count']:2d}개 ({b['pct']:4.1f}%)")

        if category in top_items:
            log(f"\n  [핵심 아이템 TOP5]")
            for idx, item in enumerate(top_items[category][:5], 1):
                log(f"    {idx}. [{item['brand']}] {item['name'][:25]:25s} {item['sale_price']}")

        if category in price_band:
            log(f"\n  [주력 가격대]")
            top_bands = sorted(price_band[category]['bands'], key=lambda x: x[1], reverse=True)[:3]
            for label, count, pct, _ in top_bands:
                log(f"    {label:12s} {count:2d}개 ({pct:.1f}%)")


# ─── 메인 ───

def main():
    log("=" * 60)
    log("  TOPTEN 랭킹 분석 도구")
    log("=" * 60)

    log("\n[1/5] 크롤링 데이터 로드")
    filepath = find_latest_ranking_file()
    if not filepath:
        return

    all_data = load_ranking_data(filepath)
    total_products = sum(len(v) for v in all_data.values())
    log(f"  총 {len(all_data)}개 시트, {total_products}개 상품 로드 완료")

    date_key = extract_date_from_filename(filepath)

    log(f"\n[2/5] 분석 실행 (날짜: {date_key})")
    type_dist = analyze_item_type_distribution(all_data)
    top_items = analyze_top_items(all_data)
    brand_dist = analyze_brand_distribution(all_data)
    discount_analysis = analyze_discount(all_data)
    price_band = analyze_price_band(all_data)

    log(f"\n[3/5] 히스토리 업데이트")
    history = update_history(all_data, date_key)
    ranking_changes = analyze_ranking_changes(history)

    log(f"\n[4/5] 분석 결과 요약")
    print_summary(type_dist, top_items, brand_dist, price_band)

    log(f"\n[5/5] 엑셀 출력")
    output_path = create_analysis_excel(
        type_dist, top_items, brand_dist, discount_analysis,
        price_band, ranking_changes, date_key
    )

    log("\n" + "=" * 60)
    log("  분석 완료!")
    log(f"  파일: {os.path.basename(output_path)}")
    log("=" * 60)


if __name__ == '__main__':
    main()
