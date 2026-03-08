# -*- coding: utf-8 -*-
"""
미쏘 이번주 베스트 크롤러 (Selenium + 스크린샷 캡처)
- URL: https://mixxo.com/product/list.html?cate_no=45
- 여성 전용 브랜드 → 시트 1개(여성)
- 2페이지까지 크롤링 (약 40개 상품)
"""

import sys
import io
import urllib.parse
import urllib.request
import urllib.robotparser
from PIL import Image as PILImage

# 이미지 설정
IMG_WIDTH = 80
IMG_HEIGHT = 107
HD_IMG_WIDTH = 400   # 대시보드용 고해상도 이미지
HD_IMG_HEIGHT = 534  # 3:4 비율
ROW_HEIGHT = 85

# 안전 크롤링 설정
LOG_FILE = "mixxo_crawler_log.txt"
SAFE_MODE = True
REQUEST_DELAY = 1.5
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

BASE_URL = "https://mixxo.com/product/list.html?cate_no=45"


def log(msg, end='\n'):
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + end)
    try:
        print(msg, end=end, flush=True)
    except:
        pass


def check_robots_allowed(url):
    if not SAFE_MODE:
        return True
    try:
        parsed = urllib.parse.urlparse(url)
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        rp = urllib.robotparser.RobotFileParser()
        with urllib.request.urlopen(robots_url, timeout=5) as resp:
            content = resp.read().decode('utf-8', 'ignore')
        rp.parse(content.splitlines())
        return rp.can_fetch(USER_AGENT, url)
    except Exception:
        return True


def safe_get(driver, url):
    if SAFE_MODE and not check_robots_allowed(url):
        log(f"  -> robots.txt 제한으로 접근 건너뜀: {url}")
        return False
    time.sleep(REQUEST_DELAY)
    driver.get(url)
    return True


# 로그 초기화
with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write("")

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import time
import os
import re
import hashlib
from datetime import datetime


def capture_image_from_element(element):
    """요소를 스크린샷 캡처하여 이미지로 반환"""
    try:
        png_data = element.screenshot_as_png
        if not png_data:
            return None
        img = PILImage.open(io.BytesIO(png_data))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        hd_img = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        hd_bytes = io.BytesIO()
        hd_img.save(hd_bytes, format='JPEG', quality=92)
        hd_bytes.seek(0)

        xl_img = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        xl_bytes = io.BytesIO()
        xl_img.save(xl_bytes, format='JPEG', quality=85)
        xl_bytes.seek(0)

        return (xl_bytes, hd_bytes)
    except Exception:
        return None


def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)

    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument(f'--user-agent={USER_AGENT}')
    options.page_load_strategy = 'eager'

    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작...")
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(60)
    driver.set_script_timeout(30)
    log("  [OK] 완료!\n")
    return driver


def extract_products_from_page(driver, start_rank=1):
    """현재 페이지에서 상품 데이터 추출"""
    products = []

    # 스크롤하여 모든 이미지 로드
    for i in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.8)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)

    # Cafe24 상품 리스트 셀렉터 시도
    selectors = [
        "ul.prdList > li",
        ".xans-product-listnormal li.xans-record-",
        "#contents ul > li[id^='anchorBoxId']",
        ".product-list li.item",
        "ul.grid4 > li",
        "ul.grid3 > li",
        "#prdList > li",
    ]

    items = []
    for sel in selectors:
        items = driver.find_elements(By.CSS_SELECTOR, sel)
        if len(items) >= 5:
            log(f"      [DEBUG] 셀렉터 '{sel}'로 {len(items)}개 상품 발견")
            break

    if not items:
        # 폴백: 링크 기반으로 상품 추출
        log("      [WARN] 리스트 셀렉터 실패, 링크 기반 추출 시도")
        items = driver.find_elements(By.CSS_SELECTOR, "div.thumbnail a, a[href*='/product/']")
        if not items:
            return products

    img_success = 0

    for idx, item in enumerate(items):
        rank = start_rank + idx
        product = {
            'rank': rank,
            'name': '',
            'sale_price': '',
            'original_price': '',
            'discount_rate': '',
            'review_count': '없음',
            'image_data': None,
            'hd_image_data': None,
        }

        try:
            # 상품명 추출
            name_selectors = [
                ".description .name a",
                ".name a",
                "p.name a",
                ".prd-name a",
                ".description a",
            ]
            for ns in name_selectors:
                try:
                    name_elem = item.find_element(By.CSS_SELECTOR, ns)
                    raw_name = name_elem.text.strip()
                    if raw_name:
                        product['name'] = raw_name
                        break
                except:
                    continue

            if not product['name']:
                # 대안: 전체 텍스트에서 상품명 추출
                try:
                    desc = item.find_element(By.CSS_SELECTOR, ".description, .info")
                    lines = desc.text.strip().split('\n')
                    if lines:
                        product['name'] = lines[0].strip()
                except:
                    pass

            if not product['name']:
                continue

            # 가격 추출 (mixxo Cafe24 구조)
            # 할인가: <span class="sale ">92,900</span>
            # 정가: <span class="price  through">99,900</span>
            try:
                sale_elem = item.find_element(By.CSS_SELECTOR, "span.sale")
                sale_text = sale_elem.text.strip()
                if sale_text:
                    nums = re.findall(r'[\d,]+', sale_text)
                    if nums:
                        product['sale_price'] = f"{int(nums[0].replace(',', '')):,}원"
            except:
                pass

            try:
                orig_elem = item.find_element(By.CSS_SELECTOR, "span.price.through, span.price")
                orig_text = orig_elem.text.strip()
                if orig_text:
                    nums = re.findall(r'[\d,]+', orig_text)
                    if nums:
                        product['original_price'] = f"{int(nums[0].replace(',', '')):,}원"
            except:
                pass

            # 가격이 아직 없으면 전체 텍스트에서 추출
            if not product['sale_price']:
                try:
                    price_area = item.find_element(By.CSS_SELECTOR, ".description, .price_box")
                    price_text = price_area.text
                    prices = re.findall(r'([\d,]+)', price_text)
                    # 숫자 중 4자리 이상만 가격으로 간주
                    price_nums = [p for p in prices if len(p.replace(',', '')) >= 4]
                    if len(price_nums) >= 2:
                        product['sale_price'] = f"{int(price_nums[0].replace(',', '')):,}원"
                        product['original_price'] = f"{int(price_nums[1].replace(',', '')):,}원"
                    elif len(price_nums) == 1:
                        product['sale_price'] = f"{int(price_nums[0].replace(',', '')):,}원"
                except:
                    pass

            # 할인율: <span class="sale_text">7%</span>
            try:
                rate_elem = item.find_element(By.CSS_SELECTOR, ".sale_text, #sale_bg .sale_text")
                rate_text = rate_elem.text.strip()
                rate_match = re.search(r'(\d+)\s*%', rate_text)
                if rate_match:
                    product['discount_rate'] = f"{rate_match.group(1)}%"
            except:
                pass

            # 할인율 계산 (추출 못한 경우)
            if not product['discount_rate'] and product['sale_price'] and product['original_price']:
                try:
                    sp = int(re.sub(r'[^\d]', '', product['sale_price']))
                    op = int(re.sub(r'[^\d]', '', product['original_price']))
                    if op > sp > 0:
                        product['discount_rate'] = f"{int((1 - sp/op) * 100)}%"
                except:
                    pass

            # 리뷰수: "리뷰 48" 형식
            try:
                item_text = item.text
                rev_match = re.search(r'리뷰\s*(\d+)', item_text)
                if rev_match:
                    product['review_count'] = rev_match.group(1)
            except:
                pass

            # 이미지 캡처
            try:
                img_selectors = [".thumbnail img", ".prdImg img", "img[src*='product']",
                                 ".thumbnail a img", "img"]
                for ims in img_selectors:
                    try:
                        img_elem = item.find_element(By.CSS_SELECTOR, ims)
                        if img_elem.size['height'] > 50:
                            driver.execute_script(
                                "arguments[0].scrollIntoView({block: 'center'});", img_elem)
                            time.sleep(0.15)
                            img_data = capture_image_from_element(img_elem)
                            if img_data:
                                product['image_data'] = img_data[0]
                                product['hd_image_data'] = img_data[1]
                                img_success += 1
                            break
                    except:
                        continue
            except:
                pass

            products.append(product)
            if (rank % 5) == 0:
                log(f"      -> {rank}번째 상품까지 처리 ({len(products)}개 수집)")

        except Exception as e:
            log(f"      [WARN] 상품 {rank} 추출 오류: {str(e)[:60]}")
            continue

    log(f"      -> {len(products)}개 수집 (이미지: {img_success}개)")

    if products:
        for p in products[:2]:
            name_short = p['name'][:20] if len(p['name']) > 20 else p['name']
            has_img = "O" if p['image_data'] else "X"
            log(f"        {p['rank']}. {name_short:20s} | {p['sale_price']:10s} | img:{has_img}")

    return products


def scrape_all(driver):
    """미쏘 베스트 상품 크롤링 (2페이지)"""
    log(f"\n{'='*60}")
    log(f"[2/4] 데이터 수집 시작")
    log(f"{'='*60}")

    all_products = []

    for page in [1, 2]:
        url = f"{BASE_URL}&page={page}" if page > 1 else BASE_URL
        log(f"\n  [페이지 {page}] {url}")

        if not safe_get(driver, url):
            continue

        log(f"  -> 페이지 로딩 대기...", end='')
        time.sleep(4)
        log(" OK!")

        # 팝업 닫기
        try:
            driver.execute_script("""
                var popups = document.querySelectorAll(
                    '[class*="popup"], [class*="modal"], [id*="popup"], [class*="layer"]');
                popups.forEach(function(p) { p.style.display = 'none'; });
            """)
        except:
            pass

        start_rank = len(all_products) + 1
        products = extract_products_from_page(driver, start_rank)
        all_products.extend(products)

        time.sleep(REQUEST_DELAY)

    return {'여성': all_products}


def create_excel(all_data, filename):
    """엑셀 파일 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성 (이미지 포함)")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['순위', '이미지', '상품명', '할인가', '정가', '할인율', '리뷰수']
    col_widths = [6, 12, 45, 15, 15, 10, 10]

    for sheet_name, products in all_data.items():
        ws = wb.create_sheet(title=sheet_name)

        log(f"  -> 시트 [{sheet_name}]: {len(products)}개 상품 (이미지 삽입 중...)")

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 헤더
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 25

        # 데이터
        for row_idx, product in enumerate(products, 2):
            ws.row_dimensions[row_idx].height = ROW_HEIGHT

            # 순위
            cell = ws.cell(row=row_idx, column=1, value=product['rank'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # 이미지
            if product.get('image_data'):
                try:
                    img = XLImage(product['image_data'])
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    ws.add_image(img, f"B{row_idx}")
                except:
                    pass
            cell = ws.cell(row=row_idx, column=2, value='')
            cell.border = border

            # 상품명
            cell = ws.cell(row=row_idx, column=3, value=product['name'])
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border

            # 할인가
            cell = ws.cell(row=row_idx, column=4, value=product['sale_price'])
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.font = Font(bold=True, color="FF0000")
            cell.border = border

            # 정가
            cell = ws.cell(row=row_idx, column=5, value=product['original_price'])
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border

            # 할인율
            cell = ws.cell(row=row_idx, column=6, value=product['discount_rate'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(color="FF6600")
            cell.border = border

            # 리뷰수
            cell = ws.cell(row=row_idx, column=7, value=product['review_count'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    wb.save(filename)
    log(f"  [OK] 저장 완료!")
    return True


def _save_hd_images(all_data, brand_name, excel_filename):
    """대시보드용 고해상도 이미지 저장"""
    hd_dir = os.path.join(os.path.dirname(os.path.abspath(excel_filename)), 'product_images_hd')
    os.makedirs(hd_dir, exist_ok=True)

    file_hash = hashlib.md5(
        f"{os.path.basename(excel_filename)}_{datetime.now().strftime('%Y%m%d')}".encode()
    ).hexdigest()[:8]

    saved = 0
    for sheet_name, products in all_data.items():
        for p in products:
            hd_data = p.get('hd_image_data')
            if hd_data:
                name = p.get('name', '')[:20]
                safe = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
                fname = f"{file_hash}_{brand_name}_{sheet_name}_{p['rank']}_{safe}.jpg"
                fpath = os.path.join(hd_dir, fname)
                if not os.path.exists(fpath):
                    try:
                        hd_data.seek(0)
                        with open(fpath, 'wb') as f:
                            f.write(hd_data.read())
                        saved += 1
                    except:
                        pass
    if saved > 0:
        log(f"  [HD] 고해상도 이미지 {saved}개 저장 → {hd_dir}")


def main():
    log("\n" + "=" * 60)
    log("  미쏘 이번주 베스트 크롤러")
    log("=" * 60)
    log("  * 이미지를 스크린샷 캡처 방식으로 수집합니다")
    log("=" * 60)

    driver = None
    try:
        driver = setup_driver()
        all_data = scrape_all(driver)

        if all_data and any(len(v) > 0 for v in all_data.values()):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"미쏘_이번주베스트_이미지포함_{timestamp}.xlsx"

            create_excel(all_data, filename)
            _save_hd_images(all_data, '미쏘', filename)

            # 통계
            log(f"\n{'='*60}")
            log(f"[4/4] 수집 완료 통계")
            log(f"{'='*60}")

            total = sum(len(p) for p in all_data.values())
            total_images = sum(1 for prods in all_data.values() for p in prods if p.get('image_data'))
            total_prices = sum(1 for prods in all_data.values() for p in prods if p['sale_price'])

            log(f"  총 상품: {total}개")
            log(f"  가격: {total_prices}/{total}개 수집됨")
            log(f"  이미지: {total_images}/{total}개 캡처됨")
            log(f"\n  파일: {filename}")
            log("=" * 60)
        else:
            log("\n[오류] 수집된 데이터가 없습니다.")

    except Exception as e:
        log(f"\n[오류] {e}")
        import traceback
        log(traceback.format_exc())
    finally:
        if driver:
            log("\n브라우저 종료")
            driver.quit()


if __name__ == "__main__":
    main()
