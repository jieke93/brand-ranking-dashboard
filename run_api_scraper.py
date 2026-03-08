"""유니클로 여성 랭킹 상품 정보 수집기 (API 직접 호출 버전)
- 법적 리스크 최소화: 공개 API 사용, 브라우저와 동일한 방식
"""

import requests
import json
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
from PIL import Image as PILImage
import time
from datetime import datetime
import os

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'ko-KR,ko;q=0.9',
    'Referer': 'https://www.uniqlo.com/kr/ko/spl/ranking/women',
    'Origin': 'https://www.uniqlo.com',
}

def fetch_ranking_data():
    """랭킹 API에서 데이터 가져오기"""
    print("🌐 유니클로 API 호출 중...")
    
    # 유니클로 랭킹 API 엔드포인트 (추정)
    # 실제 엔드포인트는 브라우저 개발자 도구에서 확인 필요
    api_urls = [
        "https://www.uniqlo.com/kr/api/commerce/v5/ko/ranking/women",
        "https://www.uniqlo.com/kr/data/ranking/women.json",
        "https://api.uniqlo.com/kr/ranking/women",
    ]
    
    for api_url in api_urls:
        try:
            print(f"  시도: {api_url}")
            response = requests.get(api_url, headers=HEADERS, timeout=10)
            
            if response.status_code == 200:
                print(f"✅ API 응답 성공!")
                return response.json()
            else:
                print(f"  ⚠️ 상태 코드: {response.status_code}")
        
        except Exception as e:
            print(f"  ⚠️ 실패: {str(e)[:50]}")
            continue
    
    return None

def parse_html_manually():
    """HTML을 직접 파싱하여 상품 정보 추출"""
    print("\n🔍 HTML 직접 파싱 시도...")
    
    url = "https://www.uniqlo.com/kr/ko/spl/ranking/women"
    
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.raise_for_status()
        
        html = response.text
        
        # 상품 코드 추출 (E로 시작하는 패턴)
        import re
        product_pattern = r'/products/(E\d{6}-\d{3})/\d{2}'
        matches = re.findall(product_pattern, html)
        
        # 중복 제거
        unique_products = []
        seen = set()
        for match in matches:
            if match not in seen:
                unique_products.append(match)
                seen.add(match)
        
        print(f"✅ {len(unique_products)}개 상품 코드 발견")
        
        products = []
        for rank, product_code in enumerate(unique_products[:50], 1):
            # 상품 상세 API 호출 (추정)
            product_api = f"https://www.uniqlo.com/kr/api/commerce/v5/ko/products/{product_code}"
            
            try:
                time.sleep(0.5)  # 요청 간격
                prod_response = requests.get(product_api, headers=HEADERS, timeout=10)
                
                if prod_response.status_code == 200:
                    prod_data = prod_response.json()
                    
                    name = prod_data.get('name', f'상품 {rank}')
                    price = prod_data.get('prices', {}).get('base', {}).get('value', '가격 정보 없음')
                    image_url = prod_data.get('images', {}).get('main', {}).get('url', '')
                    colors = [c.get('name', '') for c in prod_data.get('colors', [])]
                    
                    products.append({
                        'rank': rank,
                        'name': name,
                        'price': f"{price}원" if isinstance(price, (int, float)) else price,
                        'image_url': image_url,
                        'colors': ', '.join(colors) if colors else '정보 없음'
                    })
                    
                    print(f"  ✓ {rank}위: {name[:40]}")
                
                else:
                    # API가 없으면 기본 정보만
                    products.append({
                        'rank': rank,
                        'name': f'상품 {product_code}',
                        'price': '가격 정보 없음',
                        'image_url': f'https://image.uniqlo.com/UQ/ST3/AsianCommon/imagesgoods/{product_code.split("-")[0][1:]}/item/goods_69_{product_code.split("-")[0][1:]}.jpg',
                        'colors': '정보 없음'
                    })
                    print(f"  ✓ {rank}위: {product_code}")
            
            except Exception as e:
                print(f"  ⚠️ {rank}위 처리 실패")
                continue
        
        return products
    
    except Exception as e:
        print(f"❌ HTML 파싱 실패: {e}")
        return []

def download_image(url):
    """이미지 다운로드"""
    if not url:
        return None
    try:
        time.sleep(0.3)
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        
        if img.mode == 'RGBA':
            bg = PILImage.new('RGB', img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        
        img.thumbnail((100, 100), PILImage.Resampling.LANCZOS)
        
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        return img_bytes
    except:
        return None

def create_excel(products, filename):
    """엑셀 파일 생성"""
    print(f"\n📊 엑셀 파일 생성: {filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "여성 랭킹"
    
    # 헤더
    headers = ['순위', '상품명', '가격', '컬러 옵션', '상품 이미지']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 열 너비
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    
    # 데이터
    for product in products:
        row_idx = ws.max_row + 1
        
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product['colors'])
        
        ws.row_dimensions[row_idx].height = 85
        
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 이미지
        if product.get('image_url'):
            print(f"  📷 {product['rank']}위 이미지 다운로드...")
            img_data = download_image(product['image_url'])
            if img_data:
                try:
                    img = XLImage(img_data)
                    img.width = 80
                    img.height = 80
                    ws.add_image(img, f'E{row_idx}')
                except:
                    pass
        
        ws.cell(row_idx, 5).border = border
    
    # 메타 정보
    meta_row = ws.max_row + 2
    ws.cell(meta_row, 1, f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(meta_row + 1, 1, "출처: 유니클로 공식 온라인스토어")
    ws.cell(meta_row + 2, 1, "⚠️ 개인 사용 목적, 상업적 사용 금지")
    
    wb.save(filename)
    print(f"✅ 엑셀 파일 저장 완료!")

def main():
    print("=" * 70)
    print("📦 유니클로 여성 랭킹 상품 정보 수집기 (API 버전)")
    print("=" * 70)
    print("\n⚖️  개인 사용 목적 / 법적 리스크 최소화")
    print("⚠️  상업적 사용 금지\n")
    
    # API 시도
    data = fetch_ranking_data()
    
    products = []
    
    if data:
        print("✅ API에서 데이터 가져오기 성공")
        
        # API 응답 구조 확인
        print("\n📋 API 응답 구조 확인 중...")
        with open("api_response.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print("💾 API 응답이 'api_response.json'에 저장되었습니다.")
        
        # API 응답 파싱 시도
        try:
            if isinstance(data, dict):
                # 가능한 키들 확인
                if 'result' in data:
                    items = data['result'].get('items', [])
                elif 'products' in data:
                    items = data['products']
                elif 'data' in data:
                    items = data['data']
                elif 'items' in data:
                    items = data['items']
                else:
                    items = []
                
                for rank, item in enumerate(items[:50], 1):
                    name = item.get('name', item.get('productName', f'상품 {rank}'))
                    price = item.get('price', item.get('salePrice', item.get('originPrice', '가격 정보 없음')))
                    image_url = item.get('image', item.get('imageUrl', item.get('thumbnail', '')))
                    
                    if isinstance(image_url, dict):
                        image_url = image_url.get('url', '')
                    
                    colors = item.get('colors', item.get('colorOptions', []))
                    if isinstance(colors, list) and colors:
                        color_names = [c.get('name', c.get('displayName', '')) if isinstance(c, dict) else str(c) for c in colors[:10]]
                        color_str = ', '.join([c for c in color_names if c])
                    else:
                        color_str = '정보 없음'
                    
                    products.append({
                        'rank': rank,
                        'name': name,
                        'price': f"{price}원" if isinstance(price, (int, float)) else str(price),
                        'image_url': image_url,
                        'colors': color_str
                    })
                    print(f"  ✓ {rank}위: {name[:40]}")
        
        except Exception as e:
            print(f"⚠️ API 파싱 오류: {e}")
            import traceback
            traceback.print_exc()
    
    if not products:
        print("\n⚠️ API 접근 실패, HTML 파싱 시도...")
        products = parse_html_manually()
    
    if not products:
        print("\n❌ 데이터 수집 실패")
        print("\n💡 해결 방법:")
        print("  1. 인터넷 연결 확인")
        print("  2. Chrome/Edge 브라우저로 페이지가 정상 작동하는지 확인")
        print("  3. 브라우저 개발자 도구(F12)에서 Network 탭을 열고")
        print("     페이지 로드 시 어떤 API가 호출되는지 확인")
        return
    
    print(f"\n✅ {len(products)}개 상품 수집 완료")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"유니클로_여성랭킹_{timestamp}.xlsx"
    
    create_excel(products, filename)
    
    print("\n" + "=" * 70)
    print("🎉 작업 완료!")
    print(f"📁 파일 위치: {os.path.abspath(filename)}")
    print(f"📊 수집 상품: {len(products)}개")
    print("=" * 70)

if __name__ == "__main__":
    main()
