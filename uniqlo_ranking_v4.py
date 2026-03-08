# -*- coding: utf-8 -*-
"""
유니클로 랭킹 크롤러 V4 (완전판)
- WOMEN, MEN, KIDS, BABY 랭킹 + 하위 탭별 수집
- 컬러 코드 → 컬러명 변환
- 이미지 URL 추출
- 가격 정확히 추출
"""
import sys
sys.stdout.reconfigure(line_buffering=True)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import time
import os
from datetime import datetime
import re

# 유니클로 컬러 코드 → 컬러명 매핑
COLOR_MAP = {
    '00': 'WHITE', '01': 'OFF WHITE', '02': 'LIGHT GRAY', '03': 'GRAY',
    '04': 'DARK GRAY', '05': 'LIGHT GRAY', '06': 'GRAY', '07': 'DARK GRAY',
    '08': 'DARK GRAY', '09': 'BLACK', '10': 'PINK', '11': 'PINK',
    '12': 'LIGHT PINK', '13': 'PINK', '14': 'PINK', '15': 'LIGHT PINK',
    '16': 'DARK PINK', '17': 'PURPLE', '18': 'PURPLE', '19': 'WINE',
    '20': 'ORANGE', '21': 'LIGHT ORANGE', '22': 'ORANGE', '23': 'BROWN',
    '24': 'DARK ORANGE', '25': 'ORANGE', '26': 'BROWN', '27': 'BROWN',
    '28': 'DARK BROWN', '29': 'BROWN', '30': 'NATURAL', '31': 'BEIGE',
    '32': 'BEIGE', '33': 'BEIGE', '34': 'LIGHT BEIGE', '35': 'BEIGE',
    '36': 'BROWN', '37': 'BROWN', '38': 'DARK BROWN', '39': 'KHAKI',
    '40': 'YELLOW', '41': 'LIGHT YELLOW', '42': 'YELLOW', '43': 'MUSTARD',
    '44': 'GOLD', '45': 'LIME', '46': 'OLIVE', '47': 'KHAKI', '48': 'OLIVE',
    '49': 'DARK GREEN', '50': 'LIGHT GREEN', '51': 'GREEN', '52': 'GREEN',
    '53': 'GREEN', '54': 'DARK GREEN', '55': 'GREEN', '56': 'GREEN',
    '57': 'GREEN', '58': 'MINT', '59': 'TURQUOISE', '60': 'LIGHT BLUE',
    '61': 'LIGHT BLUE', '62': 'SKY BLUE', '63': 'LIGHT BLUE', '64': 'BLUE',
    '65': 'BLUE', '66': 'BLUE', '67': 'BLUE', '68': 'BLUE', '69': 'NAVY',
    '70': 'DARK BLUE', '71': 'NAVY', '72': 'DARK BLUE', '73': 'NAVY',
    '74': 'DARK NAVY', '75': 'NAVY', '76': 'NAVY', '77': 'INDIGO',
    '78': 'INDIGO', '79': 'DENIM', '80': 'RED', '81': 'LIGHT RED',
    '82': 'RED', '83': 'RED', '84': 'DARK RED', '85': 'WINE', '86': 'WINE',
    '87': 'BURGUNDY', '88': 'DARK RED', '89': 'WINE', '90': 'SILVER',
    '91': 'GOLD', '92': 'MULTI', '93': 'MULTI', '94': 'PATTERN',
    '95': 'PATTERN', '96': 'STRIPE', '97': 'CHECK', '98': 'PRINT',
    '99': 'OTHER'
}

# 수집할 URL + 탭 정보
CATEGORIES = {
    'WOMEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/women',
        'tabs': ['모두보기', '상의', '팬츠', '드레스 & 스커트', '아우터', '이너웨어', '홈웨어', '악세서리']
    },
    'MEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/men',
        'tabs': ['모두보기', '상의', '팬츠', '아우터', '이너웨어', '홈웨어', '악세서리']
    },
    'KIDS': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/kids',
        'tabs': ['모두보기', '상의', '팬츠', '아우터', '이너웨어']
    },
    'BABY': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/baby',
        'tabs': ['모두보기', '상의', '팬츠', '아우터', '이너웨어']
    }
}

def log(msg, end='\n'):
    print(msg, end=end, flush=True)
    sys.stdout.flush()

def get_color_name(code):
    """컬러 코드를 컬러명으로 변환"""
    code = str(code).zfill(2)
    return COLOR_MAP.get(code, f'COLOR_{code}')

def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)
    
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    
    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작...")
    driver = webdriver.Chrome(service=service, options=options)
    log("  [OK] 완료!\n")
    return driver

def click_tab(driver, tab_name):
    """하위 탭 클릭"""
    try:
        # 탭 버튼 찾기 - 여러 방법 시도
        tab_selectors = [
            f"//button[contains(text(), '{tab_name}')]",
            f"//a[contains(text(), '{tab_name}')]",
            f"//div[contains(text(), '{tab_name}')]",
            f"//*[text()='{tab_name}']"
        ]
        
        for selector in tab_selectors:
            try:
                tabs = driver.find_elements(By.XPATH, selector)
                for tab in tabs:
                    if tab.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", tab)
                        time.sleep(2)
                        return True
            except:
                continue
        
        return False
    except Exception as e:
        log(f"      [WARN] 탭 클릭 실패: {e}")
        return False

def extract_products(driver, max_products=30):
    """상품 데이터 추출"""
    products = []
    
    # 스크롤하여 상품 로드
    for i in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)
    
    # 상품 타일 찾기
    product_tiles = driver.find_elements(By.CSS_SELECTOR, ".product-tile")
    
    if not product_tiles:
        return []
    
    for idx, tile in enumerate(product_tiles[:max_products], 1):
        try:
            product = {
                'rank': idx,
                'name': '',
                'price': '',
                'color_count': 0,
                'colors': '',
                'rating': '없음',
                'review_count': '없음',
                'image_url': ''
            }
            
            # 상품명 - 이미지 alt 또는 링크 텍스트에서
            try:
                img = tile.find_element(By.CSS_SELECTOR, "img.image__img")
                product['name'] = img.get_attribute("alt") or ""
                product['image_url'] = img.get_attribute("src") or ""
            except:
                try:
                    link = tile.find_element(By.CSS_SELECTOR, "a.product-tile__link")
                    product['name'] = link.text.strip().split('\n')[0]
                except:
                    pass
            
            # 가격 - ITOTypography div에서 "원" 포함 텍스트
            try:
                price_elements = tile.find_elements(By.CSS_SELECTOR, "[data-testid='ITOTypography']")
                for elem in price_elements:
                    txt = elem.text.strip()
                    if '원' in txt and len(txt) < 20:
                        product['price'] = txt
                        break
            except:
                pass
            
            # 가격 백업 - span 태그에서
            if not product['price']:
                try:
                    spans = tile.find_elements(By.TAG_NAME, "span")
                    for span in spans:
                        txt = span.text.strip()
                        if '원' in txt and len(txt) < 20:
                            product['price'] = txt
                            break
                except:
                    pass
            
            # 컬러 정보 - 칩 이미지에서 컬러 코드 추출 후 컬러명 변환
            try:
                color_chips = tile.find_elements(By.CSS_SELECTOR, ".product-tile__image-chip-group-item img")
                color_codes = []
                for chip in color_chips:
                    alt = chip.get_attribute("alt")
                    if alt and alt.isdigit():
                        color_name = get_color_name(alt)
                        if color_name not in color_codes:
                            color_codes.append(color_name)
                
                product['color_count'] = len(color_codes)
                product['colors'] = ', '.join(color_codes) if color_codes else '정보없음'
            except:
                product['color_count'] = 0
                product['colors'] = '정보없음'
            
            # 평점 - reviews 속성에서
            try:
                rating_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-static, [role='figure']")
                reviews_attr = rating_elem.get_attribute("reviews")
                if reviews_attr:
                    product['review_count'] = reviews_attr
                
                # 별점 계산 (full star 개수)
                full_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--full")
                half_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--half")
                product['rating'] = str(len(full_stars) + len(half_stars) * 0.5) if full_stars else '없음'
            except:
                pass
            
            # 평점 텍스트에서 추출 시도
            if product['rating'] == '없음':
                try:
                    rating_text_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-average-product-tile")
                    rating_text = rating_text_elem.text.strip()
                    rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                    if rating_match:
                        product['rating'] = rating_match.group(1)
                except:
                    pass
            
            if product['name']:
                products.append(product)
                
        except Exception as e:
            continue
    
    return products

def scrape_category_with_tabs(driver, category, url, tabs):
    """카테고리와 모든 탭 크롤링"""
    all_data = {}
    
    log(f"\n{'='*60}")
    log(f"[수집] {category} 카테고리")
    log(f"{'='*60}")
    log(f"  URL: {url}")
    log(f"  탭: {', '.join(tabs)}")
    
    # 메인 페이지 접속
    driver.get(url)
    log(f"  -> 페이지 로딩 대기...", end='')
    for i in range(10):
        time.sleep(1)
        print(".", end='', flush=True)
    log(" OK!")
    
    for tab_idx, tab_name in enumerate(tabs):
        log(f"\n  [{tab_idx+1}/{len(tabs)}] 탭: {tab_name}")
        
        # 첫 번째 탭(모두보기)이 아니면 탭 클릭
        if tab_idx > 0:
            if not click_tab(driver, tab_name):
                log(f"      -> 탭 클릭 실패, 건너뜀")
                continue
            time.sleep(3)
        
        # 상품 추출
        products = extract_products(driver, max_products=30)
        
        if products:
            sheet_name = f"{category}_{tab_name}"
            all_data[sheet_name] = products
            
            log(f"      -> {len(products)}개 수집 완료")
            
            # 샘플 출력
            if len(products) >= 3:
                for i, p in enumerate(products[:3], 1):
                    name_short = p['name'][:18] if len(p['name']) > 18 else p['name']
                    log(f"        {i}. {name_short:18s} | {p['price']:10s} | 컬러:{p['color_count']}")
        else:
            log(f"      -> 상품 없음")
        
        time.sleep(1)
    
    return all_data

def create_excel(all_data, filename):
    """엑셀 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")
    
    wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, products in all_data.items():
        # 시트명 길이 제한 (Excel은 31자까지)
        safe_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(safe_name)
        log(f"  -> 시트 [{safe_name}]: {len(products)}개 상품")
        
        headers = ['순위', '상품명', '가격', '컬러수', '컬러목록', '평점', '리뷰수', '이미지URL']
        ws.append(headers)
        
        for col in range(1, 9):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 60
        
        for p in products:
            ws.append([
                p['rank'], p['name'], p['price'],
                p['color_count'], p['colors'],
                p['rating'], p['review_count'],
                p['image_url']
            ])
            
            for col in range(1, 9):
                cell = ws.cell(ws.max_row, col)
                cell.alignment = Alignment(horizontal='center' if col != 8 else 'left', 
                                         vertical='center', wrap_text=True)
                cell.border = border
                cell.font = Font(name='맑은 고딕', size=10)
    
    # 종합 요약 시트
    summary = wb.create_sheet("종합요약", 0)
    summary.append(['시트명', '상품수', '평균컬러수', '평점수집', '리뷰수집'])
    
    for col in range(1, 6):
        cell = summary.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for sheet_name, products in all_data.items():
        if products:
            avg_colors = sum(p['color_count'] for p in products) / len(products)
            has_rating = sum(1 for p in products if p['rating'] != '없음')
            has_review = sum(1 for p in products if p['review_count'] != '없음')
            safe_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
            summary.append([safe_name, len(products), f"{avg_colors:.1f}", has_rating, has_review])
    
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 12
    summary.column_dimensions['D'].width = 10
    summary.column_dimensions['E'].width = 10
    
    wb.save(filename)
    log(f"  [완료] 저장 성공!")
    return True

def main():
    log("=" * 60)
    log("유니클로 랭킹 크롤러 V4 (완전판)")
    log("=" * 60)
    log("\n개선사항:")
    log("  1. 컬러 코드 → 컬러명 변환")
    log("  2. 이미지 URL 추출")
    log("  3. 하위 탭별 데이터 수집")
    log("  4. 가격 정확히 추출")
    log("")
    
    driver = setup_driver()
    
    try:
        all_data = {}
        total_categories = len(CATEGORIES)
        
        log(f"\n{'='*60}")
        log(f"[2/4] 데이터 수집 ({total_categories}개 카테고리)")
        log(f"{'='*60}")
        
        for idx, (category, config) in enumerate(CATEGORIES.items(), 1):
            log(f"\n>>> [{idx}/{total_categories}] {category} <<<")
            
            category_data = scrape_category_with_tabs(
                driver, 
                category, 
                config['url'], 
                config['tabs']
            )
            all_data.update(category_data)
            
            log(f"\n  {category} 완료: {len(category_data)}개 탭 수집")
            
            if idx < total_categories:
                log(f"  -> 다음 카테고리로 이동 (3초)...")
                time.sleep(3)
        
        if not all_data:
            log("\n[실패] 수집된 데이터가 없습니다!")
            return
        
        # 통계
        log(f"\n{'='*60}")
        log(f"[수집 완료 통계]")
        log(f"{'='*60}")
        total_products = sum(len(products) for products in all_data.values())
        total_sheets = len(all_data)
        log(f"  총 시트: {total_sheets}개")
        log(f"  총 상품: {total_products}개")
        
        # 판매 데이터 피드백
        all_products = [p for products in all_data.values() for p in products]
        has_rating = sum(1 for p in all_products if p['rating'] != '없음')
        has_review = sum(1 for p in all_products if p['review_count'] != '없음')
        has_price = sum(1 for p in all_products if p['price'])
        has_image = sum(1 for p in all_products if p['image_url'])
        
        log(f"\n  가격: {has_price}/{len(all_products)}개 수집됨")
        log(f"  이미지: {has_image}/{len(all_products)}개 수집됨")
        log(f"  평점: {has_rating}/{len(all_products)}개 수집됨")
        log(f"  리뷰: {has_review}/{len(all_products)}개 수집됨")
        
        # 엑셀 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_전체랭킹_{timestamp}.xlsx"
        create_excel(all_data, filename)
        
        # 최종 결과
        log(f"\n{'='*60}")
        log(f"[4/4] 작업 완료!")
        log(f"{'='*60}")
        log(f"  파일: {os.path.abspath(filename)}")
        log(f"  총 시트: {total_sheets}개")
        log(f"  총 상품: {total_products}개")
        log(f"{'='*60}")
        
    except Exception as e:
        log(f"\n[오류] {e}")
        import traceback
        traceback.print_exc()
    finally:
        log("\n브라우저 종료...")
        driver.quit()
        log("[완료]")

if __name__ == "__main__":
    main()
