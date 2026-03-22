# -*- coding: utf-8 -*-
"""
유니클로 랭킹 크롤러 V6 (이미지 삽입 버전)
- 이미지 URL 대신 실제 이미지를 엑셀에 삽입
- openpyxl + Pillow + requests 사용
"""
import sys
import io
import os
from datetime import datetime
import re
import requests
from PIL import Image as PILImage

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import time

# 이미지 설정
IMG_WIDTH = 80  # 엑셀에 삽입할 이미지 너비 (픽셀)
IMG_HEIGHT = 107  # 3:4 비율로 계산된 높이
ROW_HEIGHT = 85  # 행 높이 (포인트)

# 유니클로 컬러 코드 → 컬러명 매핑
COLOR_MAP = {
    '00': 'WHITE', '01': 'OFF WHITE', '02': 'LIGHT GRAY', '03': 'GRAY',
    '04': 'DARK GRAY', '05': 'LIGHT GRAY', '06': 'GRAY', '07': 'DARK GRAY',
    '08': 'DARK GRAY', '09': 'BLACK', '10': 'PINK', '11': 'PINK',
    '12': 'LIGHT PINK', '13': 'PINK', '14': 'PINK', '15': 'LIGHT PINK',
    '16': 'DARK PINK', '17': 'PURPLE', '18': 'PURPLE', '19': 'WINE',
    '20': 'ORANGE', '21': 'LIGHT ORANGE', '22': 'ORANGE', '23': 'BROWN',
    '24': 'DARK ORANGE', '25': 'ORANGE', '26': 'BROWN', '27': 'BROWN',
    '28': 'DARK BROWN', '29': 'BROWN', '30': 'NATURAL', '31': 'BEIGE',
    '32': 'BEIGE', '33': 'BEIGE', '34': 'LIGHT BEIGE', '35': 'BEIGE',
    '36': 'BROWN', '37': 'BROWN', '38': 'DARK BROWN', '39': 'KHAKI',
    '40': 'YELLOW', '41': 'LIGHT YELLOW', '42': 'YELLOW', '43': 'MUSTARD',
    '44': 'GOLD', '45': 'LIME', '46': 'OLIVE', '47': 'KHAKI', '48': 'OLIVE',
    '49': 'DARK GREEN', '50': 'LIGHT GREEN', '51': 'GREEN', '52': 'GREEN',
    '53': 'GREEN', '54': 'DARK GREEN', '55': 'GREEN', '56': 'GREEN',
    '57': 'GREEN', '58': 'MINT', '59': 'TURQUOISE', '60': 'LIGHT BLUE',
    '61': 'LIGHT BLUE', '62': 'SKY BLUE', '63': 'LIGHT BLUE', '64': 'BLUE',
    '65': 'BLUE', '66': 'BLUE', '67': 'BLUE', '68': 'BLUE', '69': 'NAVY',
    '70': 'DARK BLUE', '71': 'NAVY', '72': 'DARK BLUE', '73': 'NAVY',
    '74': 'DARK NAVY', '75': 'NAVY', '76': 'NAVY', '77': 'INDIGO',
    '78': 'INDIGO', '79': 'DENIM', '80': 'RED', '81': 'LIGHT RED',
    '82': 'RED', '83': 'RED', '84': 'DARK RED', '85': 'WINE', '86': 'WINE',
    '87': 'BURGUNDY', '88': 'DARK RED', '89': 'WINE', '90': 'SILVER',
    '91': 'GOLD', '92': 'MULTI', '93': 'MULTI', '94': 'PATTERN',
    '95': 'PATTERN', '96': 'STRIPE', '97': 'CHECK', '98': 'PRINT',
    '99': 'OTHER'
}

# 수집할 URL + 탭 정보 (테스트용으로 간소화)
CATEGORIES = {
    'WOMEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/women',
        'tabs': ['모두보기']  # 테스트용으로 첫 탭만
    },
    'MEN': {
        'url': 'https://www.uniqlo.com/kr/ko/spl/ranking/men',
        'tabs': ['모두보기']
    }
}

def log(msg, end='\n'):
    print(msg, end=end, flush=True)

def get_color_name(code):
    code = str(code).zfill(2)
    return COLOR_MAP.get(code, f'COLOR_{code}')

def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)
    
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36')
    
    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작 (headless)...")
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(60)
    log("  [OK] 완료!\n")
    return driver

def download_image(url, max_retries=2):
    """이미지 URL에서 이미지 다운로드하여 리사이즈"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                img = PILImage.open(io.BytesIO(response.content))
                # RGB로 변환 (PNG 알파 채널 처리)
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
                # 리사이즈
                img = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
                
                # BytesIO로 반환
                img_bytes = io.BytesIO()
                img.save(img_bytes, format='JPEG', quality=85)
                img_bytes.seek(0)
                return img_bytes
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(0.5)
                continue
    return None

def extract_products(driver, max_products=10):
    """상품 데이터 추출 (이미지 포함, 테스트용 10개)"""
    products = []
    
    # 스크롤하여 상품 로드
    for i in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)
    
    product_tiles = driver.find_elements(By.CSS_SELECTOR, ".product-tile")
    
    if not product_tiles:
        return []
    
    for idx, tile in enumerate(product_tiles[:max_products], 1):
        try:
            product = {
                'rank': idx,
                'name': '',
                'price': '',
                'color_count': 0,
                'colors': '',
                'rating': '없음',
                'review_count': '없음',
                'image_url': '',
                'image_data': None  # 실제 이미지 데이터
            }
            
            # 이미지 URL 추출
            try:
                active_img = tile.find_element(By.CSS_SELECTOR, ".swiper-slide-active img.image__img")
                product['name'] = active_img.get_attribute("alt") or ""
                img_url = active_img.get_attribute("data-src") or active_img.get_attribute("src") or ""
                product['image_url'] = img_url
            except:
                pass
            
            if not product['image_url']:
                imgs = tile.find_elements(By.CSS_SELECTOR, "img.image__img")
                for img in imgs:
                    alt = img.get_attribute("alt") or ""
                    if alt and not alt.isdigit():
                        product['name'] = alt
                        img_url = img.get_attribute("data-src") or img.get_attribute("src") or ""
                        if img_url:
                            product['image_url'] = img_url
                            break
            
            if not product['image_url']:
                all_imgs = tile.find_elements(By.CSS_SELECTOR, "[data-testid='ITOImage'] img")
                for img in all_imgs:
                    img_url = img.get_attribute("data-src") or img.get_attribute("src")
                    if img_url and "uniqlo.com" in img_url:
                        product['image_url'] = img_url
                        if not product['name']:
                            product['name'] = img.get_attribute("alt") or ""
                        break
            
            # 이미지 다운로드
            if product['image_url']:
                log(f"      [{idx}] 이미지 다운로드 중...", end='')
                img_data = download_image(product['image_url'])
                if img_data:
                    product['image_data'] = img_data
                    log(" OK")
                else:
                    log(" FAIL")
            
            # 상품명 백업
            if not product['name']:
                try:
                    link = tile.find_element(By.CSS_SELECTOR, "a.product-tile__link")
                    product['name'] = link.text.strip().split('\n')[0]
                except:
                    pass
            
            # 가격
            try:
                price_elements = tile.find_elements(By.CSS_SELECTOR, "[data-testid='ITOTypography']")
                for elem in price_elements:
                    txt = elem.text.strip()
                    if '원' in txt and len(txt) < 20:
                        product['price'] = txt
                        break
            except:
                pass
            
            # 컬러 정보
            try:
                color_chips = tile.find_elements(By.CSS_SELECTOR, ".product-tile__image-chip-group-item img")
                color_codes = []
                for chip in color_chips:
                    alt = chip.get_attribute("alt")
                    if alt and alt.isdigit():
                        color_name = get_color_name(alt)
                        if color_name not in color_codes:
                            color_codes.append(color_name)
                
                product['color_count'] = len(color_codes)
                product['colors'] = ', '.join(color_codes) if color_codes else '정보없음'
            except:
                product['color_count'] = 0
                product['colors'] = '정보없음'
            
            # 평점
            try:
                rating_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-static, [role='figure']")
                reviews_attr = rating_elem.get_attribute("reviews")
                if reviews_attr:
                    product['review_count'] = reviews_attr
                
                full_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--full")
                half_stars = tile.find_elements(By.CSS_SELECTOR, ".fr-ec-star--half")
                product['rating'] = str(len(full_stars) + len(half_stars) * 0.5) if full_stars else '없음'
            except:
                pass
            
            if product['name']:
                products.append(product)
                
        except Exception as e:
            continue
    
    return products

def scrape_category(driver, category, url, tabs):
    """카테고리 크롤링"""
    all_data = {}
    
    log(f"\n{'='*60}")
    log(f"[수집] {category} 카테고리")
    log(f"{'='*60}")
    log(f"  URL: {url}")
    
    driver.get(url)
    log(f"  -> 페이지 로딩 대기...", end='')
    for i in range(10):
        time.sleep(1)
        print(".", end='', flush=True)
    log(" OK!")
    
    for tab_idx, tab_name in enumerate(tabs):
        log(f"\n  [{tab_idx+1}/{len(tabs)}] 탭: {tab_name}")
        
        products = extract_products(driver, max_products=10)
        
        if products:
            sheet_name = f"{category}_{tab_name}"
            all_data[sheet_name] = products
            
            log(f"      -> {len(products)}개 수집 완료")
            for i, p in enumerate(products[:3], 1):
                name_short = p['name'][:15] if len(p['name']) > 15 else p['name']
                has_img = "✓" if p['image_data'] else "✗"
                log(f"        {i}. {name_short:15s} | {p['price']:10s} | img:{has_img}")
    
    return all_data

def create_excel_with_images(all_data, filename):
    """이미지가 포함된 엑셀 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성 (이미지 포함)")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")
    
    wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, products in all_data.items():
        safe_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(safe_name)
        log(f"  -> 시트 [{safe_name}]: {len(products)}개 상품 (이미지 삽입 중...)")
        
        # 헤더 (이미지 컬럼 추가)
        headers = ['순위', '이미지', '상품명', '가격', '컬러수', '컬러목록', '평점', '리뷰수']
        ws.append(headers)
        
        for col in range(1, 9):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 6   # 순위
        ws.column_dimensions['B'].width = 12  # 이미지
        ws.column_dimensions['C'].width = 35  # 상품명
        ws.column_dimensions['D'].width = 12  # 가격
        ws.column_dimensions['E'].width = 8   # 컬러수
        ws.column_dimensions['F'].width = 30  # 컬러목록
        ws.column_dimensions['G'].width = 8   # 평점
        ws.column_dimensions['H'].width = 8   # 리뷰수
        
        # 데이터 행 추가
        for row_idx, p in enumerate(products, 2):
            ws.append([
                p['rank'], '',  # 이미지 셀은 비워둠
                p['name'], p['price'],
                p['color_count'], p['colors'],
                p['rating'], p['review_count']
            ])
            
            # 행 높이 설정
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
            
            # 이미지 삽입
            if p['image_data']:
                try:
                    img = XLImage(p['image_data'])
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    # B열에 이미지 삽입
                    cell_ref = f'B{row_idx}'
                    ws.add_image(img, cell_ref)
                except Exception as e:
                    log(f"    [WARN] 이미지 삽입 실패 row {row_idx}: {e}")
            
            # 셀 스타일
            for col in range(1, 9):
                cell = ws.cell(row_idx, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
    
    wb.save(filename)
    log(f"  [OK] 저장 완료!")
    return filename

def main():
    log("\n" + "=" * 60)
    log("  유니클로 랭킹 크롤러 V6 (이미지 삽입 버전)")
    log("=" * 60)
    log("  * 이미지를 엑셀에 직접 삽입합니다")
    log("  * 테스트용: 각 카테고리 10개 상품만 수집")
    log("=" * 60)
    
    driver = setup_driver()
    all_data = {}
    
    try:
        log("\n" + "=" * 60)
        log("[2/4] 데이터 수집 시작")
        log("=" * 60)
        
        for category, info in CATEGORIES.items():
            data = scrape_category(driver, category, info['url'], info['tabs'])
            all_data.update(data)
        
        # 엑셀 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_랭킹_이미지포함_V6_{timestamp}.xlsx"
        create_excel_with_images(all_data, filename)
        
        # 최종 통계
        log(f"\n{'='*60}")
        log(f"[4/4] 수집 완료 통계")
        log(f"{'='*60}")
        
        total_products = sum(len(p) for p in all_data.values())
        total_with_img = sum(1 for prods in all_data.values() for p in prods if p['image_data'])
        
        log(f"  총 시트: {len(all_data)}개")
        log(f"  총 상품: {total_products}개")
        log(f"  이미지 삽입: {total_with_img}/{total_products}개")
        log(f"\n  파일: {filename}")
        log("=" * 60)
        
    finally:
        driver.quit()
        log("\n브라우저 종료")

if __name__ == "__main__":
    main()
