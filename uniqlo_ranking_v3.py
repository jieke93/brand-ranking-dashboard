# -*- coding: utf-8 -*-
"""
유니클로 랭킹 크롤러 V3 (수정된 버전)
- WOMEN, MEN, KIDS, BABY 랭킹 수집
- 컬러 개수 및 컬러명
- 판매 데이터 (평점, 리뷰)
"""
import sys
sys.stdout.reconfigure(line_buffering=True)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import time
import os
from datetime import datetime
import re

# 수집할 URL
URLS = {
    'WOMEN': 'https://www.uniqlo.com/kr/ko/spl/ranking/women',
    'MEN': 'https://www.uniqlo.com/kr/ko/spl/ranking/men',
    'KIDS': 'https://www.uniqlo.com/kr/ko/spl/ranking/kids',
    'BABY': 'https://www.uniqlo.com/kr/ko/spl/ranking/baby'
}

def log(msg, end='\n'):
    print(msg, end=end, flush=True)
    sys.stdout.flush()

def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)
    
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    
    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작...")
    driver = webdriver.Chrome(service=service, options=options)
    log("  [OK] 완료!\n")
    return driver

def scrape_ranking(driver, category, url, max_products=30):
    """랭킹 페이지 크롤링"""
    log(f"\n{'='*60}")
    log(f"[수집] {category} 랭킹")
    log(f"{'='*60}")
    
    # 1. 페이지 접속
    log(f"  1) 페이지 접속: {url}")
    driver.get(url)
    
    # 2. 로딩 대기
    log(f"  2) 로딩 대기 (12초)...", end='')
    for i in range(12):
        time.sleep(1)
        print(".", end='', flush=True)
    log(" OK!")
    
    # 3. 스크롤
    log(f"  3) 스크롤 (상품 로딩)...")
    for i in range(5):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
        log(f"     - {i+1}/5")
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(2)
    
    # 4. 상품 찾기
    log(f"  4) 상품 요소 찾기...")
    
    product_tiles = driver.find_elements(By.CSS_SELECTOR, ".product-tile")
    log(f"     [DEBUG] .product-tile: {len(product_tiles)}개")
    
    if not product_tiles:
        product_tiles = driver.find_elements(By.CSS_SELECTOR, "[class*='product-tile']")
        log(f"     [DEBUG] [class*='product-tile']: {len(product_tiles)}개")
    
    if not product_tiles:
        log(f"     [실패] 상품을 찾을 수 없습니다!")
        return []
    
    log(f"     [성공] {len(product_tiles)}개 발견!")
    
    # 5. 데이터 추출
    log(f"  5) 데이터 추출 (최대 {max_products}개):")
    products = []
    
    for idx, tile in enumerate(product_tiles[:max_products], 1):
        try:
            log(f"     [{idx:2d}/{min(len(product_tiles), max_products)}] ", end='')
            
            product = {
                'category': category,
                'rank': idx,
                'name': '',
                'price': '',
                'color_count': 0,
                'colors': '',
                'rating': '없음',
                'review_count': '없음',
                'image_url': ''
            }
            
            # 상품명 - 이미지 alt에서 추출!
            try:
                img = tile.find_element(By.CSS_SELECTOR, "img.image__img, img[class*='image']")
                product['name'] = img.get_attribute("alt") or ""
                product['image_url'] = img.get_attribute("src") or img.get_attribute("data-src") or ""
            except:
                try:
                    link = tile.find_element(By.CSS_SELECTOR, "a")
                    product['name'] = link.text.strip().split('\n')[0]
                except:
                    pass
            
            # 가격 - span에서 '원' 포함 텍스트 찾기
            try:
                spans = tile.find_elements(By.TAG_NAME, "span")
                for span in spans:
                    txt = span.text.strip()
                    if '원' in txt and len(txt) < 20:
                        product['price'] = txt
                        break
            except:
                pass
            
            # 컬러 정보 (핵심!)
            try:
                color_chips = tile.find_elements(By.CSS_SELECTOR, ".product-tile__image-chip-group-item")
                product['color_count'] = len(color_chips)
                
                colors = []
                for chip in color_chips[:10]:
                    try:
                        chip_img = chip.find_element(By.TAG_NAME, "img")
                        color_name = chip_img.get_attribute("alt") or chip_img.get_attribute("title") or ""
                        if color_name and color_name not in colors and color_name != product['name']:
                            colors.append(color_name)
                    except:
                        pass
                product['colors'] = ', '.join(colors) if colors else f'{product["color_count"]}개 컬러'
            except:
                product['color_count'] = 0
                product['colors'] = '정보없음'
            
            # 평점 (판매 데이터)
            try:
                rating_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-average-product-tile, [class*='rating-average']")
                rating_text = rating_elem.text.strip()
                rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                product['rating'] = rating_match.group(1) if rating_match else '없음'
            except:
                product['rating'] = '없음'
            
            # 리뷰 수
            try:
                review_elem = tile.find_element(By.CSS_SELECTOR, ".fr-ec-rating-static__count-product-tile, [class*='count-product']")
                review_text = review_elem.text.strip()
                review_match = re.search(r'\(?([\d,]+)\)?', review_text)
                product['review_count'] = review_match.group(1) if review_match else '없음'
            except:
                product['review_count'] = '없음'
            
            if product['name']:
                products.append(product)
                log(f"{product['name'][:20]:20s} | 컬러:{product['color_count']:2d} | 평점:{product['rating']}")
            else:
                log("건너뜀 (이름없음)")
            
        except Exception as e:
            log(f"오류: {str(e)[:30]}")
            continue
    
    log(f"\n  [완료] {category}: {len(products)}개 수집")
    return products

def create_excel(all_products, filename):
    """엑셀 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")
    
    wb = openpyxl.Workbook()
    
    categories = {}
    for p in all_products:
        cat = p['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(p)
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cat, products in categories.items():
        ws = wb.create_sheet(cat)
        log(f"  -> 시트 [{cat}]: {len(products)}개 상품")
        
        headers = ['순위', '상품명', '가격', '컬러수', '컬러목록', '평점', '리뷰수']
        ws.append(headers)
        
        for col in range(1, 8):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        
        for p in products:
            ws.append([
                p['rank'], p['name'], p['price'],
                p['color_count'], p['colors'],
                p['rating'], p['review_count']
            ])
            
            for col in range(1, 8):
                cell = ws.cell(ws.max_row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                cell.font = Font(name='맑은 고딕', size=10)
    
    # 종합 시트
    summary = wb.create_sheet("종합요약", 0)
    summary.append(['카테고리', '상품수', '평균컬러수', '평점있음', '리뷰있음'])
    
    for col in range(1, 6):
        cell = summary.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for cat, products in categories.items():
        avg_colors = sum(p['color_count'] for p in products) / len(products) if products else 0
        has_rating = sum(1 for p in products if p['rating'] != '없음')
        has_review = sum(1 for p in products if p['review_count'] != '없음')
        summary.append([cat, len(products), f"{avg_colors:.1f}", has_rating, has_review])
    
    summary.column_dimensions['A'].width = 15
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 12
    summary.column_dimensions['D'].width = 10
    summary.column_dimensions['E'].width = 10
    
    wb.save(filename)
    log(f"  [완료] 저장 성공!")
    return True

def main():
    log("=" * 60)
    log("유니클로 랭킹 크롤러 V3 (WOMEN/MEN/KIDS/BABY)")
    log("=" * 60)
    log("\n수집 항목:")
    log("  - 상품명, 가격")
    log("  - 컬러 개수, 컬러명")
    log("  - 판매 데이터 (평점, 리뷰수)")
    log("")
    
    driver = setup_driver()
    
    try:
        all_products = []
        total = len(URLS)
        
        log(f"\n{'='*60}")
        log(f"[2/4] 데이터 수집 ({total}개 카테고리)")
        log(f"{'='*60}")
        
        for idx, (category, url) in enumerate(URLS.items(), 1):
            log(f"\n>>> [{idx}/{total}] {category} <<<")
            products = scrape_ranking(driver, category, url, max_products=30)
            all_products.extend(products)
            log(f"\n  누적: {len(all_products)}개")
            
            if idx < total:
                log(f"  -> 다음 카테고리까지 3초 대기...")
                time.sleep(3)
        
        if not all_products:
            log("\n[실패] 수집된 데이터가 없습니다!")
            return
        
        # 판매 데이터 피드백
        log(f"\n{'='*60}")
        log(f"[판매 데이터 수집 결과]")
        log(f"{'='*60}")
        has_rating = sum(1 for p in all_products if p['rating'] != '없음')
        has_review = sum(1 for p in all_products if p['review_count'] != '없음')
        log(f"  평점 정보: {has_rating}/{len(all_products)}개 상품에서 수집됨")
        log(f"  리뷰 정보: {has_review}/{len(all_products)}개 상품에서 수집됨")
        
        if has_rating > 0:
            log(f"  -> [OK] 평점 데이터 공개됨!")
        else:
            log(f"  -> [INFO] 평점 데이터 없음 또는 비공개")
        
        if has_review > 0:
            log(f"  -> [OK] 리뷰 데이터 공개됨!")
        else:
            log(f"  -> [INFO] 리뷰 데이터 없음 또는 비공개")
        
        # 엑셀 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_랭킹_{timestamp}.xlsx"
        create_excel(all_products, filename)
        
        # 최종 결과
        log(f"\n{'='*60}")
        log(f"[4/4] 작업 완료!")
        log(f"{'='*60}")
        log(f"  파일: {os.path.abspath(filename)}")
        log(f"  총 상품: {len(all_products)}개")
        
        for cat in URLS.keys():
            count = sum(1 for p in all_products if p['category'] == cat)
            avg_colors = sum(p['color_count'] for p in all_products if p['category'] == cat) / count if count else 0
            log(f"  - {cat}: {count}개 (평균 컬러: {avg_colors:.1f}개)")
        
        log(f"{'='*60}")
        
    except Exception as e:
        log(f"\n[오류] {e}")
        import traceback
        traceback.print_exc()
    finally:
        log("\n브라우저 종료...")
        driver.quit()
        log("[완료]")

if __name__ == "__main__":
    main()
