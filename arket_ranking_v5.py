# -*- coding: utf-8 -*-
"""
아르켓(ARKET) 랭킹 크롤러 V5 (빠른 버전)
- 여성/남성 인기상품 판매순 수집
- 이미지 캡처 최적화
"""
import io
import os
import re
import urllib.parse
import urllib.request
import urllib.robotparser
from PIL import Image as PILImage
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import time
from datetime import datetime

# 설정
IMG_WIDTH = 80
IMG_HEIGHT = 107
HD_IMG_WIDTH = 400   # 대시보드용 고해상도 이미지
HD_IMG_HEIGHT = 534  # 3:4 비율
ROW_HEIGHT = 85
MAX_PRODUCTS = 100  # 각 카테고리 최대값 수집
SAFE_MODE = True  # 법적 위험 최소화 모드 (robots.txt 준수)
REQUEST_DELAY = 1.5  # 요청 간격 (초)
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

CATEGORIES = {
    'WOMEN': 'https://www.arket.com/ko-kr/women/most-popular.html',
    'MEN': 'https://www.arket.com/ko-kr/men/most-popular.html'
}

def log(msg, end='\n'):
    print(msg, end=end, flush=True)

def check_robots_allowed(url):
    """robots.txt 기준으로 접근 가능 여부 확인"""
    if not SAFE_MODE:
        return True
    try:
        parsed = urllib.parse.urlparse(url)
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        rp = urllib.robotparser.RobotFileParser()
        with urllib.request.urlopen(robots_url, timeout=5) as resp:
            content = resp.read().decode('utf-8', 'ignore')
        rp.parse(content.splitlines())
        return rp.can_fetch(USER_AGENT, url)
    except Exception:
        return True

def safe_get(driver, url):
    """robots.txt 준수 + 요청 간격 적용 후 페이지 로드"""
    if SAFE_MODE and not check_robots_allowed(url):
        log(f"      -> robots.txt 제한으로 접근 건너뜀: {url}")
        return False
    time.sleep(REQUEST_DELAY)
    driver.get(url)
    return True

def capture_image(element):
    """요소를 스크린샷으로 캐처 - 엑셀용 + 대시보드 고해상도"""
    try:
        png_data = element.screenshot_as_png
        if not png_data:
            return None
        img = PILImage.open(io.BytesIO(png_data))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        hd_img = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        hd_bytes = io.BytesIO()
        hd_img.save(hd_bytes, format='JPEG', quality=92)
        hd_bytes.seek(0)
        
        xl_img = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        xl_bytes = io.BytesIO()
        xl_img.save(xl_bytes, format='JPEG', quality=85)
        xl_bytes.seek(0)
        
        return (xl_bytes, hd_bytes)
    except:
        return None

def _save_hd_images(all_data, brand_name, excel_filename):
    """대시보드용 고해상도 이미지를 product_images_hd/ 폴더에 저장"""
    import hashlib
    hd_dir = os.path.join(os.path.dirname(os.path.abspath(excel_filename)), 'product_images_hd')
    os.makedirs(hd_dir, exist_ok=True)
    
    file_hash = hashlib.md5(f"{os.path.basename(excel_filename)}_{os.path.getmtime(excel_filename)}".encode()).hexdigest()[:8]
    saved = 0
    for sheet_name, products in all_data.items():
        for p in products:
            hd_data = p.get('hd_image_data')
            if hd_data:
                name = p.get('name', '')[:20]
                safe = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
                fname = f"{file_hash}_{brand_name}_{sheet_name}_{p['rank']}_{safe}.jpg"
                fpath = os.path.join(hd_dir, fname)
                if not os.path.exists(fpath):
                    try:
                        hd_data.seek(0)
                        with open(fpath, 'wb') as f:
                            f.write(hd_data.read())
                        saved += 1
                    except:
                        pass
    if saved > 0:
        log(f"  [HD] 고해상도 이미지 {saved}개 저장 → {hd_dir}")

def main():
    log("\n" + "=" * 60)
    log("  ARKET 인기상품 크롤러 V5 (판매순)")
    log("=" * 60)
    
    # 드라이버 설정
    log("\n[1/4] Chrome 시작...")
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    options.page_load_strategy = 'eager'
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    })
    driver.set_page_load_timeout(60)
    log("  완료!")
    
    all_data = {}
    
    log(f"\n[2/4] 데이터 수집")
    
    for cat_idx, (cat_name, url) in enumerate(CATEGORIES.items(), 1):
        log(f"\n  [{cat_idx}/{len(CATEGORIES)}] {cat_name}")
        
        # 재시도 로직
        max_retries = 3
        success = False
        
        for retry in range(max_retries):
            try:
                if retry > 0:
                    log(f"      재시도 {retry}/{max_retries-1}...")
                    time.sleep(3)
                
                # 페이지 로드
                if not safe_get(driver, url):
                    break
                time.sleep(4)
                
                # 타이틀 확인
                if "Access Denied" in driver.title:
                    raise Exception("페이지 접근 거부")
                
                # 판매순 정렬
                try:
                    sort_btn = driver.find_element(By.XPATH, "//button[contains(text(), '정렬')]")
                    driver.execute_script("arguments[0].click();", sort_btn)
                    time.sleep(0.5)
                    
                    sale_btn = driver.find_element(By.CSS_SELECTOR, 'button[data-value="sale"]')
                    driver.execute_script("arguments[0].click();", sale_btn)
                    time.sleep(2)
                    log(f"      판매순 정렬 적용")
                except:
                    log(f"      기본 정렬 사용")
                
                success = True
                break
                
            except Exception as e:
                log(f"      오류 ({retry+1}/{max_retries}): {str(e)[:50]}")
                if retry == max_retries - 1:
                    log(f"      {cat_name} 카테고리 건너뜀")
                    continue
        
        if not success:
            continue
            
        try:
            # 스크롤 (모든 상품 로드를 위해 충분히 스크롤)
            for _ in range(10):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);");
                time.sleep(1)
            driver.execute_script("window.scrollTo(0, 0);");
            time.sleep(1)
            
            # 상품 추출
            tiles = driver.find_elements(By.CSS_SELECTOR, 'div.o-product')
            log(f"      {len(tiles)}개 발견")
            
            # 발견된 모든 상품 수집
            max_to_collect = len(tiles)
            log(f"      -> 모든 상품 ({max_to_collect}개) 수집 예정")
            
            products = []
            img_count = 0
            
            # 상품 정보 먼저 수집
            items_to_capture = []
            for idx, tile in enumerate(tiles[:max_to_collect], 1):
                try:
                    product = {
                        'rank': idx,
                        'name': tile.get_attribute('data-name') or '',
                        'color': tile.get_attribute('data-color') or '',
                        'price': '',
                        'image_data': None,
                        'tile': tile  # 이미지용
                    }
                    
                    price = tile.get_attribute('data-sellprc')
                    if price:
                        try:
                            product['price'] = f"{int(price):,}원"
                        except:
                            product['price'] = price
                    
                    if product['name']:
                        items_to_capture.append(product)
                except:
                    continue
            
            log(f"      {len(items_to_capture)}개 상품 정보 수집 완료")
            
            if len(items_to_capture) == 0:
                log(f"      수집할 상품이 없습니다")
                continue
            
            # 이미지 캡처
            log(f"      이미지 캡처 중...", end='')
            for product in items_to_capture:
                try:
                    tile = product['tile']
                    img_elem = tile.find_element(By.CSS_SELECTOR, 'img')
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", img_elem)
                    time.sleep(0.1)
                    img_data = capture_image(img_elem)
                    if img_data:
                        product['image_data'] = img_data[0]      # 엑셀용
                        product['hd_image_data'] = img_data[1]   # 대시보드용 고해상도
                        img_count += 1
                except:
                    pass
                del product['tile']  # tile 참조 제거
                products.append(product)
            
            log(f" {img_count}개 완료")
            
            if products:
                all_data[cat_name] = products
                for p in products[:2]:
                    name = p['name'][:20]
                    log(f"        - {name} | {p['price']}")
                    
        except Exception as e:
            log(f"      상품 추출 오류: {e}")
            continue
        
        time.sleep(1)
    
    driver.quit()
    log("\n  브라우저 종료")
    
    # 엑셀 생성
    if all_data:
        log(f"\n[3/4] 엑셀 생성")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"아르켓_인기상품_판매순_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = ['순위', '이미지', '상품명', '색상', '가격']
        col_widths = [6, 12, 45, 15, 15]
        
        for sheet_name, products in all_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            ws.row_dimensions[1].height = 25
            
            for row_idx, product in enumerate(products, 2):
                ws.row_dimensions[row_idx].height = ROW_HEIGHT
                
                ws.cell(row=row_idx, column=1, value=product['rank']).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=1).border = border
                
                if product['image_data']:
                    try:
                        img = XLImage(product['image_data'])
                        img.width = IMG_WIDTH
                        img.height = IMG_HEIGHT
                        ws.add_image(img, f"B{row_idx}")
                    except:
                        pass
                ws.cell(row=row_idx, column=2).border = border
                
                ws.cell(row=row_idx, column=3, value=product['name']).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(row=row_idx, column=3).border = border
                
                ws.cell(row=row_idx, column=4, value=product['color']).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=4).border = border
                
                ws.cell(row=row_idx, column=5, value=product['price']).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row_idx, column=5).font = Font(bold=True)
                ws.cell(row=row_idx, column=5).border = border
            
            log(f"  시트 [{sheet_name}]: {len(products)}개")
        
        wb.save(filename)
        
        # 대시보드용 고해상도 이미지 저장
        _save_hd_images(all_data, '아르켓', filename)
        
        # 결과
        log(f"\n[4/4] 완료!")
        log("=" * 60)
        total = sum(len(p) for p in all_data.values())
        total_img = sum(1 for prods in all_data.values() for p in prods if p['image_data'])
        log(f"  카테고리: {len(all_data)}개")
        log(f"  총 상품: {total}개")
        log(f"  이미지: {total_img}/{total}개")
        log(f"\n  파일: {filename}")
        log("=" * 60)
    else:
        log("\n[오류] 수집된 데이터가 없습니다.")

if __name__ == "__main__":
    main()
