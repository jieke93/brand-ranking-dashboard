"""
유니클로 여성 랭킹 상품 정보 수집기 (Edge 브라우저 버전)
"""

from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
from PIL import Image as PILImage
import time
import os
from datetime import datetime

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Edge/120.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9',
}

def setup_driver():
    """Edge 드라이버 설정"""
    print("🔧 Edge 드라이버 설정 중...")
    
    edge_options = EdgeOptions()
    edge_options.add_argument('--disable-gpu')
    edge_options.add_argument('--no-sandbox')
    edge_options.add_argument('--disable-dev-shm-usage')
    edge_options.add_argument('--window-size=1920,1080')
    edge_options.add_argument('--disable-blink-features=AutomationControlled')
    edge_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    edge_options.add_experimental_option('useAutomationExtension', False)
    
    try:
        service = EdgeService(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=edge_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        print("✅ Edge 드라이버 설정 완료")
        return driver
    except Exception as e:
        print(f"❌ Edge 드라이버 설정 실패: {e}")
        return None

def scrape_ranking_data(driver):
    """랭킹 데이터 수집"""
    url = "https://www.uniqlo.com/kr/ko/spl/ranking/women"
    print(f"\n🌐 페이지 접근: {url}")
    
    try:
        driver.get(url)
        print("⏳ 페이지 로딩 대기 (10초)...")
        time.sleep(10)  # 페이지 로딩 충분히 대기
        
        print("📜 스크롤하여 모든 상품 로드 중...")
        # 페이지 끝까지 스크롤
        for i in range(5):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            print(f"  스크롤 {i+1}/5")
        
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(2)
        
        print("\n🔍 상품 정보 추출 중...")
        
        # 모든 링크 요소 찾기 (상품 링크 포함)
        all_links = driver.find_elements(By.TAG_NAME, "a")
        product_links = []
        
        for link in all_links:
            href = link.get_attribute("href")
            if href and "/products/E" in href and "-000/" in href:
                if href not in product_links:
                    product_links.append(href)
        
        print(f"✅ {len(product_links)}개 상품 링크 발견")
        
        products = []
        
        # 랭킹 번호가 있는 요소 찾기
        for rank in range(1, min(len(product_links) + 1, 51)):  # 최대 50개
            try:
                # 여러 방법으로 상품 정보 찾기
                # 방법 1: 텍스트로 순위 찾기
                xpath_queries = [
                    f"//div[contains(text(), '{rank}')]",
                    f"//span[contains(text(), '{rank}')]",
                    f"//*[starts-with(text(), '{rank}')]"
                ]
                
                found_element = None
                for xpath in xpath_queries:
                    try:
                        elements = driver.find_elements(By.XPATH, xpath)
                        for elem in elements:
                            text = elem.text.strip()
                            if text == str(rank) or text.startswith(f"{rank}.") or text.startswith(f"{rank}위"):
                                found_element = elem
                                break
                        if found_element:
                            break
                    except:
                        continue
                
                if not found_element and rank <= len(product_links):
                    # 순위를 찾지 못하면 링크 순서대로 사용
                    try:
                        link_elem = driver.find_element(By.XPATH, f"//a[@href='{product_links[rank-1]}']")
                        parent = link_elem.find_element(By.XPATH, "..")
                        
                        # 상품명
                        name = None
                        try:
                            name_elem = parent.find_element(By.TAG_NAME, "h3")
                            name = name_elem.text.strip()
                        except:
                            try:
                                name = link_elem.get_attribute("title") or link_elem.text.strip()
                            except:
                                name = f"상품 {rank}"
                        
                        # 가격
                        price = "가격 정보 없음"
                        for sel in ["span[class*='price']", "p[class*='price']", ".price"]:
                            try:
                                price_elem = parent.find_element(By.CSS_SELECTOR, sel)
                                price = price_elem.text.strip()
                                if price:
                                    break
                            except:
                                pass
                        
                        # 이미지
                        image_url = ""
                        try:
                            img = link_elem.find_element(By.TAG_NAME, "img")
                            image_url = img.get_attribute("src") or img.get_attribute("data-src") or ""
                            if image_url and not image_url.startswith("http"):
                                image_url = "https://image.uniqlo.com" + image_url
                        except:
                            pass
                        
                        products.append({
                            'rank': rank,
                            'name': name,
                            'price': price,
                            'image_url': image_url,
                            'colors': '정보 없음'
                        })
                        
                        print(f"  ✓ {rank}위: {name[:40]}")
                    
                    except Exception as e:
                        print(f"  ⚠️ {rank}위 처리 실패")
                        continue
            
            except Exception as e:
                continue
        
        # 디버깅: 페이지 소스 일부 저장
        if len(products) < 5:
            print("\n⚠️ 상품을 충분히 찾지 못했습니다. 페이지 구조 확인 필요")
            with open("debug_page.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("💾 디버그 파일 저장: debug_page.html")
        
        return products
    
    except Exception as e:
        print(f"❌ 수집 실패: {e}")
        import traceback
        traceback.print_exc()
        return []

def download_image(url):
    """이미지 다운로드"""
    if not url:
        return None
    try:
        time.sleep(0.3)
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        
        if img.mode == 'RGBA':
            bg = PILImage.new('RGB', img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        
        img.thumbnail((100, 100), PILImage.Resampling.LANCZOS)
        
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        return img_bytes
    except Exception as e:
        print(f"    ⚠️ 이미지 다운로드 실패: {str(e)[:50]}")
        return None

def create_excel(products, filename):
    """엑셀 파일 생성"""
    print(f"\n📊 엑셀 파일 생성: {filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "여성 랭킹"
    
    # 헤더
    headers = ['순위', '상품명', '가격', '컬러 옵션', '상품 이미지']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 열 너비
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    
    # 데이터 입력
    for product in products:
        row_idx = ws.max_row + 1
        
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product.get('colors', '정보 없음'))
        
        ws.row_dimensions[row_idx].height = 85
        
        # 텍스트 정렬
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 이미지 다운로드 및 삽입
        if product.get('image_url'):
            print(f"  📷 {product['rank']}위 이미지 다운로드 중...")
            img_data = download_image(product['image_url'])
            if img_data:
                try:
                    img = XLImage(img_data)
                    img.width = 80
                    img.height = 80
                    ws.add_image(img, f'E{row_idx}')
                except Exception as e:
                    print(f"    ⚠️ 이미지 삽입 실패: {e}")
        
        ws.cell(row_idx, 5).border = border
    
    # 메타 정보
    meta_row = ws.max_row + 2
    ws.cell(meta_row, 1, f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(meta_row + 1, 1, "출처: 유니클로 공식 온라인스토어")
    ws.cell(meta_row + 2, 1, "⚠️ 개인 사용 목적, 상업적 사용 금지")
    
    for row in range(meta_row, meta_row + 3):
        ws.cell(row, 1).font = Font(size=9, italic=True, color="666666")
    
    wb.save(filename)
    print(f"✅ 엑셀 파일 저장 완료!")

def main():
    print("=" * 70)
    print("📦 유니클로 여성 랭킹 상품 정보 수집기 (Edge 브라우저)")
    print("=" * 70)
    print("\n⚖️  개인 사용 목적 / 법적 리스크 최소화")
    print("⚠️  상업적 사용 금지\n")
    
    driver = setup_driver()
    if not driver:
        print("\n❌ Edge 드라이버 실패")
        print("💡 Microsoft Edge 브라우저가 필요합니다.")
        print("   Windows에 기본 설치되어 있습니다.")
        return
    
    try:
        products = scrape_ranking_data(driver)
        
        if not products:
            print("\n❌ 데이터 수집 실패")
            print("💡 debug_page.html 파일을 확인하세요")
            return
        
        print(f"\n✅ {len(products)}개 상품 수집 완료")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"유니클로_여성랭킹_{timestamp}.xlsx"
        
        create_excel(products, filename)
        
        print("\n" + "=" * 70)
        print("🎉 작업 완료!")
        print(f"📁 파일 위치: {os.path.abspath(filename)}")
        print(f"📊 수집 상품: {len(products)}개")
        print("=" * 70)
    
    except KeyboardInterrupt:
        print("\n\n⚠️ 사용자 중단")
    except Exception as e:
        print(f"\n❌ 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            print("\n🔒 브라우저 종료 중...")
            driver.quit()
            print("✅ 종료 완료")

if __name__ == "__main__":
    main()
