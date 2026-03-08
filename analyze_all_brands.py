#!/usr/bin/env python3
"""
3개 브랜드(유니클로·아르켓·탑텐) 통합 랭킹 분석 도구
──────────────────────────────────────────────────
• 각 브랜드 크롤링 엑셀을 읽어 통일된 포맷으로 변환
• 브랜드별 개별 분석 + 3사 비교 분석
• 히스토리 누적 관리 (JSON + 백업 + 엑셀 복구)
• 결과를 *1개의 엑셀 파일*로 출력

출력 시트 구성:
  1) 종합_대시보드      : 3사 아이템타입 비중 한눈 비교
  2) 유니클로_분석      : 유니클로 상세 분석
  3) 아르켓_분석        : 아르켓 상세 분석
  4) 탑텐_분석          : 탑텐 상세 분석
  5) 가격대_비교        : 3사 가격대 분포 비교
  6) 핵심아이템_비교    : 3사 TOP10 핵심 아이템
  7) 랭킹변동_비교      : 3사 랭킹 변동 현황
  8) 아이템비중_차트    : 파이차트 데이터
  9) 수집_히스토리      : 전 브랜드 히스토리 추적
"""

import os
import glob
import json
import re
import shutil
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ─── 경로 설정 ───
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_FILE = os.path.join(WORK_DIR, 'all_brands_history.json')
HISTORY_BACKUP = os.path.join(WORK_DIR, 'all_brands_history_backup.json')

# ─── 브랜드 설정 자동 로드 (brands_config.json) ───
def _load_brands_config():
    cfg_path = os.path.join(WORK_DIR, 'brands_config.json')
    if os.path.exists(cfg_path):
        with open(cfg_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

_BRANDS_JSON = _load_brands_config()

BRAND_CONFIG = {}
for _bn, _bc in _BRANDS_JSON.items():
    BRAND_CONFIG[_bn] = {
        'file_pattern': _bc['file_pattern'],
        'color': _bc.get('color_hex', 'C41E3A'),
        'compare_sheets': _bc.get('compare_sheets', []),
        'category_map': _bc.get('category_map', {}),
    }

# ─── 스타일 ───
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
UP_FONT = Font(color='FF0000', bold=True)
DOWN_FONT = Font(color='0070C0', bold=True)
NEW_FONT = Font(color='00B050', bold=True)
BRAND_FILLS = {bn: PatternFill('solid', fgColor=bc['color'])
               for bn, bc in BRAND_CONFIG.items()}

# ─── 아이템타입 분류 규칙 (통합) ───
UNIQLO_ITEM_RULES = [
    (r'(?i)(ultra\s*light\s*down|울트라라이트다운|패딩|다운)', '패딩/다운'),
    (r'(?i)(coat|코트|트렌치)', '코트'),
    (r'(?i)(jacket|자켓|재킷|블루종|blouson|블레이저|blazer|점퍼|jumper|파카|parka)', '자켓/아우터'),
    (r'(?i)(cardigan|가디건)', '카디건'),
    (r'(?i)(vest|조끼|베스트)', '베스트'),
    (r'(?i)(fleece|후리스|플리스)', '플리스'),
    (r'(?i)(hoodie|후드|후디)', '후드'),
    (r'(?i)(sweat|맨투맨|스웨트)', '맨투맨/스웨트'),
    (r'(?i)(니트|knit|sweater|스웨터|터틀넥)', '니트/스웨터'),
    (r'(?i)(flannel|플란넬|셔츠|shirt|블라우스)', '셔츠'),
    (r'(?i)(polo|폴로)', '폴로'),
    (r'(?i)(airism|에어리즘)', '에어리즘'),
    (r'(?i)(히트텍|heattech)', '히트텍'),
    (r'(?i)(t-shirt|tee|티셔츠|반팔|크루넥T)', '티셔츠'),
    (r'(?i)(jeans|진|데님)', '진/데님'),
    (r'(?i)(chino|치노|팬츠|pants|슬랙스|이지팬츠|카고|바지|트라우저)', '팬츠'),
    (r'(?i)(jogger|조거|스웨트팬츠|트레이닝)', '조거'),
    (r'(?i)(shorts|숏|반바지)', '쇼츠'),
    (r'(?i)(skirt|스커트|치마)', '스커트'),
    (r'(?i)(legging|레깅스)', '레깅스'),
    (r'(?i)(dress|원피스|드레스)', '원피스'),
    (r'(?i)(inner|이너|속옷|브라탑|런닝|팬티)', '이너웨어'),
    (r'(?i)(sock|양말)', '양말'),
    (r'(?i)(pajama|파자마|라운지|loungewear)', '라운지웨어'),
    (r'(?i)(bag|가방|백|토트)', '가방'),
    (r'(?i)(cap|모자|hat|비니|버킷)', '모자'),
    (r'(?i)(towel|타올|수건)', '타올'),
    (r'(?i)(blanket|블랭킷|담요)', '블랭킷'),
]

ARKET_ITEM_RULES = [
    (r'(?i)(coat|코트)', '코트'),
    (r'(?i)(puffer|padded|패딩)', '패딩/다운'),
    (r'(?i)(jacket|자켓|blazer|블레이저)', '자켓/아우터'),
    (r'(?i)(cardigan|카디건)', '카디건'),
    (r'(?i)(vest|gilet)', '베스트'),
    (r'(?i)(hoodie|후디|hooded)', '후드'),
    (r'(?i)(sweatshirt|스웨트)', '맨투맨/스웨트'),
    (r'(?i)(knit|니트|sweater|스웨터|jumper)', '니트/스웨터'),
    (r'(?i)(shirt|셔츠|blouse|블라우스)', '셔츠'),
    (r'(?i)(polo)', '폴로'),
    (r'(?i)(t-shirt|tee|티셔츠)', '티셔츠'),
    (r'(?i)(top|탑|tank)', '탑'),
    (r'(?i)(jeans|jean|denim)', '진/데님'),
    (r'(?i)(trouser|팬츠|pants|chino)', '팬츠'),
    (r'(?i)(shorts)', '쇼츠'),
    (r'(?i)(skirt|스커트)', '스커트'),
    (r'(?i)(legging)', '레깅스'),
    (r'(?i)(dress|드레스)', '원피스'),
    (r'(?i)(jumpsuit|romper)', '점프수트'),
    (r'(?i)(bag|백|tote|backpack|crossbody)', '가방'),
    (r'(?i)(scarf|머플러|muffler)', '스카프/머플러'),
    (r'(?i)(hat|cap|beanie)', '모자'),
    (r'(?i)(sneaker|스니커즈)', '스니커즈'),
    (r'(?i)(boot|부츠)', '부츠'),
    (r'(?i)(sandal|샌들)', '샌들'),
    (r'(?i)(loafer|로퍼)', '로퍼'),
    (r'(?i)(shoe|슈즈|flat|trainer)', '슈즈'),
    (r'(?i)(jewel|ring|necklace|earring|bracelet)', '주얼리'),
    (r'(?i)(wallet|card\s*holder)', '지갑'),
    (r'(?i)(belt|벨트)', '벨트'),
    (r'(?i)(sock|양말)', '양말'),
    (r'(?i)(swim|bikini)', '스윔웨어'),
    (r'(?i)(pyjama|pajama|lounge)', '라운지웨어'),
]

TOPTEN_ITEM_RULES = [
    (r'(?i)(패딩|다운|덕다운|구스|웰론)', '패딩/다운'),
    (r'(?i)(코트|트렌치)', '코트'),
    (r'(?i)(자켓|재킷|점퍼|블레이저|집업|바람막이|야상|아노락)', '자켓/아우터'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(조끼|베스트|vest)', '베스트'),
    (r'(?i)(후리스|플리스|fleece|뽀글이)', '플리스'),
    (r'(?i)(후드|후디|hoodie)', '후드'),
    (r'(?i)(맨투맨|스웨트|크루넥)', '맨투맨/스웨트'),
    (r'(?i)(니트|스웨터|터틀넥|목폴라|폴라)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스|남방)', '셔츠'),
    (r'(?i)(긴팔|롱슬리브)', '긴팔티'),
    (r'(?i)(반팔|반소매|티셔츠|t-shirt|tee)', '티셔츠'),
    (r'(?i)(청바지|데님)', '진/데님'),
    (r'(?i)(슬랙스|치노|팬츠|바지|트라우저|면바지|카고)', '팬츠'),
    (r'(?i)(조거|트레이닝|스웨트팬츠|이지팬츠)', '조거'),
    (r'(?i)(반바지|숏팬츠|쇼츠|shorts)', '쇼츠'),
    (r'(?i)(스커트|치마)', '스커트'),
    (r'(?i)(레깅스)', '레깅스'),
    (r'(?i)(원피스|드레스)', '원피스'),
    (r'(?i)(세트|set|셋업)', '세트'),
    (r'(?i)(내복|내의|발열|히트텍|웜)', '내의/발열'),
    (r'(?i)(속옷|팬티|브라|런닝|캐미)', '이너웨어'),
    (r'(?i)(양말|삭스|sock)', '양말'),
    (r'(?i)(모자|캡|비니|버킷)', '모자'),
    (r'(?i)(가방|백팩|토트|크로스|에코백|숄더)', '가방'),
    (r'(?i)(머플러|스카프|목도리)', '스카프/머플러'),
    (r'(?i)(장갑)', '장갑'),
    (r'(?i)(운동화|스니커즈|신발|슈즈)', '스니커즈'),
    (r'(?i)(슬리퍼|샌들)', '샌들'),
    (r'(?i)(파자마|라운지)', '라운지웨어'),
]

# 미쏘 아이템타입 규칙 (여성 브랜드)
MIXXO_ITEM_RULES = [
    (r'(?i)(코트|하프코트|롱코트|트렌치)', '코트'),
    (r'(?i)(패딩|다운|무스탕|퍼)', '패딩/다운'),
    (r'(?i)(자켓|재킷|블루종|블레이저|워크자켓|트위드)', '자켓/아우터'),
    (r'(?i)(카디건)', '카디건'),
    (r'(?i)(베스트|조끼)', '베스트'),
    (r'(?i)(후드|후디)', '후드'),
    (r'(?i)(맨투맨|스웨트|스웻)', '맨투맨/스웨트'),
    (r'(?i)(니트|스웨터|터틀넥|폴라|풀오버)', '니트/스웨터'),
    (r'(?i)(셔츠|블라우스|남방)', '셔츠'),
    (r'(?i)(티셔츠|반팔|tee)', '티셔츠'),
    (r'(?i)(데님|청바지|진)', '진/데님'),
    (r'(?i)(팬츠|바지|슬랙스|트라우저)', '팬츠'),
    (r'(?i)(스커트|치마|플리츠)', '스커트'),
    (r'(?i)(원피스|드레스)', '원피스'),
    (r'(?i)(가방|백|토트)', '가방'),
]

BRAND_ITEM_RULES = {
    '유니클로': UNIQLO_ITEM_RULES,
    '아르켓':   ARKET_ITEM_RULES,
    '탑텐':     TOPTEN_ITEM_RULES,
    '미쏘':     MIXXO_ITEM_RULES,
}


def log(msg):
    print(msg)


def classify_item_type(name, brand):
    """상품명 + 브랜드에 맞는 규칙으로 아이템타입 분류"""
    if not name:
        return '미분류'
    rules = BRAND_ITEM_RULES.get(brand, UNIQLO_ITEM_RULES)
    for pattern, item_type in rules:
        if re.search(pattern, name):
            return item_type
    return '미분류'


def parse_price(price_str):
    if not price_str:
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(price_str)))
    except:
        return 0


def parse_review(review_str):
    if not review_str or str(review_str) == '없음':
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(review_str)))
    except:
        return 0


def parse_rating(rating_str):
    if not rating_str or str(rating_str) == '없음':
        return 0.0
    try:
        return float(str(rating_str).replace('★', '').strip())
    except:
        return 0.0


def extract_date_from_filename(filepath):
    basename = os.path.basename(filepath)
    match = re.search(r'(\d{8})_(\d{6})', basename)
    if match:
        return match.group(1)
    return datetime.now().strftime('%Y%m%d')


# ══════════════════════════════════════════════════════
#  데이터 로드
# ══════════════════════════════════════════════════════

def find_latest_file(file_pattern):
    """패턴에 맞는 최신 파일 찾기"""
    pattern = os.path.join(WORK_DIR, file_pattern)
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(reverse=True)
    return files[0]


def load_uniqlo(filepath):
    """유니클로 엑셀 → 통일 포맷 변환"""
    wb = load_workbook(filepath, data_only=True)
    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]
        products = []
        for r in range(2, ws.max_row + 1):
            rank = ws.cell(r, 1).value
            if rank is None:
                continue
            name = ws.cell(r, 3).value or ''
            item_type = ws.cell(r, 4).value or classify_item_type(name, '유니클로')
            price = ws.cell(r, 5).value or ''
            rating = ws.cell(r, 8).value or '없음'
            review_count = ws.cell(r, 9).value or '없음'
            products.append({
                'rank': rank, 'name': name, 'item_type': item_type,
                'price': price, 'rating': rating, 'review_count': review_count,
                'brand': '유니클로', 'sheet': sheet_name,
            })
        all_data[sheet_name] = products
    wb.close()
    return all_data


def load_arket(filepath):
    """아르켓 엑셀 → 통일 포맷 변환"""
    wb = load_workbook(filepath, data_only=True)
    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]
        products = []
        for r in range(2, ws.max_row + 1):
            rank = ws.cell(r, 1).value
            if rank is None:
                continue
            name = ws.cell(r, 3).value or ''
            color = ws.cell(r, 4).value or ''
            price = ws.cell(r, 5).value or ''
            item_type = classify_item_type(name, '아르켓')
            products.append({
                'rank': rank, 'name': name, 'item_type': item_type,
                'price': price, 'color': color,
                'rating': '없음', 'review_count': '없음',
                'brand': '아르켓', 'sheet': sheet_name,
            })
        all_data[sheet_name] = products
    wb.close()
    return all_data


def load_topten(filepath):
    """탑텐 엑셀 → 통일 포맷 변환"""
    wb = load_workbook(filepath, data_only=True)
    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]
        products = []
        for r in range(2, ws.max_row + 1):
            rank = ws.cell(r, 1).value
            if rank is None:
                continue
            sub_brand = ws.cell(r, 3).value or ''
            name = ws.cell(r, 4).value or ''
            sale_price = ws.cell(r, 5).value or ''
            discount_rate = ws.cell(r, 7).value or ''
            rating = ws.cell(r, 8).value or '없음'
            review_count = ws.cell(r, 9).value or '없음'
            item_type = classify_item_type(name, '탑텐')
            products.append({
                'rank': rank, 'name': name, 'item_type': item_type,
                'price': sale_price, 'discount_rate': discount_rate,
                'sub_brand': sub_brand,
                'rating': rating, 'review_count': review_count,
                'brand': '탑텐', 'sheet': sheet_name,
            })
        all_data[sheet_name] = products
    wb.close()
    return all_data


def load_mixxo(filepath):
    """미쏘 엑셀 → 통일 포맷 변환"""
    wb = load_workbook(filepath, data_only=True)
    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]
        products = []
        for r in range(2, ws.max_row + 1):
            rank = ws.cell(r, 1).value
            if rank is None:
                continue
            name = ws.cell(r, 3).value or ''
            sale_price = ws.cell(r, 4).value or ''
            original_price = ws.cell(r, 5).value or ''
            discount_rate = ws.cell(r, 6).value or ''
            review_count = ws.cell(r, 7).value or '없음'
            item_type = classify_item_type(name, '미쏘')
            products.append({
                'rank': rank, 'name': name, 'item_type': item_type,
                'price': sale_price, 'original_price': original_price,
                'discount_rate': discount_rate,
                'rating': '없음', 'review_count': review_count,
                'brand': '미쏘', 'sheet': sheet_name,
            })
        all_data[sheet_name] = products
    wb.close()
    return all_data


LOADERS = {
    '유니클로': load_uniqlo,
    '아르켓':   load_arket,
    '탑텐':     load_topten,
    '미쏘':     load_mixxo,
}


def load_all_brands():
    """등록된 모든 브랜드 데이터를 로드 (brands_config.json 기반)"""
    brand_data = {}  # brand_name → {sheet→products}
    brand_files = {}  # brand_name → filepath
    date_key = None

    for brand_name, config in BRAND_CONFIG.items():
        if brand_name not in LOADERS:
            log(f"  [{brand_name}] 로더 없음 → 건너뜀")
            continue
        filepath = find_latest_file(config['file_pattern'])
        if not filepath:
            log(f"  [{brand_name}] 크롤링 파일 없음 → 건너뜀")
            continue
        log(f"  [{brand_name}] {os.path.basename(filepath)}")
        brand_data[brand_name] = LOADERS[brand_name](filepath)
        brand_files[brand_name] = filepath
        dk = extract_date_from_filename(filepath)
        if date_key is None or dk > date_key:
            date_key = dk

    if not date_key:
        date_key = datetime.now().strftime('%Y%m%d')

    return brand_data, brand_files, date_key


# ══════════════════════════════════════════════════════
#  분석 함수
# ══════════════════════════════════════════════════════

def get_compare_products(brand_data, brand_name, gender='여성'):
    """비교용 상품 목록 추출 (여성/남성 기준)"""
    config = BRAND_CONFIG[brand_name]
    for sheet_name, mapped in config['category_map'].items():
        if mapped == gender and sheet_name in brand_data[brand_name]:
            return brand_data[brand_name][sheet_name]
    return []


def analyze_type_distribution(products):
    """상품 목록 → 아이템타입 비중 분석"""
    type_counter = Counter()
    type_products = defaultdict(list)
    for p in products:
        t = p['item_type']
        type_counter[t] += 1
        type_products[t].append(p)

    total = len(products)
    result = []
    for item_type, count in type_counter.most_common():
        pct = round(count / total * 100, 1) if total > 0 else 0
        prices = [parse_price(p['price']) for p in type_products[item_type]]
        avg_price = round(sum(prices) / len(prices)) if prices else 0
        tops = sorted(type_products[item_type], key=lambda x: x['rank'])[:3]
        top_str = ', '.join(f"{p['rank']}위:{p['name'][:15]}" for p in tops)
        result.append({
            'item_type': item_type, 'count': count, 'pct': pct,
            'avg_price': avg_price, 'top_products': top_str,
            'best_rank': min(p['rank'] for p in type_products[item_type]),
        })
    return {'total': total, 'analysis': result}


def analyze_price_bands(products, bands=None):
    """가격대별 분석"""
    if bands is None:
        bands = [
            (0, 10000, '1만원 이하'), (10000, 20000, '1~2만원'),
            (20000, 30000, '2~3만원'), (30000, 50000, '3~5만원'),
            (50000, 70000, '5~7만원'), (70000, 100000, '7~10만원'),
            (100000, 150000, '10~15만원'), (150000, 200000, '15~20만원'),
            (200000, 300000, '20~30만원'), (300000, float('inf'), '30만원 이상'),
        ]
    counter = Counter()
    items = defaultdict(list)
    for p in products:
        price = parse_price(p['price'])
        for lo, hi, label in bands:
            if lo <= price < hi:
                counter[label] += 1
                items[label].append(p['name'][:20])
                break
    total = len(products)
    return {
        'total': total,
        'bands': [(label, counter.get(label, 0),
                    round(counter.get(label, 0) / total * 100, 1) if total > 0 else 0,
                    items.get(label, [])[:3])
                   for _, _, label in bands if counter.get(label, 0) > 0],
    }


# ══════════════════════════════════════════════════════
#  히스토리
# ══════════════════════════════════════════════════════

def load_history():
    for fp in [HISTORY_FILE, HISTORY_BACKUP]:
        if os.path.exists(fp):
            try:
                with open(fp, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if data:
                    return data
            except:
                continue
    return {}


def save_history(history):
    if os.path.exists(HISTORY_FILE):
        try:
            shutil.copy2(HISTORY_FILE, HISTORY_BACKUP)
        except:
            pass
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def update_history(brand_data, date_key):
    """전체 브랜드 데이터를 히스토리에 추가"""
    log("  [히스토리] 업데이트 중...")
    history = load_history()

    for brand_name, sheets in brand_data.items():
        config = BRAND_CONFIG[brand_name]
        for sheet_name, products in sheets.items():
            cat_key = f"{brand_name}_{sheet_name}"
            if cat_key not in history:
                history[cat_key] = {}
            day_data = {}
            for p in products:
                day_data[p['name'][:40]] = {
                    'rank': p['rank'],
                    'item_type': p['item_type'],
                    'price': str(p['price']),
                }
            history[cat_key][date_key] = day_data

    save_history(history)

    all_dates = set()
    for cat in history.values():
        all_dates.update(cat.keys())
    sorted_dates = sorted(all_dates)
    log(f"  [히스토리] 총 {len(sorted_dates)}일치 데이터 보유")
    if len(sorted_dates) > 1:
        log(f"  [히스토리] 기간: {sorted_dates[0]} ~ {sorted_dates[-1]}")
    return history


def analyze_ranking_changes(history, brand_name, sheet_name):
    """특정 브랜드·시트의 랭킹 변동 분석"""
    cat_key = f"{brand_name}_{sheet_name}"
    if cat_key not in history:
        return {'status': 'no_data', 'changes': [], 'dropped': []}

    dates = sorted(history[cat_key].keys())
    if len(dates) < 2:
        changes = []
        if dates:
            cur = history[cat_key][dates[-1]]
            for name, info in cur.items():
                changes.append({
                    'name': name[:30], 'current_rank': info['rank'],
                    'rank_change': 0, 'status': '초회',
                    'item_type': info['item_type'], 'price': info['price'],
                })
            changes.sort(key=lambda x: x['current_rank'])
        return {
            'status': 'first_data',
            'message': f'첫 수집 ({dates[0] if dates else "N/A"})',
            'changes': changes, 'dropped': [],
        }

    cur_date, prev_date = dates[-1], dates[-2]
    cur, prev = history[cat_key][cur_date], history[cat_key][prev_date]

    changes = []
    for name, info in cur.items():
        cr = info['rank']
        if name in prev:
            pr = prev[name]['rank']
            rc = pr - cr
            st = '상승' if rc > 0 else ('하락' if rc < 0 else '유지')
        else:
            rc, st = 0, '신규진입'
        changes.append({
            'name': name[:30], 'current_rank': cr, 'rank_change': rc,
            'status': st, 'item_type': info['item_type'], 'price': info['price'],
        })

    dropped = [{'name': n[:30], 'prev_rank': prev[n]['rank'],
                'item_type': prev[n]['item_type']}
               for n in prev if n not in cur]
    changes.sort(key=lambda x: x['current_rank'])

    return {
        'status': 'compared', 'current_date': cur_date, 'prev_date': prev_date,
        'changes': changes, 'dropped': dropped, 'total_dates': len(dates),
    }


# ══════════════════════════════════════════════════════
#  엑셀 생성 헬퍼
# ══════════════════════════════════════════════════════

def hdr(ws, row, max_col, fill_color='2F5496'):
    fill = PatternFill('solid', fgColor=fill_color)
    font = Font(bold=True, color='FFFFFF', size=10)
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def drow(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def title_row(ws, row, text, max_col, color='2F5496', size=13):
    ws.merge_cells(f'A{row}:{get_column_letter(max_col)}{row}')
    cell = ws.cell(row, 1, text)
    cell.font = Font(bold=True, size=size, color=color)
    cell.alignment = Alignment(horizontal='center')


def section_row(ws, row, text, max_col, color='C00000'):
    ws.merge_cells(f'A{row}:{get_column_letter(max_col)}{row}')
    ws.cell(row, 1, text).font = Font(bold=True, size=11, color=color)


# ══════════════════════════════════════════════════════
#  메인 엑셀 생성
# ══════════════════════════════════════════════════════

def create_combined_excel(brand_data, history, date_key):
    _bl = list(brand_data.keys())
    _bn_str = ' \u00b7 '.join(_bl)
    log("\n============================================================")
    log(f"[출력] {len(_bl)}사 통합 분석 엑셀 생성")
    log("============================================================")

    wb = Workbook()
    wb.remove(wb.active)
    MAX = 12  # 최대 컬럼 수 (타이틀 병합용)

    # ═══════════════════════════════════════
    # 시트 1: 종합 대시보드
    # ═══════════════════════════════════════
    ws = wb.create_sheet('종합_대시보드')
    log("  -> 시트 [종합_대시보드]")

    row = 1
    title_row(ws, row, f'{len(_bl)}사 브랜드 랭킹 통합 분석 ({date_key[:4]}.{date_key[4:6]}.{date_key[6:8]})', MAX)
    row += 1
    ws.cell(row, 1, f'{_bn_str}의 여성/남성 랭킹을 비교 분석합니다.').font = Font(italic=True, color='666666')

    for gender in ['여성', '남성']:
        row += 2
        section_row(ws, row, f'■ {gender} 카테고리 - {len(_bl)}사 아이템타입 비중 비교', MAX, 'C00000')

        # 헤더
        row += 1
        cols = ['순번']
        for bn in _bl:
            cols += [f'{bn} 타입', f'{bn} 수', f'{bn} %']
        for ci, h in enumerate(cols, 1):
            ws.cell(row, ci, h)
        hdr(ws, row, len(cols), '333333')

        # 각 브랜드 분석 취합
        brand_analyses = {}
        max_rows = 0
        for bn in _bl:
            if bn in brand_data:
                prods = get_compare_products(brand_data, bn, gender)
                if prods:
                    analysis = analyze_type_distribution(prods)
                    brand_analyses[bn] = analysis['analysis']
                    max_rows = max(max_rows, len(analysis['analysis']))

        for i in range(max_rows):
            row += 1
            ws.cell(row, 1, i + 1)
            col_offset = 2
            for bn in _bl:
                items = brand_analyses.get(bn, [])
                if i < len(items):
                    ws.cell(row, col_offset, items[i]['item_type'])
                    ws.cell(row, col_offset + 1, items[i]['count'])
                    ws.cell(row, col_offset + 2, items[i]['pct'])
                col_offset += 3
            drow(ws, row, len(cols))

            # 첫 줄 하이라이트
            if i == 0:
                for c in range(1, len(cols) + 1):
                    ws.cell(row, c).fill = HIGHLIGHT_FILL

        # 소계
        row += 1
        ws.cell(row, 1, '합계').font = Font(bold=True)
        col_offset = 2
        for bn in _bl:
            items = brand_analyses.get(bn, [])
            total_count = sum(it['count'] for it in items)
            ws.cell(row, col_offset, '-')
            ws.cell(row, col_offset + 1, total_count)
            ws.cell(row, col_offset + 2, '100%')
            col_offset += 3
        drow(ws, row, len(cols))
        for c in range(1, len(cols) + 1):
            ws.cell(row, c).font = Font(bold=True)

    for ci, w in enumerate([6, 14, 5, 6, 14, 5, 6, 14, 5, 6], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════
    # 시트 2~4: 브랜드별 상세 분석
    # ═══════════════════════════════════════
    for brand_name in _bl:
        if brand_name not in brand_data:
            continue
        ws_b = wb.create_sheet(f'{brand_name}_분석')
        log(f"  -> 시트 [{brand_name}_분석]")
        color = BRAND_CONFIG[brand_name]['color']
        config = BRAND_CONFIG[brand_name]

        row = 1
        title_row(ws_b, row, f'{brand_name} 랭킹 상세 분석', 8, color)

        for sheet_name, products in brand_data[brand_name].items():
            mapped_cat = config['category_map'].get(sheet_name, sheet_name)
            if not products:
                continue

            analysis = analyze_type_distribution(products)

            row += 2
            section_row(ws_b, row, f'■ {mapped_cat} ({sheet_name}) - {analysis["total"]}개 상품', 8)

            row += 1
            for ci, h in enumerate(['순번', '아이템타입', '상품수', '비중(%)', '평균가격', '최고순위', '구성바', '대표상품'], 1):
                ws_b.cell(row, ci, h)
            hdr(ws_b, row, 8, color)

            for idx, item in enumerate(analysis['analysis'], 1):
                row += 1
                ws_b.cell(row, 1, idx)
                ws_b.cell(row, 2, item['item_type'])
                ws_b.cell(row, 3, item['count'])
                ws_b.cell(row, 4, item['pct'])
                ws_b.cell(row, 5, f"{item['avg_price']:,}원" if item['avg_price'] else '-')
                ws_b.cell(row, 6, item['best_rank'])
                bar = '█' * int(item['pct'] / 5) + '░' * max(0, 20 - int(item['pct'] / 5))
                ws_b.cell(row, 7, bar)
                ws_b.cell(row, 8, item['top_products'])
                drow(ws_b, row, 8)
                if item['pct'] >= 15:
                    for c in range(1, 9):
                        ws_b.cell(row, c).fill = HIGHLIGHT_FILL

        for ci, w in enumerate([6, 16, 8, 8, 12, 8, 22, 50], 1):
            ws_b.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════
    # 시트 5: 가격대 비교
    # ═══════════════════════════════════════
    ws5 = wb.create_sheet('가격대_비교')
    log("  -> 시트 [가격대_비교]")

    row = 1
    title_row(ws5, row, f'{len(_bl)}사 가격대 분포 비교', 10)

    for gender in ['여성', '남성']:
        row += 2
        section_row(ws5, row, f'■ {gender} 카테고리 가격대 비교', 10)

        row += 1
        cols5 = ['가격대']
        for bn in _bl:
            cols5 += [f'{bn} 수', f'{bn} %', f'{bn} 바']
        for ci, h in enumerate(cols5, 1):
            ws5.cell(row, ci, h)
        hdr(ws5, row, len(cols5), '333333')

        # 분석
        band_def = [
            (0, 10000, '1만원 이하'), (10000, 20000, '1~2만원'),
            (20000, 30000, '2~3만원'), (30000, 50000, '3~5만원'),
            (50000, 70000, '5~7만원'), (70000, 100000, '7~10만원'),
            (100000, 150000, '10~15만원'), (150000, 200000, '15~20만원'),
            (200000, 300000, '20~30만원'), (300000, float('inf'), '30만원 이상'),
        ]
        brand_bands = {}
        for bn in _bl:
            if bn in brand_data:
                prods = get_compare_products(brand_data, bn, gender)
                if prods:
                    pb = analyze_price_bands(prods, band_def)
                    # label → (count, pct)
                    brand_bands[bn] = {b[0]: (b[1], b[2]) for b in pb['bands']}

        for _, _, label in band_def:
            has_data = any(label in brand_bands.get(bn, {}) for bn in _bl)
            if not has_data:
                continue
            row += 1
            ws5.cell(row, 1, label)
            col_offset = 2
            for bn in _bl:
                bd = brand_bands.get(bn, {})
                if label in bd:
                    cnt, pct = bd[label]
                    ws5.cell(row, col_offset, cnt)
                    ws5.cell(row, col_offset + 1, pct)
                    bar = '█' * int(pct / 5)
                    ws5.cell(row, col_offset + 2, bar)
                else:
                    ws5.cell(row, col_offset, 0)
                    ws5.cell(row, col_offset + 1, 0)
                    ws5.cell(row, col_offset + 2, '')
                col_offset += 3
            drow(ws5, row, len(cols5))

    for ci, w in enumerate([12, 6, 6, 14, 6, 6, 14, 6, 6, 14], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════
    # 시트 6: 핵심 아이템 비교
    # ═══════════════════════════════════════
    ws6 = wb.create_sheet('핵심아이템_비교')
    log("  -> 시트 [핵심아이템_비교]")

    row = 1
    title_row(ws6, row, f'{len(_bl)}사 핵심 아이템 TOP 10 비교', 12)

    for gender in ['여성', '남성']:
        row += 2
        section_row(ws6, row, f'■ {gender} 카테고리 핵심 아이템', 12)

        row += 1
        cols6 = ['순위']
        for bn in _bl:
            cols6 += [f'{bn} 상품명', f'{bn} 타입', f'{bn} 가격']
        for ci, h in enumerate(cols6, 1):
            ws6.cell(row, ci, h)
        hdr(ws6, row, len(cols6), '333333')

        brand_tops = {}
        for bn in _bl:
            if bn in brand_data:
                prods = get_compare_products(brand_data, bn, gender)
                brand_tops[bn] = sorted(prods, key=lambda x: x['rank'])[:10] if prods else []

        for i in range(10):
            row += 1
            ws6.cell(row, 1, i + 1)
            col_offset = 2
            for bn in _bl:
                tops = brand_tops.get(bn, [])
                if i < len(tops):
                    ws6.cell(row, col_offset, tops[i]['name'][:25])
                    ws6.cell(row, col_offset + 1, tops[i]['item_type'])
                    ws6.cell(row, col_offset + 2, tops[i]['price'])
                col_offset += 3
            drow(ws6, row, len(cols6))
            if i < 3:
                for c in range(1, len(cols6) + 1):
                    ws6.cell(row, c).fill = HIGHLIGHT_FILL

    for ci, w in enumerate([5, 28, 12, 12, 28, 12, 12, 28, 12, 12], 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════
    # 시트 7: 랭킹 변동 비교
    # ═══════════════════════════════════════
    ws7 = wb.create_sheet('랭킹변동_비교')
    log("  -> 시트 [랭킹변동_비교]")

    row = 1
    title_row(ws7, row, f'{len(_bl)}사 랭킹 변동 현황', 7)

    for brand_name in _bl:
        if brand_name not in brand_data:
            continue
        config = BRAND_CONFIG[brand_name]
        color = config['color']

        for sheet_name in config.get('compare_sheets', []):
            if sheet_name not in brand_data[brand_name]:
                continue
            mapped = config['category_map'].get(sheet_name, sheet_name)
            rc = analyze_ranking_changes(history, brand_name, sheet_name)

            row += 2
            section_row(ws7, row, f'■ {brand_name} - {mapped}', 7, color)

            if rc['status'] == 'first_data':
                row += 1
                ws7.cell(row, 1, f'{rc.get("message", "첫 수집 데이터")} - 2회 이상 수집 시 비교 가능').font = Font(italic=True, color='666666')

            elif rc['status'] == 'compared':
                row += 1
                ups = sum(1 for c in rc['changes'] if c['status'] == '상승')
                downs = sum(1 for c in rc['changes'] if c['status'] == '하락')
                news = sum(1 for c in rc['changes'] if c['status'] == '신규진입')
                sames = sum(1 for c in rc['changes'] if c['status'] == '유지')
                summary = f'{rc["prev_date"]}→{rc["current_date"]}  상승 {ups} | 하락 {downs} | 유지 {sames} | 신규 {news} | 이탈 {len(rc["dropped"])}'
                ws7.cell(row, 1, summary).font = Font(bold=True, size=10)

            # 헤더
            row += 1
            for ci, h in enumerate(['순위', '상품명', '아이템타입', '가격', '변동', '상태', ''], 1):
                ws7.cell(row, ci, h)
            hdr(ws7, row, 6, color)

            for item in rc.get('changes', [])[:20]:
                row += 1
                ws7.cell(row, 1, item['current_rank'])
                ws7.cell(row, 2, item['name'])
                ws7.cell(row, 3, item['item_type'])
                ws7.cell(row, 4, item['price'])
                chg = item['rank_change']
                if item['status'] == '신규진입':
                    ws7.cell(row, 5, 'NEW'); ws7.cell(row, 5).font = NEW_FONT
                    ws7.cell(row, 6, '신규'); ws7.cell(row, 6).font = NEW_FONT
                elif item['status'] == '초회':
                    ws7.cell(row, 5, '-'); ws7.cell(row, 6, '초회')
                elif chg > 0:
                    ws7.cell(row, 5, f'▲{chg}'); ws7.cell(row, 5).font = UP_FONT
                    ws7.cell(row, 6, '상승'); ws7.cell(row, 6).font = UP_FONT
                elif chg < 0:
                    ws7.cell(row, 5, f'▼{abs(chg)}'); ws7.cell(row, 5).font = DOWN_FONT
                    ws7.cell(row, 6, '하락'); ws7.cell(row, 6).font = DOWN_FONT
                else:
                    ws7.cell(row, 5, '-'); ws7.cell(row, 6, '유지')
                drow(ws7, row, 6)

    ws7.column_dimensions['A'].width = 6
    ws7.column_dimensions['B'].width = 35
    ws7.column_dimensions['C'].width = 14
    ws7.column_dimensions['D'].width = 12
    ws7.column_dimensions['E'].width = 8
    ws7.column_dimensions['F'].width = 8

    # ═══════════════════════════════════════
    # 시트 8: 아이템비중 차트
    # ═══════════════════════════════════════
    ws8 = wb.create_sheet('아이템비중_차트')
    log("  -> 시트 [아이템비중_차트]")

    col_offset = 0
    for brand_name in _bl:
        if brand_name not in brand_data:
            continue
        prods = get_compare_products(brand_data, brand_name, '여성')
        if not prods:
            continue
        analysis = analyze_type_distribution(prods)

        sc = col_offset * 4 + 1
        ws8.cell(1, sc, f'{brand_name} 여성 아이템타입 비중')
        ws8.cell(1, sc).font = Font(bold=True, size=11)

        ws8.cell(2, sc, '아이템타입')
        ws8.cell(2, sc + 1, '상품수')
        ws8.cell(2, sc + 2, '비중(%)')

        for r, item in enumerate(analysis['analysis'], 3):
            ws8.cell(r, sc, item['item_type'])
            ws8.cell(r, sc + 1, item['count'])
            ws8.cell(r, sc + 2, item['pct'])

        dc = len(analysis['analysis'])
        if dc > 0:
            pie = PieChart()
            pie.title = f'{brand_name} 여성'
            pie.style = 26
            pie.width = 16
            pie.height = 12

            data_ref = Reference(ws8, min_col=sc + 1, min_row=2, max_row=2 + dc)
            cats_ref = Reference(ws8, min_col=sc, min_row=3, max_row=2 + dc)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)

            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showCatName = True

            ws8.add_chart(pie, f'{get_column_letter(sc)}{dc + 5}')

        col_offset += 1

    # ═══════════════════════════════════════
    # 시트 9: 수집 히스토리
    # ═══════════════════════════════════════
    ws9 = wb.create_sheet('수집_히스토리')
    log("  -> 시트 [수집_히스토리]")

    row = 1
    title_row(ws9, row, '3사 브랜드 수집 히스토리 (랭킹 추적)', 8)

    all_dates = set()
    for cat in history.values():
        all_dates.update(cat.keys())
    sorted_dates = sorted(all_dates)

    row += 1
    ws9.cell(row, 1, f'총 수집 횟수: {len(sorted_dates)}회').font = Font(italic=True, color='666666')

    for brand_name in _bl:
        if brand_name not in brand_data:
            continue
        config = BRAND_CONFIG[brand_name]
        color = config['color']

        for sheet_name in config.get('compare_sheets', []):
            cat_key = f"{brand_name}_{sheet_name}"
            if cat_key not in history:
                continue

            dates = sorted(history[cat_key].keys())
            mapped = config['category_map'].get(sheet_name, sheet_name)

            row += 2
            max_c = 3 + len(dates)
            section_row(ws9, row, f'■ {brand_name} - {mapped} 랭킹 추이', max_c, color)

            row += 1
            ws9.cell(row, 1, '상품명')
            ws9.cell(row, 2, '브랜드')
            ws9.cell(row, 3, '아이템타입')
            for di, d in enumerate(dates):
                ws9.cell(row, 4 + di, f'{d[4:6]}/{d[6:8]}')
            hdr(ws9, row, max_c, color)

            all_prods = {}
            for d in dates:
                for name, info in history[cat_key][d].items():
                    if name not in all_prods:
                        all_prods[name] = info['item_type']

            for name in sorted(all_prods.keys()):
                row += 1
                ws9.cell(row, 1, name[:35])
                ws9.cell(row, 2, brand_name)
                ws9.cell(row, 3, all_prods[name])
                for di, d in enumerate(dates):
                    if name in history[cat_key][d]:
                        ws9.cell(row, 4 + di, history[cat_key][d][name]['rank'])
                    else:
                        ws9.cell(row, 4 + di, '-')
                drow(ws9, row, max_c)

    ws9.column_dimensions['A'].width = 38
    ws9.column_dimensions['B'].width = 10
    ws9.column_dimensions['C'].width = 14

    # ═══════════════════════════════════════
    # 저장
    # ═══════════════════════════════════════
    filename = f'3사_통합_랭킹분석_{date_key}.xlsx'
    filepath = os.path.join(WORK_DIR, filename)
    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime('%H%M%S')
        filename = f'3사_통합_랭킹분석_{date_key}_{ts}.xlsx'
        filepath = os.path.join(WORK_DIR, filename)
        wb.save(filepath)

    log(f"\n  [OK] 저장 완료: {filename}")
    return filepath


# ══════════════════════════════════════════════════════
#  콘솔 요약
# ══════════════════════════════════════════════════════

def print_summary(brand_data):
    _bl = list(brand_data.keys())
    log("\n" + "=" * 60)
    log(f"  {len(_bl)}사 브랜드 랭킹 비교 요약")
    log("=" * 60)

    for gender in ['여성', '남성']:
        log(f"\n  ■ {gender} 카테고리")
        log(f"  {'─' * 54}")
        log(f"  {'브랜드':8s} | {'1위 타입':12s} | {'2위 타입':12s} | {'평균가격':>10s} | 총수")

        for bn in _bl:
            prods = get_compare_products(brand_data, bn, gender)
            if not prods:
                continue
            analysis = analyze_type_distribution(prods)
            items = analysis['analysis']
            t1 = items[0]['item_type'] if len(items) >= 1 else '-'
            t2 = items[1]['item_type'] if len(items) >= 2 else '-'
            prices = [parse_price(p['price']) for p in prods if parse_price(p['price']) > 0]
            avg = round(sum(prices) / len(prices)) if prices else 0
            log(f"  {bn:8s} | {t1:12s} | {t2:12s} | {avg:>8,}원 | {len(prods)}개")

        # TOP3 비교
        log(f"\n  {gender} TOP3 상품:")
        for bn in _bl:
            prods = get_compare_products(brand_data, bn, gender)
            if not prods:
                continue
            tops = sorted(prods, key=lambda x: x['rank'])[:3]
            log(f"    [{bn}]")
            for p in tops:
                log(f"      {p['rank']}위 {p['name'][:30]:30s} {p['price']}")


# ══════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════

def main():
    _bl = list(BRAND_CONFIG.keys())
    _bn_str = ' \u00b7 '.join(_bl)
    log("=" * 60)
    log(f"  {len(_bl)}사 브랜드 통합 랭킹 분석 도구")
    log(f"  {_bn_str}")
    log("=" * 60)

    # 1. 데이터 로드
    log("\n[1/4] 크롤링 데이터 로드")
    brand_data, brand_files, date_key = load_all_brands()

    if not brand_data:
        log("\n[ERROR] 로드된 브랜드 데이터가 없습니다.")
        return

    total = sum(sum(len(p) for p in sheets.values()) for sheets in brand_data.values())
    log(f"  총 {len(brand_data)}개 브랜드, {total}개 상품 로드 완료")

    # 2. 히스토리 업데이트
    log(f"\n[2/4] 히스토리 업데이트 (날짜: {date_key})")
    history = update_history(brand_data, date_key)

    # 3. 콘솔 요약
    log(f"\n[3/4] 분석 결과 요약")
    print_summary(brand_data)

    # 4. 엑셀 생성
    log(f"\n[4/4] 통합 엑셀 출력")
    output_path = create_combined_excel(brand_data, history, date_key)

    log("\n" + "=" * 60)
    log("  통합 분석 완료!")
    log(f"  파일: {os.path.basename(output_path)}")
    log("=" * 60)


if __name__ == '__main__':
    main()
