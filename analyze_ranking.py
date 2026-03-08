# -*- coding: utf-8 -*-
"""
패션 랭킹 아이템 분류 분석기
- 패션 업계 표준 아이템 분류 체계(대분류/중분류/소분류) 적용
- 크롤링된 랭킹 엑셀 파일을 읽어 자동 분류 후 분석 결과 출력 & 엑셀 저장
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter, defaultdict
from datetime import datetime
import re
import os

# ============================================================
# 1. 패션 아이템 분류 체계 (대분류 > 중분류 > 소분류)
# ============================================================
# 패션 업계에서 사용하는 표준 분류 기준:
#   대분류: 의류(Apparel) / 잡화(Accessories) / 언더웨어(Innerwear)
#   중분류: 아우터 / 상의 / 하의 / 원피스&스커트 / 니트 / 이너웨어 / 홈웨어 / 악세서리
#   소분류: 구체적 아이템 타입

ITEM_CLASSIFICATION = {
    # ── 아우터 (Outerwear) ──
    "아우터": {
        "대분류": "의류",
        "중분류": "아우터",
        "keywords": [
            # 재킷류
            ("블레이저", "재킷"), ("재킷", "재킷"), ("점퍼", "점퍼/점프수트"),
            ("블루종", "블루종"), ("해링턴", "재킷"),
            # 코트류
            ("코트", "코트"), ("트렌치", "트렌치코트"),
            # 파카/패딩
            ("파카", "파카"), ("패딩", "패딩"), ("다운", "패딩"),
            ("푸퍼", "패딩"),
            # 기타 아우터
            ("윈드블럭", "윈드브레이커"), ("윈드브레이커", "윈드브레이커"),
            ("집업", "집업"), ("후디", "후드집업"),
            ("베스트", "베스트/조끼"), ("조끼", "베스트/조끼"),
        ]
    },
    # ── 상의 (Tops) ──
    "상의": {
        "대분류": "의류",
        "중분류": "상의",
        "keywords": [
            # T셔츠류
            ("크루넥T", "티셔츠"), ("T(", "티셔츠"), ("T셔츠", "티셔츠"),
            ("UT(", "그래픽티"), ("UT ", "그래픽티"), ("그래픽T", "그래픽티"),
            ("탱크탑", "탱크탑"), ("슬리브리스", "슬리브리스"),
            # 셔츠/블라우스
            ("셔츠", "셔츠"), ("블라우스", "블라우스"),
            # 스웨트류
            ("스웨트셔츠", "스웨트셔츠"), ("맨투맨", "스웨트셔츠"),
            ("후디", "후디"), ("후드", "후디"),
            # 폴로/피케
            ("폴로", "폴로"), ("피케", "폴로"),
        ]
    },
    # ── 니트 (Knitwear) ──
    "니트": {
        "대분류": "의류",
        "중분류": "니트",
        "keywords": [
            ("가디건", "가디건"), ("스웨터", "스웨터/풀오버"),
            ("니트", "니트"), ("카디건", "가디건"),
            ("터틀넥", "터틀넥"), ("풀오버", "스웨터/풀오버"),
            ("케이블", "케이블니트"),
        ]
    },
    # ── 하의 (Bottoms) ──
    "하의": {
        "대분류": "의류",
        "중분류": "하의",
        "keywords": [
            # 진/데님
            ("진(", "데님/진"), ("진 ", "데님/진"), ("커브진", "데님/진"),
            ("스트레이트진", "데님/진"), ("배기진", "데님/진"),
            # 팬츠류
            ("팬츠", "팬츠"), ("치노", "치노팬츠"), ("카고", "카고팬츠"),
            ("조거", "조거팬츠"), ("트라우저", "트라우저"),
            ("레깅스", "레깅스"), ("이지팬츠", "이지팬츠"),
            # 쇼츠
            ("쇼츠", "쇼츠/반바지"), ("반바지", "쇼츠/반바지"),
            ("숏팬츠", "쇼츠/반바지"),
        ]
    },
    # ── 원피스 & 스커트 (Dresses & Skirts) ──
    "원피스&스커트": {
        "대분류": "의류",
        "중분류": "원피스&스커트",
        "keywords": [
            ("원피스", "원피스"), ("드레스", "원피스"),
            ("스커트", "스커트"), ("미디스커트", "스커트"),
            ("미니스커트", "스커트"), ("플리츠", "스커트"),
            ("스코츠", "스코츠"),
        ]
    },
    # ── 이너웨어 (Innerwear/Underwear) ──
    "이너웨어": {
        "대분류": "언더웨어",
        "중분류": "이너웨어",
        "keywords": [
            ("브라", "브라"), ("브래지어", "브라"),
            ("캐미솔", "캐미솔"), ("슬립", "슬립"),
            ("팬티", "팬티"), ("드로즈", "드로즈"),
            ("런닝", "런닝셔츠"), ("바디수트", "바디수트"),
            ("심리스", "심리스이너"),
        ]
    },
    # ── 홈웨어/파자마 (Loungewear) ──
    "홈웨어": {
        "대분류": "의류",
        "중분류": "홈웨어",
        "keywords": [
            ("파자마", "파자마"), ("잠옷", "파자마"),
            ("라운지", "라운지웨어"), ("룸웨어", "라운지웨어"),
            ("이지웨어", "이지웨어"),
        ]
    },
    # ── 악세서리 (Accessories) ──
    "악세서리": {
        "대분류": "잡화",
        "중분류": "악세서리",
        "keywords": [
            ("백", "가방"), ("가방", "가방"), ("토트", "토트백"),
            ("숄더", "숄더백"), ("크로스", "크로스백"),
            ("모자", "모자"), ("캡", "모자"), ("햇", "모자"),
            ("벨트", "벨트"), ("스카프", "스카프"), ("머플러", "머플러"),
            ("양말", "양말"), ("삭스", "양말"),
            ("장갑", "장갑"), ("우산", "우산"),
            ("슈즈", "슈즈"), ("스니커즈", "스니커즈"),
            ("슬리퍼", "슬리퍼"), ("샌들", "샌들"),
        ]
    },
    # ── 베이비/유아 전용 ──
    "베이비": {
        "대분류": "의류",
        "중분류": "베이비",
        "keywords": [
            ("커버올", "커버올"), ("바디수트", "바디수트"),
            ("롬퍼", "롬퍼"),
        ]
    },
}

# 셔츠재킷 같은 복합 아이템 우선 처리 규칙
PRIORITY_RULES = [
    # (키워드, 대분류, 중분류, 소분류) - 순서 중요: 먼저 매칭되면 우선 적용
    ("셔츠재킷", "의류", "아우터", "셔츠재킷"),
    ("셔츠자켓", "의류", "아우터", "셔츠재킷"),
    ("풀집파카", "의류", "아우터", "파카"),
    ("풀파카", "의류", "아우터", "파카"),
    ("풀집후디", "의류", "아우터", "후드집업"),
    ("풀집", "의류", "아우터", "집업"),
    ("스웨트셔츠", "의류", "상의", "스웨트셔츠"),
    ("오버셔츠", "의류", "아우터", "오버셔츠"),
    ("데님셔츠", "의류", "상의", "셔츠"),
    ("커버올", "의류", "베이비", "커버올"),
    ("바디수트", "언더웨어", "이너웨어", "바디수트"),
    ("파자마", "의류", "홈웨어", "파자마"),
    ("미디스커트", "의류", "원피스&스커트", "스커트"),
    ("미니스코츠", "의류", "원피스&스커트", "스코츠"),
    ("스웨트와이드팬츠", "의류", "하의", "팬츠"),
    ("스웨트스트레이트팬츠", "의류", "하의", "팬츠"),
    ("카고조거팬츠", "의류", "하의", "카고팬츠"),
    ("퍼스트브라", "언더웨어", "이너웨어", "브라"),
    ("와이어리스브라", "언더웨어", "이너웨어", "브라"),
    ("COOL-AIR 브라", "언더웨어", "이너웨어", "브라"),
    ("숄더백", "잡화", "악세서리", "숄더백"),
    ("캔버스토트", "잡화", "악세서리", "토트백"),
    ("Tote", "잡화", "악세서리", "토트백"),
    ("Trainers", "잡화", "악세서리", "스니커즈"),
    ("Boot", "잡화", "악세서리", "부츠"),
    ("Sneaker", "잡화", "악세서리", "스니커즈"),
]


def classify_item(product_name, category_hint=""):
    """
    상품명을 분석하여 (대분류, 중분류, 소분류)를 반환
    category_hint: 시트명 등에서 알 수 있는 카테고리 힌트 (예: 'WOMEN_아우터')
    """
    name = product_name.strip()
    name_upper = name.upper()

    # 1) 우선 처리 규칙 체크
    for keyword, major, mid, sub in PRIORITY_RULES:
        if keyword.upper() in name_upper or keyword in name:
            return (major, mid, sub)

    # 2) 일반 분류 체계 매칭
    for cat_name, cat_info in ITEM_CLASSIFICATION.items():
        for keyword, sub_type in cat_info["keywords"]:
            if keyword.upper() in name_upper or keyword in name:
                return (cat_info["대분류"], cat_info["중분류"], sub_type)

    # 3) 카테고리 힌트 활용
    if category_hint:
        hint = category_hint.upper()
        if "아우터" in hint:
            return ("의류", "아우터", "기타아우터")
        elif "상의" in hint:
            return ("의류", "상의", "기타상의")
        elif "팬츠" in hint or "하의" in hint:
            return ("의류", "하의", "기타하의")
        elif "드레스" in hint or "스커트" in hint:
            return ("의류", "원피스&스커트", "기타")
        elif "이너" in hint:
            return ("언더웨어", "이너웨어", "기타이너")
        elif "홈웨어" in hint:
            return ("의류", "홈웨어", "기타홈웨어")
        elif "악세서리" in hint or "액세서리" in hint:
            return ("잡화", "악세서리", "기타잡화")

    return ("미분류", "미분류", "미분류")


def load_uniqlo_data(filepath):
    """유니클로 엑셀 파일에서 데이터 로드"""
    wb = openpyxl.load_workbook(filepath)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 시트명에서 성별/카테고리 추출
        parts = sheet_name.split("_", 1)
        gender = parts[0] if len(parts) > 0 else ""
        tab_name = parts[1] if len(parts) > 1 else ""

        for row in range(2, ws.max_row + 1):
            rank = ws.cell(row, 1).value
            product_name = ws.cell(row, 3).value  # 상품명
            price_str = ws.cell(row, 4).value  # 가격
            colors = ws.cell(row, 5).value  # 컬러수
            rating = ws.cell(row, 7).value  # 평점
            reviews = ws.cell(row, 8).value  # 리뷰수

            if not product_name:
                continue

            # 가격 숫자 변환
            price = 0
            if price_str:
                digits = re.sub(r'[^\d]', '', str(price_str))
                price = int(digits) if digits else 0

            # 리뷰수 숫자 변환
            review_count = 0
            if reviews:
                digits = re.sub(r'[^\d]', '', str(reviews))
                review_count = int(digits) if digits else 0

            # 평점 변환
            rating_val = 0
            if rating:
                try:
                    rating_val = float(str(rating).replace('점', '').strip())
                except:
                    rating_val = 0

            all_data.append({
                "시트": sheet_name,
                "성별": gender,
                "탭": tab_name,
                "순위": rank,
                "상품명": product_name,
                "가격": price,
                "가격표시": price_str or "",
                "컬러수": colors,
                "평점": rating_val,
                "리뷰수": review_count,
            })

    wb.close()
    return all_data


def load_topten_data(filepath):
    """탑텐 엑셀 파일에서 데이터 로드"""
    wb = openpyxl.load_workbook(filepath)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            rank = ws.cell(row, 1).value
            product_name = ws.cell(row, 3).value
            price_str = ws.cell(row, 4).value
            rating = ws.cell(row, 5).value
            reviews = ws.cell(row, 6).value

            if not product_name:
                continue

            price = 0
            if price_str:
                price = int(re.sub(r'[^\d]', '', str(price_str)))

            review_count = 0
            if reviews:
                review_count = int(re.sub(r'[^\d]', '', str(reviews)))

            rating_val = 0
            if rating:
                try:
                    rating_val = float(str(rating).replace('점', '').strip())
                except:
                    rating_val = 0

            all_data.append({
                "시트": sheet_name,
                "성별": sheet_name,
                "탭": sheet_name,
                "순위": rank,
                "상품명": product_name,
                "가격": price,
                "가격표시": price_str or "",
                "컬러수": 0,
                "평점": rating_val,
                "리뷰수": review_count,
            })

    wb.close()
    return all_data


def load_arket_data(filepath):
    """아르켓 엑셀 파일에서 데이터 로드"""
    wb = openpyxl.load_workbook(filepath)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            rank = ws.cell(row, 1).value
            product_name = ws.cell(row, 3).value
            price_str = ws.cell(row, 4).value

            if not product_name:
                continue

            price = 0
            if price_str:
                price = int(re.sub(r'[^\d]', '', str(price_str)))

            all_data.append({
                "시트": sheet_name,
                "성별": sheet_name,
                "탭": sheet_name,
                "순위": rank,
                "상품명": product_name,
                "가격": price,
                "가격표시": price_str or "",
                "컬러수": 0,
                "평점": 0,
                "리뷰수": 0,
            })

    wb.close()
    return all_data


def classify_all(data):
    """전체 데이터에 분류 추가"""
    for item in data:
        major, mid, sub = classify_item(item["상품명"], item.get("탭", ""))
        item["대분류"] = major
        item["중분류"] = mid
        item["소분류"] = sub
    return data


def print_analysis(data, brand_name):
    """분석 결과를 콘솔에 출력"""
    print()
    print("=" * 70)
    print(f"  {brand_name} 랭킹 아이템 분류 분석")
    print(f"  분석 일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  총 상품 수: {len(data)}개")
    print("=" * 70)

    # ── 대분류별 분포 ──
    print("\n┌─────────────────────────────────────────┐")
    print("│  [1] 대분류별 분포                        │")
    print("└─────────────────────────────────────────┘")
    major_counter = Counter(d["대분류"] for d in data)
    total = len(data)
    for cat, cnt in major_counter.most_common():
        bar = "█" * int(cnt / total * 40)
        print(f"  {cat:<8} {cnt:>4}개 ({cnt/total*100:5.1f}%) {bar}")

    # ── 중분류별 분포 ──
    print("\n┌─────────────────────────────────────────┐")
    print("│  [2] 중분류별 분포                        │")
    print("└─────────────────────────────────────────┘")
    mid_counter = Counter(d["중분류"] for d in data)
    for cat, cnt in mid_counter.most_common():
        bar = "█" * int(cnt / total * 40)
        print(f"  {cat:<12} {cnt:>4}개 ({cnt/total*100:5.1f}%) {bar}")

    # ── 소분류별 TOP 15 ──
    print("\n┌─────────────────────────────────────────┐")
    print("│  [3] 소분류(아이템) TOP 15                 │")
    print("└─────────────────────────────────────────┘")
    sub_counter = Counter(d["소분류"] for d in data)
    for i, (cat, cnt) in enumerate(sub_counter.most_common(15), 1):
        bar = "█" * int(cnt / total * 40)
        print(f"  {i:>2}. {cat:<16} {cnt:>4}개 ({cnt/total*100:5.1f}%) {bar}")

    # ── 성별별 중분류 크로스탭 ──
    genders = sorted(set(d["성별"] for d in data))
    if len(genders) > 1:
        print("\n┌─────────────────────────────────────────┐")
        print("│  [4] 성별 × 중분류 크로스탭               │")
        print("└─────────────────────────────────────────┘")
        cross = defaultdict(lambda: defaultdict(int))
        for d in data:
            cross[d["성별"]][d["중분류"]] += 1

        mid_cats = [c for c, _ in mid_counter.most_common()]
        header = f"  {'':>12}" + "".join(f"{g:>8}" for g in genders) + f"{'합계':>8}"
        print(header)
        print("  " + "-" * (len(header) - 2))
        for mc in mid_cats:
            row = f"  {mc:<12}"
            row_total = 0
            for g in genders:
                v = cross[g][mc]
                row += f"{v:>8}"
                row_total += v
            row += f"{row_total:>8}"
            print(row)

    # ── 중분류별 평균 가격 ──
    print("\n┌─────────────────────────────────────────┐")
    print("│  [5] 중분류별 평균 가격                    │")
    print("└─────────────────────────────────────────┘")
    price_by_mid = defaultdict(list)
    for d in data:
        if d["가격"] > 0:
            price_by_mid[d["중분류"]].append(d["가격"])

    price_stats = []
    for cat, prices in price_by_mid.items():
        avg = sum(prices) / len(prices)
        price_stats.append((cat, avg, min(prices), max(prices), len(prices)))

    price_stats.sort(key=lambda x: -x[1])
    for cat, avg, mn, mx, cnt in price_stats:
        print(f"  {cat:<12} 평균 {avg:>10,.0f}원  (최저 {mn:>8,}원 ~ 최고 {mx:>8,}원)  {cnt}개")

    # ── 랭킹 TOP10 아이템 분류 ──
    print("\n┌─────────────────────────────────────────┐")
    print("│  [6] 전체 랭킹 TOP 10 아이템 분류          │")
    print("└─────────────────────────────────────────┘")
    # 모두보기 시트에서 순위 기준
    overview = [d for d in data if "모두보기" in d.get("탭", "") or d.get("탭", "") in ("전체", "WOMEN", "MEN")]
    if not overview:
        overview = data

    # 성별별 TOP10
    for gender in genders[:3]:  # 최대 3개 성별
        gender_data = [d for d in overview if d["성별"] == gender]
        if not gender_data:
            continue
        gender_data.sort(key=lambda x: x["순위"] if x["순위"] else 999)
        print(f"\n  [{gender}]")
        for d in gender_data[:10]:
            print(f"    {d['순위']:>2}위  {d['상품명'][:28]:<30} → {d['중분류']}/{d['소분류']}")

    # ── 미분류 아이템 ──
    unclassified = [d for d in data if d["중분류"] == "미분류"]
    if unclassified:
        print(f"\n  ⚠ 미분류 아이템: {len(unclassified)}개")
        for d in unclassified[:10]:
            print(f"    - {d['상품명']}")


def save_analysis_excel(data, brand_name, output_path):
    """분석 결과를 엑셀 파일로 저장"""
    wb = openpyxl.Workbook()

    # 스타일 정의
    header_fill = PatternFill("solid", fgColor="2B579A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sub_header_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── 시트1: 전체 상품 분류 목록 ──
    ws1 = wb.active
    ws1.title = "전체분류목록"
    headers = ["순위", "성별", "탭", "상품명", "가격", "대분류", "중분류", "소분류", "평점", "리뷰수"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, d in enumerate(data, 2):
        values = [d["순위"], d["성별"], d["탭"], d["상품명"], d["가격표시"],
                  d["대분류"], d["중분류"], d["소분류"],
                  d["평점"] if d["평점"] > 0 else "",
                  d["리뷰수"] if d["리뷰수"] > 0 else ""]
        for c, v in enumerate(values, 1):
            cell = ws1.cell(r, c, v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 4 else "left")

    # 열 너비 조정
    col_widths = [6, 8, 12, 40, 12, 8, 14, 16, 6, 8]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 자동 필터
    ws1.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # ── 시트2: 중분류별 분포 요약 ──
    ws2 = wb.create_sheet("중분류별분포")
    mid_counter = Counter(d["중분류"] for d in data)
    total = len(data)

    ws2.cell(1, 1, "중분류").fill = header_fill
    ws2.cell(1, 1).font = header_font
    ws2.cell(1, 2, "상품수").fill = header_fill
    ws2.cell(1, 2).font = header_font
    ws2.cell(1, 3, "비율(%)").fill = header_fill
    ws2.cell(1, 3).font = header_font

    for r, (cat, cnt) in enumerate(mid_counter.most_common(), 2):
        ws2.cell(r, 1, cat).border = border
        ws2.cell(r, 2, cnt).border = border
        ws2.cell(r, 3, round(cnt / total * 100, 1)).border = border

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 10

    # 파이 차트 추가
    if len(mid_counter) > 1:
        pie = PieChart()
        pie.title = f"{brand_name} 중분류별 비율"
        pie.style = 10
        labels = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(mid_counter))
        values = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(mid_counter))
        pie.add_data(values, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 18
        pie.height = 12
        ws2.add_chart(pie, "E2")

    # ── 시트3: 소분류별 분포 ──
    ws3 = wb.create_sheet("소분류별분포")
    sub_counter = Counter(d["소분류"] for d in data)

    for c, h in enumerate(["소분류(아이템)", "상품수", "비율(%)"], 1):
        cell = ws3.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font

    for r, (cat, cnt) in enumerate(sub_counter.most_common(), 2):
        ws3.cell(r, 1, cat).border = border
        ws3.cell(r, 2, cnt).border = border
        ws3.cell(r, 3, round(cnt / total * 100, 1)).border = border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 10

    # 바 차트
    if len(sub_counter) > 1:
        bar = BarChart()
        bar.title = f"{brand_name} 소분류(아이템) TOP"
        bar.style = 10
        bar.type = "col"
        max_items = min(15, len(sub_counter))
        bar_data = Reference(ws3, min_col=2, min_row=1, max_row=1 + max_items)
        bar_cats = Reference(ws3, min_col=1, min_row=2, max_row=1 + max_items)
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        bar.width = 22
        bar.height = 12
        ws3.add_chart(bar, "E2")

    # ── 시트4: 성별×중분류 크로스탭 ──
    ws4 = wb.create_sheet("성별×중분류")
    genders = sorted(set(d["성별"] for d in data))
    mid_cats = [c for c, _ in mid_counter.most_common()]

    # 헤더
    ws4.cell(1, 1, "중분류").fill = header_fill
    ws4.cell(1, 1).font = header_font
    for gc, g in enumerate(genders, 2):
        ws4.cell(1, gc, g).fill = header_fill
        ws4.cell(1, gc).font = header_font
    ws4.cell(1, len(genders) + 2, "합계").fill = header_fill
    ws4.cell(1, len(genders) + 2).font = header_font

    cross = defaultdict(lambda: defaultdict(int))
    for d in data:
        cross[d["성별"]][d["중분류"]] += 1

    for r, mc in enumerate(mid_cats, 2):
        ws4.cell(r, 1, mc).border = border
        row_total = 0
        for gc, g in enumerate(genders, 2):
            v = cross[g][mc]
            ws4.cell(r, gc, v).border = border
            row_total += v
        ws4.cell(r, len(genders) + 2, row_total).border = border

    ws4.column_dimensions['A'].width = 16

    # ── 시트5: 중분류별 가격 분석 ──
    ws5 = wb.create_sheet("가격분석")
    price_headers = ["중분류", "상품수", "평균가격", "최저가", "최고가", "가격대"]
    for c, h in enumerate(price_headers, 1):
        cell = ws5.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font

    price_by_mid = defaultdict(list)
    for d in data:
        if d["가격"] > 0:
            price_by_mid[d["중분류"]].append(d["가격"])

    price_stats = []
    for cat, prices in price_by_mid.items():
        avg = sum(prices) / len(prices)
        price_stats.append((cat, len(prices), avg, min(prices), max(prices)))
    price_stats.sort(key=lambda x: -x[2])

    for r, (cat, cnt, avg, mn, mx) in enumerate(price_stats, 2):
        ws5.cell(r, 1, cat).border = border
        ws5.cell(r, 2, cnt).border = border
        ws5.cell(r, 3, f"{avg:,.0f}원").border = border
        ws5.cell(r, 4, f"{mn:,}원").border = border
        ws5.cell(r, 5, f"{mx:,}원").border = border
        # 가격대 분류
        if avg >= 80000:
            tier = "프리미엄"
        elif avg >= 50000:
            tier = "중고가"
        elif avg >= 30000:
            tier = "중저가"
        else:
            tier = "저가"
        ws5.cell(r, 6, tier).border = border

    for i, w in enumerate([16, 8, 12, 12, 12, 10], 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 저장
    wb.save(output_path)
    print(f"\n  📊 분석 엑셀 저장: {output_path}")


def find_latest_file(pattern):
    """패턴에 맞는 가장 최근 파일 찾기"""
    import glob
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(base_dir, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def main():
    print()
    print("=" * 70)
    print("  패션 랭킹 아이템 분류 분석기")
    print("=" * 70)
    print("  분류 체계: 대분류(의류/언더웨어/잡화)")
    print("           → 중분류(아우터/상의/니트/하의/원피스&스커트/이너웨어/홈웨어/악세서리/베이비)")
    print("           → 소분류(재킷/티셔츠/데님/팬츠/스커트 등 구체 아이템)")
    print("=" * 70)

    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 유니클로 분석
    uniqlo_file = find_latest_file("유니클로_*V5*.xlsx")
    if uniqlo_file:
        print(f"\n  ▶ 유니클로 파일: {os.path.basename(uniqlo_file)}")
        data = load_uniqlo_data(uniqlo_file)

        # "모두보기" 시트 데이터만 사용 (중복 제거. 탭별 시트는 모두보기의 하위집합)
        overview_data = [d for d in data if "모두보기" in d["탭"]]
        if overview_data:
            data = overview_data

        data = classify_all(data)
        print_analysis(data, "유니클로")

        output = os.path.join(base_dir, f"유니클로_아이템분류분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        save_analysis_excel(data, "유니클로", output)
    else:
        print("\n  ⚠ 유니클로 엑셀 파일을 찾을 수 없습니다.")

    print("\n" + "=" * 70)
    print("  분석 완료!")
    print("=" * 70)


if __name__ == "__main__":
    main()
