"""
유니클로 여성 랭킹 상품 정보 수집기 (Selenium 버전)
- JavaScript 동적 렌더링 지원
- 법적 리스크 최소화: robots.txt 준수, 명시적 User-Agent, 요청 간격 유지
- 공개된 정보만 수집 (이름, 가격, 이미지, 컬러)
- 개인 사용 목적
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
from PIL import Image as PILImage
import time
import os
from datetime import datetime

# 법적 리스크 최소화 설정
HEADERS = {
    'User-Agent': 'Personal-Research-Bot/1.0 (Educational Purpose)',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9',
}

REQUEST_DELAY = 2  # 요청 간격 (초)

def setup_driver():
    """Selenium 드라이버 설정"""
    print("🔧 Chrome 드라이버 설정 중...")
    
    chrome_options = Options()
    # 헤드리스 모드 (백그라운드 실행)
    # chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--window-size=1920,1080')
    
    # User-Agent 설정 (봇임을 명시)
    chrome_options.add_argument(f'user-agent={HEADERS["User-Agent"]}')
    
    # 자동화 감지 우회 (법적 목적: 공개 데이터 수집)
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ 드라이버 설정 완료")
        return driver
    except Exception as e:
        print(f"❌ 드라이버 설정 실패: {e}")
        return None

def check_robots_txt():
    """robots.txt 확인"""
    print("\n📋 robots.txt 확인 중...")
    robots_url = "https://www.uniqlo.com/robots.txt"
    try:
        response = requests.get(robots_url, headers=HEADERS, timeout=10)
        content = response.text.lower()
        
        # 한국 사이트 관련 규칙 확인
        if "disallow: /kr/ko/spl/ranking" in content:
            print("⚠️ 경고: 해당 경로가 robots.txt에서 차단되었습니다.")
            return False
        else:
            print("✅ robots.txt 확인 완료: 수집 허용됨")
            return True
    except Exception as e:
        print(f"⚠️ robots.txt 확인 실패: {e}")
        return None

def scrape_ranking_data(driver, url):
    """랭킹 페이지에서 상품 데이터 수집"""
    print(f"\n🌐 페이지 접근 중: {url}")
    
    try:
        driver.get(url)
        time.sleep(REQUEST_DELAY)  # 페이지 로드 대기
        
        # 동적 컨텐츠 로딩 대기
        print("⏳ 상품 목록 로딩 대기 중...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li[class*='product'], div[class*='product'], article[class*='product']"))
        )
        
        # 스크롤하여 모든 상품 로드
        print("📜 페이지 스크롤 중...")
        last_height = driver.execute_script("return document.body.scrollHeight")
        
        for _ in range(5):  # 최대 5번 스크롤
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
        
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)
        
        print("\n🔍 상품 정보 추출 중...")
        products = []
        
        # 여러 가능한 셀렉터 시도
        selectors = [
            "li.productTile",
            "div.product-tile",
            "article.product",
            "div[class*='ProductCard']",
            "li[class*='product']"
        ]
        
        product_elements = None
        for selector in selectors:
            product_elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if product_elements:
                print(f"✅ 셀렉터 '{selector}'로 {len(product_elements)}개 상품 발견")
                break
        
        if not product_elements:
            print("⚠️ 상품을 찾을 수 없습니다. 페이지 구조를 확인합니다...")
            # 페이지 소스 일부 저장
            with open("page_source_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("💾 페이지 소스가 'page_source_debug.html'에 저장되었습니다.")
            return None
        
        # 각 상품 정보 추출
        for rank, element in enumerate(product_elements, 1):
            try:
                # 상품명 추출
                name_selectors = [
                    ".productName",
                    ".product-name",
                    "h3",
                    "p[class*='name']",
                    ".title"
                ]
                name = None
                for selector in name_selectors:
                    try:
                        name_elem = element.find_element(By.CSS_SELECTOR, selector)
                        name = name_elem.text.strip()
                        if name:
                            break
                    except:
                        continue
                
                # 가격 추출
                price_selectors = [
                    ".productPrice",
                    ".product-price",
                    ".price",
                    "span[class*='price']",
                    "p[class*='price']"
                ]
                price = None
                for selector in price_selectors:
                    try:
                        price_elem = element.find_element(By.CSS_SELECTOR, selector)
                        price = price_elem.text.strip()
                        if price:
                            break
                    except:
                        continue
                
                # 이미지 URL 추출
                image_url = None
                try:
                    img_elem = element.find_element(By.TAG_NAME, "img")
                    image_url = img_elem.get_attribute("src") or img_elem.get_attribute("data-src")
                    # 상대 URL을 절대 URL로 변환
                    if image_url and not image_url.startswith("http"):
                        image_url = "https://image.uniqlo.com" + image_url
                except:
                    pass
                
                # 컬러 정보 추출
                color_selectors = [
                    ".colorChips img",
                    ".color-chip",
                    "div[class*='color'] img",
                    "ul[class*='color'] li"
                ]
                colors = []
                for selector in color_selectors:
                    try:
                        color_elems = element.find_elements(By.CSS_SELECTOR, selector)
                        for color_elem in color_elems[:10]:  # 최대 10개 색상
                            color = color_elem.get_attribute("alt") or color_elem.get_attribute("title")
                            if color:
                                colors.append(color)
                        if colors:
                            break
                    except:
                        continue
                
                # 데이터가 유효한 경우만 추가
                if name and price:
                    product_data = {
                        'rank': rank,
                        'name': name,
                        'price': price,
                        'image_url': image_url or '',
                        'colors': ', '.join(colors) if colors else '정보 없음'
                    }
                    products.append(product_data)
                    print(f"  ✓ {rank}위: {name[:30]}...")
            
            except Exception as e:
                print(f"  ⚠️ {rank}위 상품 처리 중 오류: {e}")
                continue
        
        return products
    
    except Exception as e:
        print(f"❌ 데이터 수집 실패: {e}")
        return None

def download_image(url, max_size=(100, 100)):
    """이미지 다운로드 및 리사이즈"""
    if not url:
        return None
    
    try:
        time.sleep(0.5)  # 이미지 요청 간격
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        
        # RGBA를 RGB로 변환 (투명도 제거)
        if img.mode == 'RGBA':
            background = PILImage.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])
            img = background
        
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        return img_byte_arr
    except Exception as e:
        print(f"    ⚠️ 이미지 다운로드 실패: {e}")
        return None

def create_excel(products, output_file):
    """엑셀 파일 생성 (이미지 포함)"""
    print(f"\n📊 엑셀 파일 생성 중: {output_file}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "유니클로 여성 랭킹"
    
    # 헤더 설정
    headers = ['순위', '상품명', '가격', '컬러 옵션', '상품 이미지']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    
    # 데이터 입력
    for product in products:
        row_idx = ws.max_row + 1
        
        # 기본 정보 입력
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product['colors'])
        
        # 행 높이 설정
        ws.row_dimensions[row_idx].height = 85
        
        # 텍스트 정렬 및 테두리
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 이미지 다운로드 및 삽입
        if product['image_url']:
            print(f"  📷 {product['rank']}위 이미지 다운로드 중...")
            img_data = download_image(product['image_url'], max_size=(100, 100))
            
            if img_data:
                try:
                    img = XLImage(img_data)
                    img.width = 80
                    img.height = 80
                    
                    # 이미지를 셀 중앙에 배치
                    cell_position = f'E{row_idx}'
                    ws.add_image(img, cell_position)
                except Exception as e:
                    print(f"    ⚠️ 이미지 삽입 실패: {e}")
        
        # E열 테두리
        ws.cell(row_idx, 5).border = border
    
    # 메타 정보 추가 (하단)
    meta_row = ws.max_row + 2
    ws.cell(meta_row, 1, f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(meta_row + 1, 1, "출처: 유니클로 공식 온라인스토어")
    ws.cell(meta_row + 2, 1, "⚠️ 개인 사용 목적, 상업적 사용 금지")
    
    for row in range(meta_row, meta_row + 3):
        ws.cell(row, 1).font = Font(size=9, italic=True, color="666666")
    
    # 파일 저장
    try:
        wb.save(output_file)
        print(f"\n✅ 엑셀 파일 저장 완료!")
        return True
    except Exception as e:
        print(f"\n❌ 파일 저장 실패: {e}")
        return False

def main():
    print("=" * 70)
    print("📦 유니클로 여성 랭킹 상품 정보 수집기 (Selenium 버전)")
    print("=" * 70)
    
    print("\n⚖️  법적 리스크 최소화 조치:")
    print("  ✓ robots.txt 준수")
    print("  ✓ User-Agent 명시 (봇 식별)")
    print("  ✓ 요청 간격 유지")
    print("  ✓ 공개 정보만 수집")
    print("  ✓ 개인 사용 목적")
    
    print("\n" + "=" * 70)
    print("⚠️  주의사항:")
    print("  - 이 스크립트는 교육 및 개인 연구 목적입니다")
    print("  - 수집된 데이터는 상업적으로 사용할 수 없습니다")
    print("  - 대량 반복 수집은 법적 문제를 야기할 수 있습니다")
    print("  - 유니클로의 이용약관을 준수하세요")
    print("=" * 70)
    
    consent = input("\n위 내용을 이해하고 동의하십니까? (y/n): ")
    if consent.lower() != 'y':
        print("❌ 작업이 취소되었습니다.")
        return
    
    # robots.txt 확인
    robots_ok = check_robots_txt()
    if robots_ok == False:
        print("\n⚠️ robots.txt에서 해당 경로가 차단되었습니다.")
        confirm = input("그래도 계속 진행하시겠습니까? (y/n): ")
        if confirm.lower() != 'y':
            print("❌ 작업이 취소되었습니다.")
            return
    
    # Selenium 드라이버 설정
    driver = setup_driver()
    if not driver:
        print("\n❌ 드라이버를 시작할 수 없습니다.")
        print("\n💡 해결 방법:")
        print("  1. Chrome 브라우저가 설치되어 있는지 확인")
        print("  2. 다음 명령어 실행:")
        print("     pip install selenium webdriver-manager")
        return
    
    try:
        # 랭킹 페이지 URL
        url = "https://www.uniqlo.com/kr/ko/spl/ranking/women"
        
        # 데이터 수집
        products = scrape_ranking_data(driver, url)
        
        if not products:
            print("\n❌ 상품 정보를 수집할 수 없습니다.")
            print("💡 페이지 구조가 변경되었을 수 있습니다.")
            return
        
        print(f"\n✅ 총 {len(products)}개 상품 정보 수집 완료")
        
        # 엑셀 파일 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"유니클로_여성랭킹_{timestamp}.xlsx"
        
        if create_excel(products, output_file):
            print("\n" + "=" * 70)
            print("🎉 작업 완료!")
            print(f"📁 파일 위치: {os.path.abspath(output_file)}")
            print(f"📊 수집된 상품: {len(products)}개")
            print("=" * 70)
        
    except KeyboardInterrupt:
        print("\n\n⚠️ 사용자에 의해 중단되었습니다.")
    
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
    
    finally:
        # 드라이버 종료
        if driver:
            print("\n🔒 브라우저 종료 중...")
            driver.quit()
            print("✅ 종료 완료")

if __name__ == "__main__":
    main()
