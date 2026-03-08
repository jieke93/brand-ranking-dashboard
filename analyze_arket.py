#!/usr/bin/env python3
"""
아르켓(ARKET) 랭킹 분석 도구
- 크롤링된 엑셀 데이터를 읽어 아이템타입별 비중 분석
- 가격대별 분석, 핵심 아이템 도출
- 날짜별 랭킹 변동 추적 (히스토리 누적)
- 분석 결과를 엑셀로 출력
"""

import os
import glob
import json
import re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ─── 설정 ───
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_FILE = os.path.join(WORK_DIR, 'arket_history.json')
HISTORY_BACKUP = os.path.join(WORK_DIR, 'arket_history_backup.json')
BRAND_NAME = '아르켓'
FILE_PATTERN = '아르켓_인기상품_판매순_*.xlsx'
CATEGORIES = ['WOMEN', 'MEN']

# 스타일
HEADER_FILL = PatternFill('solid', fgColor='1A1A1A')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUB_HEADER_FILL = PatternFill('solid', fgColor='E8E0D6')
SUB_HEADER_FONT = Font(bold=True, size=10)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
UP_FONT = Font(color='FF0000', bold=True)
DOWN_FONT = Font(color='0070C0', bold=True)
NEW_FONT = Font(color='00B050', bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ─── 아이템타입 분류 규칙 (아르켓 특화) ───
ITEM_TYPE_RULES = [
    # 아우터
    (r'(?i)(coat|코트)', '코트'),
    (r'(?i)(puffer|padded|패딩|패디드|퍼퍼)', '패딩/퍼퍼'),
    (r'(?i)(jacket|자켓|재킷|blazer|블레이저)', '자켓/블레이저'),
    (r'(?i)(cardigan|카디건)', '카디건'),
    (r'(?i)(vest|조끼|gilet)', '베스트/조끼'),
    # 상의
    (r'(?i)(hoodie|후디|hooded)', '후디'),
    (r'(?i)(sweatshirt|스웨트셔츠|스웨트)', '스웨트셔츠'),
    (r'(?i)(knit|니트|sweater|스웨터|jumper|점퍼)', '니트/스웨터'),
    (r'(?i)(shirt|셔츠|blouse|블라우스)', '셔츠/블라우스'),
    (r'(?i)(polo)', '폴로'),
    (r'(?i)(t-shirt|tee|티셔츠|반팔)', '티셔츠'),
    (r'(?i)(top|탑|tank|탱크)', '탑'),
    # 하의
    (r'(?i)(jeans|jean|진|denim\s*(?:p|t))', '진'),
    (r'(?i)(trouser|팬츠|pants|슬랙스|chino)', '팬츠/트라우저'),
    (r'(?i)(shorts|숏팬츠|반바지)', '쇼츠'),
    (r'(?i)(skirt|스커트)', '스커트'),
    (r'(?i)(legging|레깅스)', '레깅스'),
    # 원피스/점프수트
    (r'(?i)(dress|드레스|원피스)', '드레스'),
    (r'(?i)(jumpsuit|점프수트|romper)', '점프수트'),
    # 가방/액세서리
    (r'(?i)(bag|백|tote|토트|backpack|배낭|crossbody|숄더백)', '가방'),
    (r'(?i)(scarf|스카프|머플러|muffler)', '스카프/머플러'),
    (r'(?i)(hat|cap|모자|beanie|비니)', '모자'),
    (r'(?i)(belt|벨트)', '벨트'),
    (r'(?i)(glove|장갑)', '장갑'),
    (r'(?i)(wallet|지갑|card\s*holder|카드홀더)', '지갑'),
    (r'(?i)(jewel|ring|반지|necklace|목걸이|earring|귀걸이|bracelet|팔찌)', '주얼리'),
    (r'(?i)(sunglasses|선글라스)', '선글라스'),
    (r'(?i)(accessori|액세서리)', '액세서리'),
    # 신발
    (r'(?i)(sneaker|스니커즈)', '스니커즈'),
    (r'(?i)(boot|부츠)', '부츠'),
    (r'(?i)(sandal|샌들)', '샌들'),
    (r'(?i)(loafer|로퍼)', '로퍼'),
    (r'(?i)(shoe|슈즈|flat|플랫)', '슈즈'),
    (r'(?i)(slipper|슬리퍼)', '슬리퍼'),
    # 언더웨어/라운지
    (r'(?i)(underwear|속옷|brief|브리프|boxer|팬티)', '언더웨어'),
    (r'(?i)(sock|양말)', '양말'),
    (r'(?i)(pyjama|pajama|파자마|lounge|라운지)', '라운지웨어'),
    # 수영복
    (r'(?i)(swim|수영|bikini|비키니)', '스윔웨어'),
]


def log(msg):
    print(msg)


def classify_item_type(name):
    """상품명으로 아이템타입 자동 분류"""
    if not name:
        return '미분류'
    for pattern, item_type in ITEM_TYPE_RULES:
        if re.search(pattern, name):
            return item_type
    return '미분류'


def find_latest_ranking_file():
    """가장 최근 크롤링 엑셀 파일 찾기"""
    pattern = os.path.join(WORK_DIR, FILE_PATTERN)
    files = glob.glob(pattern)
    if not files:
        log(f"[ERROR] 크롤링 결과 파일을 찾을 수 없습니다. ({FILE_PATTERN})")
        return None
    files.sort(reverse=True)
    return files[0]


def load_ranking_data(filepath):
    """엑셀 파일에서 랭킹 데이터 로드 (아르켓 컬럼 구조)"""
    log(f"  파일 로드: {os.path.basename(filepath)}")
    wb = load_workbook(filepath, data_only=True)

    all_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            continue
        ws = wb[sheet_name]

        products = []
        for row_idx in range(2, ws.max_row + 1):
            rank = ws.cell(row_idx, 1).value
            if rank is None:
                continue

            # 아르켓 컬럼: 순위(1), 이미지(2), 상품명(3), 색상(4), 가격(5)
            name = ws.cell(row_idx, 3).value or ''
            color = ws.cell(row_idx, 4).value or ''
            price = ws.cell(row_idx, 5).value or ''
            item_type = classify_item_type(name)

            products.append({
                'rank': rank,
                'name': name,
                'color': color,
                'item_type': item_type,
                'price': price,
                'sheet': sheet_name
            })

        all_data[sheet_name] = products

    wb.close()
    return all_data


def parse_price(price_str):
    """가격 문자열을 숫자로 변환"""
    if not price_str:
        return 0
    try:
        return int(re.sub(r'[^\d]', '', str(price_str)))
    except:
        return 0


def extract_date_from_filename(filepath):
    """파일명에서 날짜 추출"""
    basename = os.path.basename(filepath)
    match = re.search(r'(\d{8})_(\d{6})', basename)
    if match:
        return match.group(1)
    return datetime.now().strftime('%Y%m%d')


# ─── 분석 함수들 ───

def analyze_item_type_distribution(all_data):
    """아이템타입별 비중 분석 (카테고리별)"""
    log("\n  [분석1] 아이템타입별 비중 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        type_counter = Counter()
        type_products = defaultdict(list)

        for p in products:
            item_type = p['item_type']
            type_counter[item_type] += 1
            type_products[item_type].append(p)

        total = len(products)
        type_analysis = []
        for item_type, count in type_counter.most_common():
            pct = round(count / total * 100, 1) if total > 0 else 0
            prices = [parse_price(p['price']) for p in type_products[item_type]]
            avg_price = round(sum(prices) / len(prices)) if prices else 0

            top_products = sorted(type_products[item_type], key=lambda x: x['rank'])[:3]
            top_names = ', '.join([f"{p['rank']}위:{p['name'][:20]}" for p in top_products])

            type_analysis.append({
                'item_type': item_type,
                'count': count,
                'pct': pct,
                'avg_price': avg_price,
                'top_products': top_names,
                'best_rank': min([p['rank'] for p in type_products[item_type]]),
            })

        results[sheet_name] = {
            'total': total,
            'analysis': sorted(type_analysis, key=lambda x: x['count'], reverse=True)
        }

    return results


def analyze_top_items(all_data):
    """핵심 아이템 분석"""
    log("  [분석2] 핵심 아이템 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        unique_items = []
        for p in products:
            unique_items.append({
                'name': p['name'],
                'item_type': p['item_type'],
                'price': p['price'],
                'color': p['color'],
                'rank': p['rank'],
                'impact_score': max(101 - p['rank'], 1),
            })
        unique_items.sort(key=lambda x: x['impact_score'], reverse=True)
        results[sheet_name] = unique_items[:30]

    return results


def analyze_price_band(all_data):
    """가격대별 분석"""
    log("  [분석3] 가격대별 분석...")

    price_bands = [
        (0, 30000, '3만원 이하'),
        (30000, 50000, '3~5만원'),
        (50000, 100000, '5~10만원'),
        (100000, 150000, '10~15만원'),
        (150000, 200000, '15~20만원'),
        (200000, 300000, '20~30만원'),
        (300000, 500000, '30~50만원'),
        (500000, float('inf'), '50만원 이상'),
    ]

    results = {}
    for sheet_name, products in all_data.items():
        band_counter = Counter()
        band_items = defaultdict(list)

        for p in products:
            price = parse_price(p['price'])
            for low, high, label in price_bands:
                if low <= price < high:
                    band_counter[label] += 1
                    band_items[label].append(p['name'][:20])
                    break

        total = len(products)
        results[sheet_name] = {
            'total': total,
            'bands': [(label, band_counter.get(label, 0),
                       round(band_counter.get(label, 0) / total * 100, 1) if total > 0 else 0,
                       band_items.get(label, [])[:3])
                      for _, _, label in price_bands if band_counter.get(label, 0) > 0]
        }

    return results


def analyze_color_distribution(all_data):
    """색상 분포 분석 (아르켓 특화)"""
    log("  [분석4] 색상 분포 분석...")

    results = {}
    for sheet_name, products in all_data.items():
        color_counter = Counter()
        for p in products:
            color = p.get('color', '').strip()
            if color:
                color_counter[color] += 1

        total = len(products)
        color_analysis = []
        for color, count in color_counter.most_common(15):
            pct = round(count / total * 100, 1) if total > 0 else 0
            color_analysis.append({
                'color': color,
                'count': count,
                'pct': pct,
            })

        results[sheet_name] = {
            'total': total,
            'colors': color_analysis
        }

    return results


# ─── 히스토리 관리 ───

def load_history():
    """랭킹 히스토리 로드"""
    for fpath in [HISTORY_FILE, HISTORY_BACKUP]:
        if os.path.exists(fpath):
            try:
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if data:
                    return data
            except:
                continue
    return {}


def save_history(history):
    """랭킹 히스토리 저장 (메인 + 백업)"""
    if os.path.exists(HISTORY_FILE):
        try:
            import shutil
            shutil.copy2(HISTORY_FILE, HISTORY_BACKUP)
        except:
            pass

    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def recover_history_from_excel_files():
    """이전 크롤링 엑셀 파일들에서 히스토리 역추출"""
    log("  [히스토리 복구] 이전 크롤링 파일에서 데이터 복구 시도...")

    pattern = os.path.join(WORK_DIR, FILE_PATTERN)
    files = sorted(glob.glob(pattern))

    if not files:
        log("  [히스토리 복구] 크롤링 파일 없음 - 신규 시작")
        return {}

    recovered = {}
    for filepath in files:
        date_key = extract_date_from_filename(filepath)
        try:
            wb = load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Sheet':
                    continue

                category = sheet_name
                if category not in recovered:
                    recovered[category] = {}

                ws = wb[sheet_name]
                day_data = {}
                for row_idx in range(2, ws.max_row + 1):
                    rank = ws.cell(row_idx, 1).value
                    if rank is None:
                        continue
                    name = str(ws.cell(row_idx, 3).value or '')[:40]
                    color = str(ws.cell(row_idx, 4).value or '')
                    price = str(ws.cell(row_idx, 5).value or '')
                    item_type = classify_item_type(name)

                    if name:
                        day_data[name] = {
                            'rank': rank,
                            'item_type': item_type,
                            'color': color,
                            'price': price,
                        }

                if day_data:
                    recovered[category][date_key] = day_data

            wb.close()
            log(f"    -> {os.path.basename(filepath)} ({date_key}) 복구 완료")
        except Exception as e:
            log(f"    -> {os.path.basename(filepath)} 복구 실패: {e}")

    return recovered


def merge_history(existing, new_data):
    """두 히스토리 데이터를 병합"""
    merged = {}
    all_categories = set(list(existing.keys()) + list(new_data.keys()))

    for category in all_categories:
        merged[category] = {}
        if category in existing:
            merged[category].update(existing[category])
        if category in new_data:
            for date_key, day_data in new_data[category].items():
                if date_key not in merged[category]:
                    merged[category][date_key] = day_data

    return merged


def update_history(all_data, date_key):
    """현재 데이터를 히스토리에 추가"""
    log("  [히스토리] 랭킹 변동 데이터 업데이트...")

    history = load_history()

    existing_dates = set()
    for cat in history.values():
        existing_dates.update(cat.keys())

    if not existing_dates:
        recovered = recover_history_from_excel_files()
        if recovered:
            history = merge_history(history, recovered)
            rec_dates = set()
            for cat in recovered.values():
                rec_dates.update(cat.keys())
            log(f"  [히스토리 복구] {len(rec_dates)}일치 데이터 복구됨")

    for sheet_name, products in all_data.items():
        category = sheet_name
        if category not in history:
            history[category] = {}

        day_data = {}
        for p in products:
            day_data[p['name'][:40]] = {
                'rank': p['rank'],
                'item_type': p['item_type'],
                'color': p.get('color', ''),
                'price': str(p['price']),
            }

        history[category][date_key] = day_data

    save_history(history)

    all_dates = set()
    for cat in history.values():
        all_dates.update(cat.keys())
    sorted_dates = sorted(all_dates)

    log(f"  [히스토리] 총 {len(sorted_dates)}일치 데이터 보유")
    if len(sorted_dates) > 1:
        log(f"  [히스토리] 기간: {sorted_dates[0]} ~ {sorted_dates[-1]}")

    return history


def analyze_ranking_changes(history):
    """랭킹 변동 분석 (최근 2회 비교)"""
    log("  [분석5] 랭킹 변동 분석...")

    results = {}

    for category, date_dict in history.items():
        dates = sorted(date_dict.keys())
        if len(dates) < 2:
            results[category] = {
                'status': 'first_data',
                'message': f'첫 번째 수집 데이터 (비교 불가, {dates[0] if dates else "N/A"})',
                'changes': []
            }
            # 첫 데이터도 목록에 포함
            if dates:
                current = date_dict[dates[-1]]
                for name, info in current.items():
                    results[category]['changes'].append({
                        'name': name[:30],
                        'current_rank': info['rank'],
                        'rank_change': 0,
                        'status': '초회',
                        'item_type': info['item_type'],
                        'price': info['price'],
                    })
                results[category]['changes'].sort(key=lambda x: x['current_rank'])
            continue

        current_date = dates[-1]
        prev_date = dates[-2]
        current = date_dict[current_date]
        prev = date_dict[prev_date]

        changes = []
        for name, info in current.items():
            current_rank = info['rank']
            if name in prev:
                prev_rank = prev[name]['rank']
                rank_change = prev_rank - current_rank
                status = '상승' if rank_change > 0 else ('하락' if rank_change < 0 else '유지')
            else:
                rank_change = 0
                status = '신규진입'

            changes.append({
                'name': name[:30],
                'current_rank': current_rank,
                'rank_change': rank_change,
                'status': status,
                'item_type': info['item_type'],
                'price': info['price'],
            })

        dropped = []
        for name in prev:
            if name not in current:
                dropped.append({
                    'name': name[:30],
                    'prev_rank': prev[name]['rank'],
                    'item_type': prev[name]['item_type'],
                })

        changes.sort(key=lambda x: x['current_rank'])

        results[category] = {
            'status': 'compared',
            'current_date': current_date,
            'prev_date': prev_date,
            'changes': changes,
            'dropped': dropped,
            'total_dates': len(dates),
        }

    return results


# ─── 엑셀 출력 ───

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def create_analysis_excel(type_dist, top_items, price_band, color_dist, ranking_changes, date_key):
    """분석 결과를 엑셀로 출력"""
    log("\n============================================================")
    log("[출력] 분석 결과 엑셀 생성")
    log("============================================================")

    wb = Workbook()
    wb.remove(wb.active)

    # ═══ 시트1: 종합 대시보드 ═══
    ws = wb.create_sheet('종합_대시보드')
    log("  -> 시트 [종합_대시보드] 생성")

    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row, 1, f'ARKET 랭킹 분석 보고서 ({date_key[:4]}.{date_key[4:6]}.{date_key[6:8]})')
    ws.cell(row, 1).font = Font(bold=True, size=14, color='1A1A1A')
    ws.cell(row, 1).alignment = Alignment(horizontal='center')

    for category in CATEGORIES:
        if category not in type_dist:
            continue

        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row, 1, f'■ {category} ({type_dist[category]["total"]}개 상품)')
        ws.cell(row, 1).font = Font(bold=True, size=12, color='C00000')

        row += 1
        headers = ['순번', '아이템타입', '상품수', '비중(%)', '평균가격', '최고순위', '대표상품(순위:상품명)']
        for col, h in enumerate(headers, 1):
            ws.cell(row, col, h)
        apply_header_style(ws, row, len(headers))

        data = type_dist[category]['analysis']
        for idx, item in enumerate(data, 1):
            row += 1
            ws.cell(row, 1, idx)
            ws.cell(row, 2, item['item_type'])
            ws.cell(row, 3, item['count'])
            ws.cell(row, 4, item['pct'])
            ws.cell(row, 5, f"{item['avg_price']:,}원" if item['avg_price'] > 0 else '-')
            ws.cell(row, 6, item['best_rank'])
            ws.cell(row, 7, item['top_products'])
            apply_data_style(ws, row, len(headers))

            if item['pct'] >= 10:
                for col in range(1, len(headers) + 1):
                    ws.cell(row, col).fill = HIGHLIGHT_FILL

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 55

    # ═══ 시트2: 핵심 아이템 ═══
    ws2 = wb.create_sheet('핵심_아이템')
    log("  -> 시트 [핵심_아이템] 생성")

    row = 1
    for category in CATEGORIES:
        if category not in top_items:
            continue

        if row > 1:
            row += 1
        ws2.merge_cells(f'A{row}:F{row}')
        ws2.cell(row, 1, f'■ {category} 핵심 아이템 TOP 20')
        ws2.cell(row, 1).font = Font(bold=True, size=12, color='C00000')

        row += 1
        headers = ['순위', '상품명', '아이템타입', '색상', '가격', '영향력점수']
        for col, h in enumerate(headers, 1):
            ws2.cell(row, col, h)
        apply_header_style(ws2, row, len(headers))

        items = top_items[category][:20]
        for item in items:
            row += 1
            ws2.cell(row, 1, item['rank'])
            ws2.cell(row, 2, item['name'][:40])
            ws2.cell(row, 3, item['item_type'])
            ws2.cell(row, 4, item['color'])
            ws2.cell(row, 5, item['price'])
            ws2.cell(row, 6, item['impact_score'])
            apply_data_style(ws2, row, len(headers))

            if item['rank'] <= 5:
                for col in range(1, len(headers) + 1):
                    ws2.cell(row, col).fill = HIGHLIGHT_FILL

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 10

    # ═══ 시트3: 색상 분포 ═══
    ws3 = wb.create_sheet('색상_분석')
    log("  -> 시트 [색상_분석] 생성")

    row = 1
    ws3.merge_cells(f'A{row}:D{row}')
    ws3.cell(row, 1, '색상 분포 분석')
    ws3.cell(row, 1).font = Font(bold=True, size=12, color='1A1A1A')

    for category in CATEGORIES:
        if category not in color_dist:
            continue

        row += 2
        ws3.merge_cells(f'A{row}:D{row}')
        ws3.cell(row, 1, f'■ {category} TOP 색상')
        ws3.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        for col, h in enumerate(['색상', '상품수', '비중(%)', '구성바'], 1):
            ws3.cell(row, col, h)
        apply_header_style(ws3, row, 4)

        for c in color_dist[category]['colors']:
            row += 1
            ws3.cell(row, 1, c['color'])
            ws3.cell(row, 2, c['count'])
            ws3.cell(row, 3, c['pct'])
            bar = '█' * int(c['pct'] / 3) + '░' * max(0, 20 - int(c['pct'] / 3))
            ws3.cell(row, 4, bar)
            apply_data_style(ws3, row, 4)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 8
    ws3.column_dimensions['C'].width = 8
    ws3.column_dimensions['D'].width = 25

    # ═══ 시트4: 가격대별 분석 ═══
    ws4 = wb.create_sheet('가격대별_분석')
    log("  -> 시트 [가격대별_분석] 생성")

    row = 1
    ws4.merge_cells(f'A{row}:F{row}')
    ws4.cell(row, 1, '가격대별 랭킹 분포')
    ws4.cell(row, 1).font = Font(bold=True, size=12, color='1A1A1A')

    for category in CATEGORIES:
        if category not in price_band:
            continue

        row += 2
        ws4.merge_cells(f'A{row}:F{row}')
        ws4.cell(row, 1, f'■ {category}')
        ws4.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        for col, h in enumerate(['가격대', '상품수', '비중(%)', '구성바', '대표상품'], 1):
            ws4.cell(row, col, h)
        apply_header_style(ws4, row, 5)

        for label, count, pct, items in price_band[category]['bands']:
            row += 1
            ws4.cell(row, 1, label)
            ws4.cell(row, 2, count)
            ws4.cell(row, 3, pct)
            bar = '█' * int(pct / 5) + '░' * max(0, 20 - int(pct / 5))
            ws4.cell(row, 4, bar)
            ws4.cell(row, 5, ', '.join(items) if items else '')
            apply_data_style(ws4, row, 5)

    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 8
    ws4.column_dimensions['C'].width = 8
    ws4.column_dimensions['D'].width = 25
    ws4.column_dimensions['E'].width = 50

    # ═══ 시트5: 랭킹 변동 현황 ═══
    ws5 = wb.create_sheet('랭킹_변동현황')
    log("  -> 시트 [랭킹_변동현황] 생성")

    row = 1
    ws5.merge_cells(f'A{row}:F{row}')
    ws5.cell(row, 1, '랭킹 변동 현황')
    ws5.cell(row, 1).font = Font(bold=True, size=12, color='1A1A1A')

    for category in CATEGORIES:
        if category not in ranking_changes:
            continue

        rc = ranking_changes[category]
        row += 2
        ws5.merge_cells(f'A{row}:F{row}')

        if rc['status'] == 'first_data':
            ws5.cell(row, 1, f'■ {category} - {rc["message"]}')
            ws5.cell(row, 1).font = Font(bold=True, size=11, color='808080')

            row += 1
            ws5.cell(row, 1, '※ 2회 이상 수집 시 변동 비교가 가능합니다.')
            ws5.cell(row, 1).font = Font(italic=True, color='666666')

            row += 1
            headers = ['현재순위', '상품명', '아이템타입', '가격', '변동', '상태']
            for col, h in enumerate(headers, 1):
                ws5.cell(row, col, h)
            apply_header_style(ws5, row, len(headers))

            for item in rc.get('changes', []):
                row += 1
                ws5.cell(row, 1, item['current_rank'])
                ws5.cell(row, 2, item['name'])
                ws5.cell(row, 3, item['item_type'])
                ws5.cell(row, 4, item['price'])
                ws5.cell(row, 5, '-')
                ws5.cell(row, 6, '초회')
                apply_data_style(ws5, row, len(headers))
            continue

        ws5.cell(row, 1, f'■ {category} ({rc["prev_date"]} → {rc["current_date"]}, 총 {rc["total_dates"]}회 수집)')
        ws5.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        ups = sum(1 for c in rc['changes'] if c['status'] == '상승')
        downs = sum(1 for c in rc['changes'] if c['status'] == '하락')
        news = sum(1 for c in rc['changes'] if c['status'] == '신규진입')
        sames = sum(1 for c in rc['changes'] if c['status'] == '유지')
        drops = len(rc.get('dropped', []))

        summary = f'상승 {ups}개 | 하락 {downs}개 | 유지 {sames}개 | 신규진입 {news}개 | 이탈 {drops}개'
        ws5.cell(row, 1, summary)
        ws5.cell(row, 1).font = Font(bold=True, size=10)

        row += 1
        headers = ['현재순위', '상품명', '아이템타입', '가격', '변동', '상태']
        for col, h in enumerate(headers, 1):
            ws5.cell(row, col, h)
        apply_header_style(ws5, row, len(headers))

        for item in rc['changes']:
            row += 1
            ws5.cell(row, 1, item['current_rank'])
            ws5.cell(row, 2, item['name'])
            ws5.cell(row, 3, item['item_type'])
            ws5.cell(row, 4, item['price'])

            change = item['rank_change']
            if item['status'] == '신규진입':
                ws5.cell(row, 5, 'NEW')
                ws5.cell(row, 5).font = NEW_FONT
                ws5.cell(row, 6, '신규')
                ws5.cell(row, 6).font = NEW_FONT
            elif change > 0:
                ws5.cell(row, 5, f'▲{change}')
                ws5.cell(row, 5).font = UP_FONT
                ws5.cell(row, 6, '상승')
                ws5.cell(row, 6).font = UP_FONT
            elif change < 0:
                ws5.cell(row, 5, f'▼{abs(change)}')
                ws5.cell(row, 5).font = DOWN_FONT
                ws5.cell(row, 6, '하락')
                ws5.cell(row, 6).font = DOWN_FONT
            else:
                ws5.cell(row, 5, '-')
                ws5.cell(row, 6, '유지')

            apply_data_style(ws5, row, len(headers))

        if rc.get('dropped'):
            row += 1
            ws5.merge_cells(f'A{row}:F{row}')
            ws5.cell(row, 1, '  ▼ 이탈 상품')
            ws5.cell(row, 1).font = Font(bold=True, size=10, color='808080')

            for item in rc['dropped']:
                row += 1
                ws5.cell(row, 1, '-')
                ws5.cell(row, 2, item['name'])
                ws5.cell(row, 3, item['item_type'])
                ws5.cell(row, 4, '')
                ws5.cell(row, 5, f'이전: {item["prev_rank"]}위')
                ws5.cell(row, 6, '이탈')
                ws5.cell(row, 6).font = Font(color='808080')
                apply_data_style(ws5, row, 6)

    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 40
    ws5.column_dimensions['C'].width = 16
    ws5.column_dimensions['D'].width = 14
    ws5.column_dimensions['E'].width = 10
    ws5.column_dimensions['F'].width = 10

    # ═══ 시트6: 아이템비중 차트 ═══
    ws6 = wb.create_sheet('아이템비중_차트')
    log("  -> 시트 [아이템비중_차트] 생성")

    col_offset = 0
    for category in CATEGORIES:
        if category not in type_dist:
            continue

        start_col = col_offset * 4 + 1
        ws6.cell(1, start_col, f'{category} 아이템타입 비중')
        ws6.cell(1, start_col).font = Font(bold=True, size=11)

        ws6.cell(2, start_col, '아이템타입')
        ws6.cell(2, start_col + 1, '상품수')
        ws6.cell(2, start_col + 2, '비중(%)')

        for r, item in enumerate(type_dist[category]['analysis'], 3):
            ws6.cell(r, start_col, item['item_type'])
            ws6.cell(r, start_col + 1, item['count'])
            ws6.cell(r, start_col + 2, item['pct'])

        data_count = len(type_dist[category]['analysis'])
        if data_count > 0:
            pie = PieChart()
            pie.title = f'{category} 아이템타입 비중'
            pie.style = 26
            pie.width = 18
            pie.height = 13

            data_ref = Reference(ws6, min_col=start_col + 1, min_row=2, max_row=2 + data_count)
            cats_ref = Reference(ws6, min_col=start_col, min_row=3, max_row=2 + data_count)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)

            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showVal = False
            pie.dataLabels.showCatName = True

            chart_row = data_count + 5
            ws6.add_chart(pie, f'{get_column_letter(start_col)}{chart_row}')

        col_offset += 1

    # ═══ 시트7: 히스토리 원본 데이터 ═══
    ws7 = wb.create_sheet('수집_히스토리')
    log("  -> 시트 [수집_히스토리] 생성")

    history = load_history()
    row = 1
    ws7.cell(row, 1, '날짜별 수집 히스토리 (랭킹 추적용)')
    ws7.cell(row, 1).font = Font(bold=True, size=12, color='1A1A1A')

    row += 1
    ws7.cell(row, 1, f'총 수집 횟수: {len(set(d for cat in history.values() for d in cat))}회')
    ws7.cell(row, 1).font = Font(italic=True, color='666666')

    for category in CATEGORIES:
        if category not in history:
            continue

        dates = sorted(history[category].keys())
        row += 2
        ws7.merge_cells(f'A{row}:{get_column_letter(2 + len(dates))}{row}')
        ws7.cell(row, 1, f'■ {category} 랭킹 추이')
        ws7.cell(row, 1).font = Font(bold=True, size=11, color='C00000')

        row += 1
        ws7.cell(row, 1, '상품명')
        ws7.cell(row, 2, '아이템타입')
        for di, d in enumerate(dates):
            ws7.cell(row, 3 + di, f'{d[4:6]}/{d[6:8]}')
        apply_header_style(ws7, row, 2 + len(dates))

        all_products_set = {}
        for d in dates:
            for name, info in history[category][d].items():
                if name not in all_products_set:
                    all_products_set[name] = info['item_type']

        for name in sorted(all_products_set.keys()):
            row += 1
            ws7.cell(row, 1, name[:35])
            ws7.cell(row, 2, all_products_set[name])
            for di, d in enumerate(dates):
                if name in history[category][d]:
                    ws7.cell(row, 3 + di, history[category][d][name]['rank'])
                else:
                    ws7.cell(row, 3 + di, '-')
            apply_data_style(ws7, row, 2 + len(dates))

    ws7.column_dimensions['A'].width = 40
    ws7.column_dimensions['B'].width = 16

    # ═══ 저장 ═══
    filename = f'아르켓_랭킹분석_{date_key}.xlsx'
    filepath = os.path.join(WORK_DIR, filename)

    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime('%H%M%S')
        filename = f'아르켓_랭킹분석_{date_key}_{ts}.xlsx'
        filepath = os.path.join(WORK_DIR, filename)
        wb.save(filepath)

    log(f"\n  [OK] 저장 완료: {filename}")
    return filepath


# ─── 콘솔 요약 출력 ───

def print_summary(type_dist, top_items, price_band):
    """콘솔에 핵심 요약 출력"""
    log("\n" + "=" * 60)
    log("  ARKET 랭킹 분석 요약")
    log("=" * 60)

    for category in CATEGORIES:
        if category not in type_dist:
            continue

        log(f"\n  ■ {category} ({type_dist[category]['total']}개)")
        log(f"  {'─' * 50}")

        log("  [아이템타입 비중 TOP5]")
        for item in type_dist[category]['analysis'][:5]:
            bar = '█' * int(item['pct'] / 3)
            log(f"    {item['item_type']:16s} {item['count']:2d}개 ({item['pct']:4.1f}%) {bar}")

        if category in top_items:
            log(f"\n  [핵심 아이템 TOP5]")
            for idx, item in enumerate(top_items[category][:5], 1):
                log(f"    {idx}. {item['name'][:30]:30s} [{item['item_type']}] {item['price']}")

        if category in price_band:
            log(f"\n  [주력 가격대]")
            top_bands = sorted(price_band[category]['bands'], key=lambda x: x[1], reverse=True)[:3]
            for label, count, pct, _ in top_bands:
                log(f"    {label:12s} {count:2d}개 ({pct:.1f}%)")


# ─── 메인 ───

def main():
    log("=" * 60)
    log("  ARKET 랭킹 분석 도구")
    log("=" * 60)

    log("\n[1/5] 크롤링 데이터 로드")
    filepath = find_latest_ranking_file()
    if not filepath:
        return

    all_data = load_ranking_data(filepath)
    total_products = sum(len(v) for v in all_data.values())
    log(f"  총 {len(all_data)}개 시트, {total_products}개 상품 로드 완료")

    date_key = extract_date_from_filename(filepath)

    log(f"\n[2/5] 분석 실행 (날짜: {date_key})")
    type_dist = analyze_item_type_distribution(all_data)
    top_items = analyze_top_items(all_data)
    price_band = analyze_price_band(all_data)
    color_dist = analyze_color_distribution(all_data)

    log(f"\n[3/5] 히스토리 업데이트")
    history = update_history(all_data, date_key)
    ranking_changes = analyze_ranking_changes(history)

    log(f"\n[4/5] 분석 결과 요약")
    print_summary(type_dist, top_items, price_band)

    log(f"\n[5/5] 엑셀 출력")
    output_path = create_analysis_excel(type_dist, top_items, price_band, color_dist, ranking_changes, date_key)

    log("\n" + "=" * 60)
    log("  분석 완료!")
    log(f"  파일: {os.path.basename(output_path)}")
    log("=" * 60)


if __name__ == '__main__':
    main()
