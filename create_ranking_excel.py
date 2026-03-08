# -*- coding: utf-8 -*-
"""
유니클로 여성 랭킹 - 확인된 데이터 기반 엑셀 생성
fetch_webpage로 확인한 실제 랭킹 데이터 사용
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
from PIL import Image as PILImage
import time
from datetime import datetime
import os

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Accept': 'image/webp,image/*,*/*',
}

# fetch_webpage로 확인한 실제 랭킹 데이터
RANKING_DATA = [
    {
        'rank': 1,
        'name': 'UNISEX, XS-4XL 스웨트와이드팬츠(다리길이66~72cm)',
        'price': '49,900원',
        'colors': '여러 컬러',
        'product_code': '471809',
        'color_code': '65'
    },
    {
        'rank': 2,
        'name': 'WOMEN, 22-29 배기커브진(다리길이75cm)',
        'price': '39,900원 기간한정가격',
        'colors': 'WHITE, DARK GRAY, BLUE, NAVY',
        'product_code': '479000',
        'color_code': '00'
    },
    {
        'rank': 3,
        'name': 'WOMEN, XS-3XL 저지배럴레그팬츠(다리길이70~72Cm)',
        'price': '49,900원',
        'colors': 'DARK BROWN, BLACK, NAVY 등',
        'product_code': '475344',
        'color_code': '39'
    },
    {
        'rank': 4,
        'name': 'UNISEX, XS-3XL 스무드코튼크루넥스웨터',
        'price': '39,900원 기간한정가격',
        'colors': 'OFF WHITE, GRAY, BLACK, PINK, RED 등',
        'product_code': '475053',
        'color_code': '01'
    },
    {
        'rank': 5,
        'name': 'WOMEN, XS-3XL 립헨리넥T(긴팔)',
        'price': '29,900원',
        'colors': 'NATURAL, OFF WHITE, LIGHT GRAY, BLACK, RED 등',
        'product_code': '487908',
        'color_code': '30'
    },
]

def download_image(product_code, color_code):
    """유니클로 이미지 다운로드"""
    # 유니클로 이미지 URL 패턴
    url = f'https://image.uniqlo.com/UQ/ST3/AsianCommon/imagesgoods/{product_code}/item/goods_{color_code}_{product_code}.jpg'
    
    try:
        print(f"  이미지 다운로드: {product_code}")
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        
        if img.mode == 'RGBA':
            bg = PILImage.new('RGB', img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        
        img.thumbnail((120, 120), PILImage.Resampling.LANCZOS)
        
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        return img_bytes
    except Exception as e:
        print(f"    [WARNING] 이미지 다운로드 실패: {str(e)[:50]}")
        return None

def create_excel_with_data(products, filename):
    """엑셀 파일 생성"""
    print(f"\n엑셀 파일 생성: {filename}")
    print(f"상품 수: {len(products)}개")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "유니클로 여성 랭킹"
    
    # 헤더
    headers = ['순위', '상품명', '가격', '컬러 옵션', '상품 이미지']
    ws.append(headers)
    
    # 헤더 스타일 (유니클로 레드)
    header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    header_font = Font(bold=True, size=13, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    
    # 헤더 행 높이
    ws.row_dimensions[1].height = 35
    
    # 데이터 입력
    for product in products:
        row_idx = ws.max_row + 1
        
        # 데이터 입력
        ws.cell(row_idx, 1, product['rank'])
        ws.cell(row_idx, 2, product['name'])
        ws.cell(row_idx, 3, product['price'])
        ws.cell(row_idx, 4, product['colors'])
        
        # 행 높이 설정
        ws.row_dimensions[row_idx].height = 95
        
        # 스타일 적용
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(name='맑은 고딕', size=11)
        
        # 순위 셀 강조
        ws.cell(row_idx, 1).font = Font(bold=True, size=14, name='맑은 고딕')
        ws.cell(row_idx, 1).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # 이미지 다운로드 및 삽입
        if 'product_code' in product and 'color_code' in product:
            img_data = download_image(product['product_code'], product['color_code'])
            if img_data:
                try:
                    img = XLImage(img_data)
                    img.width = 90
                    img.height = 90
                    ws.add_image(img, f'E{row_idx}')
                    print(f"  [{product['rank']}위] 이미지 삽입 완료")
                except Exception as e:
                    print(f"  [{product['rank']}위] 이미지 삽입 실패: {e}")
        
        ws.cell(row_idx, 5).border = border
    
    # 메타 정보 추가
    meta_row = ws.max_row + 2
    ws.cell(meta_row, 1, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}")
    ws.cell(meta_row + 1, 1, "출처: 유니클로 공식 온라인스토어 (www.uniqlo.com/kr)")
    ws.cell(meta_row + 2, 1, "주의: 개인 사용 목적으로만 사용하세요. 상업적 사용 금지")
    ws.cell(meta_row + 3, 1, "법적 리스크 최소화를 위해 공개된 정보만 수집하였습니다.")
    
    for row in range(meta_row, meta_row + 4):
        ws.cell(row, 1).font = Font(size=9, italic=True, color="666666", name='맑은 고딕')
        ws.merge_cells(f'A{row}:E{row}')
    
    # 파일 저장
    try:
        wb.save(filename)
        print(f"\n[완료] 엑셀 파일 저장 완료!")
        return True
    except Exception as e:
        print(f"\n[오류] 파일 저장 실패: {e}")
        return False

def main():
    print("=" * 70)
    print("유니클로 여성 랭킹 TOP 5 - 엑셀 생성")
    print("=" * 70)
    print("\n개인 사용 목적 / 법적 리스크 최소화")
    print("출처: 유니클로 공식 홈페이지에서 확인한 실제 랭킹")
    print()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"유니클로_여성랭킹_TOP5_{timestamp}.xlsx"
    
    if create_excel_with_data(RANKING_DATA, filename):
        print("\n" + "=" * 70)
        print("[SUCCESS] 작업 완료!")
        print("=" * 70)
        print(f"파일 위치: {os.path.abspath(filename)}")
        print(f"수록 상품: {len(RANKING_DATA)}개 (TOP 5)")
        print("\n엑셀 파일에 포함된 정보:")
        print("  - 순위 (1-5위)")
        print("  - 상품명 (사이즈, 특징 포함)")
        print("  - 가격 (기간한정가격 표시)")
        print("  - 컬러 옵션")
        print("  - 상품 이미지 (실제 다운로드)")
        print("=" * 70)
    else:
        print("\n [ERROR] 파일 생성 실패")

if __name__ == "__main__":
    main()
