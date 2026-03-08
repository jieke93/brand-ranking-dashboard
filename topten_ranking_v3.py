# -*- coding: utf-8 -*-
"""
탑텐 랭킹 크롤러 V3 (스크린샷 캡처 버전)
- 이미지를 네트워크 다운로드 대신 스크린샷 캡처 방식으로 수집
- 유니클로 V5와 동일한 방식
"""
import sys
import io
import urllib.parse
import urllib.request
import urllib.robotparser
from PIL import Image as PILImage

# 이미지 설정
IMG_WIDTH = 80
IMG_HEIGHT = 107
HD_IMG_WIDTH = 200   # 대시보드용 고해상도 이미지
HD_IMG_HEIGHT = 267  # 3:4 비율
ROW_HEIGHT = 85

# 로그 파일 설정
LOG_FILE = "topten_crawler_v3_log.txt"
SAFE_MODE = True  # 법적 위험 최소화 모드 (robots.txt 준수)
REQUEST_DELAY = 1.5  # 요청 간격 (초)
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

def log(msg, end='\n'):
    """로그를 파일과 콘솔에 출력"""
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + end)
    try:
        print(msg, end=end, flush=True)
    except:
        pass

def check_robots_allowed(url):
    """robots.txt 기준으로 접근 가능 여부 확인"""
    if not SAFE_MODE:
        return True
    try:
        parsed = urllib.parse.urlparse(url)
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        rp = urllib.robotparser.RobotFileParser()
        with urllib.request.urlopen(robots_url, timeout=5) as resp:
            content = resp.read().decode('utf-8', 'ignore')
        rp.parse(content.splitlines())
        return rp.can_fetch(USER_AGENT, url)
    except Exception:
        return True

def safe_get(driver, url):
    """robots.txt 준수 + 요청 간격 적용 후 페이지 로드"""
    if SAFE_MODE and not check_robots_allowed(url):
        log(f"  -> robots.txt 제한으로 접근 건너뜀: {url}")
        return False
    time.sleep(REQUEST_DELAY)
    driver.get(url)
    return True

# 로그 파일 초기화
with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write("")

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import time
import os
from datetime import datetime
import re

# 카테고리 코드 매핑
CATEGORIES = {
    '전체': 'ALL',
    '여성': 'SSMA41',
    '남성': 'SSMA42',
    '키즈': 'SSMA43',
    '베이비': 'SSMA46'
}

def capture_image_from_element(element):
    """요소를 스크린샷 캡처하여 이미지로 반환
    반환: (excel_img_bytes, hd_img_bytes) 튜플 또는 None"""
    try:
        png_data = element.screenshot_as_png
        if not png_data:
            return None
        
        img = PILImage.open(io.BytesIO(png_data))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        hd_img = img.resize((HD_IMG_WIDTH, HD_IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        hd_bytes = io.BytesIO()
        hd_img.save(hd_bytes, format='JPEG', quality=92)
        hd_bytes.seek(0)
        
        xl_img = img.resize((IMG_WIDTH, IMG_HEIGHT), PILImage.Resampling.LANCZOS)
        xl_bytes = io.BytesIO()
        xl_img.save(xl_bytes, format='JPEG', quality=85)
        xl_bytes.seek(0)
        
        return (xl_bytes, hd_bytes)
    except Exception as e:
        return None

def close_cookie_popup(driver):
    """쿠키 동의 팝업 닫기"""
    try:
        # 일반적인 쿠키 팝업 닫기 버튼 시도
        cookie_selectors = [
            'button#onetrust-accept-btn-handler',
            'button[id*="accept"]',
            'button[class*="cookie-accept"]',
            '.cookie-agree',
            '.btn-cookie-accept'
        ]
        
        for selector in cookie_selectors:
            try:
                btn = driver.find_element(By.CSS_SELECTOR, selector)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(0.5)
                    return True
            except:
                continue
        
        # JavaScript로 쿠키 배너 숨기기
        driver.execute_script("""
            var banners = document.querySelectorAll('[class*="cookie"], [id*="cookie"], [class*="consent"]');
            banners.forEach(function(b) { b.style.display = 'none'; });
        """)
        
    except Exception:
        pass
    return False

def setup_driver():
    log("=" * 60)
    log("[1/4] Chrome 드라이버 초기화")
    log("=" * 60)
    
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    # 빠른 로딩을 위한 eager 모드
    options.page_load_strategy = 'eager'
    
    log("  -> ChromeDriver 준비...")
    service = Service(ChromeDriverManager().install())
    log("  -> Chrome 시작...")
    driver = webdriver.Chrome(service=service, options=options)
    
    # 타임아웃 설정
    driver.set_page_load_timeout(60)
    driver.set_script_timeout(30)
    
    log("  [OK] 완료!\n")
    return driver

def click_category(driver, category_code):
    """카테고리 변경 - JavaScript 함수 호출"""
    try:
        driver.execute_script(f"rankingForm.rankingAll('{category_code}');")
        time.sleep(2)
        return True
    except Exception as e:
        log(f"      [오류] 카테고리 변경 실패: {e}")
        return False

def extract_products(driver, max_products=30):
    """상품 데이터 추출 - 스크린샷 캡처 방식"""
    products = []
    
    # 스크롤하여 상품 로드
    log("      [DEBUG] 스크롤 시작")
    for i in range(5):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.6)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)
    log("      [DEBUG] 스크롤 완료")
    
    # 상품 타일 찾기
    product_tiles = driver.find_elements(By.CSS_SELECTOR, ".catalog-item.st-tile")
    
    if not product_tiles:
        log("      [경고] 상품 타일을 찾을 수 없습니다")
        return []
    
    log(f"      [DEBUG] {len(product_tiles)}개 상품 발견")
    
    img_success = 0
    
    for idx, tile in enumerate(product_tiles[:max_products], 1):
        try:
            product = {
                'rank': idx,
                'brand': '',
                'name': '',
                'original_price': '',
                'sale_price': '',
                'discount_rate': '',
                'rating': '없음',
                'review_count': '없음',
                'image_data': None
            }
            
            # 순위 (tile-ranking에서)
            try:
                rank_elem = tile.find_element(By.CSS_SELECTOR, ".tile-ranking")
                product['rank'] = int(rank_elem.text.strip())
            except:
                pass
            
            # 브랜드
            try:
                brand_elem = tile.find_element(By.CSS_SELECTOR, ".tile-brand")
                product['brand'] = brand_elem.text.strip()
            except:
                pass
            
            # 상품명
            try:
                name_elem = tile.find_element(By.CSS_SELECTOR, ".tile-goods-label-inner")
                product['name'] = name_elem.text.strip()
            except:
                pass
            
            # 할인가
            try:
                sale_price_elem = tile.find_element(By.CSS_SELECTOR, ".tile-price.st-current")
                sale_text = sale_price_elem.text.strip().replace(',', '').replace('원', '')
                if sale_text.isdigit():
                    product['sale_price'] = f"{int(sale_text):,}원"
            except:
                pass
            
            # 정가
            try:
                original_price_elem = tile.find_element(By.CSS_SELECTOR, ".tile-price.st-was")
                original_text = original_price_elem.text.strip().replace(',', '').replace('원', '')
                if original_text.isdigit():
                    product['original_price'] = f"{int(original_text):,}원"
            except:
                pass
            
            # 할인율
            try:
                rate_elem = tile.find_element(By.CSS_SELECTOR, ".tile-rate")
                product['discount_rate'] = rate_elem.text.strip()
            except:
                pass
            
            # 평점 & 리뷰
            try:
                affinity = tile.find_element(By.CSS_SELECTOR, ".tile-affinity")
                affinity_texts = affinity.find_elements(By.CSS_SELECTOR, ".tile-text")
                for text_elem in affinity_texts:
                    text = text_elem.text.strip()
                    rating_match = re.search(r'(\d+\.?\d*)\s*\((\d+)\)', text)
                    if rating_match:
                        product['rating'] = rating_match.group(1)
                        product['review_count'] = rating_match.group(2)
                        break
            except:
                pass
            
            # 이미지 캡처 (스크린샷 방식)
            try:
                # 이미지 요소로 스크롤
                img_elem = tile.find_element(By.CSS_SELECTOR, ".tile-img")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", img_elem)
                time.sleep(0.1)
                
                # 스크린샷 캡처
                img_data = capture_image_from_element(img_elem)
                if img_data:
                    product['image_data'] = img_data[0]      # 엑셀용
                    product['hd_image_data'] = img_data[1]   # 대시보드용 고해상도
                    img_success += 1
            except:
                pass
            
            if product['name']:
                products.append(product)
                
        except Exception as e:
            continue
    
    log(f"      -> {len(products)}개 수집 (이미지: {img_success}개)")
    
    # 샘플 출력
    if len(products) >= 2:
        for i, p in enumerate(products[:2], 1):
            name_short = p['name'][:15] if len(p['name']) > 15 else p['name']
            has_img = "O" if p['image_data'] else "X"
            log(f"        {i}. {name_short:15s} | {p['sale_price']:10s} | img:{has_img}")
    
    return products

def scrape_all_categories(driver):
    """전체 카테고리 크롤링"""
    all_data = {}
    url = "https://topten10.goodwearmall.com/display/ranking"
    
    log(f"\n{'='*60}")
    log(f"[2/4] 데이터 수집 시작")
    log(f"{'='*60}")
    log(f"  URL: {url}")
    log(f"  카테고리: {', '.join(CATEGORIES.keys())}")
    
    if not safe_get(driver, url):
        return {}
    log(f"  -> 페이지 로딩 대기...", end='')
    time.sleep(5)
    log(" OK!")
    
    # 쿠키 팝업 닫기
    if close_cookie_popup(driver):
        log("  -> 쿠키 팝업 닫음")
    
    category_list = list(CATEGORIES.items())
    
    for cat_idx, (category_name, category_code) in enumerate(category_list, 1):
        log(f"\n  [{cat_idx}/{len(CATEGORIES)}] 카테고리: {category_name}")
        
        # 첫 번째(전체)가 아니면 카테고리 변경
        if cat_idx > 1:
            log(f"      -> '{category_name}' 탭 전환 중...", end='')
            if click_category(driver, category_code):
                log(" OK!")
            else:
                log(" 실패!")
                continue
        
        # 상품 추출
        products = extract_products(driver, max_products=30)
        
        if products:
            all_data[category_name] = products
        else:
            log(f"      -> 상품 없음")
        
        time.sleep(1)
    
    return all_data

def create_excel(all_data, filename):
    """이미지가 포함된 엑셀 생성"""
    log(f"\n{'='*60}")
    log(f"[3/4] 엑셀 파일 생성 (이미지 포함)")
    log(f"{'='*60}")
    log(f"  파일명: {filename}")
    
    wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 탑텐 브랜드 컬러
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name='맑은 고딕')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['순위', '이미지', '브랜드', '상품명', '할인가', '정가', '할인율', '평점', '리뷰수']
    col_widths = [6, 12, 15, 45, 15, 15, 10, 8, 10]
    
    for sheet_name, products in all_data.items():
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_name)
        
        log(f"  -> 시트 [{safe_name}]: {len(products)}개 상품 (이미지 삽입 중...)")
        
        # 열 너비 설정
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # 헤더 작성
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 25
        
        # 데이터 작성
        for row_idx, product in enumerate(products, 2):
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
            
            # 순위
            cell = ws.cell(row=row_idx, column=1, value=product['rank'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # 이미지 삽입
            if product['image_data']:
                try:
                    img = XLImage(product['image_data'])
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    ws.add_image(img, f"B{row_idx}")
                except:
                    pass
            
            cell = ws.cell(row=row_idx, column=2, value='')
            cell.border = border
            
            # 브랜드
            cell = ws.cell(row=row_idx, column=3, value=product['brand'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # 상품명
            cell = ws.cell(row=row_idx, column=4, value=product['name'])
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
            
            # 할인가
            cell = ws.cell(row=row_idx, column=5, value=product['sale_price'])
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.font = Font(bold=True, color="FF0000")
            cell.border = border
            
            # 정가
            cell = ws.cell(row=row_idx, column=6, value=product['original_price'])
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            # 할인율
            cell = ws.cell(row=row_idx, column=7, value=product['discount_rate'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(color="FF6600")
            cell.border = border
            
            # 평점
            cell = ws.cell(row=row_idx, column=8, value=product['rating'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # 리뷰수
            cell = ws.cell(row=row_idx, column=9, value=product['review_count'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    wb.save(filename)
    log(f"  [OK] 저장 완료!")
    return True

def main():
    log("\n" + "=" * 60)
    log("  탑텐 주간베스트 랭킹 크롤러 V3 (스크린샷 캡처 버전)")
    log("=" * 60)
    log("  * 이미지를 스크린샷 캡처 방식으로 수집합니다")
    log("=" * 60)
    
    driver = None
    try:
        driver = setup_driver()
        all_data = scrape_all_categories(driver)
        
        if all_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"탑텐_주간베스트_이미지포함_V3_{timestamp}.xlsx"
            
            create_excel(all_data, filename)
            
            # 대시보드용 고해상도 이미지 저장
            _save_hd_images(all_data, '탑텐', filename)
            
            # 통계 출력
            log(f"\n{'='*60}")
            log(f"[4/4] 수집 완료 통계")
            log(f"{'='*60}")
            
            total_products = sum(len(products) for products in all_data.values())
            total_images = sum(1 for products in all_data.values() for p in products if p['image_data'])
            total_prices = sum(1 for products in all_data.values() for p in products if p['sale_price'])
            total_ratings = sum(1 for products in all_data.values() for p in products if p['rating'] != '없음')
            
            log(f"  총 카테고리: {len(all_data)}개")
            log(f"  총 상품: {total_products}개")
            log(f"  가격: {total_prices}/{total_products}개 수집됨")
            log(f"  이미지 삽입: {total_images}/{total_products}개")
            log(f"  평점: {total_ratings}/{total_products}개 수집됨")
            log(f"\n  파일: {filename}")
            log("=" * 60)
        else:
            log("\n[오류] 수집된 데이터가 없습니다.")
            
    except Exception as e:
        log(f"\n[오류] {e}")
        import traceback
        log(traceback.format_exc())
    finally:
        if driver:
            log("\n브라우저 종료")
            driver.quit()


def _save_hd_images(all_data, brand_name, excel_filename):
    """대시보드용 고해상도 이미지를 product_images_hd/ 폴더에 저장"""
    import hashlib
    hd_dir = os.path.join(os.path.dirname(os.path.abspath(excel_filename)), 'product_images_hd')
    os.makedirs(hd_dir, exist_ok=True)
    
    file_hash = hashlib.md5(f"{os.path.basename(excel_filename)}_{os.path.getmtime(excel_filename)}".encode()).hexdigest()[:8]
    saved = 0
    for sheet_name, products in all_data.items():
        for p in products:
            hd_data = p.get('hd_image_data')
            if hd_data:
                name = p.get('name', '')[:20]
                safe = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')
                fname = f"{file_hash}_{brand_name}_{sheet_name}_{p['rank']}_{safe}.jpg"
                fpath = os.path.join(hd_dir, fname)
                if not os.path.exists(fpath):
                    try:
                        hd_data.seek(0)
                        with open(fpath, 'wb') as f:
                            f.write(hd_data.read())
                        saved += 1
                    except:
                        pass
    if saved > 0:
        log(f"  [HD] 고해상도 이미지 {saved}개 저장 → {hd_dir}")

if __name__ == "__main__":
    main()
